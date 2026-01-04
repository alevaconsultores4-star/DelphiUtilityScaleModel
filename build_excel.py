"""
Build Excel financial model with formulas for renewable project.
Generates excel_template.xlsx with 8 tabs and all formulas.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from datetime import date, timedelta
import calendar

# Colors
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
CHECK_PASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CHECK_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_workbook():
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create all sheets
    ws_inputs = wb.create_sheet("Inputs")
    ws_timeline = wb.create_sheet("Timeline_M")
    ws_capex = wb.create_sheet("CAPEX_M")
    ws_revenue = wb.create_sheet("Revenue_M")
    ws_opex = wb.create_sheet("OPEX_M")
    ws_debt = wb.create_sheet("Debt_M")
    ws_cashflow = wb.create_sheet("Cashflow_M")
    ws_outputs = wb.create_sheet("Outputs")
    
    return wb, {
        "Inputs": ws_inputs,
        "Timeline_M": ws_timeline,
        "CAPEX_M": ws_capex,
        "Revenue_M": ws_revenue,
        "OPEX_M": ws_opex,
        "Debt_M": ws_debt,
        "Cashflow_M": ws_cashflow,
        "Outputs": ws_outputs
    }

def setup_inputs_tab(ws):
    """Create Inputs tab with all assumptions."""
    row = 1
    
    # Title
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "RENEWABLE PROJECT FINANCIAL MODEL - INPUTS"
    ws[f"A{row}"].font = Font(bold=True, size=14)
    row += 2
    
    # Timeline
    ws[f"A{row}"] = "TIMELINE"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "Start Date"
    ws[f"B{row}"] = date(2024, 1, 1)
    ws[f"B{row}"].fill = INPUT_FILL
    ws[f"B{row}"].number_format = "mm/dd/yyyy"
    row += 1
    ws[f"A{row}"] = "Development Months"
    ws[f"B{row}"] = 18
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Construction Months"
    ws[f"B{row}"] = 12
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Operation Years"
    ws[f"B{row}"] = 25
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # Generation
    ws[f"A{row}"] = "GENERATION"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "MWac"
    ws[f"B{row}"] = 100
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "MWp"
    ws[f"B{row}"] = 130
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "P50 MWh/Year"
    ws[f"B{row}"] = 220000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Degradation %/Year"
    ws[f"B{row}"] = 0.5
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # Revenue
    ws[f"A{row}"] = "REVENUE"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "PPA Price (COP/kWh)"
    ws[f"B{row}"] = 320
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "PPA Term (Years)"
    ws[f"B{row}"] = 12
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Merchant Price (COP/kWh)"
    ws[f"B{row}"] = 250
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Indexation"
    ws[f"B{row}"] = "Colombia CPI"
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Colombia CPI %"
    ws[f"B{row}"] = 6.0
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # CAPEX
    ws[f"A{row}"] = "CAPEX"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "Development (COP)"
    ws[f"B{row}"] = 5000000000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Construction (COP)"
    ws[f"B{row}"] = 80000000000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "At COD (COP)"
    ws[f"B{row}"] = 5000000000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Distribution"
    ws[f"B{row}"] = "Straight-line"
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # OPEX
    ws[f"A{row}"] = "OPEX"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "Fixed OM (COP/MWac/Year)"
    ws[f"B{row}"] = 5000000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Variable OM (COP/MWh)"
    ws[f"B{row}"] = 5000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Insurance (COP/MWac/Year)"
    ws[f"B{row}"] = 2000000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Grid Fees (COP/MWh)"
    ws[f"B{row}"] = 10000
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "OPEX Indexation"
    ws[f"B{row}"] = "Colombia CPI"
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # Debt
    ws[f"A{row}"] = "DEBT"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "Debt Enabled"
    ws[f"B{row}"] = "Yes"
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Debt % of CAPEX"
    ws[f"B{row}"] = 70
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Tenor (Years)"
    ws[f"B{row}"] = 15
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "All-in Interest Rate %"
    ws[f"B{row}"] = 8.5
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Amortization Method"
    ws[f"B{row}"] = "Equal Principal"
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Upfront Fee (bps)"
    ws[f"B{row}"] = 100
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Commitment Fee % of Margin"
    ws[f"B{row}"] = 50
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # Tax
    ws[f"A{row}"] = "TAX"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    ws[f"A{row}"] = "Corporate Tax Rate %"
    ws[f"B{row}"] = 35
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Depreciation Years"
    ws[f"B{row}"] = 20
    ws[f"B{row}"].fill = INPUT_FILL
    row += 1
    ws[f"A{row}"] = "Depreciation % of CAPEX"
    ws[f"B{row}"] = 100
    ws[f"B{row}"].fill = INPUT_FILL
    row += 2
    
    # Create named ranges (modern API)
    wb = ws.parent
    named_ranges = {
        "StartDate": f"Inputs!$B$3",
        "DevMonths": f"Inputs!$B$4",
        "ConMonths": f"Inputs!$B$5",
        "OpYears": f"Inputs!$B$6",
        "MWac": f"Inputs!$B$9",
        "MWp": f"Inputs!$B$10",
        "P50MWh": f"Inputs!$B$11",
        "DegradPct": f"Inputs!$B$12",
        "PPAPrice": f"Inputs!$B$15",
        "PPATerm": f"Inputs!$B$16",
        "MerchantPrice": f"Inputs!$B$17",
        "Indexation": f"Inputs!$B$18",
        "ColCPI": f"Inputs!$B$19",
        "CapexDev": f"Inputs!$B$22",
        "CapexCon": f"Inputs!$B$23",
        "CapexCOD": f"Inputs!$B$24",
        "CapexDist": f"Inputs!$B$25",
        "FixedOM": f"Inputs!$B$28",
        "VarOM": f"Inputs!$B$29",
        "Insurance": f"Inputs!$B$30",
        "GridFees": f"Inputs!$B$31",
        "OpexIndex": f"Inputs!$B$32",
        "DebtEnabled": f"Inputs!$B$35",
        "DebtPct": f"Inputs!$B$36",
        "DebtTenor": f"Inputs!$B$37",
        "DebtRate": f"Inputs!$B$38",
        "DebtAmort": f"Inputs!$B$39",
        "UpfrontFee": f"Inputs!$B$40",
        "CommitFee": f"Inputs!$B$41",
        "TaxRate": f"Inputs!$B$44",
        "DepYears": f"Inputs!$B$45",
        "DepPct": f"Inputs!$B$46",
    }
    
    for name, ref in named_ranges.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
    
    # Format columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

def setup_timeline_tab(ws):
    """Create Timeline_M tab with monthly dates."""
    # Headers
    headers = ["Month", "Year", "Phase", "MonthNum"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Calculate total months
    # Formula: =DevMonths+ConMonths+OpYears*12
    ws["E1"] = "Total Months"
    ws["E2"] = "=DevMonths+ConMonths+OpYears*12"
    ws["E2"].fill = FORMULA_FILL
    
    # Month 1 formula
    ws["A2"] = "=StartDate"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["A2"].fill = FORMULA_FILL
    
    # Year formula
    ws["B2"] = "=YEAR(A2)"
    ws["B2"].fill = FORMULA_FILL
    
    # Phase helper (Development if month <= DevMonths, Construction if month <= DevMonths+ConMonths, else Operation)
    ws["C2"] = '=IF(MonthNum<=DevMonths,"Development",IF(MonthNum<=DevMonths+ConMonths,"Construction","Operation"))'
    ws["C2"].fill = FORMULA_FILL
    
    # MonthNum helper
    ws["D2"] = "=ROW()-1"
    ws["D2"].fill = FORMULA_FILL
    
    # Copy formulas down (assume max 500 months = ~41 years)
    max_rows = 500
    for row in range(3, max_rows + 1):
        # Month = previous month + 1 month
        ws[f"A{row}"] = f"=EDATE(A{row-1},1)"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"A{row}"].fill = FORMULA_FILL
        
        # Year
        ws[f"B{row}"] = f"=YEAR(A{row})"
        ws[f"B{row}"].fill = FORMULA_FILL
        
        # Phase
        ws[f"C{row}"] = f'=IF(D{row}<=DevMonths,"Development",IF(D{row}<=DevMonths+ConMonths,"Construction","Operation"))'
        ws[f"C{row}"].fill = FORMULA_FILL
        
        # MonthNum
        ws[f"D{row}"] = f"=ROW()-1"
        ws[f"D{row}"].fill = FORMULA_FILL
    
    # Format columns
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10

def setup_capex_tab(ws, timeline_ws):
    """Create CAPEX_M tab with monthly CAPEX schedule."""
    # Headers
    headers = ["Month", "Year", "Phase", "CAPEX_Dev", "CAPEX_Con", "CAPEX_COD", "CAPEX_Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Link to Timeline_M
    ws["A2"] = "=Timeline_M!A2"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["B2"] = "=Timeline_M!B2"
    ws["C2"] = "=Timeline_M!C2"
    
    # CAPEX formulas
    # Development: spread over DevMonths
    ws["D2"] = '=IF(C2="Development",CapexDev/DevMonths,0)'
    ws["D2"].fill = FORMULA_FILL
    
    # Construction: spread over ConMonths (only during Construction phase)
    ws["E2"] = '=IF(C2="Construction",CapexCon/ConMonths,0)'
    ws["E2"].fill = FORMULA_FILL
    
    # At COD: only in the month that is DevMonths+ConMonths+1
    ws["F2"] = '=IF(AND(C2="Operation",Timeline_M!D2=DevMonths+ConMonths+1),CapexCOD,0)'
    ws["F2"].fill = FORMULA_FILL
    
    # Total
    ws["G2"] = "=SUM(D2:F2)"
    ws["G2"].fill = FORMULA_FILL
    
    # Copy down
    max_rows = 500
    for row in range(3, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        ws[f"D{row}"] = f'=IF(C{row}="Development",CapexDev/DevMonths,0)'
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"E{row}"] = f'=IF(C{row}="Construction",CapexCon/ConMonths,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"F{row}"] = f'=IF(AND(C{row}="Operation",Timeline_M!D{row}=DevMonths+ConMonths+1),CapexCOD,0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"G{row}"] = f"=SUM(D{row}:F{row})"
        ws[f"G{row}"].fill = FORMULA_FILL
    
    # Format columns
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18
        for row in range(2, max_rows + 1):
            cell = ws[f"{col}{row}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = "#,##0"

def setup_revenue_tab(ws, timeline_ws):
    """Create Revenue_M tab with monthly revenue."""
    # Headers
    headers = ["Month", "Year", "Phase", "OperatingYear", "Energy_MWh", "Price_COP_kWh", "Revenue_COP"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Link to Timeline_M
    ws["A2"] = "=Timeline_M!A2"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["B2"] = "=Timeline_M!B2"
    ws["C2"] = "=Timeline_M!C2"
    
    # OperatingYear: 0 for pre-operation, 1+ for operation years
    # COD year = year of month (DevMonths+ConMonths+1)
    ws["D2"] = '=IF(C2="Operation",YEAR(A2)-YEAR(EDATE(StartDate,DevMonths+ConMonths))+1,0)'
    ws["D2"].fill = FORMULA_FILL
    
    # Energy: only during operation, with degradation
    # Base energy = P50MWh, degraded by (1-DegradPct/100)^(OperatingYear-1)
    # Monthly = annual / 12
    ws["E2"] = '=IF(C2="Operation",(P50MWh*(1-DegradPct/100)^(D2-1))/12,0)'
    ws["E2"].fill = FORMULA_FILL
    
    # Price: PPA price if OperatingYear <= PPATerm, else MerchantPrice
    # Apply indexation: base price * (1+ColCPI/100)^(OperatingYear-1)
    ws["F2"] = '=IF(C2="Operation",IF(D2<=PPATerm,PPAPrice,MerchantPrice)*(1+ColCPI/100)^(D2-1),0)'
    ws["F2"].fill = FORMULA_FILL
    
    # Revenue = Energy * Price * 1000 (MWh to kWh)
    ws["G2"] = "=E2*F2*1000"
    ws["G2"].fill = FORMULA_FILL
    
    # Copy down
    max_rows = 500
    for row in range(3, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        ws[f"D{row}"] = f'=IF(C{row}="Operation",YEAR(A{row})-YEAR(EDATE(StartDate,DevMonths+ConMonths))+1,0)'
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"E{row}"] = f'=IF(C{row}="Operation",(P50MWh*(1-DegradPct/100)^(D{row}-1))/12,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"F{row}"] = f'=IF(C{row}="Operation",IF(D{row}<=PPATerm,PPAPrice,MerchantPrice)*(1+ColCPI/100)^(D{row}-1),0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"G{row}"] = f"=E{row}*F{row}*1000"
        ws[f"G{row}"].fill = FORMULA_FILL
    
    # Format columns
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    for col in ["E", "F", "G"]:
        for row in range(2, max_rows + 1):
            cell = ws[f"{col}{row}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = "#,##0.00" if col == "F" else "#,##0"

def setup_opex_tab(ws, timeline_ws, revenue_ws):
    """Create OPEX_M tab with monthly OPEX."""
    # Headers
    headers = ["Month", "Year", "Phase", "OperatingYear", "FixedOM", "VarOM", "Insurance", "GridFees", "OPEX_Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Link to Timeline_M
    ws["A2"] = "=Timeline_M!A2"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["B2"] = "=Timeline_M!B2"
    ws["C2"] = "=Timeline_M!C2"
    ws["D2"] = "=Revenue_M!D2"
    
    # OPEX components (only during operation)
    # FixedOM: per MWac per year, monthly, indexed
    ws["E2"] = '=IF(C2="Operation",(FixedOM*MWac*(1+ColCPI/100)^(D2-1))/12,0)'
    ws["E2"].fill = FORMULA_FILL
    
    # VarOM: per MWh, indexed
    ws["F2"] = '=IF(C2="Operation",Revenue_M!E2*VarOM*(1+ColCPI/100)^(D2-1),0)'
    ws["F2"].fill = FORMULA_FILL
    
    # Insurance: per MWac per year, monthly, indexed
    ws["G2"] = '=IF(C2="Operation",(Insurance*MWac*(1+ColCPI/100)^(D2-1))/12,0)'
    ws["G2"].fill = FORMULA_FILL
    
    # GridFees: per MWh
    ws["H2"] = '=IF(C2="Operation",Revenue_M!E2*GridFees,0)'
    ws["H2"].fill = FORMULA_FILL
    
    # Total
    ws["I2"] = "=SUM(E2:H2)"
    ws["I2"].fill = FORMULA_FILL
    
    # Copy down
    max_rows = 500
    for row in range(3, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        ws[f"D{row}"] = f"=Revenue_M!D{row}"
        ws[f"E{row}"] = f'=IF(C{row}="Operation",(FixedOM*MWac*(1+ColCPI/100)^(D{row}-1))/12,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"F{row}"] = f'=IF(C{row}="Operation",Revenue_M!E{row}*VarOM*(1+ColCPI/100)^(D{row}-1),0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"G{row}"] = f'=IF(C{row}="Operation",(Insurance*MWac*(1+ColCPI/100)^(D{row}-1))/12,0)'
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"H{row}"] = f'=IF(C{row}="Operation",Revenue_M!E{row}*GridFees,0)'
        ws[f"H{row}"].fill = FORMULA_FILL
        ws[f"I{row}"] = f"=SUM(E{row}:H{row})"
        ws[f"I{row}"].fill = FORMULA_FILL
    
    # Format columns
    for col in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 15
        for row in range(2, max_rows + 1):
            cell = ws[f"{col}{row}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = "#,##0"

def setup_debt_tab(ws, timeline_ws, capex_ws):
    """Create Debt_M tab with debt draws and service."""
    # Headers
    headers = ["Month", "Year", "Phase", "DebtDraw", "Interest", "Principal", "DebtService", "DebtFees", "DebtBalance"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Link to Timeline_M
    ws["A2"] = "=Timeline_M!A2"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["B2"] = "=Timeline_M!B2"
    ws["C2"] = "=Timeline_M!C2"
    
    # Debt Draw: only during construction (before COD), = CAPEX * DebtPct
    # COD month = DevMonths+ConMonths+1
    ws["D2"] = '=IF(AND(C2<>"Operation",DebtEnabled="Yes"),CAPEX_M!G2*DebtPct/100,0)'
    ws["D2"].fill = FORMULA_FILL
    
    # Interest: only during operation, on outstanding balance
    # Monthly rate = DebtRate/100/12
    # Interest = Previous Balance * Monthly Rate
    ws["E2"] = "=0"
    ws["E2"].fill = FORMULA_FILL
    
    # Principal: only during operation, equal principal amortization
    # Total debt = sum of draws
    # Monthly principal = Total Debt / (DebtTenor * 12)
    ws["F2"] = "=0"
    ws["F2"].fill = FORMULA_FILL
    
    # Debt Service
    ws["G2"] = "=E2+F2"
    ws["G2"].fill = FORMULA_FILL
    
    # Debt Fees: upfront fee at COD, commitment fees during construction
    ws["H2"] = "=0"
    ws["H2"].fill = FORMULA_FILL
    
    # Debt Balance: cumulative draws - cumulative principal
    ws["I2"] = "=D2"
    ws["I2"].fill = FORMULA_FILL
    
    # Row 3 onwards: need cumulative calculations
    # Interest = Previous Balance * Monthly Rate (only during operation)
    ws["E3"] = '=IF(AND(C3="Operation",DebtEnabled="Yes"),I2*DebtRate/100/12,0)'
    ws["E3"].fill = FORMULA_FILL
    
    # Principal: equal monthly payment during operation
    # Total debt amount (sum of all draws)
    ws["J1"] = "TotalDebt"
    ws["J2"] = "=SUM(D:D)"
    ws["J2"].fill = FORMULA_FILL
    
    # COD month number (helper)
    ws["L1"] = "CODMonth"
    ws["L2"] = "=DevMonths+ConMonths+1"
    ws["L2"].fill = FORMULA_FILL
    
    # Monthly principal payment (helper)
    ws["K1"] = "MonthlyPrincipal"
    ws["K2"] = '=IF(DebtEnabled="Yes",TotalDebt/(DebtTenor*12),0)'
    ws["K2"].fill = FORMULA_FILL
    
    # Principal in row 3: only during operation, within tenor, and balance > 0
    ws["F3"] = '=IF(AND(C3="Operation",DebtEnabled="Yes",Timeline_M!D3>=L2,Timeline_M!D3<L2+DebtTenor*12,I2>0),MIN(K2,I2),0)'
    ws["F3"].fill = FORMULA_FILL
    
    # Debt Service
    ws["G3"] = "=E3+F3"
    ws["G3"].fill = FORMULA_FILL
    
    # Debt Fees: upfront at COD month
    ws["H3"] = '=IF(AND(C3="Operation",Timeline_M!D3=L2,DebtEnabled="Yes"),TotalDebt*UpfrontFee/10000,0)'
    ws["H3"].fill = FORMULA_FILL
    
    # Debt Balance: previous balance + draw - principal
    ws["I3"] = "=I2+D3-F3"
    ws["I3"].fill = FORMULA_FILL
    
    # Copy down
    max_rows = 500
    for row in range(4, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        ws[f"D{row}"] = f'=IF(AND(C{row}<>"Operation",DebtEnabled="Yes"),CAPEX_M!G{row}*DebtPct/100,0)'
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"E{row}"] = f'=IF(AND(C{row}="Operation",DebtEnabled="Yes"),I{row-1}*DebtRate/100/12,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"F{row}"] = f'=IF(AND(C{row}="Operation",DebtEnabled="Yes",Timeline_M!D{row}>=L2,Timeline_M!D{row}<L2+DebtTenor*12,I{row-1}>0),MIN(K2,I{row-1}),0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"G{row}"] = f"=E{row}+F{row}"
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"H{row}"] = f'=IF(AND(C{row}="Operation",Timeline_M!D{row}=L2,DebtEnabled="Yes"),TotalDebt*UpfrontFee/10000,0)'
        ws[f"H{row}"].fill = FORMULA_FILL
        ws[f"I{row}"] = f"=I{row-1}+D{row}-F{row}"
        ws[f"I{row}"].fill = FORMULA_FILL
    
    # Format columns
    for col in ["D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 15
        for row in range(2, max_rows + 1):
            cell = ws[f"{col}{row}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = "#,##0"

def setup_cashflow_tab(ws, timeline_ws, capex_ws, revenue_ws, opex_ws, debt_ws):
    """Create Cashflow_M tab with unlevered and equity cash flows."""
    # Headers
    headers = ["Month", "Year", "Phase", "Revenue", "OPEX", "CAPEX", "UnleveredCF", "DebtDraw", "DebtService", "DebtFees", "EquityCF"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Link to Timeline_M
    ws["A2"] = "=Timeline_M!A2"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["B2"] = "=Timeline_M!B2"
    ws["C2"] = "=Timeline_M!C2"
    
    # Revenue
    ws["D2"] = "=Revenue_M!G2"
    ws["D2"].fill = FORMULA_FILL
    
    # OPEX
    ws["E2"] = "=OPEX_M!I2"
    ws["E2"].fill = FORMULA_FILL
    
    # CAPEX
    ws["F2"] = "=CAPEX_M!G2"
    ws["F2"].fill = FORMULA_FILL
    
    # Unlevered CF = Revenue - OPEX - CAPEX
    ws["G2"] = "=D2-E2-F2"
    ws["G2"].fill = FORMULA_FILL
    
    # Debt Draw
    ws["H2"] = "=Debt_M!D2"
    ws["H2"].fill = FORMULA_FILL
    
    # Debt Service
    ws["I2"] = "=Debt_M!G2"
    ws["I2"].fill = FORMULA_FILL
    
    # Debt Fees
    ws["J2"] = "=Debt_M!H2"
    ws["J2"].fill = FORMULA_FILL
    
    # Equity CF = Unlevered CF + Debt Draw - Debt Service - Debt Fees
    ws["K2"] = "=G2+H2-I2-J2"
    ws["K2"].fill = FORMULA_FILL
    
    # Copy down
    max_rows = 500
    for row in range(3, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        ws[f"D{row}"] = f"=Revenue_M!G{row}"
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"E{row}"] = f"=OPEX_M!I{row}"
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"F{row}"] = f"=CAPEX_M!G{row}"
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"G{row}"] = f"=D{row}-E{row}-F{row}"
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"H{row}"] = f"=Debt_M!D{row}"
        ws[f"H{row}"].fill = FORMULA_FILL
        ws[f"I{row}"] = f"=Debt_M!G{row}"
        ws[f"I{row}"].fill = FORMULA_FILL
        ws[f"J{row}"] = f"=Debt_M!H{row}"
        ws[f"J{row}"].fill = FORMULA_FILL
        ws[f"K{row}"] = f"=G{row}+H{row}-I{row}-J{row}"
        ws[f"K{row}"].fill = FORMULA_FILL
    
    # Format columns
    for col in ["D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 15
        for row in range(2, max_rows + 1):
            cell = ws[f"{col}{row}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = "#,##0"

def setup_outputs_tab(ws, timeline_ws, cashflow_ws, capex_ws, debt_ws):
    """Create Outputs tab with summary metrics, IRRs, and checks."""
    row = 1
    
    # Title
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "OUTPUTS & CHECKS"
    ws[f"A{row}"].font = Font(bold=True, size=14)
    row += 2
    
    # Summary Metrics
    ws[f"A{row}"] = "SUMMARY METRICS"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    
    ws[f"A{row}"] = "Total CAPEX (COP)"
    ws[f"B{row}"] = "=CapexDev+CapexCon+CapexCOD"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"B{row}"].number_format = "#,##0"
    row += 1
    
    ws[f"A{row}"] = "Total Debt (COP)"
    ws[f"B{row}"] = "=SUM(Debt_M!D:D)"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"B{row}"].number_format = "#,##0"
    row += 1
    
    ws[f"A{row}"] = "Total Equity (COP)"
    ws[f"B{row}"] = "=B4-B5"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"B{row}"].number_format = "#,##0"
    row += 2
    
    # IRRs
    ws[f"A{row}"] = "IRRs"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    
    # Find last row with data in Cashflow_M
    ws[f"A{row}"] = "Unlevered IRR (Pre-tax)"
    # XIRR needs dates and values
    # Create helper range for dates and cash flows
    ws["D1"] = "UnleveredCF_Dates"
    ws["D2"] = "=Cashflow_M!A2"
    ws["E1"] = "UnleveredCF_Values"
    ws["E2"] = "=Cashflow_M!G2"
    # Copy down to row 500
    for r in range(3, 502):
        ws[f"D{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",Cashflow_M!A{r},\"\")"
        ws[f"E{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",Cashflow_M!G{r},\"\")"
    
    # XIRR formula (Excel 365)
    ws[f"B{row}"] = "=XIRR(E2:E501,D2:D501)"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"B{row}"].number_format = "0.00%"
    row += 1
    
    ws[f"A{row}"] = "Equity IRR (After-tax)"
    # Equity CF helper
    ws["F1"] = "EquityCF_Dates"
    ws["F2"] = "=Cashflow_M!A2"
    ws["G1"] = "EquityCF_Values"
    ws["G2"] = "=Cashflow_M!K2"
    for r in range(3, 502):
        ws[f"F{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",Cashflow_M!A{r},\"\")"
        ws[f"G{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",Cashflow_M!K{r},\"\")"
    
    ws[f"B{row}"] = "=XIRR(G2:G501,F2:F501)"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"B{row}"].number_format = "0.00%"
    row += 2
    
    # Checks
    ws[f"A{row}"] = "CHECKS"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1
    
    # Check 1: Sources = Uses each month
    ws[f"A{row}"] = "Sources = Uses (Monthly)"
    # Sources: Revenue + Debt Draw
    # Uses: OPEX + CAPEX + Debt Service + Debt Fees
    # Check: max absolute difference should be 0
    ws["H1"] = "Sources"
    ws["H2"] = "=Cashflow_M!D2+Cashflow_M!H2"
    ws["I1"] = "Uses"
    ws["I2"] = "=Cashflow_M!E2+Cashflow_M!F2+Cashflow_M!I2+Cashflow_M!J2"
    ws["J1"] = "Diff"
    ws["J2"] = "=H2-I2"
    for r in range(3, 502):
        ws[f"H{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",Cashflow_M!D{r}+Cashflow_M!H{r},\"\")"
        ws[f"I{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",Cashflow_M!E{r}+Cashflow_M!F{r}+Cashflow_M!I{r}+Cashflow_M!J{r},\"\")"
        ws[f"J{r}"] = f"=IF(Cashflow_M!A{r}<>\"\",H{r}-I{r},\"\")"
    
    ws[f"B{row}"] = "=IF(MAX(ABS(J2:J501))<0.01,\"PASS\",\"FAIL\")"
    ws[f"B{row}"].fill = FORMULA_FILL
    # Color code
    ws[f"C{row}"] = '=IF(B{row}="PASS","✓","✗")'
    ws[f"C{row}"].fill = FORMULA_FILL
    row += 1
    
    # Check 2: Debt draw after COD = 0
    ws[f"A{row}"] = "Debt Draw After COD = 0"
    # COD month = DevMonths+ConMonths+1
    # Check all draws in Operation phase = 0
    ws["K1"] = "DebtDraw_Check"
    ws["K2"] = '=IF(Cashflow_M!C2="Operation",ABS(Cashflow_M!H2),0)'
    for r in range(3, 502):
        ws[f"K{r}"] = f'=IF(Cashflow_M!C{r}="Operation",ABS(Cashflow_M!H{r}),0)'
    
    ws[f"B{row}"] = "=IF(MAX(K2:K501)<0.01,\"PASS\",\"FAIL\")"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"C{row}"] = '=IF(B{row}="PASS","✓","✗")'
    ws[f"C{row}"].fill = FORMULA_FILL
    row += 1
    
    # Check 3: Debt balance never < 0
    ws[f"A{row}"] = "Debt Balance >= 0"
    ws["L1"] = "DebtBalance_Check"
    ws["L2"] = "=Debt_M!I2"
    for r in range(3, 502):
        ws[f"L{r}"] = f"=IF(Debt_M!A{r}<>\"\",Debt_M!I{r},\"\")"
    
    ws[f"B{row}"] = "=IF(MIN(L2:L501)>=0,\"PASS\",\"FAIL\")"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"C{row}"] = '=IF(B{row}="PASS","✓","✗")'
    ws[f"C{row}"].fill = FORMULA_FILL
    row += 1
    
    # Check 4: CAPEX total matches Inputs
    ws[f"A{row}"] = "CAPEX Total Matches Inputs"
    ws[f"B{row}"] = "=IF(ABS(SUM(CAPEX_M!G:G)-(CapexDev+CapexCon+CapexCOD))<1000,\"PASS\",\"FAIL\")"
    ws[f"B{row}"].fill = FORMULA_FILL
    ws[f"C{row}"] = '=IF(B{row}="PASS","✓","✗")'
    ws[f"C{row}"].fill = FORMULA_FILL
    
    # Format columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8

def main():
    """Main function to build the Excel model."""
    print("Creating Excel financial model...")
    wb, sheets = create_workbook()
    
    print("Setting up Inputs tab...")
    setup_inputs_tab(sheets["Inputs"])
    
    print("Setting up Timeline_M tab...")
    setup_timeline_tab(sheets["Timeline_M"])
    
    print("Setting up CAPEX_M tab...")
    setup_capex_tab(sheets["CAPEX_M"], sheets["Timeline_M"])
    
    print("Setting up Revenue_M tab...")
    setup_revenue_tab(sheets["Revenue_M"], sheets["Timeline_M"])
    
    print("Setting up OPEX_M tab...")
    setup_opex_tab(sheets["OPEX_M"], sheets["Timeline_M"], sheets["Revenue_M"])
    
    print("Setting up Debt_M tab...")
    setup_debt_tab(sheets["Debt_M"], sheets["Timeline_M"], sheets["CAPEX_M"])
    
    print("Setting up Cashflow_M tab...")
    setup_cashflow_tab(
        sheets["Cashflow_M"],
        sheets["Timeline_M"],
        sheets["CAPEX_M"],
        sheets["Revenue_M"],
        sheets["OPEX_M"],
        sheets["Debt_M"]
    )
    
    print("Setting up Outputs tab...")
    setup_outputs_tab(
        sheets["Outputs"],
        sheets["Timeline_M"],
        sheets["Cashflow_M"],
        sheets["CAPEX_M"],
        sheets["Debt_M"]
    )
    
    print("Saving excel_template.xlsx...")
    wb.save("excel_template.xlsx")
    print("Done! excel_template.xlsx created successfully.")

if __name__ == "__main__":
    main()

