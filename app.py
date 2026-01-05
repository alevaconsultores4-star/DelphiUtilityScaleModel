# Delphi Utility-Scale Financial Model (No Excel)
# Streamlit single-file app with: Projects + Scenarios, Macro, Timeline, Generation, Revenues,
# CAPEX, OPEX, SG&A, Depreciation, Debt & Covenants, Unlevered Base Cash Flow, Compare
# All inputs in COP; outputs selectable COP/USD (USD via FX path).

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple
from io import BytesIO

import numpy as np
import pandas as pd
# Try to import plotly - catch ALL exceptions since plotly.io might fail internally
PLOTLY_AVAILABLE = False
px = None
go = None
try:
    import plotly.express as px
    import plotly.graph_objects as go
    # Test that plotly.io is available (it's imported internally by plotly.express)
    import plotly.io
    PLOTLY_AVAILABLE = True
except Exception as e:
    # Plotly not available - app will work but charts won't display
    px = None
    go = None
    PLOTLY_AVAILABLE = False
    import sys
    print(f"Warning: Plotly not available: {type(e).__name__}: {e}", file=sys.stderr)
import streamlit as st

# Monkey-patch st.plotly_chart to check PLOTLY_AVAILABLE
_original_plotly_chart = st.plotly_chart
def _safe_plotly_chart(*args, **kwargs):
    if not PLOTLY_AVAILABLE:
        st.warning("⚠️ Charts unavailable - Plotly installation issue. All other functionality works normally.")
        return None
    return _original_plotly_chart(*args, **kwargs)
st.plotly_chart = _safe_plotly_chart

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Check for Kaleido (required for Plotly image export)
try:
    import kaleido
    KALEIDO_AVAILABLE = True
except ImportError:
    KALEIDO_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    try:
        from pdf2image import convert_from_path
        PDF2IMAGE_AVAILABLE = True
    except ImportError:
        PDF2IMAGE_AVAILABLE = False


# -----------------------------
# Storage
# -----------------------------
# Check for data folder first, then root directory
if os.path.exists("data/delphi_projects.json"):
    PROJECTS_FILE = "data/delphi_projects.json"
elif os.path.exists("delphi_projects.json"):
    PROJECTS_FILE = "delphi_projects.json"
else:
    PROJECTS_FILE = "delphi_projects.json"  # Default location


def _load_db() -> dict:
    if os.path.exists(PROJECTS_FILE):
        try:
            with open(PROJECTS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"projects": {}}
    return {"projects": {}}


def _save_db(db: dict) -> None:
    with open(PROJECTS_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2, ensure_ascii=False)


# -----------------------------
# Formatting helpers (thousand separators)
# -----------------------------
def _fmt_num(x: float, decimals: int = 0) -> str:
    if x is None:
        return "—"
    try:
        if isinstance(x, float) and not np.isfinite(x):
            return "—"
    except Exception:
        pass
    return f"{x:,.{decimals}f}"


def _fmt_cop(x: float) -> str:
    return f"COP {_fmt_num(float(x), 0)}"


def _fmt_usd(x: float) -> str:
    return f"USD {_fmt_num(float(x), 0)}"


def _df_format_money(df: pd.DataFrame, cols: List[str], decimals: int = 0) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda v: _fmt_num(float(v), decimals) if pd.notnull(v) else "")
    return out


def _transpose_annual_table(df: pd.DataFrame, year_col: str = "Year") -> pd.DataFrame:
    """Transpose annual table so years are column headers instead of rows."""
    if df.empty or year_col not in df.columns:
        return df
    
    # Set Year as index
    df_transposed = df.set_index(year_col).T
    
    # Reset index to make the original column names into a column
    df_transposed = df_transposed.reset_index()
    df_transposed.rename(columns={"index": "Metric"}, inplace=True)
    
    # Convert all column names to strings to avoid Arrow serialization issues
    # Streamlit requires string column names for Arrow compatibility
    df_transposed.columns = [str(col) for col in df_transposed.columns]
    
    return df_transposed


def _load_logo_image():
    """Load and convert Delphi logo from PDF to image.
    
    Checks for logo.png first, then converts LOGO V2.pdf if available.
    Returns the path to the logo image file, or None if not found.
    
    Note: Not using @st.cache_data to allow dynamic checking of file existence.
    """
    logo_pdf_path = "assets/images/LOGO V2.pdf"
    logo_png_path = "assets/images/logo.png"
    
    # Ensure assets/images directory exists
    os.makedirs("assets/images", exist_ok=True)
    
    # Check if PNG already exists (check file system directly, not cached)
    if os.path.exists(logo_png_path):
        return logo_png_path
    
    # If PDF exists, try to convert to PNG
    if os.path.exists(logo_pdf_path):
        # Try PyMuPDF (fitz) first - it's pure Python and easier to deploy
        try:
            import fitz
            doc = fitz.open(logo_pdf_path)
            page = doc[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for quality
            pix.save(logo_png_path)
            doc.close()
            if os.path.exists(logo_png_path):
                return logo_png_path
        except ImportError:
            # PyMuPDF not installed - try pdf2image
            pass
        except Exception as e:
            # Conversion failed, but continue to try other methods
            pass
        
        # Fallback to pdf2image if available
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(logo_pdf_path, dpi=200)
            if images:
                images[0].save(logo_png_path, "PNG")
                if os.path.exists(logo_png_path):
                    return logo_png_path
        except ImportError:
            # pdf2image not installed
            pass
        except Exception as e:
            # Conversion failed
            pass
    
    return None


def _metric_row(items):
    # Filter out empty labels to avoid warnings
    filtered_items = [(k, v) for k, v in items if k and v]
    if not filtered_items:
        return
    cols = st.columns(len(items))
    for i, (k, v) in enumerate(items):
        if k and v:
            cols[i].metric(k, v)
        # Skip empty metrics (they were just for spacing)


# -----------------------------
# Date utilities
# -----------------------------
def _today() -> date:
    return date.today()


def _parse_date_iso(s: str) -> date:
    return date.fromisoformat(s)


def _add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    day = min(
        d.day,
        [31, 29 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1],
    )
    return date(y, m, day)


def _month_starts(start: date, months: int) -> List[date]:
    return [_add_months(start, i) for i in range(months)]


def build_timeline(t: "TimelineInputs") -> dict:
    start = _parse_date_iso(t.start_date) if t.start_date else _today()
    rtb = _add_months(start, int(t.dev_months))
    cod = _add_months(rtb, int(t.capex_months))
    end_op = _add_months(cod, int(t.operation_years) * 12)
    return {"start": start, "rtb": rtb, "cod": cod, "end_op": end_op}


# -----------------------------
# Inputs (dataclasses)
# -----------------------------
INDEX_CHOICES = ["Colombia CPI", "Colombia PPI", "US CPI", "Custom"]


@dataclass
class MacroInputs:
    col_cpi: float = 6.0
    col_ppi: float = 5.0
    us_cpi: float = 3.0
    custom_index_rate: float = 5.0

    fx_cop_per_usd_start: float = 4000.0
    fx_method: str = "Inflation differential (PPP approx.)"  # or "Flat"
    fx_flat: float = 4000.0


@dataclass
class TimelineInputs:
    start_date: str = str(_today())
    dev_months: int = 18
    capex_months: int = 12
    operation_years: int = 25


@dataclass
class GenerationInputs:
    mwac: float = 100.0
    mwp: float = 130.0
    p50_mwh_yr: float = 220000.0
    p75_mwh_yr: float = 210000.0
    p90_mwh_yr: float = 200000.0
    production_choice: str = "P50"
    degradation_pct: float = 0.5  # %/yr


@dataclass
class RevenueOption1PPA:
    ppa_price_cop_per_kwh: float = 320.0
    ppa_term_years: int = 12
    merchant_price_cop_per_kwh: float = 250.0
    indexation: str = "Colombia CPI"


def _default_manual_prices() -> Dict[int, float]:
    return {i: 300.0 for i in range(1, 26)}


@dataclass
class RevenueOption2Manual:
    prices_constant_cop_per_kwh: Dict[int, float] = field(default_factory=_default_manual_prices)
    indexation: str = "Colombia CPI"


@dataclass
class PPAContractConfig:
    """PPA Contract Type Configuration - Contract structure independent from price schedule."""
    # Contract type identifier
    type_id: str = "pay_as_generated_100"  # "pay_as_generated_100", "pay_as_generated_cap", "fixed_block", "floor_cap"
    ui_label_es: str = "Pague lo generado (100%)"
    ui_label_en: str = "Pay-as-Generated (100%)"
    
    # Type 2: Pay-as-Generated with Cap
    cap_mode: str = "% de P50"  # "MWh/año" or "% de P50"
    cap_value: float = 100.0  # Cap value (percentage or MWh/year)
    excess_treatment: str = "Spot"  # "Spot", "% del PPA", "0 (no pagado)"
    excess_discount_pct: float = 80.0  # Only used if excess_treatment == "% del PPA"
    
    # Type 3: Fixed Energy Block (Baseload)
    block_mode: str = "% de P50"  # "MWh/año" or "% de P50"
    block_value: float = 100.0  # Block size (percentage or MWh/year)
    settlement_price_mode: str = "Spot (promedio)"  # "Spot (promedio)" or "Fijo"
    settlement_price_fixed: float = 0.0  # Fixed settlement price (only used if settlement_price_mode == "Fijo")
    
    # Type 4: Floor & Cap (Collar)
    floor_pct_of_p50_revenue: float = 90.0  # Floor as % of P50 revenue
    cap_pct_of_p50_revenue: float = 110.0  # Cap as % of P50 revenue


CAPEX_PHASES = ["Development", "Construction", "At COD"]
CAPEX_DISTS = ["Straight-line (monthly)", "Front-loaded", "Back-loaded"]


def _default_capex_lines() -> List[Dict[str, object]]:
    return [
        {"Item": "EPC (modules + BOS + installation)", "Amount_COP": 0.0, "Phase": "Construction"},
        {"Item": "Interconnection / Substation / Line", "Amount_COP": 0.0, "Phase": "Construction"},
        {"Item": "Development costs", "Amount_COP": 0.0, "Phase": "Development"},
    ]


@dataclass
class CapexInputs:
    lines: List[Dict[str, object]] = field(default_factory=_default_capex_lines)
    distribution: str = "Straight-line (monthly)"


PHASES = ["Development", "Construction", "Operation"]


def _default_opex_other_items() -> List[Dict[str, object]]:
    return [
        {"Item": "Other OPEX item", "Amount_COP_per_year": 0.0, "Phase": "Operation", "Indexation": "Colombia CPI"},
    ]


@dataclass
class OpexInputs:
    fixed_om_cop_per_mwac_year: float = 0.0
    fixed_om_indexation: str = "Colombia CPI"

    variable_om_cop_per_mwh: float = 0.0
    variable_om_indexation: str = "Colombia CPI"

    insurance_cop_per_mwac_year: float = 0.0
    insurance_indexation: str = "Colombia CPI"

    grid_fees_cop_per_mwh: float = 0.0

    land_hectares: float = 0.0
    land_price_dev_cop_per_ha_year: float = 0.0
    land_price_con_cop_per_ha_year: float = 0.0
    land_price_op_cop_per_ha_year: float = 0.0
    land_indexation: str = "Colombia CPI"

    ica_pct_of_revenue: float = 0.0
    gmf_pct_of_outflows: float = 0.4  # default 0.4%

    other_items: List[Dict[str, object]] = field(default_factory=_default_opex_other_items)


def _default_sga_items() -> List[Dict[str, object]]:
    return [
        {"Item": "Project management / Owner team", "Amount_COP_per_year": 0.0, "Phase": "Development", "Indexation": "Colombia CPI"},
        {"Item": "Permitting / legal / compliance", "Amount_COP_per_year": 0.0, "Phase": "Development", "Indexation": "Colombia CPI"},
        {"Item": "Corporate overhead allocation", "Amount_COP_per_year": 0.0, "Phase": "Operation", "Indexation": "Colombia CPI"},
    ]


@dataclass
class SgaInputs:
    items: List[Dict[str, object]] = field(default_factory=_default_sga_items)


@dataclass
class DepreciationInputs:
    pct_of_capex_depreciated: float = 100.0
    dep_years: int = 20  # 3–25


@dataclass
class TaxInputs:
    corporate_tax_rate_pct: float = 35.0
    allow_loss_carryforward: bool = True

@dataclass
class RenewableIncentivesInputs:
    # Special deduction (Ley 1715 / 2099 style): up to 50% of eligible investment
    # usable over up to 15 years, but each year capped at 50% of taxable income.
    enable_special_deduction: bool = True
    special_deduction_pct_of_capex: float = 50.0          # 0–50 (%)
    special_deduction_years: int = 15                     # typically 15
    special_deduction_max_pct_of_taxable_income: float = 50.0  # cap per year

    # VAT: model as either excluded (default) or refunded as a one-time cash inflow
    vat_mode: str = "Excluded"                            # "Excluded" or "Refund"
    vat_pct_of_capex: float = 0.0                         # if you prefer % input
    vat_fixed_cop: float = 0.0                            # or fixed COP input
    vat_refund_year_index: int = 1                        # refund in op Year 1 by default

@dataclass
class WorkingCapitalInputs:
    ar_days: int = 90   # revenue collection lag
    ap_days: int = 60   # expense payment lag
    apply_ap_to_opex: bool = True
    apply_ap_to_sga: bool = True

@dataclass
class DebtInputs:
    enabled: bool = False
    debt_pct_of_capex: float = 70.0  # %
    tenor_years: int = 7            # allow 5–10
    grace_years: int = 0            # 0–(tenor-1)

    # Pricing (Natural COP)
    base_rate_pct: float = 9.0      # e.g., IBR
    margin_pct: float = 7.0         # spread
    # Fees
    upfront_fee_bps: float = 175.0  # bps of total debt, paid at COD/first draw
    commitment_fee_pct_of_margin: float = 30.0  # % of margin, on undrawn during construction

    # Covenants (for indicators)
    target_dscr: float = 1.20
    min_dscr_covenant: float = 1.20
    lockup_dscr: float = 1.15

    amortization_type: str = "Sculpted to DSCR"  # "Sculpted to DSCR", "Equal principal", or "Typical amortization"
    balloon_pct: float = 0.0   # NEW: % of original debt left as balloon (0 = fully amortizing)

@dataclass
class RenewableTaxInputs:
    enable_special_deduction: bool = True
    special_deduction_pct_of_capex: float = 50.0   # % of CAPEX
    special_deduction_years: int = 15               # max carryforward

    enable_vat_refund: bool = True
    vat_refund_mode: str = "percent"                # "percent" or "fixed"
    vat_pct_of_capex: float = 19.0                  # if percent
    vat_fixed_cop: float = 0.0                      # if fixed
    vat_refund_year: int = 1                         # years after COD


@dataclass
class ProjectOverviewInputs:
    """Project Overview metadata - descriptive fields only, no calculations."""
    project_name: str = ""
    country: str = ""
    region_department: str = ""
    technology: str = ""
    installed_capacity_mw: float = 0.0
    short_description: str = ""
    optional_notes: str = ""


@dataclass
class ScenarioInputs:
    name: str = "Base"
    project_overview: ProjectOverviewInputs = field(default_factory=ProjectOverviewInputs)
    macro: MacroInputs = field(default_factory=MacroInputs)
    timeline: TimelineInputs = field(default_factory=TimelineInputs)
    generation: GenerationInputs = field(default_factory=GenerationInputs)

    revenue_mode: str = "Standard PPA Parameters"
    revenue1: RevenueOption1PPA = field(default_factory=RevenueOption1PPA)
    revenue2: RevenueOption2Manual = field(default_factory=RevenueOption2Manual)
    ppa_contract: PPAContractConfig = field(default_factory=PPAContractConfig)

    capex: CapexInputs = field(default_factory=CapexInputs)
    opex: OpexInputs = field(default_factory=OpexInputs)
    sga: SgaInputs = field(default_factory=SgaInputs)

    depreciation: DepreciationInputs = field(default_factory=DepreciationInputs)
    tax: TaxInputs = field(default_factory=TaxInputs)
    wc: WorkingCapitalInputs = field(default_factory=WorkingCapitalInputs)

    debt: DebtInputs = field(default_factory=DebtInputs)
    renewable_tax: RenewableTaxInputs = field(default_factory=RenewableTaxInputs)
    incentives: RenewableIncentivesInputs = field(default_factory=RenewableIncentivesInputs)
    
    uploaded_files: Dict[str, List[Dict[str, str]]] = field(default_factory=dict)  # tab_name -> list of file metadata



def _scenario_to_dict(s: ScenarioInputs) -> dict:
    return asdict(s)


def _scenario_from_dict(d: dict) -> ScenarioInputs:
    macro = MacroInputs(**d.get("macro", {}))
    timeline = TimelineInputs(**d.get("timeline", {}))
    generation = GenerationInputs(**d.get("generation", {}))

    revenue_mode = d.get("revenue_mode", "Standard PPA Parameters")
    revenue1 = RevenueOption1PPA(**d.get("revenue1", {}))
    # Handle revenue2 with proper integer key conversion for prices_constant_cop_per_kwh
    revenue2_dict = d.get("revenue2", {})
    if "prices_constant_cop_per_kwh" in revenue2_dict:
        # Convert string keys back to integers (JSON serializes dict keys as strings)
        prices_dict = revenue2_dict["prices_constant_cop_per_kwh"]
        if isinstance(prices_dict, dict) and len(prices_dict) > 0:
            # Check if keys are strings and convert to int
            first_key = next(iter(prices_dict.keys()))
            if isinstance(first_key, str):
                revenue2_dict["prices_constant_cop_per_kwh"] = {int(k): float(v) for k, v in prices_dict.items()}
    revenue2 = RevenueOption2Manual(**revenue2_dict)

    capex = CapexInputs(**d.get("capex", {})) if "capex" in d else CapexInputs()
    opex = OpexInputs(**d.get("opex", {})) if "opex" in d else OpexInputs()
    sga = SgaInputs(**d.get("sga", {})) if "sga" in d else SgaInputs()

    depreciation = DepreciationInputs(**d.get("depreciation", {})) if "depreciation" in d else DepreciationInputs()
    tax = TaxInputs(**d.get("tax", {})) if "tax" in d else TaxInputs()
    wc = WorkingCapitalInputs(**d.get("wc", {})) if "wc" in d else WorkingCapitalInputs()
    debt = DebtInputs(**d.get("debt", {})) if "debt" in d else DebtInputs()
    incentives = RenewableIncentivesInputs(**d.get("incentives", {}))
    
    # PPA Contract Config (backward compatible - use defaults if missing)
    ppa_contract = PPAContractConfig(**d.get("ppa_contract", {})) if "ppa_contract" in d else PPAContractConfig()

    # Project Overview (robust backward compatibility: handle missing keys, wrong types, malformed data)
    project_overview_dict = d.get("project_overview", {})
    if not isinstance(project_overview_dict, dict):
        project_overview_dict = {}
    
    try:
        project_overview = ProjectOverviewInputs(
            project_name=str(project_overview_dict.get("project_name", "")),
            country=str(project_overview_dict.get("country", "")),
            region_department=str(project_overview_dict.get("region_department", "")),
            technology=str(project_overview_dict.get("technology", "")),
            installed_capacity_mw=float(project_overview_dict.get("installed_capacity_mw", 0.0)) if project_overview_dict.get("installed_capacity_mw") is not None else 0.0,
            short_description=str(project_overview_dict.get("short_description", "")),
            optional_notes=str(project_overview_dict.get("optional_notes", ""))
        )
    except (ValueError, TypeError, KeyError):
        # Fall back to defaults if parsing fails
        project_overview = ProjectOverviewInputs()

    uploaded_files = d.get("uploaded_files", {})
    if not isinstance(uploaded_files, dict):
        uploaded_files = {}
    
    return ScenarioInputs(
        name=d.get("name", "Base"),
        project_overview=project_overview,
        macro=macro,
        timeline=timeline,
        generation=generation,
        revenue_mode=revenue_mode,
        revenue1=revenue1,
        revenue2=revenue2,
        ppa_contract=ppa_contract,
        capex=capex,
        opex=opex,
        sga=sga,
        depreciation=depreciation,
        tax=tax,
        wc=wc,
        debt=debt,
        incentives=incentives,
        uploaded_files=uploaded_files,
    )


# -----------------------------
# File upload helpers
# -----------------------------
def _get_tab_name(tab_var) -> str:
    """Map tab variable to directory-friendly name."""
    tab_mapping = {
        "tab_macro": "macro",
        "tab_timeline": "timeline",
        "tab_gen": "generation",
        "tab_rev": "revenues",
        "tab_capex": "capex",
        "tab_opex": "opex",
        "tab_sga": "sga",
        "tab_dep": "depreciation",
        "tab_incent": "incentives",
        "tab_ucf": "unlevered_cashflow",
        "tab_debt": "debt",
        "tab_levered": "levered_cashflow",
        "tab_compare": "compare",
        "tab_sensitivity": "sensitivity",
        "tab_summary": "summary",
    }
    # Get the variable name as string by checking locals/globals
    for name, value in {**globals(), **locals()}.items():
        if value is tab_var and name.startswith("tab_"):
            return tab_mapping.get(name, name.replace("tab_", ""))
    return "unknown"


def _get_upload_dir(project_name: str, scenario_name: str, tab_name: str) -> str:
    """Get upload directory path for a tab, creating it if it doesn't exist."""
    # Sanitize names for filesystem (remove invalid characters)
    safe_project = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_scenario = "".join(c for c in scenario_name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_tab = "".join(c for c in tab_name if c.isalnum() or c in (' ', '-', '_')).strip()
    
    upload_dir = os.path.join("data", safe_project, safe_scenario, safe_tab, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    return upload_dir


def _format_file_size(bytes_size: int) -> str:
    """Format file size for display."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_size < 1024.0:
            return f"{bytes_size:.1f} {unit}"
        bytes_size /= 1024.0
    return f"{bytes_size:.1f} TB"


def _save_uploaded_file(uploaded_file, project_name: str, scenario_name: str, tab_name: str) -> dict:
    """Save uploaded file to disk and return metadata."""
    upload_dir = _get_upload_dir(project_name, scenario_name, tab_name)
    
    # Handle duplicate filenames by appending timestamp
    original_filename = uploaded_file.name
    filepath = os.path.join(upload_dir, original_filename)
    
    # If file exists, append timestamp
    if os.path.exists(filepath):
        name, ext = os.path.splitext(original_filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"{name}_{timestamp}{ext}"
        filepath = os.path.join(upload_dir, new_filename)
        original_filename = new_filename
    
    # Save file
    try:
        with open(filepath, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        file_size = os.path.getsize(filepath)
        upload_date = datetime.now().isoformat()
        
        return {
            "filename": original_filename,
            "filepath": filepath,
            "upload_date": upload_date,
            "file_size": str(file_size)
        }
    except Exception as e:
        raise Exception(f"Failed to save file: {str(e)}")


def _delete_uploaded_file(filepath: str) -> bool:
    """Delete file from disk."""
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
        return False
    except Exception:
        return False


def _render_file_upload_section(s: ScenarioInputs, project_name: str, scenario_name: str, tab_name: str, section_title: str) -> None:
    """Reusable UI component for file uploads."""
    st.markdown(f"#### {section_title}")
    st.info("ℹ️ Files uploaded here are for reference only and do not affect any financial calculations.")
    
    # Initialize uploaded_files dict if needed
    if not hasattr(s, 'uploaded_files') or s.uploaded_files is None:
        s.uploaded_files = {}
    if tab_name not in s.uploaded_files:
        s.uploaded_files[tab_name] = []
    
    # File uploader (accepts all file types)
    uploaded_file = st.file_uploader(
        "Upload file",
        type=None,  # Accept all file types
        key=f"file_upload_{tab_name}_{project_name}_{scenario_name}",
        help="Upload any file type for reference"
    )
    
    # Handle file upload
    if uploaded_file is not None:
        # Track processed files using session state to prevent duplicates
        upload_state_key = f"last_uploaded_{tab_name}_{project_name}_{scenario_name}"
        file_id = f"{uploaded_file.name}_{uploaded_file.size}"
        last_processed = st.session_state.get(upload_state_key, None)
        
        # Only process if this is a new/different file
        if last_processed != file_id:
            try:
                file_metadata = _save_uploaded_file(uploaded_file, project_name, scenario_name, tab_name)
                s.uploaded_files[tab_name].append(file_metadata)
                
                # Save to database
                db = _load_db()
                proj = db["projects"].setdefault(project_name, {"scenarios": {}})
                proj["scenarios"][scenario_name] = _scenario_to_dict(s)
                _save_db(db)
                
                # Update session state to track this file as processed
                st.session_state[upload_state_key] = file_id
                
                st.success(f"✓ File '{file_metadata['filename']}' uploaded successfully.")
                st.rerun()
            except Exception as e:
                st.error(f"Error uploading file: {str(e)}")
    
    # Display uploaded files
    files = s.uploaded_files.get(tab_name, [])
    
    # Clean up metadata for files that no longer exist
    valid_files = []
    for file_info in files:
        filepath = file_info.get("filepath", "")
        if os.path.exists(filepath):
            valid_files.append(file_info)
        else:
            st.warning(f"⚠️ File '{file_info.get('filename', 'unknown')}' not found on disk. It will be removed from the list.")
    
    # Update if files were removed
    if len(valid_files) != len(files):
        s.uploaded_files[tab_name] = valid_files
        db = _load_db()
        proj = db["projects"].setdefault(project_name, {"scenarios": {}})
        proj["scenarios"][scenario_name] = _scenario_to_dict(s)
        _save_db(db)
        files = valid_files
    
    if files:
        st.markdown("**Uploaded files:**")
        for idx, file_info in enumerate(files):
            col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
            with col1:
                st.text(file_info.get("filename", "Unknown"))
            with col2:
                file_size = int(file_info.get("file_size", 0))
                st.text(_format_file_size(file_size))
            with col3:
                upload_date = file_info.get("upload_date", "")
                if upload_date:
                    try:
                        dt = datetime.fromisoformat(upload_date.replace('Z', '+00:00'))
                        st.text(dt.strftime("%Y-%m-%d"))
                    except:
                        st.text("—")
                else:
                    st.text("—")
            with col4:
                filepath = file_info.get("filepath", "")
                delete_key = f"delete_{tab_name}_{idx}_{project_name}_{scenario_name}"
                
                # Download button
                if os.path.exists(filepath):
                    with open(filepath, "rb") as f:
                        file_bytes = f.read()
                    st.download_button(
                        "📥",
                        file_bytes,
                        file_name=file_info.get("filename", "file"),
                        key=f"download_{tab_name}_{idx}",
                        help="Download file"
                    )
                
                # Delete button
                if st.button("🗑️", key=delete_key, help="Delete file"):
                    if _delete_uploaded_file(filepath):
                        # Remove from metadata
                        s.uploaded_files[tab_name] = [f for f in files if f.get("filepath") != filepath]
                        
                        # Save to database
                        db = _load_db()
                        proj = db["projects"].setdefault(project_name, {"scenarios": {}})
                        proj["scenarios"][scenario_name] = _scenario_to_dict(s)
                        _save_db(db)
                        
                        st.success(f"✓ File '{file_info.get('filename', 'unknown')}' deleted.")
                        st.rerun()
                    else:
                        st.error("Failed to delete file.")
    else:
        st.info("No files uploaded yet.")


# -----------------------------
# PPA Contract Validation Helpers
# -----------------------------
def _clamp_ppa_value(value: float, min_val: float, max_val: float, field_name: str) -> Tuple[float, Optional[str]]:
    """Clamp a value to a range and return (clamped_value, warning_message)."""
    if value < min_val:
        return min_val, f"⚠️ {field_name} was {value:.2f}, clamped to minimum {min_val:.2f}"
    elif value > max_val:
        return max_val, f"⚠️ {field_name} was {value:.2f}, clamped to maximum {max_val:.2f}"
    return value, None


def _validate_ppa_contract_config(config: PPAContractConfig, current_ppa_price: float) -> List[str]:
    """Validate PPA contract configuration and return list of warning messages."""
    warnings = []
    
    # Validate percent values (0-200 range)
    percent_fields = [
        ("cap_value", config.cap_value) if config.cap_mode == "% de P50" else None,
        ("block_value", config.block_value) if config.block_mode == "% de P50" else None,
        ("excess_discount_pct", config.excess_discount_pct) if config.excess_treatment == "% del PPA" else None,
        ("floor_pct_of_p50_revenue", config.floor_pct_of_p50_revenue),
        ("cap_pct_of_p50_revenue", config.cap_pct_of_p50_revenue),
    ]
    
    for field_info in percent_fields:
        if field_info is not None:
            field_name, value = field_info
            clamped_val, warning = _clamp_ppa_value(value, 0.0, 200.0, field_name.replace("_", " ").title())
            if warning:
                warnings.append(warning)
            # Auto-update the value
            setattr(config, field_name, clamped_val)
    
    # Validate MWh/year values (>= 0)
    mwh_fields = [
        ("cap_value", config.cap_value) if config.cap_mode == "MWh/año" else None,
        ("block_value", config.block_value) if config.block_mode == "MWh/año" else None,
    ]
    
    for field_info in mwh_fields:
        if field_info is not None:
            field_name, value = field_info
            clamped_val, warning = _clamp_ppa_value(value, 0.0, float('inf'), field_name.replace("_", " ").title())
            if warning:
                warnings.append(warning)
            # Auto-update the value
            setattr(config, field_name, clamped_val)
    
    # Validate Floor & Cap: floor <= cap
    if config.type_id == "floor_cap":
        if config.floor_pct_of_p50_revenue > config.cap_pct_of_p50_revenue:
            warnings.append(f"⚠️ Floor ({config.floor_pct_of_p50_revenue:.2f}%) cannot exceed Cap ({config.cap_pct_of_p50_revenue:.2f}%). Auto-adjusting floor to cap value.")
            config.floor_pct_of_p50_revenue = config.cap_pct_of_p50_revenue
    
    # Validate settlement_price_fixed (should be set if mode is "Fijo")
    if config.type_id == "fixed_block" and config.settlement_price_mode == "Fijo":
        if config.settlement_price_fixed <= 0.0:
            # Default to current PPA price if not set
            if current_ppa_price > 0:
                config.settlement_price_fixed = current_ppa_price
                warnings.append(f"⚠️ Fixed settlement price was not set. Defaulted to current PPA price: {current_ppa_price:.4f} COP/kWh")
            else:
                warnings.append("⚠️ Fixed settlement price mode selected but price is 0. Please set a valid price.")
    
    return warnings


# -----------------------------
# Macro series
# -----------------------------
def _idx_key(choice: str) -> str:
    if choice == "Colombia CPI":
        return "col_cpi"
    if choice == "Colombia PPI":
        return "col_ppi"
    if choice == "US CPI":
        return "us_cpi"
    return "custom_index_rate"


def annual_index_series(macro: MacroInputs, base_year: int, years: List[int], index_key: str) -> pd.Series:
    r = float(getattr(macro, index_key)) / 100.0
    out = {}
    for y in years:
        n = y - base_year
        out[y] = (1.0 + r) ** n
    return pd.Series(out)


def fx_series(macro: MacroInputs, base_year: int, years: List[int]) -> pd.Series:
    fx0 = float(macro.fx_cop_per_usd_start)
    if macro.fx_method == "Flat":
        return pd.Series({y: float(macro.fx_flat or fx0) for y in years})
    col = float(macro.col_cpi) / 100.0
    us = float(macro.us_cpi) / 100.0
    drift = (1.0 + col) / (1.0 + us) - 1.0
    return pd.Series({y: fx0 * ((1.0 + drift) ** (y - base_year)) for y in years})


def index_factor_for_year(macro: MacroInputs, year: int, base_year: int, index_choice: str) -> float:
    years = list(range(min(base_year, year), max(base_year, year) + 1))
    s = annual_index_series(macro, base_year, years, _idx_key(index_choice))
    return float(s.loc[year])


# -----------------------------
# Generation & Revenues (annual)
# -----------------------------
def operating_year_table(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    end_op = tl["end_op"]
    op_years = int(s.timeline.operation_years)

    # Get all calendar years from COD to end of operation
    years = list(range(cod.year, end_op.year + 1))
    base_year = cod.year

    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    base_mwh = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    degr = float(gen.degradation_pct) / 100.0
    
    # Calculate prorated generation for each calendar year
    mwh_list = []
    price_base_list = []
    
    for year in years:
        # Determine start and end dates for this calendar year
        if year == cod.year:
            # First year: from COD date to end of year
            year_start = cod
            year_end = date(year, 12, 31)
        elif year == end_op.year:
            # Last year: from start of year to end_op date
            year_start = date(year, 1, 1)
            year_end = end_op
        else:
            # Full year
            year_start = date(year, 1, 1)
            year_end = date(year, 12, 31)
        
        # Calculate months of operation in this calendar year
        months_in_year = (year_end.year - year_start.year) * 12 + (year_end.month - year_start.month) + 1
        fraction_of_year = months_in_year / 12.0
        
        # Calculate which operating year the middle of this calendar year falls into
        # Use the midpoint of the calendar year to determine operating year
        if year == cod.year:
            mid_month = (cod.month + 12) / 2.0
        elif year == end_op.year:
            mid_month = (1 + end_op.month) / 2.0
        else:
            mid_month = 6.5  # Middle of year
        
        # Calculate months from COD to midpoint of this calendar year
        months_from_cod = (year - cod.year) * 12 + (mid_month - cod.month)
        operating_year_num = int(months_from_cod // 12)  # 0-indexed operating year
        
        # Calculate annual generation for this operating year (with degradation)
        annual_mwh = base_mwh * ((1.0 - degr) ** operating_year_num)
        
        # Prorate by fraction of year
        prorated_mwh = annual_mwh * fraction_of_year
        
        # Calculate price based on operating year
        if s.revenue_mode == "Standard PPA Parameters":
            r = s.revenue1
            term = int(r.ppa_term_years)
            p0 = float(r.ppa_price_cop_per_kwh)
            pm = float(r.merchant_price_cop_per_kwh)
            price = p0 if (operating_year_num < term) else pm
        else:
            r = s.revenue2
            # Look up price by operating year (1-indexed)
            # operating_year_num is 0-indexed (0 = first operating year, 1 = second, etc.)
            # Manual prices are stored with 1-indexed keys (1 = first operating year, 2 = second, etc.)
            op_year_key = operating_year_num + 1
            price = float(r.prices_constant_cop_per_kwh.get(op_year_key, 0.0))
            
            # Debug: if price is 0, check if the key exists in the dictionary
            # This helps identify if it's a lookup issue or missing data
            if price == 0.0 and op_year_key not in r.prices_constant_cop_per_kwh:
                # Price not found for this operating year - will show warning in UI
                pass
        
        mwh_list.append(prorated_mwh)
        price_base_list.append(price)
    
    # Apply indexation
    index_choice = s.revenue1.indexation if s.revenue_mode == "Standard PPA Parameters" else s.revenue2.indexation
    idx = annual_index_series(s.macro, base_year, years, _idx_key(index_choice))
    price_indexed = [price_base_list[i] * float(idx.loc[years[i]]) for i in range(len(years))]

    df = pd.DataFrame({"Year": years, "Energy (MWh)": mwh_list})
    df["Price (COP/kWh)"] = price_indexed
    df["Revenue (COP)"] = df["Energy (MWh)"] * 1000.0 * df["Price (COP/kWh)"]
    return df


# -----------------------------
# CAPEX
# -----------------------------
def _weights(n: int, mode: str) -> np.ndarray:
    if n <= 0:
        return np.array([])
    if mode == "Straight-line (monthly)":
        w = np.ones(n)
    elif mode == "Front-loaded":
        w = np.linspace(n, 1, n)
    else:
        w = np.linspace(1, n, n)
    return w / w.sum()


def capex_monthly_schedule(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    start = tl["start"]
    rtb = tl["rtb"]
    cod = tl["cod"]

    dev_n = int(s.timeline.dev_months)
    con_n = int(s.timeline.capex_months)

    dev_months = _month_starts(start, dev_n)
    con_months = _month_starts(rtb, con_n)

    lines = s.capex.lines or []
    dev_total = sum(float(x.get("Amount_COP", 0.0) or 0.0) for x in lines if x.get("Phase") == "Development")
    con_total = sum(float(x.get("Amount_COP", 0.0) or 0.0) for x in lines if x.get("Phase") == "Construction")
    cod_total = sum(float(x.get("Amount_COP", 0.0) or 0.0) for x in lines if x.get("Phase") == "At COD")

    w_dev = _weights(dev_n, s.capex.distribution)
    w_con = _weights(con_n, s.capex.distribution)

    rows = []
    for i, m in enumerate(dev_months):
        # Normalize to 1st of month to match opex_monthly_schedule format
        month_normalized = date(m.year, m.month, 1)
        weight = w_dev[i] if (len(w_dev) > i) else (1.0 / dev_n if dev_n > 0 else 0.0)
        capex_value = dev_total * weight
        rows.append({"Month": month_normalized, "Phase": "Development", "CAPEX (COP)": capex_value})
    
    for i, m in enumerate(con_months):
        # Normalize to 1st of month to match opex_monthly_schedule format
        month_normalized = date(m.year, m.month, 1)
        weight = w_con[i] if (len(w_con) > i) else (1.0 / con_n if con_n > 0 else 0.0)
        capex_value = con_total * weight
        rows.append({"Month": month_normalized, "Phase": "Construction", "CAPEX (COP)": capex_value})
    
    rows.append({"Month": date(cod.year, cod.month, 1), "Phase": "At COD", "CAPEX (COP)": cod_total})

    df = pd.DataFrame(rows).sort_values("Month").reset_index(drop=True)
    df["Year"] = df["Month"].apply(lambda d: d.year)
    return df


# -----------------------------
# OPEX monthly schedule (includes Revenue + Energy + CAPEX merge)
# -----------------------------
def _phase_for_month(tl: dict, m: date) -> str:
    if m < date(tl["rtb"].year, tl["rtb"].month, 1):
        return "Development"
    if m < date(tl["cod"].year, tl["cod"].month, 1):
        return "Construction"
    return "Operation"


def opex_monthly_schedule(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    start = tl["start"]
    end_op = tl["end_op"]

    months = []
    cur = date(start.year, start.month, 1)
    endm = date(end_op.year, end_op.month, 1)
    while cur < endm:
        months.append(cur)
        cur = _add_months(cur, 1)

    df = pd.DataFrame({"Month": months})
    df["Year"] = df["Month"].apply(lambda d: d.year)
    df["Phase"] = df["Month"].apply(lambda m: _phase_for_month(tl, m))

    op = operating_year_table(s)
    op_map_mwh = {int(r["Year"]): float(r["Energy (MWh)"]) for _, r in op.iterrows()}
    op_map_rev = {int(r["Year"]): float(r["Revenue (COP)"]) for _, r in op.iterrows()}
    
    # Calculate operating months per calendar year for proper proration
    cod = tl["cod"]
    operating_months_per_year = {}
    for year in op["Year"].unique():
        year_int = int(year)
        if year_int == cod.year:
            # First year: from COD month to end of year
            operating_months_per_year[year_int] = 13 - cod.month
        elif year_int == end_op.year:
            # Last year: from start of year to end_op month
            operating_months_per_year[year_int] = end_op.month
        else:
            # Full year
            operating_months_per_year[year_int] = 12

    def get_monthly_value(row, annual_map, months_map):
        if row["Phase"] != "Operation":
            return 0.0
        year = int(row["Year"])
        annual_val = annual_map.get(year, 0.0)
        months = months_map.get(year, 12)
        return annual_val / months if months > 0 else 0.0

    df["Energy (MWh)"] = df.apply(lambda r: get_monthly_value(r, op_map_mwh, operating_months_per_year), axis=1)
    df["Revenue (COP)"] = df.apply(lambda r: get_monthly_value(r, op_map_rev, operating_months_per_year), axis=1)

    cap = capex_monthly_schedule(s)[["Month", "CAPEX (COP)"]].copy()
    # Ensure Month column is normalized to 1st of month for both dataframes
    df["Month"] = df["Month"].apply(lambda d: date(d.year, d.month, 1) if isinstance(d, date) else d)
    cap["Month"] = cap["Month"].apply(lambda d: date(d.year, d.month, 1) if isinstance(d, date) else d)
    df = df.merge(cap, on="Month", how="left")
    df["CAPEX (COP)"] = df["CAPEX (COP)"].fillna(0.0)

    mwac = float(s.generation.mwac or 0.0)

    base_year = tl["cod"].year
    idx_fixed = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, s.opex.fixed_om_indexation))
    idx_ins = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, s.opex.insurance_indexation))
    idx_land = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, s.opex.land_indexation))

    df["Fixed O&M"] = 0.0
    df.loc[df["Phase"] == "Operation", "Fixed O&M"] = (float(s.opex.fixed_om_cop_per_mwac_year) * mwac / 12.0) * idx_fixed[df["Phase"] == "Operation"].values

    df["Insurance"] = 0.0
    df.loc[df["Phase"] == "Operation", "Insurance"] = (float(s.opex.insurance_cop_per_mwac_year) * mwac / 12.0) * idx_ins[df["Phase"] == "Operation"].values

    df["Variable O&M"] = float(s.opex.variable_om_cop_per_mwh) * df["Energy (MWh)"]
    df["Grid fees"] = float(s.opex.grid_fees_cop_per_mwh) * df["Energy (MWh)"]

    ha = float(s.opex.land_hectares or 0.0)
    df["Land lease"] = 0.0
    df.loc[df["Phase"] == "Development", "Land lease"] = (ha * float(s.opex.land_price_dev_cop_per_ha_year) / 12.0) * idx_land[df["Phase"] == "Development"].values
    df.loc[df["Phase"] == "Construction", "Land lease"] = (ha * float(s.opex.land_price_con_cop_per_ha_year) / 12.0) * idx_land[df["Phase"] == "Construction"].values
    df.loc[df["Phase"] == "Operation", "Land lease"] = (ha * float(s.opex.land_price_op_cop_per_ha_year) / 12.0) * idx_land[df["Phase"] == "Operation"].values

    other = pd.DataFrame(s.opex.other_items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in other.columns:
            other[col] = "" if col != "Amount_COP_per_year" else 0.0
    other["Phase"] = other["Phase"].where(other["Phase"].isin(PHASES), "Operation")
    other["Indexation"] = other["Indexation"].where(other["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    for _, r in other.iterrows():
        name = str(r.get("Item", "")).strip() or "Other"
        amt = float(r.get("Amount_COP_per_year", 0.0) or 0.0)
        ph = str(r.get("Phase", "Operation"))
        idx_choice = str(r.get("Indexation", "Colombia CPI"))
        if name not in df.columns:
            df[name] = 0.0
        idx_series = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, idx_choice))
        df.loc[df["Phase"] == ph, name] = (amt / 12.0) * idx_series[df["Phase"] == ph].values

    df["ICA"] = 0.0
    df.loc[df["Phase"] == "Operation", "ICA"] = (float(s.opex.ica_pct_of_revenue) / 100.0) * df.loc[df["Phase"] == "Operation", "Revenue (COP)"]

    meta = {"Month", "Year", "Phase", "Energy (MWh)", "Revenue (COP)", "CAPEX (COP)"}
    fixed_cols = ["Fixed O&M", "Insurance", "Variable O&M", "Grid fees", "Land lease", "ICA"]
    dyn_cols = [c for c in df.columns if c not in meta and c not in fixed_cols and c not in {"OPEX subtotal", "GMF"}]
    opex_cols = fixed_cols + dyn_cols
    df["OPEX subtotal"] = df[opex_cols].sum(axis=1) if opex_cols else 0.0

    # GMF on outgoing cash: CAPEX + OPEX subtotal
    gmf = float(s.opex.gmf_pct_of_outflows) / 100.0
    df["GMF"] = gmf * (df["CAPEX (COP)"].fillna(0.0) + df["OPEX subtotal"].fillna(0.0))

    return df


# -----------------------------
# SG&A schedules
# -----------------------------
def sga_monthly_schedule(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    start = tl["start"]
    end_op = tl["end_op"]

    months = []
    cur = date(start.year, start.month, 1)
    endm = date(end_op.year, end_op.month, 1)
    while cur < endm:
        months.append(cur)
        cur = _add_months(cur, 1)

    df = pd.DataFrame({"Month": months})
    df["Year"] = df["Month"].apply(lambda d: d.year)
    df["Phase"] = df["Month"].apply(lambda m: _phase_for_month(tl, m))

    items = pd.DataFrame(s.sga.items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in items.columns:
            items[col] = "" if col != "Amount_COP_per_year" else 0.0
    items["Phase"] = items["Phase"].where(items["Phase"].isin(PHASES), "Development")
    items["Indexation"] = items["Indexation"].where(items["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    base_year = tl["cod"].year
    for _, r in items.iterrows():
        name = str(r.get("Item", "")).strip() or "SGA"
        amt = float(r.get("Amount_COP_per_year", 0.0) or 0.0)
        ph = str(r.get("Phase", "Development"))
        idx_choice = str(r.get("Indexation", "Colombia CPI"))
        if name not in df.columns:
            df[name] = 0.0
        idx_series = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, idx_choice))
        df.loc[df["Phase"] == ph, name] = (amt / 12.0) * idx_series[df["Phase"] == ph].values

    return df.fillna(0.0)


def sga_annual_by_item(s: ScenarioInputs) -> pd.DataFrame:
    dfm = sga_monthly_schedule(s).copy()
    meta = {"Month", "Year", "Phase"}
    item_cols = [c for c in dfm.columns if c not in meta]
    annual = dfm.groupby("Year", as_index=False)[item_cols].sum() if item_cols else dfm.groupby("Year", as_index=False).size()[["Year"]]
    annual["Total SG&A (COP)"] = annual[item_cols].sum(axis=1) if item_cols else 0.0
    return annual


# -----------------------------
# Depreciation (annual)
# -----------------------------
def depreciation_annual_table(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    cod_year = cod.year

    cap_df = pd.DataFrame(s.capex.lines or [])
    total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0

    pct = float(s.depreciation.pct_of_capex_depreciated or 0.0) / 100.0
    dep_years = int(s.depreciation.dep_years)
    dep_months = dep_years * 12  # Total months of depreciation

    dep_base = total_capex * pct
    monthly_dep = dep_base / dep_months if dep_months > 0 else 0.0
    annual_dep = monthly_dep * 12.0  # Full year equivalent

    # Calculate end date of depreciation period
    dep_end = _add_months(cod, dep_months)
    
    # Get all calendar years from COD to end of depreciation
    years = list(range(cod_year, dep_end.year + 1))
    
    # Calculate prorated depreciation for each calendar year
    dep_list = []
    cumulative_dep = 0.0
    
    for year in years:
        # Determine start and end dates for this calendar year
        if year == cod_year:
            # First year: from COD date to end of year
            year_start = cod
            year_end = date(year, 12, 31)
        elif year == dep_end.year:
            # Last year: from start of year to dep_end date
            year_start = date(year, 1, 1)
            year_end = dep_end
        else:
            # Full year
            year_start = date(year, 1, 1)
            year_end = date(year, 12, 31)
        
        # Calculate months of depreciation in this calendar year
        months_in_year = (year_end.year - year_start.year) * 12 + (year_end.month - year_start.month) + 1
        
        # Calculate depreciation for this year
        year_dep = monthly_dep * months_in_year
        dep_list.append(year_dep)
        cumulative_dep += year_dep
    
    df = pd.DataFrame({
        "Year": years,
        "Depreciable CAPEX (COP)": [dep_base] * len(years),
        "Depreciation (COP)": dep_list
    })
    df["Cumulative Depreciation (COP)"] = df["Depreciation (COP)"].cumsum()
    df["Remaining Book Value (COP)"] = np.maximum(dep_base - df["Cumulative Depreciation (COP)"], 0.0)
    return df

# -----------------------------
# CAPEX deduction for renewable incentives
# -----------------------------

def renewable_tax_annual_table(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    cod_year = tl["cod"].year

    cf = cashflow_annual_table(s).copy()

    # Base taxable income BEFORE incentives
    cf["Taxable Income (Pre-Incentives)"] = (
        cf["Operating CF (COP)"]
        - cf.get("Depreciation (COP)", 0.0)
    )

    # -------------------------
    # Special CAPEX deduction
    # -------------------------
    eligible_capex = _eligible_capex_for_tax(s)
    deduction_pool = eligible_capex
    max_years = int(s.renewable_tax.special_deduction_years)

    deductions = []
    remaining = deduction_pool

    for _, r in cf.iterrows():
        year = int(r["Year"])
        ti = max(0.0, float(r["Taxable Income (Pre-Incentives)"]))

        if (
            not s.renewable_tax.enable_special_deduction
            or remaining <= 0
            or year < cod_year
            or year >= cod_year + max_years
        ):
            used = 0.0
        else:
            used = min(remaining, 0.5 * ti)

        remaining -= used
        deductions.append(used)

    cf["Special Deduction Used (COP)"] = deductions
    cf["Remaining Deduction Balance (COP)"] = (
        deduction_pool - pd.Series(deductions).cumsum()
    ).clip(lower=0.0)

    # -------------------------
    # Final taxable income
    # -------------------------
    cf["Taxable Income (After Incentives)"] = (
        cf["Taxable Income (Pre-Incentives)"]
        - cf["Special Deduction Used (COP)"]
    ).clip(lower=0.0)

    tax_rate = float(s.tax.corporate_tax_rate_pct) / 100.0
    cf["Income Tax (COP)"] = cf["Taxable Income (After Incentives)"] * tax_rate

    # -------------------------
    # VAT refund (cash inflow)
    # -------------------------
    cf["VAT Refund (COP)"] = 0.0

    if s.renewable_tax.enable_vat_refund:
        if s.renewable_tax.vat_refund_mode == "percent":
            vat_amt = (
                _total_capex_from_lines(s)
                * float(s.renewable_tax.vat_pct_of_capex)
                / 100.0
            )
        else:
            vat_amt = float(s.renewable_tax.vat_fixed_cop)

        refund_year = cod_year + int(s.renewable_tax.vat_refund_year)
        cf.loc[cf["Year"] == refund_year, "VAT Refund (COP)"] = vat_amt

    return cf


# -----------------------------
# CASH FLOW (monthly + annual)
# -----------------------------
def _sga_monthly_total(s: ScenarioInputs) -> pd.DataFrame:
    df = sga_monthly_schedule(s).copy()
    meta = {"Month", "Year", "Phase"}
    item_cols = [c for c in df.columns if c not in meta]
    df["SG&A (COP)"] = df[item_cols].sum(axis=1) if item_cols else 0.0
    return df[["Month", "Year", "Phase", "SG&A (COP)"]]

def _shift_by_months(values: pd.Series, months: int) -> pd.Series:
    """Shift a monthly series forward by N months (cash received/paid later).
    Missing months are filled with 0.0."""
    months = int(max(0, months))
    if months == 0:
        return values.astype(float)
    return values.shift(months, fill_value=0.0).astype(float)

def cashflow_monthly_table(s: ScenarioInputs) -> pd.DataFrame:
    base = opex_monthly_schedule(s).copy()

    sga_m = _sga_monthly_total(s).copy()
    base = base.merge(sga_m[["Month", "SG&A (COP)"]], on="Month", how="left")
    base["SG&A (COP)"] = base["SG&A (COP)"].fillna(0.0)

    base["Total OPEX (COP)"] = base["OPEX subtotal"].fillna(0.0) + base["GMF"].fillna(0.0)

    # --- Accrual (non-cash timing) operating CF + unlevered CF (this is what debt sizing will use initially) ---
    base["Operating CF (COP)"] = (
        base["Revenue (COP)"].fillna(0.0)
        - base["Total OPEX (COP)"].fillna(0.0)
        - base["SG&A (COP)"].fillna(0.0)
    )

    # -----------------------------
    # Working Capital (simple AR/AP lags) + Cash CF
    # -----------------------------
    # You said you use s.wc (not s.working_capital)
    ar_days = float(getattr(s.wc, "ar_days", 90.0))   # revenue collected in ~90 days
    ap_days = float(getattr(s.wc, "ap_days", 60.0))   # expenses paid in ~60 days

    # Convert days to whole months (simple approximation)
    ar_lag_m = int(round(ar_days / 30.0))
    ap_lag_m = int(round(ap_days / 30.0))

    # What "cash paid expenses" means here: OPEX + SG&A (exclude CAPEX)
    base["Total OpEx+SGA (COP)"] = base["Total OPEX (COP)"].fillna(0.0) + base["SG&A (COP)"].fillna(0.0)

    # Cash timing via shifts (simple, fast, stable)
    base["Cash Collected (COP)"]   = base["Revenue (COP)"].fillna(0.0).shift(ar_lag_m, fill_value=0.0)
    base["Cash Paid OPEX (COP)"]   = base["Total OPEX (COP)"].fillna(0.0).shift(ap_lag_m, fill_value=0.0)
    base["Cash Paid SG&A (COP)"]   = base["SG&A (COP)"].fillna(0.0).shift(ap_lag_m, fill_value=0.0)

    # AR/AP balances implied by the lag model
    # AR(t) = AR(t-1) + Revenue(t) - CashCollected(t)
    # AP(t) = AP(t-1) + Expenses(t) - CashPaid(t)
    ar = []
    ap = []
    ar_prev = 0.0
    ap_prev = 0.0

    rev = base["Revenue (COP)"].fillna(0.0).to_numpy()
    cash_in = base["Cash Collected (COP)"].fillna(0.0).to_numpy()

    opex = base["Total OPEX (COP)"].fillna(0.0).to_numpy()
    sga  = base["SG&A (COP)"].fillna(0.0).to_numpy()
    cash_opex = base["Cash Paid OPEX (COP)"].fillna(0.0).to_numpy()
    cash_sga  = base["Cash Paid SG&A (COP)"].fillna(0.0).to_numpy()

    for i in range(len(base)):
        ar_now = ar_prev + rev[i] - cash_in[i]
        ap_now = ap_prev + (opex[i] + sga[i]) - (cash_opex[i] + cash_sga[i])
        ar.append(ar_now)
        ap.append(ap_now)
        ar_prev, ap_prev = ar_now, ap_now

    base["AR Balance (COP)"] = ar
    base["AP Balance (COP)"] = ap

    base["Net Working Capital (COP)"] = base["AR Balance (COP)"] - base["AP Balance (COP)"]
    base["ΔNWC (COP)"] = base["Net Working Capital (COP)"].diff().fillna(base["Net Working Capital (COP)"])

    # Update Unlevered CF (COP) to include working capital changes
    # FCF = Operating CF - CAPEX - ΔNWC
    base["Unlevered CF (COP)"] = base["Operating CF (COP)"] - base["CAPEX (COP)"].fillna(0.0) - base["ΔNWC (COP)"].fillna(0.0)
    base["Cumulative Unlevered CF (COP)"] = base["Unlevered CF (COP)"].cumsum()

    # Cash operating CF and cash unlevered CF
    base["Operating CF (Cash, COP)"] = (
        base["Cash Collected (COP)"].fillna(0.0)
        - base["Cash Paid OPEX (COP)"].fillna(0.0)
        - base["Cash Paid SG&A (COP)"].fillna(0.0)
    )

    # CAPEX assumed paid when incurred (no lag)
    # FCF = Operating CF (Cash) - CAPEX - ΔNWC
    base["Unlevered CF (Cash, COP)"] = base["Operating CF (Cash, COP)"] - base["CAPEX (COP)"].fillna(0.0) - base["ΔNWC (COP)"].fillna(0.0)
    base["Cumulative Unlevered CF (Cash, COP)"] = base["Unlevered CF (Cash, COP)"].cumsum()

    # -----------------------------
    # Debt fees (cash) — only when there is debt
    # -----------------------------
    base["Debt Fees (COP)"] = 0.0

    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * _total_capex_from_lines(s)
    debt_amt = max(debt_amt, 0.0)

    if debt_amt > 0:
        # 1) Commitment fee (annual -> spread monthly)
        cf_ann = debt_commitment_fee_annual(s).copy()  # expects columns: Year, Commitment Fee (COP)
        if (not cf_ann.empty) and ("Year" in cf_ann.columns) and ("Commitment Fee (COP)" in cf_ann.columns):
            fee_map = {int(r["Year"]): float(r["Commitment Fee (COP)"]) for _, r in cf_ann.iterrows()}
            base["Debt Fees (COP)"] += base["Year"].map(lambda y: fee_map.get(int(y), 0.0) / 12.0)

        # 2) Upfront fee at COD month (if you have upfront_fee_pct in s.debt)
        tl = build_timeline(s.timeline)
        cod_m = date(tl["cod"].year, tl["cod"].month, 1)
        upfront_pct = float(getattr(s.debt, "upfront_fee_pct", 0.0)) / 100.0
        if upfront_pct > 0:
            base.loc[base["Month"] == cod_m, "Debt Fees (COP)"] += debt_amt * upfront_pct

    # Apply fees to CASH unlevered CF (equity cash)
    base["Unlevered CF (Cash, COP)"] = base["Unlevered CF (Cash, COP)"].fillna(0.0) - base["Debt Fees (COP)"].fillna(0.0)
    base["Cumulative Unlevered CF (Cash, COP)"] = base["Unlevered CF (Cash, COP)"].cumsum()


    cols = [
        "Month", "Year", "Phase",
        "Energy (MWh)", "Revenue (COP)", "Cash Collected (COP)",
        "CAPEX (COP)", "Debt Fees (COP)",
        "Total OPEX (COP)", "Cash Paid OPEX (COP)",
        "SG&A (COP)", "Cash Paid SG&A (COP)",
        "Operating CF (COP)", "Operating CF (Cash, COP)",
        "AR Balance (COP)", "AP Balance (COP)", "Net Working Capital (COP)", "ΔNWC (COP)",
        "Unlevered CF (COP)", "Unlevered CF (Cash, COP)",
        "Cumulative Unlevered CF (COP)", "Cumulative Unlevered CF (Cash, COP)",
    ]

    # IMPORTANT: prevents KeyError if some columns are not created yet
    cols = [c for c in cols if c in base.columns]

    return base.loc[:, cols].copy()


def cashflow_annual_table(s: ScenarioInputs) -> pd.DataFrame:
    m = cashflow_monthly_table(s).copy()
    annual = (
        m.groupby("Year", as_index=False)[
            ["Energy (MWh)", "Revenue (COP)", "CAPEX (COP)", "Total OPEX (COP)", "SG&A (COP)", "Operating CF (COP)", "Unlevered CF (COP)", "ΔNWC (COP)"]
        ]
        .sum()
    )
    annual["Cumulative Unlevered CF (COP)"] = annual["Unlevered CF (COP)"].cumsum()
    return annual


def unlevered_base_cashflow_annual(s: ScenarioInputs) -> pd.DataFrame:
    a = cashflow_annual_table(s).copy()
    dep = depreciation_annual_table(s)[["Year", "Depreciation (COP)"]].copy()
    out = a.merge(dep, on="Year", how="left").fillna({"Depreciation (COP)": 0.0})
    
    # Ensure CAPEX is preserved in the output (it should already be in 'a' from cashflow_annual_table)
    # The "Unlevered CF (COP)" column already includes CAPEX: Operating CF - CAPEX
    # Verify CAPEX column exists, if not add it from the annual table
    if "CAPEX (COP)" not in out.columns and "CAPEX (COP)" in a.columns:
        out["CAPEX (COP)"] = a["CAPEX (COP)"]

    out["EBITDA (COP)"] = out["Operating CF (COP)"]
    out["Taxable Income (COP)"] = out["EBITDA (COP)"] - out["Depreciation (COP)"]

    rate = float(s.tax.corporate_tax_rate_pct) / 100.0
    allow_nol = bool(s.tax.allow_loss_carryforward)

    # ---- Renewable incentive (Law 1715 / 2099): special deduction ----
    # Pool = 50% of total investment (CAPEX). Usable up to 15 years starting year after COD.
    # Annual cap: deduction <= 50% of renta líquida (before this deduction).

    tl = build_timeline(s.timeline)
    cod_year = int(tl["cod"].year)

    total_capex = float(_total_capex_from_lines(s))  # you already have this helper for debt
    special_pool_total = 0.50 * total_capex          # 50% of total investment
    special_pool_remaining = special_pool_total

    special_years = 15
    special_start_year = cod_year + 1
    special_end_year = special_start_year + special_years - 1

    # ---- Incentives + tax calc ----
    inc = getattr(s, "incentives", RenewableIncentivesInputs())

    # Total eligible CAPEX (same base you already use for depreciation)
    cap_df = pd.DataFrame(s.capex.lines or [])
    total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0

    # Special deduction pool
    pool_total = (float(inc.special_deduction_pct_of_capex) / 100.0) * total_capex if bool(inc.enable_special_deduction) else 0.0
    pool_remaining = pool_total

    # VAT refund (cash only)
    vat_amount = (float(inc.vat_pct_of_capex) / 100.0) * total_capex + float(inc.vat_fixed_cop)
    vat_amount = max(0.0, vat_amount) if inc.vat_mode == "Refund" else 0.0

    nol = 0.0
    taxes = []
    nol_end = []

    ded_used_list = []
    ded_rem_list = []
    taxable_after_all_list = []
    vat_refund_list = []

    # Operating-year index so "refund in op Year #1" maps correctly
    # Operating year 1 = COD calendar year, operating year 2 = COD+1, etc.
    # Years before COD are not operating years (op_year = 0)
    # This calculation works regardless of when COD occurs (early/late in year, any calendar year)
    max_years = int(max(1, inc.special_deduction_years))

    for _, r in out.iterrows():
        year_i = int(r["Year"])
        # Calculate operating year: 1 for COD year, 2 for COD+1, etc. 0 for pre-COD years
        # This works for any COD date (e.g., COD in Jan 2029 = op_year 1 for 2029, op_year 2 for 2030)
        op_year = (year_i - cod_year + 1) if year_i >= cod_year else 0

        # Base taxable income = EBITDA - Depreciation
        ti_base = float(r["Taxable Income (COP)"])

        # Apply CAPEX deduction FIRST (to minimize taxes)
        # The deduction applies in the FIRST year with positive taxable income (regardless of timeline)
        ded_used = 0.0
        if (
            bool(inc.enable_special_deduction) 
            and (1 <= op_year <= max_years)  # Only during operating years 1 through max_years
            and ti_base > 0  # Apply deduction if there's any taxable income (before NOL)
            and pool_remaining > 0  # Only if pool hasn't been exhausted
        ):
            # Annual cap is based on taxable income BEFORE NOL (as per tax law)
            # The cap is a percentage of taxable income, ensuring we don't exceed the annual limit
            annual_cap = (float(inc.special_deduction_max_pct_of_taxable_income) / 100.0) * ti_base
            ded_used = min(pool_remaining, annual_cap)  # Use the smaller of: remaining pool or annual cap
            pool_remaining -= ded_used

        # Taxable income after CAPEX deduction: EBITDA - Depreciation - CAPEX Tax Deduction
        ti_after_ded = ti_base - ded_used

        # Loss carryforward: accumulate losses (negative taxable income after deduction)
        # If negative, add to NOL. If positive, use NOL to offset it.
        if allow_nol:
            if ti_after_ded < 0:
                # Loss: add to NOL carryforward
                nol = nol + (-ti_after_ded)
                taxable_after_all = 0.0  # No taxable income when there's a loss
            else:
                # Profit: use NOL to offset taxable income
                used = min(nol, ti_after_ded)
                nol -= used
                taxable_after_all = max(ti_after_ded - used, 0.0)
        else:
            # No NOL: taxable income is what's left after deduction (can be negative)
            taxable_after_all = max(ti_after_ded, 0.0)
            if ti_after_ded < 0:
                nol = nol + (-ti_after_ded)  # Still track losses even if NOL not allowed (for display)

        taxes_payable = taxable_after_all * rate
        taxes.append(taxes_payable)
        nol_end.append(nol)

        ded_used_list.append(ded_used)
        ded_rem_list.append(pool_remaining)
        taxable_after_all_list.append(taxable_after_all)

        # VAT refund: one-time cash inflow in operating year N (default year 1)
        vat_refund = vat_amount if (vat_amount > 0 and op_year == int(inc.vat_refund_year_index)) else 0.0
        vat_refund_list.append(vat_refund)

    out["CAPEX Tax Deduction (COP)"] = ded_used_list
    out["Loss Carryforward End (COP)"] = nol_end  # Moved before Taxable Income
    out["Special Deduction Remaining (COP)"] = ded_rem_list
    # Taxable Income is what remains after: EBITDA - Depreciation - CAPEX Deduction - NOL
    out["Taxable Income (COP)"] = taxable_after_all_list
    out["Taxable Income After Incentives (COP)"] = taxable_after_all_list

    out["Taxes Payable (COP)"] = taxes

    # After-tax CF: taxes reduce cash; VAT refund adds cash (does not affect taxes)
    out["VAT Refund (COP)"] = vat_refund_list
    # Ensure "Unlevered CF (COP)" includes CAPEX and ΔNWC (it should from cashflow_annual_table)
    # If for some reason CAPEX is missing, recalculate: Operating CF - CAPEX - ΔNWC
    if "CAPEX (COP)" in out.columns:
        # Verify Unlevered CF includes CAPEX and ΔNWC: it should be Operating CF - CAPEX - ΔNWC
        # ΔNWC is already included from cashflow_annual_table, but ensure it's preserved
        if "ΔNWC (COP)" in out.columns:
            out["Unlevered CF (COP)"] = out["Operating CF (COP)"] - out["CAPEX (COP)"].fillna(0.0) - out["ΔNWC (COP)"].fillna(0.0)
        else:
            out["Unlevered CF (COP)"] = out["Operating CF (COP)"] - out["CAPEX (COP)"].fillna(0.0)
    # Rename pre-tax column for clarity (this already includes CAPEX and ΔNWC from the calculation above)
    out["Unlevered CF Pre-tax (COP)"] = out["Unlevered CF (COP)"]
    out["Unlevered CF After Tax (COP)"] = out["Unlevered CF Pre-tax (COP)"] - out["Taxes Payable (COP)"] + out["VAT Refund (COP)"]

    return out


# -----------------------------
# IRR / Payback helpers
# -----------------------------
def _irr_bisection(cashflows: List[float], low=-0.95, high=5.0, tol=1e-7, max_iter=200) -> float:
    def npv(rate: float) -> float:
        denom_base = 1.0 + rate
        if denom_base <= 1e-12:
            return float("inf")
        total = 0.0
        for i, cf in enumerate(cashflows):
            if i == 0:
                total += cf
                continue
            den = denom_base ** i
            if den == 0.0:
                return float("inf")
            total += cf / den
        return total

    if len(cashflows) < 2:
        return float("nan")

    f_low = npv(low)
    f_high = npv(high)

    tries = 0
    while f_low * f_high > 0 and tries < 20:
        high *= 2
        f_high = npv(high)
        tries += 1

    if f_low * f_high > 0:
        return float("nan")

    mid = float("nan")
    for _ in range(max_iter):
        mid = (low + high) / 2
        f_mid = npv(mid)
        if abs(f_mid) < tol:
            return mid
        if f_low * f_mid <= 0:
            high = mid
            f_high = f_mid
        else:
            low = mid
            f_low = f_mid
    return mid


def _payback_months(months: List[date], unlevered_cf: List[float]) -> float:
    """
    Calculate payback period in months using cumulative cash flows.
    Payback is the time (in months) from the start until cumulative CF becomes positive.
    Uses linear interpolation in the period where cumulative crosses zero.
    
    Args:
        months: List of dates (not used in calculation, but kept for consistency)
        unlevered_cf: List of monthly cash flows (should be after-tax for after-tax payback)
    
    Returns:
        Payback period in months (as float, can be fractional), or NaN if never pays back
    """
    cum = 0.0
    for i, cf in enumerate(unlevered_cf):
        prev = cum  # Cumulative at end of previous period
        cum += cf   # Cumulative at end of current period
        if cum >= 0 and i > 0:  # Cumulative has crossed zero
            if cf == 0:
                # Edge case: zero cash flow, payback is at end of this period
                return float(i)
            # Linear interpolation: what fraction of current period's CF is needed
            # to go from prev (negative) to 0?
            # frac = amount needed / cash flow in period = (0 - prev) / cf
            frac = (0 - prev) / cf
            # Payback = (i-1) full periods + fraction of current period
            return (i - 1) + frac
    return float("nan")  # Never pays back


# -----------------------------
# Debt engine (annual sculpting + fees)
# -----------------------------
def _total_capex_from_lines(s: ScenarioInputs) -> float:
    cap_df = pd.DataFrame(s.capex.lines or [])
    if cap_df.empty or "Amount_COP" not in cap_df.columns:
        return 0.0
    return float(pd.to_numeric(cap_df["Amount_COP"], errors="coerce").fillna(0.0).sum())

def _eligible_capex_for_tax(s: ScenarioInputs) -> float:
    total_capex = _total_capex_from_lines(s)
    pct = max(0.0, min(100.0, s.renewable_tax.special_deduction_pct_of_capex))
    return total_capex * pct / 100.0

def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)

def debt_commitment_fee_annual(s: ScenarioInputs) -> pd.DataFrame:
    """
    Commitment fee on undrawn debt during construction:
    Commitment rate = margin * commitment_fee_pct_of_margin
    Applied monthly on undrawn = TotalDebt - CumulativeDrawn
    """
    tl = build_timeline(s.timeline)
    cod_m = date(tl["cod"].year, tl["cod"].month, 1)

    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex
    debt_amt = max(debt_amt, 0.0)

    # Monthly CAPEX schedule up to COD month (exclude COD month)
    capm = capex_monthly_schedule(s).copy()
    capm["Month"] = pd.to_datetime(capm["Month"])
    capm = capm[capm["Month"] < pd.to_datetime(cod_m)].copy()
    if capm.empty or debt_amt <= 0:
        return pd.DataFrame({"Year": [], "Commitment Fee (COP)": []})

    # Assume debt draws pro-rata with CAPEX spend during dev+construction (pre-COD)
    draw_m = capm.copy()
    draw_m["Draw (COP)"] = (float(s.debt.debt_pct_of_capex) / 100.0) * pd.to_numeric(draw_m["CAPEX (COP)"], errors="coerce").fillna(0.0)

    draw_m["Cum Draw (COP)"] = draw_m["Draw (COP)"].cumsum()
    draw_m["Undrawn (COP)"] = np.maximum(debt_amt - draw_m["Cum Draw (COP)"], 0.0)

    margin = float(s.debt.margin_pct) / 100.0
    commit_mult = float(s.debt.commitment_fee_pct_of_margin) / 100.0
    commit_rate_annual = margin * commit_mult
    commit_rate_monthly = commit_rate_annual / 12.0

    draw_m["Commitment Fee (COP)"] = draw_m["Undrawn (COP)"] * commit_rate_monthly
    draw_m["Year"] = draw_m["Month"].dt.year

    ann = draw_m.groupby("Year", as_index=False)[["Commitment Fee (COP)"]].sum()
    return ann


def debt_schedule_annual(s: ScenarioInputs) -> pd.DataFrame:
    """
    Annual debt schedule starting at COD year for tenor years.
    Amortization:
      - Sculpted to Target DSCR,
      - Equal principal, or
      - Typical amortization (fixed payment/annuity)
    DSCR calculated as: Operating CF / Debt Service
    """
    tl = build_timeline(s.timeline)
    cod_year = tl["cod"].year

    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex
    debt_amt = max(debt_amt, 0.0)

    tenor = int(s.debt.tenor_years)
    tenor = max(5, min(10, tenor))
    grace = int(s.debt.grace_years)
    grace = max(0, min(tenor - 1, grace))

    years = list(range(cod_year, cod_year + tenor))

    # Operating CF annual from model (EBITDA proxy)
    a = cashflow_annual_table(s).copy()
    opcf_map = {int(r["Year"]): float(r["Operating CF (COP)"]) for _, r in a.iterrows()}
    cf = [float(opcf_map.get(y, 0.0)) for y in years]

    all_in = (float(s.debt.base_rate_pct) + float(s.debt.margin_pct)) / 100.0
    target = float(s.debt.target_dscr) if float(s.debt.target_dscr) > 0 else 1.20

    # NEW: balloon target (COP). % of original debt you want left outstanding at maturity.
    balloon_pct = float(getattr(s.debt, "balloon_pct", 0.0)) / 100.0
    balloon_target = debt_amt * balloon_pct
    balloon_target = max(0.0, min(balloon_target, debt_amt))

    outstanding = debt_amt
    rows = []
    
    # Pre-calculate fixed payment for typical amortization (if applicable)
    fixed_payment_typical = 0.0
    if s.debt.amortization_type == "Typical amortization":
        # Calculate fixed payment to amortize (debt_amt - balloon_target) over (tenor - grace) years
        principal_to_amortize = debt_amt - balloon_target
        amortization_years = tenor - grace
        if amortization_years > 0 and all_in > 0 and principal_to_amortize > 1e-9:
            r = all_in
            n = amortization_years
            if (1.0 + r) ** n > 1.0:
                fixed_payment_typical = principal_to_amortize * (r * (1.0 + r) ** n) / ((1.0 + r) ** n - 1.0)
            else:
                fixed_payment_typical = principal_to_amortize / n

    for i, y in enumerate(years, start=1):
        operating_cf = float(cf[i - 1])
        interest_due = outstanding * all_in  # Interest due on current outstanding

        if outstanding <= 1e-9:
            principal = 0.0
            debt_service = 0.0
            dscr = float("nan")
            rows.append({
                "Year": y,
                "Operating CF (COP)": operating_cf,
                "Interest (COP)": 0.0,
                "Principal (COP)": principal,
                "Debt Service (COP)": debt_service,
                "DSCR": dscr,
                "Outstanding End (COP)": 0.0,
            })
            outstanding = 0.0
            continue

        if i <= grace:
            principal = 0.0
            # During grace period:
            # - No principal payments
            # - Interest is capitalized (added to principal) if cash flow is insufficient
            # - Interest is only paid in cash if there's sufficient operating cash flow
            if operating_cf >= interest_due:
                # Sufficient cash flow: pay all interest in cash
                interest_paid = interest_due
                interest_capitalized = 0.0
            else:
                # Insufficient cash flow: pay what we can, capitalize the rest
                interest_paid = max(0.0, operating_cf)  # Pay what we can from operating CF (can't be negative)
                interest_capitalized = interest_due - interest_paid  # Capitalize the rest
                outstanding = outstanding + interest_capitalized  # Add capitalized interest to principal
            
            # Interest paid in cash (for levered CF calculation)
            interest = interest_paid
        else:
            # After grace period: pay full interest in cash
            interest = interest_due
            
            # Calculate remaining amortization years (tenor minus grace, minus years already amortized)
            # Years already amortized = (i - 1 - grace) since grace years don't count
            years_amortized = max(0, i - 1 - grace)
            remaining_years = max(tenor - grace - years_amortized, 1)

            if s.debt.amortization_type == "Equal principal":
                # Amortize evenly down to balloon_target over remaining years
                principal = max(0.0, (outstanding - balloon_target) / remaining_years)

                # Final year: force ending outstanding = balloon_target
                if i == tenor:
                    principal = max(0.0, outstanding - balloon_target)

            elif s.debt.amortization_type == "Typical amortization":
                # Fixed payment (annuity) amortization
                # Use pre-calculated fixed payment, but only during amortization period (after grace)
                if fixed_payment_typical > 0:
                    # Payment = Interest + Principal, so Principal = Payment - Interest
                    principal = max(0.0, fixed_payment_typical - interest)
                    
                    # Ensure we don't pay more than outstanding (minus balloon)
                    principal = min(principal, max(0.0, outstanding - balloon_target))
                else:
                    principal = 0.0
                
                # Final year: force ending outstanding = balloon_target
                if i == tenor:
                    principal = max(0.0, outstanding - balloon_target)

            else:
                # Sculpt to target DSCR (CFADS proxy = Operating CF)
                max_ds = operating_cf / target if target > 0 else 0.0
                principal = max(0.0, max_ds - interest)

                # Constrain: don't pay more than what would amortize evenly over remaining years
                # This ensures we respect the tenor (don't pay off too early)
                max_principal_by_tenor = max(0.0, (outstanding - balloon_target) / remaining_years) if remaining_years > 0 else outstanding - balloon_target
                principal = min(principal, max(0.0, max_principal_by_tenor))

                # Do not amortize below balloon target (unless final year)
                if i < tenor:
                    principal = min(principal, max(0.0, outstanding - balloon_target))
                else:
                    principal = max(0.0, outstanding - balloon_target)

        principal = min(principal, outstanding)
        debt_service = interest + principal
        dscr = (operating_cf / debt_service) if debt_service > 0 else float("nan")

        outstanding = outstanding - principal

        rows.append({
            "Year": y,
            "Operating CF (COP)": operating_cf,
            "Interest (COP)": interest,
            "Principal (COP)": principal,
            "Debt Service (COP)": debt_service,
            "DSCR": dscr,
            "Outstanding End (COP)": outstanding,
        })

    df = pd.DataFrame(rows)

    balloon = float(df["Outstanding End (COP)"].iloc[-1]) if len(df) else 0.0
    df["Balloon at Maturity (COP)"] = 0.0
    if len(df):
        df.loc[df.index[-1], "Balloon at Maturity (COP)"] = balloon

    # Upfront fee (paid at COD, show in first year)
    upfront_fee = debt_amt * (float(s.debt.upfront_fee_bps) / 10_000.0)
    df["Upfront Fee (COP)"] = 0.0
    if len(df):
        df.loc[df.index[0], "Upfront Fee (COP)"] = upfront_fee

    # Covenant / lock-up flags
    min_covenant = float(s.debt.min_dscr_covenant)
    lockup = float(s.debt.lockup_dscr)
    df["Lock-up? (DSCR < lock-up)"] = df["DSCR"].apply(lambda x: bool(np.isfinite(x) and x < lockup))
    df["Breach? (DSCR < min)"] = df["DSCR"].apply(lambda x: bool(np.isfinite(x) and x < min_covenant))

    return df


# -----------------------------
# Levered Cash Flow (Equity Cash Flow)
# -----------------------------
def _recalculate_taxes_with_interest(s: ScenarioInputs, unlevered_df: pd.DataFrame, interest_by_year: dict) -> pd.DataFrame:
    """
    Recalculate tax benefits and taxes for levered cash flow.
    Taxable income = EBITDA - Depreciation - Interest (interest is tax-deductible).
    Then recalculate CAPEX deduction, loss carryforward, and taxes payable.
    """
    out = unlevered_df.copy()
    
    rate = float(s.tax.corporate_tax_rate_pct) / 100.0
    allow_nol = bool(s.tax.allow_loss_carryforward)
    
    # Get incentives
    inc = getattr(s, "incentives", RenewableIncentivesInputs())
    tl = build_timeline(s.timeline)
    cod_year = int(tl["cod"].year)
    
    # Total eligible CAPEX for deduction pool
    cap_df = pd.DataFrame(s.capex.lines or [])
    total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0
    
    # Special deduction pool
    pool_total = (float(inc.special_deduction_pct_of_capex) / 100.0) * total_capex if bool(inc.enable_special_deduction) else 0.0
    pool_remaining = pool_total
    
    # VAT refund (same as unlevered)
    vat_amount = (float(inc.vat_pct_of_capex) / 100.0) * total_capex + float(inc.vat_fixed_cop)
    vat_amount = max(0.0, vat_amount) if inc.vat_mode == "Refund" else 0.0
    
    nol = 0.0
    taxes = []
    nol_end = []
    ded_used_list = []
    taxable_after_all_list = []
    
    max_years = int(max(1, inc.special_deduction_years))
    
    for _, r in out.iterrows():
        year_i = int(r["Year"])
        op_year = (year_i - cod_year + 1) if year_i >= cod_year else 0
        
        # Levered taxable income = EBITDA - Depreciation - Interest
        ebitda = float(r.get("EBITDA (COP)", 0.0))
        depreciation = float(r.get("Depreciation (COP)", 0.0))
        interest = float(interest_by_year.get(year_i, 0.0))
        
        ti_base = ebitda - depreciation - interest  # Interest reduces taxable income
        
        # Apply CAPEX deduction FIRST (to minimize taxes)
        ded_used = 0.0
        if (
            bool(inc.enable_special_deduction) 
            and (1 <= op_year <= max_years) 
            and ti_base > 0  # Apply deduction if there's any taxable income (before NOL)
            and pool_remaining > 0
        ):
            annual_cap = (float(inc.special_deduction_max_pct_of_taxable_income) / 100.0) * ti_base
            ded_used = min(pool_remaining, annual_cap)
            pool_remaining -= ded_used
        
        # Taxable income after CAPEX deduction
        ti_after_ded = ti_base - ded_used
        
        # Apply NOL on remaining taxable income (after deduction)
        if allow_nol:
            if ti_after_ded < 0:
                nol = nol + (-ti_after_ded)
                taxable_after_all = 0.0
            else:
                used = min(nol, ti_after_ded)
                nol -= used
                taxable_after_all = max(ti_after_ded - used, 0.0)
        else:
            taxable_after_all = max(ti_after_ded, 0.0)
            if ti_after_ded < 0:
                nol = nol + (-ti_after_ded)  # Still track losses even if NOL not allowed
        
        taxes_payable = taxable_after_all * rate
        taxes.append(taxes_payable)
        nol_end.append(nol)
        ded_used_list.append(ded_used)
        taxable_after_all_list.append(taxable_after_all)
    
    # Update columns with recalculated values
    out["Levered Taxable Income (COP)"] = taxable_after_all_list
    out["Levered CAPEX Tax Deduction (COP)"] = ded_used_list
    out["Levered Loss Carryforward End (COP)"] = nol_end
    out["Levered Taxes Payable (COP)"] = taxes
    
    # Calculate Net Income After Tax = Taxable Income - Taxes Payable
    # (Taxable Income is already after all deductions and NOL)
    out["Levered Net Income After Tax (COP)"] = out["Levered Taxable Income (COP)"] - out["Levered Taxes Payable (COP)"]
    
    return out


def levered_cashflow_annual(s: ScenarioInputs) -> pd.DataFrame:
    """
    Calculate levered (equity) cash flow after-tax.
    Recalculates taxes with interest expense included (interest is tax-deductible).
    If debt is disabled or not enabled, assumes no debt (returns unlevered CF).
    """
    # Start with unlevered base cash flow (pre-tax)
    unlevered = unlevered_base_cashflow_annual(s).copy()
    
    # Check if debt is enabled
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex if debt_enabled else 0.0
    debt_amt = max(debt_amt, 0.0)
    
    # Initialize debt-related columns
    unlevered["Debt Draw (COP)"] = 0.0
    unlevered["Interest (COP)"] = 0.0
    unlevered["Principal (COP)"] = 0.0
    unlevered["Debt Service (COP)"] = 0.0
    unlevered["Debt Fees (COP)"] = 0.0
    
    if debt_enabled and debt_amt > 0:
        # Get debt schedule (annual interest, principal, upfront fee)
        ds = debt_schedule_annual(s).copy()
        
        if not ds.empty and "Year" in ds.columns:
            # Merge debt schedule - drop initialized columns first to avoid suffix conflicts
            merge_cols = ["Year"]
            for col in ["Interest (COP)", "Principal (COP)", "Upfront Fee (COP)"]:
                if col in ds.columns:
                    merge_cols.append(col)
                    # Drop initialized column if it exists to avoid merge conflicts
                    if col in unlevered.columns:
                        unlevered = unlevered.drop(columns=[col])
            
            if len(merge_cols) > 1:  # More than just "Year"
                unlevered = unlevered.merge(ds[merge_cols], on="Year", how="left")
                
                # Ensure columns exist and fill NaN values (after merge, columns should exist)
                for col in ["Interest (COP)", "Principal (COP)"]:
                    if col in unlevered.columns:
                        unlevered[col] = unlevered[col].fillna(0.0)
                    else:
                        unlevered[col] = 0.0
                
                unlevered["Debt Service (COP)"] = unlevered["Interest (COP)"] + unlevered["Principal (COP)"]
                
                # Upfront fee
                if "Upfront Fee (COP)" in unlevered.columns:
                    unlevered["Debt Fees (COP)"] = unlevered["Debt Fees (COP)"] + unlevered["Upfront Fee (COP)"].fillna(0.0)
            
            # Recalculate taxes with interest expense included
            # Create interest by year map
            interest_by_year = {}
            for _, row in unlevered.iterrows():
                year = int(row["Year"])
                interest_by_year[year] = float(row.get("Interest (COP)", 0.0))
            
            # Recalculate tax benefits and taxes with interest expense
            unlevered = _recalculate_taxes_with_interest(s, unlevered, interest_by_year)
            
            # Debt draws: pro-rata with CAPEX during construction (up to and including COD)
            tl = build_timeline(s.timeline)
            cod_year = tl["cod"].year
            cod_m = date(tl["cod"].year, tl["cod"].month, 1)
            
            # Get monthly CAPEX schedule
            capm = capex_monthly_schedule(s).copy()
            capm["Year"] = capm["Month"].apply(lambda d: d.year)
            
            # Calculate debt draws (pro-rata with CAPEX, up to and including COD month)
            if not capm.empty:
                # Include CAPEX up to and including COD month
                up_to_cod = capm[capm["Month"] <= cod_m].copy()
                if not up_to_cod.empty:
                    up_to_cod["Debt Draw (COP)"] = (float(s.debt.debt_pct_of_capex) / 100.0) * pd.to_numeric(up_to_cod["CAPEX (COP)"], errors="coerce").fillna(0.0)
                    
                    # Ensure total debt draws equal debt amount (adjust last draw if needed)
                    total_draws = up_to_cod["Debt Draw (COP)"].sum()
                    if total_draws > 0 and abs(total_draws - debt_amt) > 1.0:  # Allow small rounding differences
                        # Adjust proportionally or add remainder to COD month
                        adjustment = debt_amt - total_draws
                        if cod_m in up_to_cod["Month"].values:
                            # Add adjustment to COD month
                            cod_idx = up_to_cod[up_to_cod["Month"] == cod_m].index[0]
                            up_to_cod.loc[cod_idx, "Debt Draw (COP)"] += adjustment
                        else:
                            # Scale proportionally
                            if total_draws > 0:
                                scale_factor = debt_amt / total_draws
                                up_to_cod["Debt Draw (COP)"] = up_to_cod["Debt Draw (COP)"] * scale_factor
                    
                    # Aggregate to annual
                    annual_draws = up_to_cod.groupby("Year", as_index=False)["Debt Draw (COP)"].sum()
                    # Merge into unlevered (drop the initialized column first to avoid suffix conflicts)
                    if "Debt Draw (COP)" in unlevered.columns:
                        unlevered = unlevered.drop(columns=["Debt Draw (COP)"])
                    unlevered = unlevered.merge(annual_draws, on="Year", how="left")
                    # Ensure column exists and fill NaN
                    if "Debt Draw (COP)" in unlevered.columns:
                        unlevered["Debt Draw (COP)"] = unlevered["Debt Draw (COP)"].fillna(0.0)
                    else:
                        unlevered["Debt Draw (COP)"] = 0.0
                # If no CAPEX up to COD, Debt Draw stays at initialized 0.0
            
            # Commitment fees (during construction)
            cf_ann = debt_commitment_fee_annual(s).copy()
            if not cf_ann.empty and "Year" in cf_ann.columns and "Commitment Fee (COP)" in cf_ann.columns:
                unlevered = unlevered.merge(cf_ann[["Year", "Commitment Fee (COP)"]], on="Year", how="left")
                unlevered["Debt Fees (COP)"] = unlevered["Debt Fees (COP)"] + unlevered["Commitment Fee (COP)"].fillna(0.0)
    
    # Calculate levered (equity) cash flow
    # If debt is enabled, use recalculated taxes (with interest expense)
    # If no debt, use unlevered after-tax CF
    if debt_enabled and debt_amt > 0:
        # Levered CF = Unlevered Pre-tax CF + Debt Draws - Interest - Principal - Debt Fees - Levered Taxes + VAT Refund
        unlevered["Levered CF (After-tax, COP)"] = (
            unlevered["Unlevered CF Pre-tax (COP)"].fillna(0.0)
            + unlevered["Debt Draw (COP)"].fillna(0.0)
            - unlevered["Interest (COP)"].fillna(0.0)
            - unlevered["Principal (COP)"].fillna(0.0)
            - unlevered["Debt Fees (COP)"].fillna(0.0)
            - unlevered["Levered Taxes Payable (COP)"].fillna(0.0)
            + unlevered["VAT Refund (COP)"].fillna(0.0)
        )
    else:
        # No debt: levered CF = unlevered after-tax CF
        # Also set levered tax columns to match unlevered (no interest, so taxes are same)
        unlevered["Levered CF (After-tax, COP)"] = unlevered["Unlevered CF After Tax (COP)"].fillna(0.0)
        unlevered["Levered Taxable Income (COP)"] = unlevered["Taxable Income (COP)"].fillna(0.0)
        unlevered["Levered CAPEX Tax Deduction (COP)"] = unlevered["CAPEX Tax Deduction (COP)"].fillna(0.0)
        unlevered["Levered Loss Carryforward End (COP)"] = unlevered["Loss Carryforward End (COP)"].fillna(0.0)
        unlevered["Levered Taxes Payable (COP)"] = unlevered["Taxes Payable (COP)"].fillna(0.0)
        unlevered["Levered Net Income After Tax (COP)"] = unlevered["Levered Taxable Income (COP)"] - unlevered["Levered Taxes Payable (COP)"]
    
    unlevered["Cumulative Levered CF (COP)"] = unlevered["Levered CF (After-tax, COP)"].cumsum()
    
    return unlevered


def levered_cashflow_monthly(s: ScenarioInputs) -> pd.DataFrame:
    """
    Calculate monthly levered (equity) cash flow after-tax.
    Starts with monthly unlevered CF, adds debt draws, subtracts debt service and fees.
    """
    # Start with monthly cash flow table
    monthly = cashflow_monthly_table(s).copy()
    
    # Get annual levered CF to extract debt components
    annual_levered = levered_cashflow_annual(s).copy()
    
    # Initialize debt columns
    monthly["Debt Draw (COP)"] = 0.0
    monthly["Interest (COP)"] = 0.0
    monthly["Principal (COP)"] = 0.0
    monthly["Debt Service (COP)"] = 0.0
    monthly["Debt Fees (COP)"] = 0.0
    
    # Check if debt is enabled
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex if debt_enabled else 0.0
    debt_amt = max(debt_amt, 0.0)
    
    if debt_enabled and debt_amt > 0:
        tl = build_timeline(s.timeline)
        cod_m = date(tl["cod"].year, tl["cod"].month, 1)
        
        # Debt draws: pro-rata with CAPEX during construction (pre-COD)
        capm = capex_monthly_schedule(s).copy()
        if not capm.empty:
            capm["Debt Draw (COP)"] = (float(s.debt.debt_pct_of_capex) / 100.0) * pd.to_numeric(capm["CAPEX (COP)"], errors="coerce").fillna(0.0)
            # Only draws before COD
            capm.loc[capm["Month"] >= cod_m, "Debt Draw (COP)"] = 0.0
            # Merge into monthly (drop initialized column first to avoid suffix conflicts)
            if "Debt Draw (COP)" in monthly.columns:
                monthly = monthly.drop(columns=["Debt Draw (COP)"])
            monthly = monthly.merge(capm[["Month", "Debt Draw (COP)"]], on="Month", how="left")
            # Ensure column exists and fill NaN
            if "Debt Draw (COP)" in monthly.columns:
                monthly["Debt Draw (COP)"] = monthly["Debt Draw (COP)"].fillna(0.0)
            else:
                monthly["Debt Draw (COP)"] = 0.0
        
        # Debt service: spread annual interest and principal evenly across months
        if not annual_levered.empty:
            # Create a map of annual debt service components
            interest_map = {int(r["Year"]): float(r["Interest (COP)"]) for _, r in annual_levered.iterrows()}
            principal_map = {int(r["Year"]): float(r["Principal (COP)"]) for _, r in annual_levered.iterrows()}
            
            monthly["Interest (COP)"] = monthly["Year"].map(lambda y: interest_map.get(int(y), 0.0) / 12.0)
            monthly["Principal (COP)"] = monthly["Year"].map(lambda y: principal_map.get(int(y), 0.0) / 12.0)
            monthly["Debt Service (COP)"] = monthly["Interest (COP)"] + monthly["Principal (COP)"]
            
            # Only apply debt service during operation (after COD)
            monthly.loc[monthly["Month"] < cod_m, "Interest (COP)"] = 0.0
            monthly.loc[monthly["Month"] < cod_m, "Principal (COP)"] = 0.0
            monthly.loc[monthly["Month"] < cod_m, "Debt Service (COP)"] = 0.0
        
        # Debt fees: upfront fee at COD, commitment fees during construction
        # Initialize Debt Fees column if not exists
        if "Debt Fees (COP)" not in monthly.columns:
            monthly["Debt Fees (COP)"] = 0.0
        
        # Upfront fee
        upfront_fee = debt_amt * (float(s.debt.upfront_fee_bps) / 10_000.0)
        monthly.loc[monthly["Month"] == cod_m, "Debt Fees (COP)"] = upfront_fee
        
        # Commitment fees: get from annual and spread
        cf_ann = debt_commitment_fee_annual(s).copy()
        if not cf_ann.empty and "Year" in cf_ann.columns:
            cf_map = {int(r["Year"]): float(r["Commitment Fee (COP)"]) for _, r in cf_ann.iterrows()}
            monthly["Commitment Fee (COP)"] = monthly["Year"].map(lambda y: cf_map.get(int(y), 0.0) / 12.0)
            monthly["Debt Fees (COP)"] = monthly["Debt Fees (COP)"].fillna(0.0) + monthly["Commitment Fee (COP)"].fillna(0.0)
        
        monthly["Debt Fees (COP)"] = monthly["Debt Fees (COP)"].fillna(0.0)
    
    # Calculate monthly levered CF from annual levered CF
    # Use the annual levered CF (which has taxes recalculated with interest) and spread proportionally
    if not annual_levered.empty:
        # Get annual levered CF (this already includes all debt components and recalculated taxes)
        levered_cf_annual_map = {int(r["Year"]): float(r["Levered CF (After-tax, COP)"]) for _, r in annual_levered.iterrows()}
        
        # Calculate monthly levered CF by prorating annual levered CF
        # For operation months: prorate based on operating months
        # For construction months: use monthly unlevered pre-tax CF + debt draws - debt service - fees
        tl = build_timeline(s.timeline)
        cod = tl["cod"]
        end_op = tl["end_op"]
        
        monthly["Levered CF (After-tax, COP)"] = 0.0
        
        for _, row in monthly.iterrows():
            month_date = row["Month"]
            year = month_date.year
            phase = _phase_for_month(tl, month_date)
            
            if phase == "Operation" and year in levered_cf_annual_map:
                # Operation months: prorate annual levered CF
                if year == cod.year:
                    operating_months = 13 - cod.month
                elif year == end_op.year:
                    operating_months = end_op.month
                else:
                    operating_months = 12
                
                if operating_months > 0:
                    annual_levered_cf = levered_cf_annual_map[year]
                    monthly.loc[monthly.index == row.name, "Levered CF (After-tax, COP)"] = annual_levered_cf / operating_months
            else:
                # Construction/non-operation months: calculate from components
                # Use unlevered pre-tax CF + debt draws - debt service - fees
                # (No taxes during construction, so this is simpler)
                monthly.loc[monthly.index == row.name, "Levered CF (After-tax, COP)"] = (
                    row.get("Unlevered CF (COP)", 0.0)
                    + row.get("Debt Draw (COP)", 0.0)
                    - row.get("Debt Service (COP)", 0.0)
                    - row.get("Debt Fees (COP)", 0.0)
                )
    else:
        # Fallback: use unlevered CF + debt components
        monthly["Levered CF (After-tax, COP)"] = (
            monthly["Unlevered CF (COP)"].fillna(0.0)
            + monthly["Debt Draw (COP)"].fillna(0.0)
            - monthly["Debt Service (COP)"].fillna(0.0)
            - monthly["Debt Fees (COP)"].fillna(0.0)
        )
    
    monthly["Cumulative Levered CF (COP)"] = monthly["Levered CF (After-tax, COP)"].cumsum()
    
    return monthly


# -----------------------------
# APP UI
# -----------------------------
st.set_page_config(page_title="Delphi Utility-Scale Model", page_icon="⚡", layout="wide")

# Check and display Kaleido status (for PDF chart export)
if not KALEIDO_AVAILABLE:
    import sys
    python_path = sys.executable
    st.warning(f"⚠️ Kaleido is not installed. PDF charts will not be available. Install with: `{python_path} -m pip install kaleido`")

# Display Delphi logo at top
logo_path = _load_logo_image()
if logo_path:
    # Center the logo
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image(logo_path, width=400)
    st.markdown("<br>", unsafe_allow_html=True)  # Add spacing
elif os.path.exists("assets/images/LOGO V2.pdf"):
    # PDF exists but conversion failed - show helpful message
    st.info("💡 Logo file found but conversion failed. Please install PyMuPDF: `pip install PyMuPDF`")

st.title("Delphi Utility-Scale Project Model (COP inputs, COP/USD outputs)")
st.caption(f"Projects + scenarios stored at: {PROJECTS_FILE}")

# User Manual Section
with st.expander("How to Read This Model / Cómo Leer Este Modelo", expanded=False):
    st.markdown("""
# HOW TO READ THIS MODEL
### Utility-Scale Power Financial Model

## 🇬🇧 ENGLISH

### 1. What this model is
This app is a decision-support tool to evaluate the economics of a utility-scale power project across its full life cycle.
It allows users to test commercial structures, macroeconomic assumptions, costs, tax benefits, and financing within a consistent framework and compare scenarios quickly.
It is designed for structuring, discussion, and iteration, not as a final bank-grade model.

### 2. What problem it solves
At early and mid stages of project development, the key questions are:
- Which commercial structure works best?
- Which assumptions drive value over time?
- Where are the main risks across the project life cycle?
This model helps answer those questions clearly and consistently.

### 3. Project life-cycle logic (from idea to operation)
The system thinks about a project from its very conception, not just from operations.
Key milestones include:
- Development Phase: permits, land, interconnection, and contracts are secured.
- RTB – Ready to Build: environmental and construction licenses obtained; construction can begin immediately.
- COD – Commercial Operation Date: approvals obtained (regulatory, environmental, grid connection, RETIE, etc.) and the project can begin producing power the next day.
All timelines, costs, and revenues are aligned around these milestones.

### 4. Macroeconomic assumptions
The Macroeconomic tab allows the user to select common macro inputs that affect long-term performance:
- CPI (inflation)
- PPI (producer price index)
- User-defined growth rates
These assumptions are used to grow or adjust specific variables over time (prices, costs, etc.).
Some variables may respond differently depending on the macro input selected.

### 5. Power generation
The Power Generation section uses external technical inputs from tools such as PVSyst or other energy simulation software.
These inputs define:
- Expected annual generation levels (e.g., P50, P90)
- Annual degradation rates
The financial model does not simulate energy; it consumes these technical outputs and translates them into revenues.

### 6. Power resources / commercial structure
Power Resources defines how energy is sold over the life of the project:
- Whether the project is contracted, partially contracted, or merchant
- The type of PPA contract applied over time (or lack thereof)
Common structures include pay-as-generated, capped PPAs, fixed blocks, or collar-type contracts.
Important (Beta): At this stage, contract structures may be captured conceptually. Unless explicitly stated, revenue calculations may still assume 100% pay-as-generated while the commercial structure inputs are prepared for future versions.

### 7. Revenue logic: pricing vs contract structure
The model separates:
A) PPA Price Schedule (how price evolves)
- Indexed PPA for a defined term, then Merchant price
- Annual PPA price table (year-by-year)
B) PPA Contract Structure (how energy is paid)
This defines how much energy is paid, capped, settled, or structured under the contract.

### 8. CAPEX
CAPEX defines the project's investment profile.
CAPEX items are entered for visualization and structuring purposes.
The model allows up to three alternative CAPEX investment scenarios for comparison.

### 9. OPEX and SG&A
The user can define operating cost metrics, add/remove cost items, and adjust escalation logic where applicable.

### 10. Depreciation
The user defines the depreciation term and profile.
In countries like Colombia, renewable projects may benefit from accelerated depreciation, which can be defined by the user according to applicable regulations.

### 11. Renewable tax benefits
This section reflects country-specific renewable tax incentives.
For Colombia, the user can define applicable tax benefits, timing, and applicability according to regulation.

### 12. How to read the outputs
- Revenues reflect generation and pricing assumptions
- EBITDA shows operating profitability
- Cash flows reflect CAPEX, taxes, and financing (if enabled)
- IRR / NPV reflect overall project attractiveness
If results look unexpected, review: (1) timeline, (2) generation, (3) price/macro assumptions, (4) CAPEX timing.

### 13. What this model is / is not
This model is:
- A scenario and structuring tool
- A framework for commercial discussions
- A way to compare alternatives consistently
This model is not:
- A final bank model
- A legal or contractual document
- A substitute for technical or regulatory diligence

---

## 🇪🇸 ESPAÑOL

### 1. Qué es este modelo
Esta app es una herramienta de apoyo a la toma de decisiones para evaluar proyectos de generación eléctrica utility-scale a lo largo de todo su ciclo de vida.
Permite probar estructuras comerciales, supuestos macroeconómicos, costos, beneficios tributarios y financiamiento dentro de un marco consistente y comparar escenarios rápidamente.
No está diseñada como un modelo bancario final.

### 2. Qué problema resuelve
En etapas tempranas e intermedias, las preguntas clave son:
- ¿Qué estructura comercial funciona mejor?
- ¿Qué supuestos generan mayor impacto en el tiempo?
- ¿Dónde se concentran los principales riesgos?
Este modelo ayuda a responder esas preguntas.

### 3. Lógica de ciclo de vida del proyecto
El sistema concibe el proyecto desde su origen, no solo desde la operación.
Hitos clave:
- Desarrollo
- RTB – Ready to Build: proyecto con licencias ambientales y de construcción, listo para iniciar obras.
- COD – Fecha de Operación Comercial: proyecto con aprobaciones (regulatorias, ambientales, conexión, RETIE, etc.) para comenzar a generar energía.

### 4. Supuestos macroeconómicos
La pestaña Macroeconómica permite seleccionar supuestos como:
- IPC (inflación)
- IPP (índice de precios al productor)
- Ajustes definidos por el usuario
Estos supuestos se usan para proyectar el crecimiento de variables en el tiempo.
Algunas variables pueden reaccionar de forma distinta según el insumo seleccionado.

### 5. Generación de energía
Utiliza insumos técnicos externos (por ejemplo PVSyst) para definir:
- Producción anual esperada (P50, P90)
- Degradación anual
El modelo financiero no simula energía; utiliza estos resultados como insumo.

### 6. Recursos de energía / estructura comercial
Define cómo se vende la energía durante la vida del proyecto:
- Contratado, parcialmente contratado o merchant
- Tipo de contrato PPA aplicado en el tiempo
Importante (Beta): por ahora, las estructuras contractuales se capturan a nivel conceptual. Salvo que se indique lo contrario, los cálculos asumen 100% pague lo generado.

### 7. CAPEX
Permite definir rubros de inversión y hasta tres alternativas de CAPEX para comparación.

### 8. OPEX y SG&A
El usuario puede definir métricas, agregar/eliminar rubros y ajustar el crecimiento de costos.

### 9. Depreciación
Permite definir plazo y esquema de depreciación. En Colombia, puede existir depreciación acelerada configurable por el usuario.

### 10. Beneficios tributarios renovables
Sección para beneficios tributarios aplicables según la regulación colombiana, definidos por el usuario.

### 11. Qué es / qué no es este modelo
Es: herramienta de escenarios y estructuración.
No es: modelo bancario final ni reemplazo de debida diligencia.
""")

db = _load_db()

# Sidebar: project & scenario management
with st.sidebar:
    st.header("Project & Scenario")

    projects = sorted(list(db.get("projects", {}).keys()))
    project_name = st.selectbox("Project", ["(New project)"] + projects, index=0)

    if project_name == "(New project)":
        new_project = st.text_input("New project name", value="")
        if st.button("Create project", type="primary", width='stretch'):
            if new_project.strip():
                db.setdefault("projects", {}).setdefault(new_project.strip(), {"scenarios": {}})
                _save_db(db)
                st.rerun()
            else:
                st.warning("Enter a project name.")
        st.stop()

    proj = db["projects"].setdefault(project_name, {"scenarios": {}})

    scen_names = sorted(list(proj.get("scenarios", {}).keys()))
    scenario_name = st.selectbox("Scenario", ["(New scenario)"] + scen_names, index=0)

    if scenario_name == "(New scenario)":
        new_s = st.text_input("New scenario name", value="Base")
        if st.button("Create scenario", type="primary", width='stretch'):
            nm = new_s.strip() or "Base"
            proj["scenarios"][nm] = _scenario_to_dict(ScenarioInputs(name=nm))
            _save_db(db)
            st.rerun()
        st.stop()

    # Load scenario
    s = _scenario_from_dict(proj["scenarios"][scenario_name])
    s.name = scenario_name

    st.divider()
    cdel1, cdel2 = st.columns(2)
    with cdel1:
        if st.button("Save scenario", width='stretch'):
            proj["scenarios"][scenario_name] = _scenario_to_dict(s)
            _save_db(db)
            st.success("Saved.")
    with cdel2:
        if st.button("Delete scenario", width='stretch'):
            try:
                del proj["scenarios"][scenario_name]
                _save_db(db)
                st.rerun()
            except Exception:
                st.error("Could not delete.")

    st.divider()
    st.subheader("Compare")
    compare_scenarios = st.multiselect("Select scenarios", scen_names, default=[])


# Tabs (top-down)
tab_overview, tab_macro, tab_timeline, tab_gen, tab_rev, tab_capex, tab_opex, tab_sga, tab_dep, tab_incent, tab_ucf, tab_debt, tab_levered, tab_compare, tab_sensitivity, tab_summary = st.tabs(
    [
        "Project Overview",
        "A) Macroeconomic",
        "B) Timeline",
        "C) Power Generation",
        "D) Power Revenues",
        "E) CAPEX",
        "F) OPEX",
        "G) SG&A",
        "H) Depreciation",
        "I) Renewable tax benefits",
        "J) Unlevered Base Cash Flow",
        "K) Debt & Covenants",
        "L) Levered Cash Flow",
        "M) Compare",
        "N) Sensitivity",
        "O) Summary",
    ]
)

# -----------------------------
# Project Overview
# -----------------------------
with tab_overview:
    st.subheader("Project Overview")
    st.caption("Descriptive metadata for this project/scenario. These fields are for reference only and do not affect any calculations.")
    
    # Create safe scenario identifier for widget keys (no project_name to avoid key churn)
    safe_scenario = re.sub(r"[^a-zA-Z0-9_-]+", "_", scenario_name or "default")
    
    c1, c2 = st.columns(2)
    with c1:
        s.project_overview.project_name = st.text_input(
            "Project Name",
            value=s.project_overview.project_name,
            key=f"overview_project_name_{safe_scenario}"
        )
    with c2:
        s.project_overview.country = st.text_input(
            "Country",
            value=s.project_overview.country,
            key=f"overview_country_{safe_scenario}"
        )
    
    c3, c4 = st.columns(2)
    with c3:
        s.project_overview.region_department = st.text_input(
            "Region / Department",
            value=s.project_overview.region_department,
            key=f"overview_region_{safe_scenario}"
        )
    with c4:
        s.project_overview.technology = st.text_input(
            "Technology",
            value=s.project_overview.technology,
            key=f"overview_technology_{safe_scenario}"
        )
    
    s.project_overview.installed_capacity_mw = st.number_input(
        "Installed Capacity (MW)",
        value=float(s.project_overview.installed_capacity_mw),
        min_value=0.0,
        step=0.1,
        format="%.2f",
        key=f"overview_capacity_{safe_scenario}"
    )
    
    s.project_overview.short_description = st.text_area(
        "Short Project Description",
        value=s.project_overview.short_description,
        height=100,
        key=f"overview_description_{safe_scenario}"
    )
    
    s.project_overview.optional_notes = st.text_area(
        "Optional Notes",
        value=s.project_overview.optional_notes,
        height=150,
        key=f"overview_notes_{safe_scenario}"
    )
    
    # Auto-save when fields change (scenario is saved automatically when switching tabs/scenarios)
    # The scenario data structure is updated in-place, so changes persist when the scenario is saved

# -----------------------------
# A) Macro
# -----------------------------
with tab_macro:
    st.subheader("Macroeconomic inputs (annual rates, %)")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        s.macro.col_cpi = st.number_input("Colombian CPI (%)", value=float(s.macro.col_cpi), step=0.1, format="%.2f")
    with c2:
        s.macro.col_ppi = st.number_input("Colombian PPI (%)", value=float(s.macro.col_ppi), step=0.1, format="%.2f")
    with c3:
        s.macro.us_cpi = st.number_input("US CPI (%)", value=float(s.macro.us_cpi), step=0.1, format="%.2f")
    with c4:
        s.macro.custom_index_rate = st.number_input("Custom index (%)", value=float(s.macro.custom_index_rate), step=0.1, format="%.2f")

    st.markdown("#### FX (COP per USD)")
    fx1, fx2, fx3 = st.columns([1.2, 1, 1])
    with fx1:
        s.macro.fx_cop_per_usd_start = st.number_input("Starting FX (COP/USD)", value=float(s.macro.fx_cop_per_usd_start), step=10.0, format="%.0f")
        st.caption(f"Formatted: {_fmt_cop(s.macro.fx_cop_per_usd_start)} / USD")
    with fx2:
        s.macro.fx_method = st.selectbox("FX method", ["Inflation differential (PPP approx.)", "Flat"], index=0 if s.macro.fx_method != "Flat" else 1)
    with fx3:
        s.macro.fx_flat = st.number_input("Flat FX (if selected)", value=float(s.macro.fx_flat or s.macro.fx_cop_per_usd_start), step=10.0, format="%.0f")
        st.caption(f"Formatted: {_fmt_cop(s.macro.fx_flat)} / USD")

    st.info("FX path default uses a simple PPP approximation: FX grows with (Col CPI / US CPI). You can switch to Flat FX.")


# -----------------------------
# B) Timeline
# -----------------------------
with tab_timeline:
    st.subheader("Project timeline (Development → CAPEX → Operation)")

    c1, c2, c3 = st.columns(3)
    with c1:
        s.timeline.start_date = st.date_input("Project start date", value=_parse_date_iso(s.timeline.start_date)).isoformat()
    with c2:
        s.timeline.dev_months = int(st.number_input("Development duration (months)", value=int(s.timeline.dev_months), min_value=0, step=1, format="%d"))
    with c3:
        s.timeline.capex_months = int(st.number_input("Construction/CAPEX duration (months)", value=int(s.timeline.capex_months), min_value=0, step=1, format="%d"))

    s.timeline.operation_years = int(st.number_input("Operation duration (years)", value=int(s.timeline.operation_years), min_value=1, step=1, format="%d"))

    tl = build_timeline(s.timeline)
    _metric_row(
        [
            ("Start", str(tl["start"])),
            ("RTB (end of dev)", str(tl["rtb"])),
            ("COD (start ops)", str(tl["cod"])),
            ("End of operation", str(tl["end_op"])),
        ]
    )

    st.markdown("#### Timeline visual (calendar, years)")

    if not PLOTLY_AVAILABLE or px is None:
        st.warning("⚠️ Charts unavailable - Plotly installation issue. All other functionality works normally.")
    else:
        gantt = pd.DataFrame(
            [
                {"Stage": "Development", "Start": date(tl["start"].year, tl["start"].month, 1), "Finish": date(tl["rtb"].year, tl["rtb"].month, 1)},
                {"Stage": "Construction", "Start": date(tl["rtb"].year, tl["rtb"].month, 1), "Finish": date(tl["cod"].year, tl["cod"].month, 1)},
                {"Stage": "Operation", "Start": date(tl["cod"].year, tl["cod"].month, 1), "Finish": date(tl["end_op"].year, tl["end_op"].month, 1)},
            ]
        )

        fig = px.timeline(
            gantt,
            x_start="Start",
            x_end="Finish",
            y="Stage",
            color="Stage",
            category_orders={"Stage": ["Development", "Construction", "Operation"]},
        )
        fig.update_yaxes(autorange="reversed")
        fig.update_xaxes(dtick="M12", tickformat="%Y")
        fig.update_layout(height=280, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
        st.plotly_chart(fig, width='stretch')

    # CAPEX Task Schedule (Visual Reference Only)
    st.markdown("#### CAPEX Task Schedule (Visual Reference - Not Used in Calculations)")
    st.info("ℹ️ This schedule is for visualization and reference only. It does not affect any financial calculations.")
    
    # Initialize task schedule in session state if not exists (keyed by scenario name to keep separate per scenario)
    scenario_key = f"capex_tasks_{s.name}"
    if scenario_key not in st.session_state:
        st.session_state[scenario_key] = []
    
    # Task management UI
    with st.expander("Manage CAPEX Tasks", expanded=False):
        # Add new task
        col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
        with col1:
            new_task_name = st.text_input("Task Name", key=f"new_task_name_{scenario_key}", placeholder="e.g., Site Preparation")
        with col2:
            new_task_start = st.date_input("Start Date", value=tl["rtb"], key=f"new_task_start_{scenario_key}")
        with col3:
            new_task_duration = st.number_input("Duration (months)", min_value=1, value=3, step=1, key=f"new_task_duration_{scenario_key}")
        with col4:
            new_task_depends_on = st.text_input("Depends On (task #)", key=f"new_task_depends_on_{scenario_key}", placeholder="e.g., 1,2", help="Enter task numbers separated by commas")
        
        col_add, col_clear = st.columns([1, 1])
        with col_add:
            if st.button("Add Task", key=f"add_task_btn_{scenario_key}"):
                if new_task_name:
                    task_num = len(st.session_state[scenario_key]) + 1
                    from datetime import timedelta
                    new_task_end = new_task_start + timedelta(days=30 * new_task_duration - 1)
                    
                    task = {
                        "Task #": task_num,
                        "Task Name": new_task_name,
                        "Start Date": new_task_start.isoformat(),
                        "Duration (months)": int(new_task_duration),
                        "End Date": new_task_end.isoformat(),
                        "Depends On": new_task_depends_on.strip() if new_task_depends_on else ""
                    }
                    st.session_state[scenario_key].append(task)
                    st.rerun()
                else:
                    st.warning("Please enter a task name.")
        with col_clear:
            if st.button("Clear All Tasks", key=f"clear_tasks_btn_{scenario_key}"):
                st.session_state[scenario_key] = []
                st.rerun()
    
    # Display and edit tasks table
    if st.session_state[scenario_key]:
        tasks_df = pd.DataFrame(st.session_state[scenario_key])
        
        # Convert date strings to date objects for editing
        tasks_df["Start Date"] = pd.to_datetime(tasks_df["Start Date"]).dt.date
        tasks_df["End Date"] = pd.to_datetime(tasks_df["End Date"]).dt.date
        
        edited_tasks = st.data_editor(
            tasks_df,
            width='stretch',
            hide_index=True,
            num_rows="fixed",
            key=f"capex_tasks_editor_{scenario_key}",
            column_config={
                "Task #": st.column_config.NumberColumn("Task #", format="%d", disabled=True),
                "Task Name": st.column_config.TextColumn("Task Name"),
                "Start Date": st.column_config.DateColumn("Start Date"),
                "Duration (months)": st.column_config.NumberColumn("Duration (months)", min_value=1, step=1, format="%d"),
                "End Date": st.column_config.DateColumn("End Date"),
                "Depends On": st.column_config.TextColumn("Depends On (task #s)", help="Enter task numbers separated by commas, e.g., 1,2"),
            },
        )
        
        # Update end dates based on start date and duration, and renumber tasks
        from datetime import timedelta
        updated_tasks = []
        for idx, row in edited_tasks.iterrows():
            start_date = pd.to_datetime(row["Start Date"])
            duration_months = int(row["Duration (months)"])
            calculated_end = start_date + timedelta(days=30 * duration_months - 1)
            
            task_num = int(row["Task #"])
            updated_tasks.append({
                "Task #": task_num,
                "Task Name": str(row["Task Name"]),
                "Start Date": start_date.date().isoformat(),
                "Duration (months)": duration_months,
                "End Date": calculated_end.date().isoformat(),
                "Depends On": str(row["Depends On"]) if pd.notna(row["Depends On"]) else ""
            })
        
        # Save back to session state
        st.session_state[scenario_key] = updated_tasks
        
        # Delete task functionality
        if len(updated_tasks) > 0:
            task_nums = [t["Task #"] for t in updated_tasks]
            selected_task_to_delete = st.selectbox(
                "Delete Task",
                options=[""] + task_nums,
                format_func=lambda x: f"Task {x}" if x else "Select task to delete...",
                key=f"delete_task_select_{scenario_key}"
            )
            if selected_task_to_delete and st.button("Delete Selected Task", key=f"delete_task_btn_{scenario_key}"):
                st.session_state[scenario_key] = [t for t in updated_tasks if t["Task #"] != selected_task_to_delete]
                # Renumber remaining tasks
                for i, task in enumerate(st.session_state[scenario_key], 1):
                    task["Task #"] = i
                st.rerun()
        
        # Create Gantt chart for tasks
        gantt_tasks = []
        for task in st.session_state[scenario_key]:
            start = pd.to_datetime(task["Start Date"]).date()
            end = pd.to_datetime(task["End Date"]).date()
            gantt_tasks.append({
                "Task": f"Task {task['Task #']}: {task['Task Name']}",
                "Start": date(start.year, start.month, 1),
                "Finish": date(end.year, end.month, 1),
                "Depends On": task.get("Depends On", "")
            })
        
        if gantt_tasks:
            tasks_gantt_df = pd.DataFrame(gantt_tasks)
            
            # Create Gantt chart
            if PLOTLY_AVAILABLE and px is not None:
                fig_tasks = px.timeline(
                    tasks_gantt_df,
                    x_start="Start",
                    x_end="Finish",
                    y="Task",
                    color="Task",
                    title="CAPEX Task Schedule"
                )
                fig_tasks.update_yaxes(autorange="reversed")
                fig_tasks.update_xaxes(dtick="M1", tickformat="%b %Y")
                fig_tasks.update_layout(
                    height=max(300, len(gantt_tasks) * 40),
                    margin=dict(l=10, r=10, t=40, b=10),
                    legend_title_text="",
                    showlegend=False
                )
                st.plotly_chart(fig_tasks, width='stretch')
            else:
                st.warning("⚠️ Chart unavailable")
            
            # Show dependency relationships
            has_dependencies = any(task.get("Depends On", "").strip() for task in st.session_state[scenario_key])
            if has_dependencies:
                st.markdown("**Task Dependencies:**")
                dep_text = []
                for task in st.session_state[scenario_key]:
                    if task.get("Depends On", "").strip():
                        dep_text.append(f"• **Task {task['Task #']}** ({task['Task Name']}) depends on: {task['Depends On']}")
                if dep_text:
                    st.markdown("\n".join(dep_text))
                else:
                    st.info("No dependencies defined.")
    else:
        st.info("No CAPEX tasks defined. Use the 'Manage CAPEX Tasks' section above to add tasks.")


# -----------------------------
# C) Generation
# -----------------------------
with tab_gen:
    st.subheader("Power generation inputs")
    
    # Show current revenue mode and note that it affects prices/revenues here
    revenue_mode_display = s.revenue_mode if hasattr(s, 'revenue_mode') else "Standard PPA Parameters"
    st.info(f"ℹ️ **Revenue mode:** {revenue_mode_display}. Prices and revenues shown below are calculated based on the revenue mode selected in the 'Power Revenues' tab. Change the revenue mode there to update prices and revenues here.")

    c1, c2, c3 = st.columns(3)
    with c1:
        s.generation.mwac = st.number_input("Capacity (MWac)", value=float(s.generation.mwac), step=1.0, format="%.2f")
    with c2:
        s.generation.mwp = st.number_input("Capacity (MWp)", value=float(s.generation.mwp), step=1.0, format="%.2f")
    with c3:
        s.generation.degradation_pct = st.number_input("Annual degradation (%/yr)", value=float(s.generation.degradation_pct), step=0.05, format="%.2f")

    c4, c5, c6, c7 = st.columns(4)
    with c4:
        s.generation.p50_mwh_yr = st.number_input("P50 (MWh/year)", value=float(s.generation.p50_mwh_yr), step=1000.0, format="%.0f")
    with c5:
        s.generation.p75_mwh_yr = st.number_input("P75 (MWh/year)", value=float(s.generation.p75_mwh_yr), step=1000.0, format="%.0f")
    with c6:
        s.generation.p90_mwh_yr = st.number_input("P90 (MWh/year)", value=float(s.generation.p90_mwh_yr), step=1000.0, format="%.0f")
    with c7:
        s.generation.production_choice = st.selectbox("Production choice", ["P50", "P75", "P90"], index=["P50", "P75", "P90"].index(s.generation.production_choice))

    # Force recalculation by calling operating_year_table with current scenario state
    op = operating_year_table(s)
    
    # Check if prices are zero when in manual mode (might indicate prices not set)
    if revenue_mode_display == "Manual annual series":
        # Check if prices dictionary has any non-zero values
        r = s.revenue2
        has_prices = any(v > 0 for v in r.prices_constant_cop_per_kwh.values()) if r.prices_constant_cop_per_kwh else False
        zero_prices = op["Price (COP/kWh)"].fillna(0.0) == 0.0
        if zero_prices.any():
            if not has_prices:
                st.warning(f"⚠️ **Warning:** Prices are showing as 0.0 because no prices have been entered yet. Please go to the 'Power Revenues' tab and enter prices for all operating years in the Manual annual series table. The table will update automatically after you save the prices.")
            else:
                # Show which operating years have prices vs which are missing
                missing_years = []
                for idx, row in op.iterrows():
                    if row["Price (COP/kWh)"] == 0.0:
                        # Try to determine which operating year this corresponds to
                        year = int(row["Year"])
                        # This is approximate - we can't easily reverse-engineer the operating year from calendar year
                        missing_years.append(year)
                if missing_years:
                    st.warning(f"⚠️ **Warning:** Some prices are showing as 0.0 for years {missing_years[:5]}{'...' if len(missing_years) > 5 else ''}. Please check that prices are entered for all operating years in the 'Power Revenues' tab. After entering prices, navigate back to this tab to see the updated values.")
    
    if PLOTLY_AVAILABLE and px is not None:
        fig = px.line(op, x="Year", y="Energy (MWh)")
        fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")
    
    # File upload section
    _render_file_upload_section(s, project_name, scenario_name, "generation", "Reference Files (PVSyst / Simulation / Documents)")


# -----------------------------
# D) Revenues
# -----------------------------
with tab_rev:
    st.subheader("Power revenues (indexed, annual)")
    
    st.warning("⚠️ **Important:** The revenue mode selected here affects prices and revenues throughout the entire model, including the Power Generation tab, Unlevered Cash Flow, and all downstream calculations.")

    s.revenue_mode = st.radio("Revenue mode", ["Standard PPA Parameters", "Manual annual series"], horizontal=True, index=0 if s.revenue_mode == "Standard PPA Parameters" else 1)

    if s.revenue_mode == "Standard PPA Parameters":
        r = s.revenue1
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            r.ppa_price_cop_per_kwh = st.number_input("PPA price at COD (COP/kWh, current COP)", value=float(r.ppa_price_cop_per_kwh), step=1.0, format="%.2f")
        with c2:
            r.ppa_term_years = int(st.number_input("PPA term (years)", value=int(r.ppa_term_years), min_value=1, step=1, format="%d"))
        with c3:
            r.merchant_price_cop_per_kwh = st.number_input("Post-term merchant price (COP/kWh)", value=float(r.merchant_price_cop_per_kwh), step=1.0, format="%.2f")
        with c4:
            r.indexation = st.selectbox("Indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(r.indexation) if r.indexation in INDEX_CHOICES else 0)
        
        st.divider()
        
        # PPA Contract Type Selector (independent from price schedule)
        st.markdown("#### PPA Contract Structure")
        st.caption("Contract structure is independent from price schedule. Price schedule (PPA + Merchant or Manual table) remains unchanged.")
        
        # Contract type options with Spanish/English labels
        contract_type_options = [
            ("pay_as_generated_100", "Pague lo generado (100%) / Pay-as-Generated (100%)"),
            ("pay_as_generated_cap", "Pague lo generado hasta un tope / Pay-as-Generated with Cap"),
            ("fixed_block", "Bloque de energía fijo / Fixed Energy Block (Baseload)"),
            ("floor_cap", "Piso y techo / Floor & Cap (Collar)"),
        ]
        
        # Find current selection index
        current_type = s.ppa_contract.type_id
        current_index = next((i for i, (tid, _) in enumerate(contract_type_options) if tid == current_type), 0)
        
        # Dropdown for contract type
        selected_label = st.selectbox(
            "PPA Contract Type",
            options=[label for _, label in contract_type_options],
            index=current_index,
            key="ppa_contract_type_selector"
        )
        
        # Update type_id and labels based on selection
        selected_type_id = next(tid for tid, label in contract_type_options if label == selected_label)
        s.ppa_contract.type_id = selected_type_id
        if selected_type_id == "pay_as_generated_100":
            s.ppa_contract.ui_label_es = "Pague lo generado (100%)"
            s.ppa_contract.ui_label_en = "Pay-as-Generated (100%)"
        elif selected_type_id == "pay_as_generated_cap":
            s.ppa_contract.ui_label_es = "Pague lo generado hasta un tope"
            s.ppa_contract.ui_label_en = "Pay-as-Generated with Cap"
        elif selected_type_id == "fixed_block":
            s.ppa_contract.ui_label_es = "Bloque de energía fijo"
            s.ppa_contract.ui_label_en = "Fixed Energy Block (Baseload)"
        elif selected_type_id == "floor_cap":
            s.ppa_contract.ui_label_es = "Piso y techo"
            s.ppa_contract.ui_label_en = "Floor & Cap (Collar)"
        
        # Info note
        st.info("ℹ️ **PPA structure captured for V2; calculations currently assume pay-as-generated. Price schedule (PPA + Merchant or Manual table) remains unchanged and independent from contract structure.**")
        
        # Conditional fields based on contract type
        if selected_type_id == "pay_as_generated_cap":
            st.markdown("##### Type 2: Pay-as-Generated with Cap Parameters")
            cap_col1, cap_col2 = st.columns(2)
            with cap_col1:
                s.ppa_contract.cap_mode = st.selectbox(
                    "Cap Mode",
                    options=["% de P50", "MWh/año"],
                    index=0 if s.ppa_contract.cap_mode == "% de P50" else 1,
                    key="ppa_cap_mode"
                )
            with cap_col2:
                if s.ppa_contract.cap_mode == "% de P50":
                    s.ppa_contract.cap_value = st.number_input(
                        "Cap Value (% of P50)",
                        value=float(s.ppa_contract.cap_value),
                        min_value=0.0,
                        max_value=200.0,
                        step=1.0,
                        format="%.2f",
                        key="ppa_cap_value_pct"
                    )
                else:
                    s.ppa_contract.cap_value = st.number_input(
                        "Cap Value (MWh/year)",
                        value=float(s.ppa_contract.cap_value),
                        min_value=0.0,
                        step=1000.0,
                        format="%.0f",
                        key="ppa_cap_value_mwh"
                    )
            
            excess_col1, excess_col2 = st.columns(2)
            with excess_col1:
                s.ppa_contract.excess_treatment = st.selectbox(
                    "Excess Treatment",
                    options=["Spot", "% del PPA", "0 (no pagado)"],
                    index=["Spot", "% del PPA", "0 (no pagado)"].index(s.ppa_contract.excess_treatment) if s.ppa_contract.excess_treatment in ["Spot", "% del PPA", "0 (no pagado)"] else 0,
                    key="ppa_excess_treatment"
                )
            with excess_col2:
                if s.ppa_contract.excess_treatment == "% del PPA":
                    s.ppa_contract.excess_discount_pct = st.number_input(
                        "Excess Discount (% of PPA)",
                        value=float(s.ppa_contract.excess_discount_pct),
                        min_value=0.0,
                        max_value=200.0,
                        step=1.0,
                        format="%.2f",
                        key="ppa_excess_discount"
                    )
        
        elif selected_type_id == "fixed_block":
            st.markdown("##### Type 3: Fixed Energy Block Parameters")
            block_col1, block_col2 = st.columns(2)
            with block_col1:
                s.ppa_contract.block_mode = st.selectbox(
                    "Block Mode",
                    options=["% de P50", "MWh/año"],
                    index=0 if s.ppa_contract.block_mode == "% de P50" else 1,
                    key="ppa_block_mode"
                )
            with block_col2:
                if s.ppa_contract.block_mode == "% de P50":
                    s.ppa_contract.block_value = st.number_input(
                        "Block Value (% of P50)",
                        value=float(s.ppa_contract.block_value),
                        min_value=0.0,
                        max_value=200.0,
                        step=1.0,
                        format="%.2f",
                        key="ppa_block_value_pct"
                    )
                else:
                    s.ppa_contract.block_value = st.number_input(
                        "Block Value (MWh/year)",
                        value=float(s.ppa_contract.block_value),
                        min_value=0.0,
                        step=1000.0,
                        format="%.0f",
                        key="ppa_block_value_mwh"
                    )
            
            settlement_col1, settlement_col2 = st.columns(2)
            with settlement_col1:
                s.ppa_contract.settlement_price_mode = st.selectbox(
                    "Settlement Price Mode",
                    options=["Spot (promedio)", "Fijo"],
                    index=0 if s.ppa_contract.settlement_price_mode == "Spot (promedio)" else 1,
                    key="ppa_settlement_mode"
                )
            with settlement_col2:
                if s.ppa_contract.settlement_price_mode == "Fijo":
                    # Default to current PPA price if not set
                    if s.ppa_contract.settlement_price_fixed <= 0.0:
                        s.ppa_contract.settlement_price_fixed = float(r.ppa_price_cop_per_kwh)
                    s.ppa_contract.settlement_price_fixed = st.number_input(
                        "Settlement Price Fixed (COP/kWh)",
                        value=float(s.ppa_contract.settlement_price_fixed),
                        min_value=0.0,
                        step=1.0,
                        format="%.4f",
                        key="ppa_settlement_fixed"
                    )
        
        elif selected_type_id == "floor_cap":
            st.markdown("##### Type 4: Floor & Cap (Collar) Parameters")
            collar_col1, collar_col2 = st.columns(2)
            with collar_col1:
                s.ppa_contract.floor_pct_of_p50_revenue = st.number_input(
                    "Floor (% of P50 Revenue)",
                    value=float(s.ppa_contract.floor_pct_of_p50_revenue),
                    min_value=0.0,
                    max_value=200.0,
                    step=1.0,
                    format="%.2f",
                    key="ppa_floor_pct"
                )
            with collar_col2:
                s.ppa_contract.cap_pct_of_p50_revenue = st.number_input(
                    "Cap (% of P50 Revenue)",
                    value=float(s.ppa_contract.cap_pct_of_p50_revenue),
                    min_value=0.0,
                    max_value=200.0,
                    step=1.0,
                    format="%.2f",
                    key="ppa_cap_pct"
                )
        
        # Validate and show warnings
        validation_warnings = _validate_ppa_contract_config(s.ppa_contract, float(r.ppa_price_cop_per_kwh))
        if validation_warnings:
            for warning in validation_warnings:
                st.warning(warning)
        
        st.divider()
    else:
        r = s.revenue2
        c1, c2 = st.columns([1, 2])
        with c1:
            r.indexation = st.selectbox("Indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(r.indexation) if r.indexation in INDEX_CHOICES else 0)
        with c2:
            st.info("Enter constant COP/kWh by operating year (Year 1 = first year at COD).")

        prices_df = pd.DataFrame({"OpYear": list(range(1, int(s.timeline.operation_years) + 1))})
        prices_df["COP_per_kWh_constant"] = prices_df["OpYear"].apply(lambda i: float(r.prices_constant_cop_per_kwh.get(int(i), 0.0)))
        edited = st.data_editor(
            prices_df,
            width='stretch',
            hide_index=True,
            num_rows="fixed",
            key="manual_prices_editor",  # Add key to ensure proper state management
            column_config={
                "OpYear": st.column_config.NumberColumn("Operating year", format="%d", disabled=True),
                "COP_per_kWh_constant": st.column_config.NumberColumn("Price (COP/kWh, constant)", step=1.0, format="%.2f"),
            },
        )
        # Save prices immediately after editing
        r.prices_constant_cop_per_kwh = {int(row.OpYear): float(row.COP_per_kWh_constant) for row in edited.itertuples(index=False)}
        
        # Show confirmation that prices are saved
        if len(r.prices_constant_cop_per_kwh) > 0:
            non_zero_prices = sum(1 for v in r.prices_constant_cop_per_kwh.values() if v > 0)
            if non_zero_prices > 0:
                st.success(f"✓ Prices saved for {non_zero_prices} operating year(s). Scenario will be auto-saved and prices will persist when you reload.")

    op = operating_year_table(s)

    c1, c2 = st.columns(2)
    with c1:
        if PLOTLY_AVAILABLE and px is not None:
            fig1 = px.bar(op, x="Year", y="Energy (MWh)")
            fig1.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig1, width='stretch')
        else:
            st.warning("⚠️ Chart unavailable")
    with c2:
        if PLOTLY_AVAILABLE and px is not None:
            fig2 = px.line(op, x="Year", y="Revenue (COP)")
            fig2.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig2, width='stretch')
        else:
            st.warning("⚠️ Chart unavailable")

    disp = op.copy()
    disp = _df_format_money(disp, ["Energy (MWh)", "Price (COP/kWh)", "Revenue (COP)"], decimals=0)
    disp = _transpose_annual_table(disp)
    st.dataframe(disp, width='stretch', hide_index=True)
    
    # File upload section
    _render_file_upload_section(s, project_name, scenario_name, "revenues", "Reference Files (PPA / Contracts / Documents)")


# -----------------------------
# E) CAPEX
# -----------------------------
with tab_capex:
    st.subheader("CAPEX (COP)")

    s.capex.distribution = st.selectbox("Construction spend distribution", CAPEX_DISTS, index=CAPEX_DISTS.index(s.capex.distribution) if s.capex.distribution in CAPEX_DISTS else 0)

    capex_df = pd.DataFrame(s.capex.lines or [])
    for col in ["Item", "Amount_COP", "Phase"]:
        if col not in capex_df.columns:
            capex_df[col] = "" if col != "Amount_COP" else 0.0
    capex_df = capex_df[["Item", "Amount_COP", "Phase"]].copy()
    capex_df["Phase"] = capex_df["Phase"].where(capex_df["Phase"].isin(CAPEX_PHASES), "Construction")

    edited = st.data_editor(
        capex_df,
        width='stretch',
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Item": st.column_config.TextColumn("CAPEX line item"),
            "Amount_COP": st.column_config.NumberColumn("Amount (COP)", min_value=0.0, step=1_000_000.0, format="%.0f"),
            "Phase": st.column_config.SelectboxColumn("Phase", options=CAPEX_PHASES),
        },
    )
    s.capex.lines = edited.to_dict(orient="records")

    total_capex = float(pd.to_numeric(edited["Amount_COP"], errors="coerce").fillna(0.0).sum())
    mwac = float(s.generation.mwac or 0.0)
    mwp = float(s.generation.mwp or 0.0)
    capex_per_mwac = (total_capex / mwac) if mwac > 0 else np.nan
    capex_per_mwp = (total_capex / mwp) if mwp > 0 else np.nan

    c1, c2, c3 = st.columns(3)
    c1.metric("Total CAPEX (COP)", _fmt_cop(total_capex))
    c2.metric("CAPEX / MWac (COP)", _fmt_cop(capex_per_mwac) if np.isfinite(capex_per_mwac) else "—")
    c3.metric("CAPEX / MWp (COP)", _fmt_cop(capex_per_mwp) if np.isfinite(capex_per_mwp) else "—")

    st.markdown("#### CAPEX breakdown (by line item)")
    capex_pie = edited.copy()
    capex_pie["Item"] = capex_pie["Item"].fillna("").astype(str).str.strip()
    capex_pie["Amount_COP"] = pd.to_numeric(capex_pie["Amount_COP"], errors="coerce").fillna(0.0)
    capex_pie = capex_pie[capex_pie["Amount_COP"] > 0].copy()

    if capex_pie.empty:
        st.info("Enter CAPEX amounts to see the breakdown chart.")
    else:
        share = capex_pie["Amount_COP"] / capex_pie["Amount_COP"].sum()
        small = share < 0.03
        if small.any() and (~small).any():
            other_amt = float(capex_pie.loc[small, "Amount_COP"].sum())
            capex_pie = capex_pie.loc[~small, ["Item", "Amount_COP"]]
            capex_pie = pd.concat(
                [capex_pie, pd.DataFrame([{"Item": "Other (<3% each)", "Amount_COP": other_amt}])],
                ignore_index=True
            )
        if PLOTLY_AVAILABLE and px is not None:
            fig_pie = px.pie(capex_pie, names="Item", values="Amount_COP", hole=0.45)
            fig_pie.update_traces(textinfo="percent+label")
            fig_pie.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
            st.plotly_chart(fig_pie, width='stretch', key="capex_breakdown_pie")
        else:
            st.warning("⚠️ Chart unavailable")
    
    st.markdown("#### CAPEX schedule (monthly, aligned to timeline)")
    sched = capex_monthly_schedule(s)
    # Ensure CAPEX (COP) column is numeric and fill any NaN with 0
    sched["CAPEX (COP)"] = pd.to_numeric(sched["CAPEX (COP)"], errors="coerce").fillna(0.0)
    sched_disp = _df_format_money(sched.copy(), ["CAPEX (COP)"], decimals=0)
    st.dataframe(sched_disp[["Month", "Phase", "CAPEX (COP)"]], width='stretch', hide_index=True)

    if PLOTLY_AVAILABLE and px is not None:
        fig = px.bar(sched, x="Month", y="CAPEX (COP)", color="Phase")
        fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")

    st.markdown("#### Annual CAPEX (calendar years)")
    ann = sched.groupby("Year", as_index=False)["CAPEX (COP)"].sum()
    ann_disp = _df_format_money(ann.copy(), ["CAPEX (COP)"], decimals=0)
    st.dataframe(ann_disp, width='stretch', hide_index=True)
    
    # File upload section
    _render_file_upload_section(s, project_name, scenario_name, "capex", "Reference Files (Quotes / Contracts / Documents)")


# -----------------------------
# F) OPEX
# -----------------------------
with tab_opex:
    st.subheader("OPEX (COP) — operating costs, land lease, taxes & levies")

    c1, c2, c3 = st.columns(3)
    with c1:
        s.opex.fixed_om_cop_per_mwac_year = st.number_input("Fixed O&M (COP/MWac-year)", value=float(s.opex.fixed_om_cop_per_mwac_year), step=1_000_000.0, format="%.0f")
        s.opex.fixed_om_indexation = st.selectbox("Fixed O&M indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(s.opex.fixed_om_indexation) if s.opex.fixed_om_indexation in INDEX_CHOICES else 0)
    with c2:
        s.opex.variable_om_cop_per_mwh = st.number_input("Variable O&M (COP/MWh)", value=float(s.opex.variable_om_cop_per_mwh), step=1_000.0, format="%.0f")
        s.opex.grid_fees_cop_per_mwh = st.number_input("Grid fees (COP/MWh)", value=float(s.opex.grid_fees_cop_per_mwh), step=1_000.0, format="%.0f")
    with c3:
        s.opex.insurance_cop_per_mwac_year = st.number_input("Insurance (COP/MWac-year)", value=float(s.opex.insurance_cop_per_mwac_year), step=1_000_000.0, format="%.0f")
        s.opex.insurance_indexation = st.selectbox("Insurance indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(s.opex.insurance_indexation) if s.opex.insurance_indexation in INDEX_CHOICES else 0)

    st.markdown("#### Land lease")
    lc1, lc2, lc3, lc4, lc5 = st.columns([1, 1, 1, 1, 1.2])
    with lc1:
        s.opex.land_hectares = st.number_input("Hectares leased (Ha)", value=float(s.opex.land_hectares), step=1.0, format="%.2f")
    with lc2:
        s.opex.land_price_dev_cop_per_ha_year = st.number_input("Dev lease (COP/Ha-year)", value=float(s.opex.land_price_dev_cop_per_ha_year), step=1_000_000.0, format="%.0f")
    with lc3:
        s.opex.land_price_con_cop_per_ha_year = st.number_input("Const lease (COP/Ha-year)", value=float(s.opex.land_price_con_cop_per_ha_year), step=1_000_000.0, format="%.0f")
    with lc4:
        s.opex.land_price_op_cop_per_ha_year = st.number_input("Op lease (COP/Ha-year)", value=float(s.opex.land_price_op_cop_per_ha_year), step=1_000_000.0, format="%.0f")
    with lc5:
        s.opex.land_indexation = st.selectbox("Land lease indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(s.opex.land_indexation) if s.opex.land_indexation in INDEX_CHOICES else 0)

    st.markdown("#### Taxes & levies")
    tc1, tc2 = st.columns(2)
    with tc1:
        s.opex.ica_pct_of_revenue = st.number_input("ICA (% of revenue)", value=float(s.opex.ica_pct_of_revenue), step=0.01, format="%.3f")
    with tc2:
        s.opex.gmf_pct_of_outflows = st.number_input("GMF (% of outgoing cash)", value=float(s.opex.gmf_pct_of_outflows), step=0.01, format="%.3f")

    st.markdown("#### Other OPEX items (dynamic)")
    o_df = pd.DataFrame(s.opex.other_items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in o_df.columns:
            o_df[col] = "" if col != "Amount_COP_per_year" else 0.0
    o_df = o_df[["Item", "Amount_COP_per_year", "Phase", "Indexation"]].copy()
    o_df["Phase"] = o_df["Phase"].where(o_df["Phase"].isin(PHASES), "Operation")
    o_df["Indexation"] = o_df["Indexation"].where(o_df["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    o_edited = st.data_editor(
        o_df,
        width='stretch',
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Item": st.column_config.TextColumn("OPEX item"),
            "Amount_COP_per_year": st.column_config.NumberColumn("Amount (COP/year)", min_value=0.0, step=1_000_000.0, format="%.0f"),
            "Phase": st.column_config.SelectboxColumn("Phase", options=PHASES),
            "Indexation": st.column_config.SelectboxColumn("Indexation", options=INDEX_CHOICES),
        },
    )
    s.opex.other_items = o_edited.to_dict(orient="records")

    st.divider()
    st.markdown("## Outputs")

    om_full = opex_monthly_schedule(s).copy()

    meta_cols = {"Month", "Year", "Phase", "Energy (MWh)", "Revenue (COP)", "CAPEX (COP)", "OPEX subtotal", "GMF"}
    fixed_cols = {"Fixed O&M", "Insurance", "Variable O&M", "Grid fees", "Land lease", "ICA"}
    dyn_cols = [c for c in om_full.columns if c not in meta_cols and c not in fixed_cols]
    opex_item_cols = list(fixed_cols) + dyn_cols

    annual_items = om_full.groupby("Year", as_index=False)[opex_item_cols + ["GMF"]].sum()
    annual_items["Total OPEX (COP)"] = annual_items[opex_item_cols].sum(axis=1) + annual_items["GMF"]

    long = annual_items.melt(id_vars=["Year"], value_vars=opex_item_cols + ["GMF"], var_name="Item", value_name="OPEX (COP)")
    if PLOTLY_AVAILABLE and px is not None:
        fig = px.bar(long, x="Year", y="OPEX (COP)", color="Item", barmode="stack")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
        st.plotly_chart(fig, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")

    annual = om_full.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
    annual["Total OPEX (COP)"] = annual["OPEX subtotal"] + annual["GMF"]

    op = operating_year_table(s)[["Year", "Energy (MWh)"]].copy()
    annual = annual.merge(op, on="Year", how="left").fillna({"Energy (MWh)": 0.0})
    annual["OPEX per MWh (COP/MWh)"] = np.where(annual["Energy (MWh)"] > 0, annual["Total OPEX (COP)"] / annual["Energy (MWh)"], 0.0)

    disp = _df_format_money(annual.copy(), ["OPEX subtotal", "GMF", "Total OPEX (COP)", "OPEX per MWh (COP/MWh)", "Energy (MWh)"], decimals=0)
    disp = _transpose_annual_table(disp)
    st.dataframe(disp, width='stretch', hide_index=True)
    
    # File upload section
    _render_file_upload_section(s, project_name, scenario_name, "opex", "Reference Files (Quotes / Contracts / Documents)")


# -----------------------------
# G) SG&A
# -----------------------------
with tab_sga:
    st.subheader("SG&A (COP) — Development, Construction, and Operation")

    sga_df = pd.DataFrame(s.sga.items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in sga_df.columns:
            sga_df[col] = "" if col != "Amount_COP_per_year" else 0.0
    sga_df = sga_df[["Item", "Amount_COP_per_year", "Phase", "Indexation"]].copy()
    sga_df["Phase"] = sga_df["Phase"].where(sga_df["Phase"].isin(PHASES), "Development")
    sga_df["Indexation"] = sga_df["Indexation"].where(sga_df["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    sga_edited = st.data_editor(
        sga_df,
        width='stretch',
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Item": st.column_config.TextColumn("SG&A item"),
            "Amount_COP_per_year": st.column_config.NumberColumn("Amount (COP/year)", min_value=0.0, step=1_000_000.0, format="%.0f"),
            "Phase": st.column_config.SelectboxColumn("Phase", options=PHASES),
            "Indexation": st.column_config.SelectboxColumn("Indexation", options=INDEX_CHOICES),
        },
    )
    s.sga.items = sga_edited.to_dict(orient="records")

    st.divider()
    st.markdown("## Outputs")

    annual_sga = sga_annual_by_item(s)
    item_cols = [c for c in annual_sga.columns if c not in ["Year", "Total SG&A (COP)"]]
    if item_cols:
        annual_long = annual_sga.melt(id_vars=["Year"], value_vars=item_cols, var_name="Item", value_name="SG&A (COP)")
        if PLOTLY_AVAILABLE and px is not None:
            fig = px.bar(annual_long, x="Year", y="SG&A (COP)", color="Item", barmode="stack")
            fig.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig, width='stretch')
        else:
            st.warning("⚠️ Chart unavailable")
    else:
        st.info("Add SG&A line items to see the chart.")

    annual_disp = _df_format_money(annual_sga.copy(), [c for c in annual_sga.columns if c != "Year"], decimals=0)
    annual_disp = _transpose_annual_table(annual_disp)
    st.dataframe(annual_disp, width='stretch', hide_index=True)


# -----------------------------
# H) Depreciation
# -----------------------------
with tab_dep:
    st.subheader("Depreciation (linear, starting at COD)")

    c1, c2 = st.columns(2)
    with c1:
        s.depreciation.pct_of_capex_depreciated = st.number_input(
            "% of CAPEX depreciated",
            value=float(s.depreciation.pct_of_capex_depreciated),
            min_value=0.0,
            max_value=100.0,
            step=1.0,
            format="%.1f",
        )
    with c2:
        s.depreciation.dep_years = int(
            st.number_input(
                "Depreciation period (years after COD)",
                value=int(s.depreciation.dep_years),
                min_value=3,
                max_value=25,
                step=1,
                format="%d",
            )
        )

    dep = depreciation_annual_table(s)

    if PLOTLY_AVAILABLE and px is not None:
        fig = px.bar(dep, x="Year", y="Depreciation (COP)")
        fig.update_layout(height=340, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")

    dep_disp = _df_format_money(dep.copy(), [c for c in dep.columns if c != "Year"], decimals=0)
    dep_disp = _transpose_annual_table(dep_disp)
    st.dataframe(dep_disp, width='stretch', hide_index=True)


# -----------------------------
# K) Debt & Covenants
# -----------------------------
with tab_debt:
    st.subheader("Debt & Covenants (tenor 5–10 years, sculpted amortization)")

    s.debt.enabled = st.checkbox("Enable debt", value=bool(s.debt.enabled))
    s.debt.balloon_pct = st.slider("Balloon (% of original debt at maturity)", 0.0, 50.0, float(getattr(s.debt, "balloon_pct", 0.0)), 1.0)

    if not s.debt.enabled:
        st.info("Debt is disabled. Enable it to see debt sizing, amortization, DSCR, and covenant indicators.")
    else:
        total_capex = _total_capex_from_lines(s)
        s.debt.debt_pct_of_capex = st.slider("Debt as % of CAPEX", min_value=0.0, max_value=90.0, value=float(s.debt.debt_pct_of_capex), step=1.0)
        debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex

        c1, c2, c3 = st.columns(3)
        with c1:
            s.debt.tenor_years = int(st.number_input("Tenor (years)", value=int(s.debt.tenor_years), min_value=5, max_value=10, step=1, format="%d"))
        with c2:
            s.debt.grace_years = int(st.number_input("Grace period (years)", value=int(s.debt.grace_years), min_value=0, max_value=max(0, int(s.debt.tenor_years) - 1), step=1, format="%d"))
        with c3:
            amort_options = ["Sculpted to DSCR", "Equal principal", "Typical amortization"]
            current_index = 0
            if s.debt.amortization_type == "Equal principal":
                current_index = 1
            elif s.debt.amortization_type == "Typical amortization":
                current_index = 2
            s.debt.amortization_type = st.selectbox("Amortization", amort_options, index=current_index)

        st.markdown("#### Pricing (Natural COP)")
        p1, p2, p3 = st.columns(3)
        with p1:
            s.debt.base_rate_pct = st.number_input("Base rate (e.g., IBR) %", value=float(s.debt.base_rate_pct), step=0.25, format="%.2f")
        with p2:
            s.debt.margin_pct = st.number_input("Margin (spread) %", value=float(s.debt.margin_pct), step=0.25, format="%.2f")
        with p3:
            all_in = float(s.debt.base_rate_pct) + float(s.debt.margin_pct)
            st.metric("All-in interest rate", f"{all_in:,.2f}%")

        st.markdown("#### Fees")
        f1, f2 = st.columns(2)
        with f1:
            s.debt.upfront_fee_bps = st.number_input("Upfront / structuring fee (bps of debt)", value=float(s.debt.upfront_fee_bps), step=5.0, format="%.0f")
        with f2:
            s.debt.commitment_fee_pct_of_margin = st.number_input("Commitment fee (% of margin) on undrawn", value=float(s.debt.commitment_fee_pct_of_margin), step=1.0, format="%.0f")

        st.markdown("#### Covenants")
        k1, k2, k3 = st.columns(3)
        with k1:
            s.debt.target_dscr = st.number_input("Target DSCR (for sculpting)", value=float(s.debt.target_dscr), step=0.01, format="%.2f")
        with k2:
            s.debt.min_dscr_covenant = st.number_input("Minimum DSCR covenant", value=float(s.debt.min_dscr_covenant), step=0.01, format="%.2f")
        with k3:
            s.debt.lockup_dscr = st.number_input("Lock-up DSCR threshold", value=float(s.debt.lockup_dscr), step=0.01, format="%.2f")

        _metric_row([
            ("Total CAPEX", _fmt_cop(total_capex)),
            ("Debt amount", _fmt_cop(debt_amt)),
            ("Equity (implied)", _fmt_cop(max(total_capex - debt_amt, 0.0))),
        ])

        st.divider()

        ds = debt_schedule_annual(s)
        com = debt_commitment_fee_annual(s)

        # KPIs over debt life where debt service > 0
        ds_valid = ds[pd.to_numeric(ds["Debt Service (COP)"], errors="coerce").fillna(0.0) > 0].copy()
        min_dscr = float(ds_valid["DSCR"].min()) if not ds_valid.empty else float("nan")
        lockup_years = int(ds["Lock-up? (DSCR < lock-up)"].sum()) if "Lock-up? (DSCR < lock-up)" in ds.columns else 0
        breach_years = int(ds["Breach? (DSCR < min)"].sum()) if "Breach? (DSCR < min)" in ds.columns else 0
        balloon = float(ds["Balloon at Maturity (COP)"].sum()) if "Balloon at Maturity (COP)" in ds.columns else 0.0
        upfront_fee = float(ds["Upfront Fee (COP)"].sum()) if "Upfront Fee (COP)" in ds.columns else 0.0
        commit_total = float(com["Commitment Fee (COP)"].sum()) if (not com.empty and "Commitment Fee (COP)" in com.columns) else 0.0

        # Calculate breach years with warning
        breach_warning = breach_years > 1
        
        # Display metrics - highlight breach years if > 1
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Min DSCR (debt life)", f"{min_dscr:,.2f}x" if np.isfinite(min_dscr) else "—")
        with col2:
            st.metric("Lock-up years", str(lockup_years))
        with col3:
            if breach_warning:
                # Display in red box for visibility
                st.markdown(
                    f'<div style="background-color: #ffebee; padding: 8px; border-radius: 5px; border-left: 4px solid #f44336; margin-top: 8px;">'
                    f'<div style="font-size: 0.8rem; color: #666;">Breach years</div>'
                    f'<div style="font-size: 1.5rem; font-weight: bold; color: #c62828;">{breach_years}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            else:
                st.metric("Breach years", str(breach_years))
        with col4:
            st.metric("Balloon", _fmt_cop(balloon))
        
        # Warning for DSCR breaches exceeding 1 year
        if breach_warning:
            st.error(f"⚠️ **WARNING: DSCR covenant breaches exceed 1 year ({breach_years} years).** The project may be overleveraged and unable to service debt from operating cash flow.")
        
        _metric_row([
            ("Upfront fee", _fmt_cop(upfront_fee)),
            ("Commitment fees (total)", _fmt_cop(commit_total)),
            ("Tenor", f"{int(s.debt.tenor_years)} years"),
            ("Grace", f"{int(s.debt.grace_years)} years"),
        ])
        
        # Calculate actual equity investment from levered cash flow and compare to implied equity
        try:
            annual_levered = levered_cashflow_annual(s)
            actual_equity_investment = 0.0
            for _, row in annual_levered.iterrows():
                levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
                if levered_cf < 0:
                    actual_equity_investment += abs(levered_cf)
            
            implied_equity = max(total_capex - debt_amt, 0.0)
            equity_difference = actual_equity_investment - implied_equity
            equity_difference_pct = (equity_difference / implied_equity * 100.0) if implied_equity > 0 else 0.0
            
            # Warning if actual equity significantly exceeds implied equity (more than 10% difference)
            if equity_difference > 0 and equity_difference_pct > 10.0:
                st.warning(
                    f"⚠️ **EQUITY SHORTFALL WARNING:**\n\n"
                    f"- **Implied Equity** (from debt tab): {_fmt_cop(implied_equity)}\n"
                    f"- **Actual Equity Investment** (from levered cash flow): {_fmt_cop(actual_equity_investment)}\n"
                    f"- **Additional Equity Required**: {_fmt_cop(equity_difference)} ({equity_difference_pct:.1f}% more)\n\n"
                    f"This indicates that debt service and fees exceed operating cash flow, requiring the investor to contribute more equity than initially expected. "
                    f"The project may be overleveraged."
                )
        except Exception as e:
            # If levered cash flow calculation fails, skip the warning
            pass

        st.markdown("### Debt schedule (annual)")
        disp = ds.copy()
        money_cols = ["Operating CF (COP)", "Interest (COP)", "Principal (COP)", "Debt Service (COP)", "Outstanding End (COP)", "Balloon at Maturity (COP)", "Upfront Fee (COP)"]
        disp = _df_format_money(disp, money_cols, decimals=0)
        disp = _transpose_annual_table(disp)
        st.dataframe(disp, width='stretch', hide_index=True)

        st.markdown("### DSCR vs covenant thresholds")
        ds_plot = ds.copy()
        ds_plot["Min covenant"] = float(s.debt.min_dscr_covenant)
        ds_plot["Lock-up"] = float(s.debt.lockup_dscr)
        ds_long = ds_plot.melt(id_vars=["Year"], value_vars=["DSCR", "Min covenant", "Lock-up"], var_name="Line", value_name="Value")
        if PLOTLY_AVAILABLE and px is not None:
            fig = px.line(ds_long, x="Year", y="Value", color="Line")
            fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig, width='stretch')
        else:
            st.warning("⚠️ Chart unavailable")

        st.markdown("### Commitment fees (during construction)")
        if com.empty:
            st.info("No commitment fees computed (likely no construction months or debt amount is 0).")
        else:
            com_disp = _df_format_money(com.copy(), ["Commitment Fee (COP)"], decimals=0)
            com_disp = _transpose_annual_table(com_disp)
            st.dataframe(com_disp, width='stretch', hide_index=True)
            if PLOTLY_AVAILABLE and px is not None:
                figc = px.bar(com, x="Year", y="Commitment Fee (COP)")
                figc.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10))
                st.plotly_chart(figc, width='stretch')
            else:
                st.warning("⚠️ Chart unavailable")


# -----------------------------
# I) Renewable tax benefits (moved before Unlevered CF)
# -----------------------------
with tab_incent:
    st.subheader("Renewable tax benefits (Colombia)")

    inc = s.incentives

    st.markdown("### Special deduction (up to 50% of eligible CAPEX over up to 15 years)")
    c1, c2, c3 = st.columns(3)
    with c1:
        inc.enable_special_deduction = st.checkbox("Enable special deduction", value=bool(inc.enable_special_deduction))
    with c2:
        inc.special_deduction_pct_of_capex = st.number_input(
            "Deduction pool (% of CAPEX, max 50%)",
            value=float(inc.special_deduction_pct_of_capex),
            min_value=0.0, max_value=50.0, step=1.0, format="%.1f"
        )
    with c3:
        inc.special_deduction_years = int(st.number_input(
            "Carryforward window (years)",
            value=int(inc.special_deduction_years),
            min_value=1, max_value=30, step=1, format="%d"
        ))

    inc.special_deduction_max_pct_of_taxable_income = st.number_input(
        "Annual cap (% of taxable income, typically 50%)",
        value=float(inc.special_deduction_max_pct_of_taxable_income),
        min_value=0.0, max_value=100.0, step=1.0, format="%.1f"
    )

    # Calculate and display total CAPEX deduction pool
    if bool(inc.enable_special_deduction):
        cap_df = pd.DataFrame(s.capex.lines or [])
        total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0
        pool_total = (float(inc.special_deduction_pct_of_capex) / 100.0) * total_capex
        st.info(f"**Total CAPEX Deduction Pool:** {_fmt_cop(pool_total)} ({inc.special_deduction_pct_of_capex:.1f}% of {_fmt_cop(total_capex)} total CAPEX)")
    else:
        st.info("Special deduction is disabled. Enable it above to see the deduction pool.")

    st.divider()
    st.markdown("### VAT treatment")
    inc.vat_mode = st.selectbox("VAT mode", ["Excluded", "Refund"], index=0 if inc.vat_mode != "Refund" else 1)

    c4, c5, c6 = st.columns(3)
    with c4:
        inc.vat_pct_of_capex = st.number_input(
            "VAT as % of CAPEX (optional)",
            value=float(inc.vat_pct_of_capex),
            min_value=0.0, max_value=100.0, step=0.5, format="%.2f"
        )
    with c5:
        inc.vat_fixed_cop = st.number_input(
            "VAT fixed COP (optional)",
            value=float(inc.vat_fixed_cop),
            min_value=0.0, step=1_000_000.0, format="%.0f"
        )
    with c6:
        inc.vat_refund_year_index = int(st.number_input(
            "Refund in operating year #",
            value=int(inc.vat_refund_year_index),
            min_value=1, max_value=10, step=1, format="%d"
        ))

    st.caption("If VAT is Excluded, leave VAT inputs at 0. If Refund, you can use % or fixed; both will be added (so use one).")
    st.info("ℹ️ These tax benefits are automatically applied in the Unlevered Base Cash Flow calculation below.")


# -----------------------------
# J) Unlevered Base Cash Flow
# -----------------------------
with tab_ucf:
    st.subheader("Unlevered Base Cash Flow (includes tax benefits, pre-debt)")

    st.info("This tab calculates unlevered cash flow before and after tax. Tax benefits from the Renewable tax benefits tab are automatically included. This is the base for calculating debt capacity and equity returns.")

    tx1, tx2 = st.columns([1, 2])
    with tx1:
        s.tax.corporate_tax_rate_pct = st.number_input("Corporate income tax rate (%)", value=float(s.tax.corporate_tax_rate_pct), step=0.5, format="%.2f")
    with tx2:
        s.tax.allow_loss_carryforward = st.checkbox("Apply loss carryforward (NOL) so taxes are zero until losses are used", value=bool(s.tax.allow_loss_carryforward))

    st.markdown("### Working Capital (timing)")
    wc1, wc2, wc3, wc4 = st.columns([1, 1, 1, 1])
    with wc1:
        s.wc.ar_days = int(st.number_input("AR days (revenue collection)", value=int(s.wc.ar_days), min_value=0, step=15, format="%d"))
    with wc2:
        s.wc.ap_days = int(st.number_input("AP days (expense payment)", value=int(s.wc.ap_days), min_value=0, step=15, format="%d"))
    with wc3:
        s.wc.apply_ap_to_opex = st.checkbox("Apply AP lag to OPEX", value=bool(s.wc.apply_ap_to_opex))
    with wc4:
        s.wc.apply_ap_to_sga = st.checkbox("Apply AP lag to SG&A", value=bool(s.wc.apply_ap_to_sga))

    currency = st.radio("Display currency", ["COP", "USD"], horizontal=True, index=0, key="currency_unlevered")

    m = cashflow_monthly_table(s)
    a = unlevered_base_cashflow_annual(s)

    tl = build_timeline(s.timeline)
    years = list(a["Year"].astype(int).tolist())
    fx = fx_series(s.macro, tl["cod"].year, years)

    def _conv(series: pd.Series, year_col: pd.Series) -> pd.Series:
        if currency == "COP":
            return series
        return series / year_col.map(lambda y: float(fx.loc[int(y)]) if int(y) in fx.index else float(s.macro.fx_cop_per_usd_start))

    st.divider()

    # --- KPIs ---
    mm = cashflow_monthly_table(s).copy()
    for col in ["CAPEX (COP)", "Unlevered CF (COP)", "Unlevered CF Pre-tax (COP)", "Cumulative Unlevered CF (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)

    # Calculate total CAPEX from line items (more reliable than summing monthly table)
    cap_total = _total_capex_from_lines(s)
    
    # Pre-tax cash flows - use Pre-tax column if available, otherwise fall back to original
    if "Unlevered CF Pre-tax (COP)" in mm.columns:
        monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist()
    else:
        monthly_cf_pre_tax = mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    monthly_dates = mm["Month"].tolist() if "Month" in mm.columns else []

    # After-tax cash flows: build from annual "Unlevered CF After Tax" values
    # Create a map of annual after-tax CF by year
    annual_after_tax_map = {}
    annual_pre_tax_map = {}
    for _, row in a.iterrows():
        year = int(row["Year"])
        annual_after_tax_map[year] = float(row.get("Unlevered CF After Tax (COP)", 0.0))
        annual_pre_tax_map[year] = float(row.get("Unlevered CF Pre-tax (COP)", row.get("Unlevered CF (COP)", 0.0)))
    
    # Calculate operating months per year for proration
    cod = tl["cod"]
    end_op = tl["end_op"]
    operating_months_map = {}
    for year in annual_after_tax_map.keys():
        if year == cod.year:
            operating_months_map[year] = 13 - cod.month
        elif year == end_op.year:
            operating_months_map[year] = end_op.month
        else:
            operating_months_map[year] = 12
    
    # Build monthly after-tax CF series from annual "Unlevered CF After Tax" values
    # This matches Excel: use annual after-tax CF values, prorated to monthly
    # For operation months: prorate annual after-tax CF evenly across operating months
    # For construction months: use pre-tax CF (no taxes during construction)
    monthly_cf_after_tax = []
    for i, month_date in enumerate(monthly_dates):
        year = month_date.year
        phase = _phase_for_month(tl, month_date)
        
        if phase == "Operation" and year in annual_after_tax_map:
            # Operation months: use prorated annual after-tax CF from "Unlevered CF After Tax" column
            operating_months = operating_months_map.get(year, 12)
            if operating_months > 0:
                # Distribute annual after-tax CF evenly across operating months in the calendar year
                # This ensures the sum of monthly values equals the annual "Unlevered CF After Tax"
                monthly_after_tax = annual_after_tax_map[year] / operating_months
            else:
                monthly_after_tax = 0.0
            monthly_cf_after_tax.append(monthly_after_tax)
        else:
            # Construction/non-operation months: use pre-tax CF (no tax impact during construction)
            # This ensures all negative cashflows (CAPEX) are included in the IRR calculation
            monthly_cf_after_tax.append(monthly_cf_pre_tax[i] if i < len(monthly_cf_pre_tax) else 0.0)

    # Pre-tax IRR
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")

    # After-tax IRR - calculate from monthly cash flows
    has_pos_after = any(cf > 0 for cf in monthly_cf_after_tax)
    has_neg_after = any(cf < 0 for cf in monthly_cf_after_tax)
    irr_m_after = _irr_bisection(monthly_cf_after_tax) if (has_pos_after and has_neg_after) else float("nan")
    irr_annual_after = (1.0 + irr_m_after) ** 12 - 1.0 if np.isfinite(irr_m_after) else float("nan")
    
    # Also calculate annual IRR directly from annual "Unlevered CF After Tax" values for verification
    # This should match Excel when using annual cash flows
    annual_cf_after_tax = []
    for _, row in a.iterrows():
        # Get the annual "Unlevered CF After Tax (COP)" value
        annual_cf_after_tax.append(float(row.get("Unlevered CF After Tax (COP)", 0.0)))
    
    # Calculate annual IRR directly (this matches Excel annual calculation)
    has_pos_annual = any(cf > 0 for cf in annual_cf_after_tax)
    has_neg_annual = any(cf < 0 for cf in annual_cf_after_tax)
    irr_annual_direct = _irr_bisection(annual_cf_after_tax) if (has_pos_annual and has_neg_annual) else float("nan")
    
    # Use the annual direct calculation (matches Excel)
    if np.isfinite(irr_annual_direct):
        irr_annual_after = irr_annual_direct

    # Payback (after-tax)
    payback_m_after = _payback_months(monthly_dates, monthly_cf_after_tax) if monthly_cf_after_tax else float("nan")
    payback_years_after = payback_m_after / 12.0 if np.isfinite(payback_m_after) else float("nan")

    # Peak funding (from pre-tax cumulative CF)
    cum = mm["Cumulative Unlevered CF (COP)"].astype(float) if "Cumulative Unlevered CF (COP)" in mm.columns else pd.Series([0.0])
    min_cum = float(cum.min()) if len(cum) else 0.0
    peak_funding = float(max(0.0, -min_cum)) if np.isfinite(min_cum) else 0.0

    if currency == "COP":
        _metric_row([
            ("Total Investment (CAPEX)", _fmt_cop(cap_total)),
            ("Unlevered IRR (annualized, pre-tax)", f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"),
            ("Unlevered IRR (annualized, after-tax)", f"{irr_annual_after*100:,.2f}%" if np.isfinite(irr_annual_after) else "—"),
            ("Payback (years, after-tax)", f"{payback_years_after:,.2f}" if np.isfinite(payback_years_after) else "—"),
        ])
        _metric_row([
            ("Peak Funding Need", _fmt_cop(peak_funding)),
            ("", ""),
            ("", ""),
            ("", ""),
        ])
    else:
        fx0 = float(s.macro.fx_cop_per_usd_start)
        _metric_row([
            ("Total Investment (CAPEX)", _fmt_usd(cap_total / fx0)),
            ("Unlevered IRR (annualized, pre-tax)", f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"),
            ("Unlevered IRR (annualized, after-tax)", f"{irr_annual_after*100:,.2f}%" if np.isfinite(irr_annual_after) else "—"),
            ("Payback (years, after-tax)", f"{payback_years_after:,.2f}" if np.isfinite(payback_years_after) else "—"),
        ])
        _metric_row([
            ("Peak Funding Need", _fmt_usd(peak_funding / fx0)),
            ("", ""),
            ("", ""),
            ("", ""),
        ])

    # Annual table (with currency switch)
    annual_view = a.copy()
    money_cols = [
        "Revenue (COP)", "Total OPEX (COP)", "SG&A (COP)", "EBITDA (COP)", "Depreciation (COP)",
        "CAPEX Tax Deduction (COP)", "Loss Carryforward End (COP)", "Taxable Income (COP)", "Taxes Payable (COP)", "CAPEX (COP)", "ΔNWC (COP)", "Unlevered CF Pre-tax (COP)", "Unlevered CF After Tax (COP)",
        "Cumulative Unlevered CF (COP)",
    ]
    for c in money_cols:
        if c in annual_view.columns:
            annual_view[c] = _conv(annual_view[c], annual_view["Year"])

    if currency == "USD":
        ren = {c: c.replace("(COP)", "(USD)") for c in annual_view.columns if "(COP)" in c}
        annual_view = annual_view.rename(columns=ren)

    st.markdown("### Annual summary (calendar years)")
    display_cols = [
        "Year",
        "CAPEX (COP)" if currency == "COP" else "CAPEX (USD)",
        "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "Total OPEX (COP)" if currency == "COP" else "Total OPEX (USD)",
        "SG&A (COP)" if currency == "COP" else "SG&A (USD)",
        "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "CAPEX Tax Deduction (COP)" if currency == "COP" else "CAPEX Tax Deduction (USD)",
        "Loss Carryforward End (COP)" if currency == "COP" else "Loss Carryforward End (USD)",
        "Taxable Income (COP)" if currency == "COP" else "Taxable Income (USD)",
        "Taxes Payable (COP)" if currency == "COP" else "Taxes Payable (USD)",
        "ΔNWC (COP)" if currency == "COP" else "ΔNWC (USD)",
        "Unlevered CF Pre-tax (COP)" if currency == "COP" else "Unlevered CF Pre-tax (USD)",
        "Unlevered CF After Tax (COP)" if currency == "COP" else "Unlevered CF After Tax (USD)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    disp = annual_view[display_cols].copy()
    disp = _df_format_money(disp, [c for c in disp.columns if c != "Year"], decimals=0)
    disp = _transpose_annual_table(disp)
    st.dataframe(disp, width='stretch', hide_index=True)

    y_after = "Unlevered CF After Tax (COP)" if currency == "COP" else "Unlevered CF After Tax (USD)"
    if PLOTLY_AVAILABLE and px is not None:
        fig = px.bar(annual_view, x="Year", y=y_after)
        fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")

    st.markdown("### Monthly cash flow (detailed, pre-tax)")
    m_disp = mm.copy()
    m_money = [c for c in m_disp.columns if c not in ["Month", "Year", "Phase"]]
    m_disp = _df_format_money(m_disp, m_money, decimals=0)
    st.dataframe(m_disp, width='stretch', hide_index=True)


# -----------------------------
# K) Debt & Covenants
# -----------------------------
with tab_levered:
    st.subheader("Levered Cash Flow (Equity Cash Flow After-Tax)")
    
    st.info("This tab calculates the levered after-tax cash flow available to equity investors. "
            "If debt is disabled or not enabled, the model assumes no debt and returns unlevered cash flow. "
            "This cash flow will be used to calculate Investor Equity IRR in the summary tab.")
    
    currency = st.radio("Display currency", ["COP", "USD"], horizontal=True, index=0, key="currency_levered")
    
    # Calculate levered cash flows
    annual_levered = levered_cashflow_annual(s)
    monthly_levered = levered_cashflow_monthly(s)
    
    # Check debt status
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex if debt_enabled else 0.0
    
    if not debt_enabled or debt_amt <= 0:
        st.warning("⚠️ Debt is not enabled or debt amount is zero. Showing unlevered cash flow (no debt impact).")
    
    # FX conversion helper
    tl = build_timeline(s.timeline)
    years = list(annual_levered["Year"].astype(int).tolist())
    fx = fx_series(s.macro, tl["cod"].year, years)
    
    def _conv(series: pd.Series, year_col: pd.Series) -> pd.Series:
        if currency == "COP":
            return series
        return series / year_col.map(lambda y: float(fx.loc[int(y)]) if int(y) in fx.index else float(s.macro.fx_cop_per_usd_start))
    
    # Key metrics
    total_debt_draws = float(annual_levered["Debt Draw (COP)"].sum())
    total_debt_service = float(annual_levered["Debt Service (COP)"].sum())
    total_debt_fees = float(annual_levered["Debt Fees (COP)"].sum())
    
    # Calculate all-in interest rate (same as debt tab)
    all_in_rate = float(s.debt.base_rate_pct) + float(s.debt.margin_pct) if debt_enabled else 0.0
    
    if currency == "COP":
        _metric_row([
            ("Total Debt Draws", _fmt_cop(total_debt_draws)),
            ("Total Debt Service", _fmt_cop(total_debt_service)),
            ("Total Debt Fees", _fmt_cop(total_debt_fees)),
            ("All-in interest rate", f"{all_in_rate:,.2f}%" if debt_enabled else "—"),
        ])
        _metric_row([
            ("Debt Status", "Enabled" if debt_enabled and debt_amt > 0 else "No Debt"),
            ("Debt Amount", _fmt_cop(debt_amt) if debt_enabled else "—"),
            ("", ""),
            ("", ""),
        ])
    else:
        fx0 = float(s.macro.fx_cop_per_usd_start)
        _metric_row([
            ("Total Debt Draws", _fmt_usd(total_debt_draws / fx0)),
            ("Total Debt Service", _fmt_usd(total_debt_service / fx0)),
            ("Total Debt Fees", _fmt_usd(total_debt_fees / fx0)),
            ("All-in interest rate", f"{all_in_rate:,.2f}%" if debt_enabled else "—"),
        ])
        _metric_row([
            ("Debt Status", "Enabled" if debt_enabled and debt_amt > 0 else "No Debt"),
            ("Debt Amount", _fmt_usd(debt_amt / fx0) if debt_enabled else "—"),
            ("", ""),
            ("", ""),
        ])
    
    st.divider()
    
    # Calculate Equity IRR from annual levered CF after-tax (matches table values)
    annual_cf_levered = []
    for _, row in annual_levered.iterrows():
        annual_cf_levered.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
    
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    # Payback - still use monthly for precision
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    # Total equity investment (sum of all negative cash flows from annual table, after fees)
    # This represents total equity contributions over the project life
    # This should match the sum of all negative values in the "Levered CF After-tax" column
    total_equity_investment = 0.0
    for _, row in annual_levered.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)  # Sum of absolute values of negative CFs
    
    _metric_row([
        ("Equity IRR (annualized, after-tax)", f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—"),
        ("Payback (years, after-tax)", f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—"),
        ("Total Equity Investment (After Fees)", _fmt_cop(total_equity_investment) if currency == "COP" else _fmt_usd(total_equity_investment / fx0)),
    ])
    
    st.divider()
    
    # Annual table - Income Statement format
    annual_view = annual_levered.copy()
    money_cols = [
        "Revenue (COP)", "EBITDA (COP)", "Depreciation (COP)", "Interest (COP)",
        "Levered CAPEX Tax Deduction (COP)", "Levered Loss Carryforward End (COP)",
        "Levered Taxable Income (COP)", "Levered Taxes Payable (COP)", "Levered Net Income After Tax (COP)",
        "CAPEX (COP)", "Debt Draw (COP)", "Principal (COP)", "Debt Fees (COP)", "VAT Refund (COP)", "ΔNWC (COP)",
        "Levered CF (After-tax, COP)", "Cumulative Levered CF (COP)",
    ]
    for c in money_cols:
        if c in annual_view.columns:
            annual_view[c] = _conv(annual_view[c], annual_view["Year"])
    
    if currency == "USD":
        ren = {c: c.replace("(COP)", "(USD)") for c in annual_view.columns if "(COP)" in c}
        annual_view = annual_view.rename(columns=ren)
    
    st.markdown("### Income Statement (levered, with debt)")
    display_cols = [
        "Year",
        "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Levered CAPEX Tax Deduction (COP)" if currency == "COP" else "Levered CAPEX Tax Deduction (USD)",
        "Levered Loss Carryforward End (COP)" if currency == "COP" else "Levered Loss Carryforward End (USD)",
        "Levered Taxable Income (COP)" if currency == "COP" else "Levered Taxable Income (USD)",
        "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
        "ΔNWC (COP)" if currency == "COP" else "ΔNWC (USD)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    disp = annual_view[display_cols].copy()
    disp = _df_format_money(disp, [c for c in disp.columns if c != "Year"], decimals=0)
    disp = _transpose_annual_table(disp)
    st.dataframe(disp, width='stretch', hide_index=True)
    
    # Detailed Levered Free Cash Flow Calculation Table
    st.markdown("### Levered Free Cash Flow Calculation (Step-by-Step)")
    st.caption("This table shows the detailed calculation of levered after-tax free cash flow from Revenue to final cash flow.")
    
    fcf_cols = ["Year"]
    fcf_labels = []
    
    # Build the calculation table step by step
    step_cols = {
        "Revenue": "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA": "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation": "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "Interest Expense": "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Debt Service": "Debt Service (COP)" if currency == "COP" else "Debt Service (USD)",
        "CAPEX Tax Deduction": "Levered CAPEX Tax Deduction (COP)" if currency == "COP" else "Levered CAPEX Tax Deduction (USD)",
        "Loss Carryforward": "Levered Loss Carryforward End (COP)" if currency == "COP" else "Levered Loss Carryforward End (USD)",
        "Taxable Income": "Levered Taxable Income (COP)" if currency == "COP" else "Levered Taxable Income (USD)",
        "Taxes Payable": "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Working Capital Change": "ΔNWC (COP)" if currency == "COP" else "ΔNWC (USD)",
        "Levered CF After-tax": "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)",
    }
    
    for label, col in step_cols.items():
        if col in annual_view.columns:
            fcf_cols.append(col)
            fcf_labels.append(label)
    
    # Create the detailed calculation table
    fcf_df = annual_view[fcf_cols].copy()
    
    # Rename columns for clarity
    rename_dict = {"Year": "Year"}
    for i, col in enumerate(fcf_cols[1:], 1):  # Skip "Year"
        if i <= len(fcf_labels):
            rename_dict[col] = fcf_labels[i-1]
    fcf_df = fcf_df.rename(columns=rename_dict)
    
    # Format the table
    fcf_df = _df_format_money(fcf_df, [c for c in fcf_df.columns if c != "Year"], decimals=0)
    fcf_df = _transpose_annual_table(fcf_df)
    st.dataframe(fcf_df, width='stretch', hide_index=True)
    
    # Income Statement Graph
    st.markdown("### Income Statement Overview")
    graph_cols = {
        "Revenue": "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA": "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Interest Expense": "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Taxes Payable": "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Net Income After Tax": "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    }
    
    # Create a melted dataframe for the graph
    graph_df = annual_view[["Year"]].copy()
    for label, col in graph_cols.items():
        if col in annual_view.columns:
            graph_df[label] = annual_view[col]
    
    # Melt for grouped bar chart
    graph_long = graph_df.melt(
        id_vars=["Year"],
        value_vars=[label for label, col in graph_cols.items() if col in annual_view.columns],
        var_name="Metric",
        value_name="Amount"
    )
    
    if PLOTLY_AVAILABLE and px is not None:
        fig_income = px.bar(
            graph_long,
            x="Year",
            y="Amount",
            color="Metric",
            barmode="group",
            title="Revenue, EBITDA, Interest, Taxes, and Net Income"
        )
        
        # Add period indicators
        tl = build_timeline(s.timeline)
        rtb_year = tl["rtb"].year
        cod_year = tl["cod"].year
        end_op_year = tl["end_op"].year
        
        # Add vertical lines for period boundaries
        fig_income.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
        fig_income.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
        fig_income.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
        
        fig_income.update_layout(height=400, margin=dict(l=10, r=10, t=40, b=10))
        st.plotly_chart(fig_income, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")
    
    # Levered Free Cash Flow (Equity Cash Flow)
    st.markdown("### Levered Free Cash Flow (Equity/After-Tax)")
    
    # Graph of levered free cash flow
    y_fcf = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if y_fcf not in annual_view.columns:
        y_fcf = "Levered CF (After-tax, COP)"
    
    fig_fcf = px.bar(annual_view, x="Year", y=y_fcf)
    
    # Add period indicators
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    # Add vertical lines for period boundaries
    fig_fcf.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig_fcf.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig_fcf.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig_fcf.update_layout(
        height=400,
        margin=dict(l=10, r=10, t=40, b=10),
        title="Equity/Levered After-Tax Free Cash Flow"
    )
    st.plotly_chart(fig_fcf, width='stretch')
    
    # Calculation table showing how levered FCF is derived
    st.markdown("#### Levered Free Cash Flow Calculation")
    
    # Build calculation table
    calc_cols = ["Year"]
    calc_labels = []
    
    # Net Income After Tax (starting point)
    net_income_col = "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)"
    if net_income_col in annual_view.columns:
        calc_cols.append(net_income_col)
        calc_labels.append("Net Income After Tax")
    
    # Add back Depreciation (non-cash)
    dep_col = "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)"
    if dep_col in annual_view.columns:
        calc_cols.append(dep_col)
        calc_labels.append("+ Depreciation (add back)")
    
    # Subtract CAPEX
    capex_col = "CAPEX (COP)" if currency == "COP" else "CAPEX (USD)"
    if capex_col in annual_view.columns:
        calc_cols.append(capex_col)
        calc_labels.append("- CAPEX")
    
    # Add Debt Draws
    debt_draw_col = "Debt Draw (COP)" if currency == "COP" else "Debt Draw (USD)"
    if debt_draw_col in annual_view.columns:
        calc_cols.append(debt_draw_col)
        calc_labels.append("+ Debt Draws")
    
    # Subtract Principal payments
    principal_col = "Principal (COP)" if currency == "COP" else "Principal (USD)"
    if principal_col in annual_view.columns:
        calc_cols.append(principal_col)
        calc_labels.append("- Principal Payments")
    
    # Subtract Debt Fees
    debt_fees_col = "Debt Fees (COP)" if currency == "COP" else "Debt Fees (USD)"
    if debt_fees_col in annual_view.columns:
        calc_cols.append(debt_fees_col)
        calc_labels.append("- Debt Fees")
    
    # Add VAT Refund
    vat_col = "VAT Refund (COP)" if currency == "COP" else "VAT Refund (USD)"
    if vat_col in annual_view.columns:
        calc_cols.append(vat_col)
        calc_labels.append("+ VAT Refund")
    
    # Subtract Working Capital Change
    nwc_col = "ΔNWC (COP)" if currency == "COP" else "ΔNWC (USD)"
    if nwc_col in annual_view.columns:
        calc_cols.append(nwc_col)
        calc_labels.append("- Working Capital Change")
    
    # Final: Levered Free Cash Flow
    calc_cols.append(y_fcf)
    calc_labels.append("= Levered Free Cash Flow (After-Tax)")
    
    # Create calculation dataframe
    calc_df = annual_view[calc_cols].copy()
    
    # Rename columns for display
    rename_dict = {}
    for i, col in enumerate(calc_cols):
        if col != "Year":
            rename_dict[col] = calc_labels[i - 1] if i > 0 else calc_labels[i]
    calc_df = calc_df.rename(columns=rename_dict)
    
    # Format the table
    calc_df = _df_format_money(calc_df, [c for c in calc_df.columns if c != "Year"], decimals=0)
    calc_df = _transpose_annual_table(calc_df)
    st.dataframe(calc_df, width='stretch', hide_index=True)
    
    # Cumulative Levered CF Chart (larger)
    y_cum = "Cumulative Levered CF (COP)" if currency == "COP" else "Cumulative Levered CF (USD)"
    # Verify column exists, fallback to COP version if USD version doesn't exist
    if y_cum not in annual_view.columns:
        y_cum = "Cumulative Levered CF (COP)"
    if PLOTLY_AVAILABLE and px is not None:
        fig2 = px.line(annual_view, x="Year", y=y_cum)
        
        # Add period indicators (vertical lines for key dates)
        tl = build_timeline(s.timeline)
        rtb_year = tl["rtb"].year
        cod_year = tl["cod"].year
        end_op_year = tl["end_op"].year
        
        # Add vertical lines for period boundaries
        fig2.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB (Construction Start)", annotation_position="top")
        fig2.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD (Operation Start)", annotation_position="top")
        fig2.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End of Operation", annotation_position="top")
        
        fig2.update_layout(height=500, margin=dict(l=10, r=10, t=40, b=10), title="Cumulative Levered Cash Flow (After-Tax)")
        st.plotly_chart(fig2, width='stretch')
    else:
        st.warning("⚠️ Chart unavailable")
    
    # Comparison chart: Unlevered After-Tax vs Levered After-Tax
    st.markdown("### Unlevered After-Tax vs Levered After-Tax Cash Flow Comparison")
    compare_df = annual_view[["Year"]].copy()
    unlevered_col = "Unlevered CF After Tax (COP)" if currency == "COP" else "Unlevered CF After Tax (USD)"
    if unlevered_col not in annual_view.columns:
        unlevered_col = "Unlevered CF After Tax (COP)"
    levered_col = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if levered_col not in annual_view.columns:
        levered_col = "Levered CF (After-tax, COP)"
    compare_df["Unlevered After-Tax"] = annual_view[unlevered_col]
    compare_df["Levered After-Tax"] = annual_view[levered_col]
    compare_long = compare_df.melt(id_vars=["Year"], value_vars=["Unlevered After-Tax", "Levered After-Tax"], var_name="Type", value_name="Cash Flow")
    fig3 = px.bar(compare_long, x="Year", y="Cash Flow", color="Type", barmode="group")
    
    # Add period indicators
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    # Add vertical lines for period boundaries
    fig3.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig3.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig3.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig3.update_layout(height=360, margin=dict(l=10, r=10, t=40, b=10))
    st.plotly_chart(fig3, width='stretch')
    
    # Monthly table (optional, collapsed)
    with st.expander("Monthly levered cash flow (detailed)"):
        m_disp = monthly_levered.copy()
        m_money = [c for c in m_disp.columns if c not in ["Month", "Year", "Phase"] and "(COP)" in c]
        for c in m_money:
            if c in m_disp.columns:
                m_disp[c] = _conv(m_disp[c], m_disp["Year"])
                if currency == "USD":
                    new_name = c.replace("(COP)", "(USD)")
                    m_disp = m_disp.rename(columns={c: new_name})
        
        m_disp = _df_format_money(m_disp, [c for c in m_disp.columns if c not in ["Month", "Year", "Phase"]], decimals=0)
        st.dataframe(m_disp, width='stretch', hide_index=True)


# -----------------------------
# M) Compare
# -----------------------------
with tab_compare:
    st.subheader("Scenario comparison")

    if not compare_scenarios:
        st.warning("Select scenarios to compare in the sidebar.")
    else:
        rows = []
        for nm in compare_scenarios:
            sd = _scenario_from_dict(proj["scenarios"][nm])

            rev = operating_year_table(sd)
            total_cap = _total_capex_from_lines(sd)

            om = opex_monthly_schedule(sd)
            annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
            annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
            total_opex_op = float(annual_opex["Total OPEX (COP)"].sum())

            rows.append(
                {
                    "Scenario": nm,
                    "Total CAPEX (COP)": total_cap,
                    "CAPEX/MWac (COP)": total_cap / float(sd.generation.mwac) if float(sd.generation.mwac) > 0 else np.nan,
                    "Total OPEX (COP)": total_opex_op,
                    "Total Revenue (COP)": float(rev["Revenue (COP)"].sum()),
                }
            )

        summary = pd.DataFrame(rows).sort_values("Scenario")
        disp = summary.copy()
        for col in ["Total CAPEX (COP)", "CAPEX/MWac (COP)", "Total OPEX (COP)", "Total Revenue (COP)"]:
            if col in disp.columns:
                disp[col] = disp[col].apply(lambda v: _fmt_num(float(v), 0) if pd.notnull(v) else "")
        st.dataframe(disp, width='stretch', hide_index=True)


# -----------------------------
# PDF Chart Generation Helpers
# -----------------------------
def _get_delphi_chart_style() -> dict:
    """Get consistent Delphi chart styling configuration.
    
    Returns:
        Dictionary with styling parameters matching Plotly appearance
    """
    return {
        'accent_color': '#1f4e79',  # Delphi blue
        'secondary_colors': ['#4a90a4', '#7db3c1'],  # Lighter blues
        'font_family': 'Arial',  # or 'sans-serif'
        'title_fontsize': 14,
        'axis_label_fontsize': 11,
        'tick_fontsize': 9,
        'line_width': 2.5,
        'marker_size': 8,
        'grid_color': '#e0e0e0',
        'grid_alpha': 0.3,
        'background_color': 'white',
        'text_color': '#333333',
        'tick_color': '#666666',
        'spine_color': '#cccccc',
    }


def _create_chart_image(fig, width_inches=5.0, dpi=275) -> BytesIO:
    """Convert matplotlib figure to PNG image for PDF embedding.
    
    Args:
        fig: matplotlib figure object
        width_inches: Desired width in inches for the chart
        
    Returns:
        BytesIO object containing PNG image data, or None if matplotlib unavailable
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    # Note: Figure size should be set before calling this function
    # Save to BytesIO with high resolution
    img_buffer = BytesIO()
    fig.savefig(img_buffer, format='png', dpi=dpi, bbox_inches='tight', 
                pad_inches=0.05, facecolor='white', edgecolor='none')
    img_buffer.seek(0)
    plt.close(fig)  # Close figure to free memory
    
    return img_buffer


def _style_chart(ax, title, xlabel, ylabel, accent_color=None):
    """Apply Delphi visual identity styling to matplotlib chart.
    
    Args:
        ax: matplotlib axes object
        title: Chart title
        xlabel: X-axis label
        ylabel: Y-axis label
        accent_color: Accent color for title (defaults to Delphi blue from style config)
    """
    style = _get_delphi_chart_style()
    if accent_color is None:
        accent_color = style['accent_color']
    
    ax.set_title(title, fontsize=style['title_fontsize'], fontweight='bold', 
                 color=accent_color, pad=10, fontfamily=style['font_family'])
    ax.set_xlabel(xlabel, fontsize=style['axis_label_fontsize'], 
                  color=style['text_color'], fontfamily=style['font_family'])
    ax.set_ylabel(ylabel, fontsize=style['axis_label_fontsize'], 
                  color=style['text_color'], fontfamily=style['font_family'])
    ax.grid(True, alpha=style['grid_alpha'], linestyle='--', linewidth=0.5, 
            color=style['grid_color'])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color(style['spine_color'])
    ax.spines['bottom'].set_color(style['spine_color'])
    ax.tick_params(colors=style['tick_color'], labelsize=style['tick_fontsize'])
    ax.set_facecolor(style['background_color'])


# -----------------------------
# Shared Chart Builder Functions (Plotly - Exact App Charts)
# -----------------------------
def _build_timeline_chart_plotly(s: ScenarioInputs) -> 'go.Figure':
    """Build timeline Gantt chart - exact copy of app chart code.
    
    Returns:
        Plotly figure object
    """
    tl = build_timeline(s.timeline)
    
    gantt = pd.DataFrame(
        [
            {"Stage": "Development", "Start": date(tl["start"].year, tl["start"].month, 1), "Finish": date(tl["rtb"].year, tl["rtb"].month, 1)},
            {"Stage": "Construction", "Start": date(tl["rtb"].year, tl["rtb"].month, 1), "Finish": date(tl["cod"].year, tl["cod"].month, 1)},
            {"Stage": "Operation", "Start": date(tl["cod"].year, tl["cod"].month, 1), "Finish": date(tl["end_op"].year, tl["end_op"].month, 1)},
        ]
    )
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    fig = px.timeline(
        gantt,
        x_start="Start",
        x_end="Finish",
        y="Stage",
        color="Stage",
        category_orders={"Stage": ["Development", "Construction", "Operation"]},
    )
    fig.update_yaxes(autorange="reversed")
    fig.update_xaxes(dtick="M12", tickformat="%Y")
    fig.update_layout(height=280, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
    
    return fig


def _build_generation_chart_plotly(s: ScenarioInputs) -> 'go.Figure':
    """Build annual energy generation chart - exact copy of app chart code.
    
    Returns:
        Plotly figure object
    """
    op = operating_year_table(s)
    if len(op) == 0 or "Energy (MWh)" not in op.columns:
        return None
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    fig = px.line(op, x="Year", y="Energy (MWh)")
    fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
    
    return fig


def _build_price_chart_plotly(s: ScenarioInputs, currency: str) -> 'go.Figure':
    """Build PPA price evolution chart.
    
    Returns:
        Plotly figure object
    """
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    op = operating_year_table(s)
    if len(op) == 0 or "Price (COP/kWh)" not in op.columns:
        return None
    
    fig = px.line(op, x="Year", y="Price (COP/kWh)")
    fig.update_layout(
        height=320,
        margin=dict(l=10, r=10, t=40, b=10),
        title=f"PPA Price Evolution ({currency})"
    )
    
    return fig


def _build_revenue_chart_plotly(s: ScenarioInputs, currency: str, _to_usd_pdf) -> 'go.Figure':
    """Build annual revenue bar chart.
    
    Returns:
        Plotly figure object
    """
    op = operating_year_table(s)
    if len(op) == 0 or "Revenue (COP)" not in op.columns:
        return None
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    revenue_col = "Revenue (COP)" if currency == "COP" else "Revenue (USD)"
    if currency == "USD":
        op["Revenue (USD)"] = op.apply(lambda row: _to_usd_pdf(float(row["Revenue (COP)"]), int(row["Year"])), axis=1)
    
    fig = px.bar(op, x="Year", y=revenue_col)
    fig.update_layout(
        height=320,
        margin=dict(l=10, r=10, t=40, b=10),
        title=f"Annual Revenue ({currency})"
    )
    
    return fig


def _build_capex_pie_chart_plotly(s: ScenarioInputs) -> 'go.Figure':
    """Build CAPEX breakdown pie chart - exact copy of app chart code.
    
    Returns:
        Plotly figure object
    """
    capex_df = pd.DataFrame(s.capex.lines or [])
    if "Amount_COP" not in capex_df.columns:
        capex_df["Amount_COP"] = 0.0
    capex_pie = capex_df.copy()
    capex_pie["Item"] = capex_pie["Item"].fillna("").astype(str).str.strip()
    capex_pie["Amount_COP"] = pd.to_numeric(capex_pie["Amount_COP"], errors="coerce").fillna(0.0)
    capex_pie = capex_pie[capex_pie["Amount_COP"] > 0].copy()
    
    if capex_pie.empty:
        return None
    
    share = capex_pie["Amount_COP"] / capex_pie["Amount_COP"].sum()
    small = share < 0.03
    if small.any() and (~small).any():
        other_amt = float(capex_pie.loc[small, "Amount_COP"].sum())
        capex_pie = capex_pie.loc[~small, ["Item", "Amount_COP"]]
        capex_pie = pd.concat(
            [capex_pie, pd.DataFrame([{"Item": "Other (<3% each)", "Amount_COP": other_amt}])],
            ignore_index=True
        )
    
    fig_pie = px.pie(capex_pie, names="Item", values="Amount_COP", hole=0.45)
    fig_pie.update_traces(textinfo="percent+label")
    fig_pie.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
    
    return fig_pie


def _build_ebitda_chart_plotly(s: ScenarioInputs, currency: str, _to_usd_pdf) -> 'go.Figure':
    """Build annual EBITDA bar chart.
    
    Returns:
        Plotly figure object
    """
    annual_unlevered = unlevered_base_cashflow_annual(s)
    if len(annual_unlevered) == 0 or "EBITDA (COP)" not in annual_unlevered.columns:
        return None
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    ebitda_col = "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)"
    if currency == "USD":
        annual_unlevered["EBITDA (USD)"] = annual_unlevered.apply(
            lambda row: _to_usd_pdf(float(row["EBITDA (COP)"]), int(row["Year"])), axis=1)
    
    fig = px.bar(annual_unlevered, x="Year", y=ebitda_col)
    fig.update_layout(
        height=320,
        margin=dict(l=10, r=10, t=40, b=10),
        title=f"Annual EBITDA ({currency})"
    )
    
    return fig


def _build_cashflow_chart_plotly(s: ScenarioInputs, currency: str, _to_usd_pdf) -> 'go.Figure':
    """Build cumulative levered cash flow chart - based on app FCF chart code.
    
    Returns:
        Plotly figure object
    """
    annual_levered = levered_cashflow_annual(s)
    if len(annual_levered) == 0:
        return None
    
    # Prepare annual view for charts
    annual_view = annual_levered.copy()
    money_cols = [c for c in annual_view.columns if c != "Year" and "(COP)" in c]
    for col in money_cols:
        annual_view[col] = pd.to_numeric(annual_view[col], errors="coerce").fillna(0.0)
        if currency == "USD":
            usd_col = col.replace("(COP)", "(USD)")
            annual_view[usd_col] = annual_view.apply(
                lambda row: _to_usd_pdf(float(row[col]), int(row["Year"])), axis=1
            )
    
    y_fcf = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if y_fcf not in annual_view.columns:
        y_fcf = "Levered CF (After-tax, COP)"
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    # Calculate cumulative
    annual_view["Cumulative CF"] = annual_view[y_fcf].cumsum()
    
    fig = px.line(annual_view, x="Year", y="Cumulative CF")
    
    # Add period indicators
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    fig.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig.update_layout(
        height=320,
        margin=dict(l=10, r=10, t=40, b=10),
        title=f"Cumulative Levered Cash Flow ({currency})"
    )
    
    return fig


def _build_dscr_chart_plotly(s: ScenarioInputs) -> 'go.Figure':
    """Build DSCR chart with threshold line.
    
    Returns:
        Plotly figure object
    """
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    if not debt_enabled:
        return None
    
    debt_sched = debt_schedule_annual(s)
    if "DSCR" not in debt_sched.columns or len(debt_sched) == 0:
        return None
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    # Filter out non-operation years (DSCR <= 0)
    dscr_df = debt_sched[debt_sched["DSCR"] > 0].copy()
    if len(dscr_df) == 0:
        return None
    
    fig = px.line(dscr_df, x="Year", y="DSCR", markers=True)
    
    # Add threshold line at 1.0
    fig.add_hline(y=1.0, line_dash="dash", line_color="red", 
                  annotation_text="Minimum (1.0)", annotation_position="right")
    
    fig.update_layout(
        height=320,
        margin=dict(l=10, r=10, t=40, b=10),
        title="Debt Service Coverage Ratio (DSCR)"
    )
    
    return fig


def _build_capital_structure_chart_plotly(s: ScenarioInputs) -> 'go.Figure':
    """Build capital structure pie chart (Debt vs Equity).
    
    Returns:
        Plotly figure object
    """
    total_capex = _total_capex_from_lines(s)
    if total_capex <= 0:
        return None
    
    annual_levered_for_equity = levered_cashflow_annual(s)
    total_equity_investment = 0.0
    for _, row in annual_levered_for_equity.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)
    
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_debt = 0.0
    if debt_enabled:
        debt_pct_of_capex = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
        total_debt = (debt_pct_of_capex / 100.0) * total_capex
    
    labels = []
    sizes = []
    if total_debt > 0:
        labels.append('Debt')
        sizes.append(total_debt)
    if total_equity_investment > 0:
        labels.append('Equity')
        sizes.append(total_equity_investment)
    
    if len(sizes) == 0:
        return None
    
    if not PLOTLY_AVAILABLE or px is None:
        return None
    
    fig = px.pie(values=sizes, names=labels, hole=0.45)
    fig.update_traces(textinfo="percent+label")
    fig.update_layout(
        height=400,
        margin=dict(l=10, r=10, t=40, b=10),
        title="Capital Structure",
        legend_title_text=""
    )
    
    return fig


def _export_plotly_to_image(fig, width_px=1200, height_px=None) -> BytesIO:
    """Export Plotly figure to PNG image for PDF embedding.
    
    Args:
        fig: Plotly figure object
        width_px: Image width in pixels (default 1200 for high quality)
        height_px: Image height in pixels (None = auto from aspect ratio)
        
    Returns:
        BytesIO object containing PNG image data, or None if export fails
    """
    if fig is None:
        return None
    
    if not KALEIDO_AVAILABLE:
        return None
    
    try:
        # Update figure size for export
        if height_px is None:
            # Calculate height from figure's aspect ratio
            if hasattr(fig.layout, 'height') and fig.layout.height:
                fig_width = fig.layout.width if hasattr(fig.layout, 'width') and fig.layout.width else 800
                aspect = fig.layout.height / fig_width
                height_px = int(width_px * aspect)
            else:
                height_px = int(width_px * 0.6)  # Default 5:3 ratio
        
        # Temporarily update layout for export (don't modify original)
        original_width = fig.layout.width if hasattr(fig.layout, 'width') else None
        original_height = fig.layout.height if hasattr(fig.layout, 'height') else None
        
        fig.update_layout(width=width_px, height=height_px)
        
        # Export to BytesIO
        img_buffer = BytesIO()
        fig.write_image(img_buffer, format='png', width=width_px, height=height_px, scale=2)
        img_buffer.seek(0)
        
        # Restore original layout
        if original_width is not None or original_height is not None:
            fig.update_layout(width=original_width, height=original_height)
        
        return img_buffer
    except Exception as e:
        # Kaleido not installed or export failed
        # Log the error for debugging (but don't crash the app)
        import sys
        print(f"Warning: Plotly image export failed: {type(e).__name__}: {e}", file=sys.stderr)
        return None


# -----------------------------
# Shared Chart Builder Functions (Matplotlib for PDF - DEPRECATED, use Plotly instead)
# -----------------------------
def _build_timeline_chart_mpl(s: ScenarioInputs) -> 'matplotlib.figure.Figure':
    """Build timeline Gantt chart matching Plotly px.timeline() appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 2.5)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    tl = build_timeline(s.timeline)
    style = _get_delphi_chart_style()
    
    fig, ax = plt.subplots(figsize=(8.0, 2.5))
    
    # Calculate phase durations
    dev_start = tl["start"]
    rtb = tl["rtb"]
    cod = tl["cod"]
    end_op = tl["end_op"]
    
    # Convert dates to months from start
    def months_from_start(d):
        return (d.year - dev_start.year) * 12 + (d.month - dev_start.month)
    
    dev_months = months_from_start(rtb) - months_from_start(dev_start)
    con_months = months_from_start(cod) - months_from_start(rtb)
    op_months = months_from_start(end_op) - months_from_start(cod)
    
    # Create Gantt bars (reversed order to match Plotly)
    phases = ['Operation', 'Construction', 'Development']  # Reversed for matplotlib
    starts = [dev_months + con_months, dev_months, 0]
    durations = [op_months, con_months, dev_months]
    colors_list = [style['secondary_colors'][1], style['secondary_colors'][0], style['accent_color']]
    
    ax.barh(phases, durations, left=starts, color=colors_list, height=0.6, 
            edgecolor='white', linewidth=1)
    ax.invert_yaxis()  # Match Plotly's reversed y-axis
    
    _style_chart(ax, "Project Timeline", "Months from Start", "Phase")
    ax.set_xlim(0, max(starts) + max(durations))
    
    return fig


def _build_generation_chart_mpl(s: ScenarioInputs) -> 'matplotlib.figure.Figure':
    """Build annual energy generation chart matching Plotly px.line() appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 3.5)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    op = operating_year_table(s)
    if len(op) == 0 or "Energy (MWh)" not in op.columns:
        return None
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(8.0, 3.5))
    
    years = op["Year"].astype(int).tolist()
    energy = op["Energy (MWh)"].astype(float).tolist()
    
    ax.plot(years, energy, marker='o', linewidth=style['line_width'], 
            markersize=style['marker_size'], color=style['accent_color'])
    ax.fill_between(years, energy, alpha=0.2, color=style['accent_color'])
    
    _style_chart(ax, "Annual Energy Generation", "Year", "Energy (MWh)")
    
    return fig


def _build_price_chart_mpl(s: ScenarioInputs, currency: str) -> 'matplotlib.figure.Figure':
    """Build PPA price evolution chart matching Plotly line chart appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 3.5)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    op = operating_year_table(s)
    if len(op) == 0 or "Price (COP/kWh)" not in op.columns:
        return None
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(8.0, 3.5))
    
    years = op["Year"].astype(int).tolist()
    prices = op["Price (COP/kWh)"].astype(float).tolist()
    
    ax.plot(years, prices, marker='o', linewidth=style['line_width'], 
            markersize=style['marker_size'], color=style['accent_color'])
    
    _style_chart(ax, f"PPA Price Evolution ({currency})", "Year", "Price (COP/kWh)")
    
    return fig


def _build_revenue_chart_mpl(s: ScenarioInputs, currency: str, _to_usd_pdf) -> 'matplotlib.figure.Figure':
    """Build annual revenue bar chart matching Plotly px.bar() appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 3.5)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    op = operating_year_table(s)
    if len(op) == 0 or "Revenue (COP)" not in op.columns:
        return None
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(8.0, 3.5))
    
    years = op["Year"].astype(int).tolist()
    revenue_col = "Revenue (COP)" if currency == "COP" else "Revenue (USD)"
    if currency == "USD":
        op["Revenue (USD)"] = op.apply(lambda row: _to_usd_pdf(float(row["Revenue (COP)"]), int(row["Year"])), axis=1)
    revenues = op[revenue_col].astype(float).tolist()
    
    ax.bar(years, revenues, color=style['accent_color'], alpha=0.7, 
           edgecolor='white', linewidth=0.5)
    
    _style_chart(ax, f"Annual Revenue ({currency})", "Year", f"Revenue ({currency})")
    
    return fig


def _build_capex_pie_chart_mpl(s: ScenarioInputs) -> 'matplotlib.figure.Figure':
    """Build CAPEX breakdown pie chart matching Plotly px.pie() with hole=0.45.
    
    Returns:
        matplotlib Figure object with figsize=(6.0, 6.0)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    capex_df = pd.DataFrame(s.capex.lines or [])
    if "Amount_COP" not in capex_df.columns or len(capex_df) == 0:
        return None
    
    capex_df["Amount_COP"] = pd.to_numeric(capex_df["Amount_COP"], errors="coerce").fillna(0.0)
    capex_df = capex_df[capex_df["Amount_COP"] > 0].copy()
    if len(capex_df) == 0:
        return None
    
    # Handle small items (<3%) like in app
    share = capex_df["Amount_COP"] / capex_df["Amount_COP"].sum()
    small = share < 0.03
    if small.any() and (~small).any():
        other_amt = float(capex_df.loc[small, "Amount_COP"].sum())
        capex_df = capex_df.loc[~small, ["Item", "Amount_COP"]]
        capex_df = pd.concat(
            [capex_df, pd.DataFrame([{"Item": "Other (<3% each)", "Amount_COP": other_amt}])],
            ignore_index=True
        )
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(6.0, 6.0))
    
    items = capex_df["Item"].astype(str).tolist()
    amounts = capex_df["Amount_COP"].astype(float).tolist()
    
    # Use Delphi color palette
    colors_list = plt.cm.Blues(np.linspace(0.4, 0.9, len(items)))
    
    # Create donut chart (pie with hole)
    wedges, texts, autotexts = ax.pie(amounts, labels=items, autopct='%1.1f%%', 
                                      colors=colors_list, startangle=90,
                                      textprops={'fontsize': style['tick_fontsize'], 
                                                'color': style['text_color'],
                                                'fontfamily': style['font_family']},
                                      pctdistance=0.85,  # Position percentages
                                      wedgeprops=dict(width=0.45))  # Create donut hole
    
    ax.set_title("CAPEX Breakdown", fontsize=style['title_fontsize'], 
                 fontweight='bold', color=style['accent_color'], pad=10,
                 fontfamily=style['font_family'])
    
    return fig


def _build_ebitda_chart_mpl(s: ScenarioInputs, currency: str, _to_usd_pdf) -> 'matplotlib.figure.Figure':
    """Build annual EBITDA bar chart matching Plotly bar chart appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 3.5)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    annual_unlevered = unlevered_base_cashflow_annual(s)
    if len(annual_unlevered) == 0 or "EBITDA (COP)" not in annual_unlevered.columns:
        return None
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(8.0, 3.5))
    
    years = annual_unlevered["Year"].astype(int).tolist()
    ebitda_col = "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)"
    if currency == "USD":
        annual_unlevered["EBITDA (USD)"] = annual_unlevered.apply(
            lambda row: _to_usd_pdf(float(row["EBITDA (COP)"]), int(row["Year"])), axis=1)
    ebitda_values = annual_unlevered[ebitda_col].astype(float).tolist()
    
    ax.bar(years, ebitda_values, color=style['accent_color'], alpha=0.7, 
           edgecolor='white', linewidth=0.5)
    
    _style_chart(ax, f"Annual EBITDA ({currency})", "Year", f"EBITDA ({currency})")
    
    return fig


def _build_cashflow_chart_mpl(s: ScenarioInputs, currency: str, _to_usd_pdf) -> 'matplotlib.figure.Figure':
    """Build cumulative levered cash flow chart matching Plotly appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 4.0)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    annual_levered = levered_cashflow_annual(s)
    if len(annual_levered) == 0 or "Levered CF (After-tax, COP)" not in annual_levered.columns:
        return None
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(8.0, 4.0))
    
    years = annual_levered["Year"].astype(int).tolist()
    cf_col = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if currency == "USD":
        annual_levered["Levered CF (After-tax, USD)"] = annual_levered.apply(
            lambda row: _to_usd_pdf(float(row["Levered CF (After-tax, COP)"]), int(row["Year"])), axis=1)
    
    cf_values = annual_levered[cf_col].astype(float).tolist()
    cumulative_cf = pd.Series(cf_values).cumsum().tolist()
    
    ax.plot(years, cumulative_cf, marker='o', linewidth=style['line_width'], 
            markersize=style['marker_size'], color=style['accent_color'])
    ax.axhline(y=0, color=style['spine_color'], linestyle='--', linewidth=1)
    ax.fill_between(years, cumulative_cf, 0, alpha=0.2, color=style['accent_color'], 
                     where=pd.Series(cumulative_cf) >= 0)
    ax.fill_between(years, cumulative_cf, 0, alpha=0.2, color='#d32f2f', 
                     where=pd.Series(cumulative_cf) < 0)
    
    _style_chart(ax, f"Cumulative Levered Cash Flow ({currency})", "Year", 
                 f"Cumulative CF ({currency})")
    
    return fig


def _build_dscr_chart_mpl(s: ScenarioInputs) -> 'matplotlib.figure.Figure':
    """Build DSCR chart with threshold line matching Plotly appearance.
    
    Returns:
        matplotlib Figure object with figsize=(8.0, 3.5)
    """
    if not MATPLOTLIB_AVAILABLE:
        return None
    
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    if not debt_enabled:
        return None
    
    debt_sched = debt_schedule_annual(s)
    if "DSCR" not in debt_sched.columns or len(debt_sched) == 0:
        return None
    
    style = _get_delphi_chart_style()
    fig, ax = plt.subplots(figsize=(8.0, 3.5))
    
    dscr_years = debt_sched["Year"].astype(int).tolist()
    dscr_values = debt_sched["DSCR"].astype(float).tolist()
    dscr_values = [v if v > 0 else None for v in dscr_values]  # Filter out non-operation years
    
    ax.plot(dscr_years, dscr_values, marker='o', linewidth=style['line_width'], 
            markersize=style['marker_size'], color=style['accent_color'])
    ax.axhline(y=1.0, color='#d32f2f', linestyle='--', linewidth=1.5, label='Minimum (1.0)')
    ax.fill_between(dscr_years, dscr_values, 1.0, alpha=0.2, color=style['accent_color'], 
                     where=pd.Series(dscr_values) >= 1.0)
    ax.fill_between(dscr_years, dscr_values, 1.0, alpha=0.2, color='#d32f2f', 
                     where=pd.Series(dscr_values) < 1.0)
    
    _style_chart(ax, "Debt Service Coverage Ratio (DSCR)", "Year", "DSCR")
    ax.legend(loc='best', fontsize=style['tick_fontsize'], 
              fontfamily=style['font_family'])
    
    return fig


# -----------------------------
# PDF Export Function
# -----------------------------
def generate_summary_pdf(project_name: str, scenario_name: str, s: ScenarioInputs, 
                         currency: str, sensitivity_data: dict = None) -> BytesIO:
    """Generate a professional PDF report following Delphi's visual identity with locked page structure."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")
    
    # ============================================
    # VISUAL IDENTITY CONSTANTS
    # ============================================
    ACCENT_COLOR = colors.HexColor('#1f4e79')  # Delphi brand blue
    FONT_FAMILY = 'Helvetica'
    SECTION_SPACING = 0.4*inch
    BOX_BORDER_WIDTH = 1.5
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                           rightMargin=0.75*inch, leftMargin=0.75*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Define paragraph styles with visual identity
    cover_title_style = ParagraphStyle(
        'CoverTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=ACCENT_COLOR,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName=FONT_FAMILY
    )
    
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=ACCENT_COLOR,
        spaceAfter=15,
        alignment=TA_LEFT,
        fontName=FONT_FAMILY
    )
    
    heading_style = ParagraphStyle(
        'Heading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=ACCENT_COLOR,
        spaceAfter=10,
        spaceBefore=15,
        fontName=f'{FONT_FAMILY}-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'Subheading',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=colors.black,
        spaceAfter=8,
        spaceBefore=10,
        fontName=f'{FONT_FAMILY}-Bold'
    )
    
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=10,
        fontName=FONT_FAMILY,
        spaceAfter=6
    )
    
    # Helper function for table styling (no vertical gridlines)
    def _apply_table_style(table, header_bg=ACCENT_COLOR):
        """Apply Delphi visual identity table styling: no vertical gridlines."""
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), f'{FONT_FAMILY}-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            # Horizontal line under header only
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.white),
            # Body
            ('FONTNAME', (0, 1), (-1, -1), FONT_FAMILY),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            # Text alignment
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # First column left
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),  # Other columns right
        ]))
    
    # Get timeline and FX for currency conversion
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    a = unlevered_base_cashflow_annual(s)
    years_all = list(a["Year"].astype(int).tolist()) if len(a) > 0 else [cod.year]
    fx_series_all = fx_series(s.macro, cod.year, years_all)
    
    def _to_usd_pdf(value_cop: float, year: int = None) -> float:
        if currency == "COP":
            return value_cop
        if year is not None and year in fx_series_all.index:
            return value_cop / float(fx_series_all.loc[year])
        return value_cop / float(s.macro.fx_cop_per_usd_start)
    
    def _fmt_money_pdf(val: float) -> str:
        """Format money with explicit currency."""
        if currency == "COP":
            return f"COP {val:,.0f}"
        else:
            return f"USD {val:,.0f}"
    
    # ============================================
    # PAGE 1: COVER
    # ============================================
    elements.append(Spacer(1, 2*inch))
    
    # Logo (if available) - match app format (simple centered image)
    logo_path = _load_logo_image()
    if logo_path and os.path.exists(logo_path):
        try:
            # Match app: width=400px. At 96 DPI, 400px = 4.17 inches
            # Use 3.5 inches for better fit on letter page (matches app proportion)
            logo_img = Image(logo_path, width=3.5*inch)
            # Center using a simple table with full width and centered alignment
            logo_table = Table([[logo_img]], colWidths=[6.5*inch])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ]))
            elements.append(logo_table)
            elements.append(Spacer(1, 0.3*inch))
        except Exception as e:
            # Silently fail if logo can't be loaded
            pass
    
    # Title
    project_display_name = s.project_overview.project_name if s.project_overview.project_name else project_name
    elements.append(Paragraph(project_display_name, cover_title_style))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph("Financial Model Report", ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.HexColor('#666666'),
        alignment=TA_CENTER,
        fontName=FONT_FAMILY
    )))
    elements.append(Spacer(1, 0.5*inch))
    
    # Metadata
    cover_data = [
        ['Scenario:', scenario_name],
        ['Date:', datetime.now().strftime('%B %d, %Y')],
        ['Currency:', currency]
    ]
    cover_table = Table(cover_data, colWidths=[1.5*inch, 4*inch])
    cover_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), f'{FONT_FAMILY}-Bold'),
        ('FONTNAME', (1, 0), (1, -1), FONT_FAMILY),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
    ]))
    elements.append(cover_table)
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 2: EXECUTIVE SUMMARY
    # ============================================
    elements.append(Paragraph("Executive Summary", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Calculate all executive summary metrics
    total_capex = _total_capex_from_lines(s)
    total_capex_display = _fmt_money_pdf(total_capex if currency == "COP" else _to_usd_pdf(total_capex, cod.year))
    
    annual_levered = levered_cashflow_annual(s)
    total_equity_investment = 0.0
    for _, row in annual_levered.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)
    total_equity_display = _fmt_money_pdf(total_equity_investment if currency == "COP" else _to_usd_pdf(total_equity_investment, cod.year))
    
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_debt = 0.0
    if debt_enabled:
        debt_pct_of_capex = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
        total_debt = (debt_pct_of_capex / 100.0) * total_capex
    total_debt_display = _fmt_money_pdf(total_debt if currency == "COP" else _to_usd_pdf(total_debt, cod.year)) if debt_enabled else "N/A"
    
    # Calculate IRRs
    mm = cashflow_monthly_table(s).copy()
    for col in ["Unlevered CF Pre-tax (COP)", "Unlevered CF (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)
    
    monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist() if "Unlevered CF Pre-tax (COP)" in mm.columns else mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    project_irr = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")
    project_irr_display = f"{project_irr*100:,.2f}%" if np.isfinite(project_irr) else "N/A"
    
    annual_cf_levered = [float(row.get("Levered CF (After-tax, COP)", 0.0)) for _, row in annual_levered.iterrows()]
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    equity_irr_display = f"{irr_annual_equiv*100:,.2f}%" if (np.isfinite(irr_annual_equiv) and debt_enabled) else "N/A"
    
    # Average EBITDA
    op = operating_year_table(s)
    annual_unlevered = unlevered_base_cashflow_annual(s)
    if "EBITDA (COP)" in annual_unlevered.columns:
        ebitda_values = annual_unlevered["EBITDA (COP)"].astype(float)
        avg_ebitda = ebitda_values.mean() if len(ebitda_values) > 0 else 0.0
        avg_ebitda_display = _fmt_money_pdf(avg_ebitda if currency == "COP" else _to_usd_pdf(avg_ebitda, cod.year))
        
        # EBITDA margin
        if "Revenue (COP)" in annual_unlevered.columns:
            revenue_values = annual_unlevered["Revenue (COP)"].astype(float)
            avg_revenue = revenue_values.mean() if len(revenue_values) > 0 else 1.0
            avg_ebitda_margin = (avg_ebitda / avg_revenue * 100.0) if avg_revenue > 0 else 0.0
            avg_ebitda_margin_display = f"{avg_ebitda_margin:.1f}%"
        else:
            avg_ebitda_margin_display = "N/A"
    else:
        avg_ebitda_display = "N/A"
        avg_ebitda_margin_display = "N/A"
    
    # Average DSCR (if leveraged)
    avg_dscr_display = "N/A"
    if debt_enabled:
        debt_sched = debt_schedule_annual(s)
        if "DSCR" in debt_sched.columns:
            dscr_values = debt_sched["DSCR"].astype(float)
            dscr_values = dscr_values[dscr_values > 0]  # Only operation years
            avg_dscr = dscr_values.mean() if len(dscr_values) > 0 else 0.0
            avg_dscr_display = f"{avg_dscr:.2f}" if avg_dscr > 0 else "N/A"
    
    # Initial annual energy production
    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    initial_production = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    initial_production_display = f"{initial_production:,.0f} MWh"
    
    # Create boxed metrics (2 columns, 2 rows)
    exec_metrics = [
        [("Total CAPEX", total_capex_display), ("Total Equity Investment", total_equity_display)],
        [("Total Debt", total_debt_display), ("Equity IRR", equity_irr_display)],
        [("Project IRR", project_irr_display), ("Avg Annual EBITDA", avg_ebitda_display)],
        [("Avg EBITDA Margin", avg_ebitda_margin_display), ("Avg DSCR", avg_dscr_display)],
        [("Initial Annual Production", initial_production_display), ("", "")]
    ]
    
    # Create metric boxes
    for row_metrics in exec_metrics:
        row_elements = []
        for label, value in row_metrics:
            if label:  # Skip empty boxes
                box_data = [[label], [value]]
                box_table = Table(box_data, colWidths=[2.5*inch])
                box_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), ACCENT_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), f'{FONT_FAMILY}-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), FONT_FAMILY),
                    ('FONTSIZE', (0, 1), (-1, -1), 11),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), BOX_BORDER_WIDTH, ACCENT_COLOR),
                ]))
                row_elements.append(box_table)
        
        if row_elements:
            # Create two-column layout
            metric_row = Table([[row_elements[0] if len(row_elements) > 0 else "", 
                                row_elements[1] if len(row_elements) > 1 else ""]], 
                              colWidths=[2.75*inch, 2.75*inch])
            metric_row.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(metric_row)
            elements.append(Spacer(1, 0.15*inch))
    
    elements.append(PageBreak())
    
    # Get timeline and FX
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    a = unlevered_base_cashflow_annual(s)
    years_all = list(a["Year"].astype(int).tolist()) if len(a) > 0 else [cod.year]
    fx_series_all = fx_series(s.macro, cod.year, years_all)
    
    def _to_usd_pdf(value_cop: float, year: int = None) -> float:
        if currency == "COP":
            return value_cop
        if year is not None and year in fx_series_all.index:
            return value_cop / float(fx_series_all.loc[year])
        return value_cop / float(s.macro.fx_cop_per_usd_start)
    
    # Helper to format money
    def _fmt_money_pdf(val: float) -> str:
        if currency == "COP":
            return f"COP {val:,.0f}"
        else:
            return f"USD {val:,.0f}"
    
    # ============================================
    # PAGE 3: PROJECT OVERVIEW & TIMELINE
    # ============================================
    elements.append(Paragraph("Project Overview & Timeline", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Project Overview metadata
    overview_data = []
    if s.project_overview.project_name:
        overview_data.append(['Project Name:', s.project_overview.project_name])
    if s.project_overview.country:
        overview_data.append(['Country:', s.project_overview.country])
    if s.project_overview.region_department:
        overview_data.append(['Region/Department:', s.project_overview.region_department])
    if s.project_overview.technology:
        overview_data.append(['Technology:', s.project_overview.technology])
    if s.project_overview.installed_capacity_mw > 0:
        overview_data.append(['Installed Capacity:', f"{s.project_overview.installed_capacity_mw:.2f} MW"])
    
    if overview_data:
        overview_table = Table(overview_data, colWidths=[2*inch, 4*inch])
        _apply_table_style(overview_table)
        elements.append(overview_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Timeline summary
    timeline_data = [
        ['Development Start:', tl["start"].strftime('%B %Y')],
        ['RTB (Ready to Build):', tl["rtb"].strftime('%B %Y')],
        ['COD (Commercial Operation):', cod.strftime('%B %Y')],
        ['Operation End:', tl["end_op"].strftime('%B %Y')],
    ]
    timeline_table = Table(timeline_data, colWidths=[2*inch, 4*inch])
    _apply_table_style(timeline_table)
    elements.append(timeline_table)
    
    # Timeline Gantt Chart
    try:
        fig = _build_timeline_chart_plotly(s)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
            else:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Paragraph("<i>Timeline chart unavailable (kaleido not installed)</i>", normal_style))
        else:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("<i>Timeline chart unavailable</i>", normal_style))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>Timeline chart unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 4: MACROECONOMIC & GENERATION ASSUMPTIONS
    # ============================================
    elements.append(Paragraph("Macroeconomic & Generation Assumptions", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Macroeconomic assumptions
    elements.append(Paragraph("Macroeconomic Assumptions", subheading_style))
    macro_data = [
        ['Colombia CPI:', f"{s.macro.col_cpi:.2f}%"],
        ['Colombia PPI:', f"{s.macro.col_ppi:.2f}%"],
        ['FX Rate (Start):', f"{s.macro.fx_cop_per_usd_start:,.2f} COP/USD"],
        ['FX Method:', s.macro.fx_method],
    ]
    if s.macro.fx_method == "Flat":
        macro_data.append(['FX Flat Rate:', f"{s.macro.fx_flat:,.2f} COP/USD"])
    macro_table = Table(macro_data, colWidths=[2.5*inch, 3.5*inch])
    _apply_table_style(macro_table)
    elements.append(macro_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Generation assumptions
    elements.append(Paragraph("Generation Assumptions", subheading_style))
    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    base_mwh = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    degr_pct = float(gen.degradation_pct)
    
    gen_data = [
        ['Production Scenario:', gen.production_choice],
        ['Initial Production:', f"{base_mwh:,.0f} MWh/yr"],
        ['Annual Degradation:', f"{degr_pct:.2f}%/yr"]
    ]
    gen_table = Table(gen_data, colWidths=[2.5*inch, 3.5*inch])
    _apply_table_style(gen_table)
    elements.append(gen_table)
    
    # Generation Chart
    try:
        fig = _build_generation_chart_plotly(s)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
            else:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Paragraph("<i>Generation chart unavailable (kaleido not installed)</i>", normal_style))
        else:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("<i>Generation data unavailable</i>", normal_style))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>Generation chart unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 5: COMMERCIAL STRUCTURE & REVENUES
    # ============================================
    elements.append(Paragraph("Commercial Structure & Revenues", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Commercial structure
    elements.append(Paragraph("Commercial Structure", subheading_style))
    if s.revenue_mode == "Standard PPA Parameters":
        r = s.revenue1
        ppa_price = float(r.ppa_price_cop_per_kwh)
        indexation = str(r.indexation)
        ppa_term = int(r.ppa_term_years)
        merchant_price = float(r.merchant_price_cop_per_kwh)
        contract_type = s.ppa_contract.ui_label_en if hasattr(s.ppa_contract, 'ui_label_en') else "Pay-as-Generated (100%)"
        
        comm_data = [
            ['PPA Contract Type:', contract_type],
            ['PPA Price:', f"{ppa_price:,.4f} COP/kWh"],
            ['PPA Term:', f"{ppa_term} years"],
            ['Indexation:', indexation],
            ['Merchant Price:', f"{merchant_price:,.4f} COP/kWh"],
        ]
    else:
        r = s.revenue2
        indexation = str(r.indexation)
        comm_data = [
            ['Revenue Mode:', 'Manual Annual Series'],
            ['Indexation:', indexation],
        ]
    
    comm_table = Table(comm_data, colWidths=[2.5*inch, 3.5*inch])
    _apply_table_style(comm_table)
    elements.append(comm_table)
    
    # Price and Revenue Charts
    try:
        # Price Chart
        fig = _build_price_chart_plotly(s, currency)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
        
        # Revenue Chart
        fig = _build_revenue_chart_plotly(s, currency, _to_usd_pdf)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>Price/revenue charts unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 6: CAPEX SUMMARY
    # ============================================
    elements.append(Paragraph("CAPEX Summary", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    total_capex = _total_capex_from_lines(s)
    capex_display = _fmt_money_pdf(total_capex if currency == "COP" else _to_usd_pdf(total_capex, cod.year))
    elements.append(Paragraph(f"Total CAPEX: {capex_display}", subheading_style))
    elements.append(Spacer(1, 0.15*inch))
    
    # CAPEX breakdown table
    capex_df = pd.DataFrame(s.capex.lines or [])
    if "Amount_COP" in capex_df.columns and len(capex_df) > 0:
        capex_df["Amount_COP"] = pd.to_numeric(capex_df["Amount_COP"], errors="coerce").fillna(0.0)
        capex_df = capex_df[capex_df["Amount_COP"] > 0].copy()
        if len(capex_df) > 0:
            capex_table_data = [['Item', f'Amount ({currency})']]
            for _, row in capex_df.iterrows():
                item = str(row.get("Item", ""))
                amt = float(row.get("Amount_COP", 0.0))
                amt_display = _fmt_money_pdf(amt if currency == "COP" else _to_usd_pdf(amt, cod.year))
                capex_table_data.append([item, amt_display])
            
            capex_table = Table(capex_table_data, colWidths=[4*inch, 2*inch])
            _apply_table_style(capex_table)
            elements.append(capex_table)
    
    # CAPEX Breakdown Pie Chart
    try:
        fig = _build_capex_pie_chart_plotly(s)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=800)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=4.5*inch, height=4.5*inch, preserveAspectRatio=True))
            else:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Paragraph("<i>CAPEX chart unavailable (kaleido not installed)</i>", normal_style))
        else:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("<i>No CAPEX data available</i>", normal_style))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>CAPEX chart unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 7: OPEX, DEPRECIATION & TAX INCENTIVES
    # ============================================
    elements.append(Paragraph("OPEX, Depreciation & Tax Incentives", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # OPEX summary
    elements.append(Paragraph("Operating Costs Summary", subheading_style))
    op = operating_year_table(s)
    om = opex_monthly_schedule(s)
    annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
    annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
    
    if "Revenue (COP)" in op.columns and "Total OPEX (COP)" in annual_opex.columns:
        merged = op[["Year", "Revenue (COP)"]].merge(annual_opex[["Year", "Total OPEX (COP)"]], on="Year", how="inner")
        total_revenue = merged["Revenue (COP)"].sum()
        total_opex = merged["Total OPEX (COP)"].sum()
        avg_opex_pct = (total_opex / total_revenue * 100.0) if total_revenue > 0 else 0.0
    else:
        avg_opex_pct = 0.0
    
    annual_sga = sga_annual_by_item(s)
    if "Total SG&A (COP)" in annual_sga.columns and "Revenue (COP)" in op.columns:
        merged_sga = op[["Year", "Revenue (COP)"]].merge(annual_sga[["Year", "Total SG&A (COP)"]], on="Year", how="inner")
        total_revenue_sga = merged_sga["Revenue (COP)"].sum()
        total_sga = merged_sga["Total SG&A (COP)"].sum()
        avg_sga_pct = (total_sga / total_revenue_sga * 100.0) if total_revenue_sga > 0 else 0.0
    else:
        avg_sga_pct = 0.0
    
    opex_data = [
        ['Average OPEX / Revenue:', f"{avg_opex_pct:.2f}%"],
        ['Average SG&A / Revenue:', f"{avg_sga_pct:.2f}%"]
    ]
    opex_table = Table(opex_data, colWidths=[2.5*inch, 3.5*inch])
    _apply_table_style(opex_table)
    elements.append(opex_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Depreciation
    elements.append(Paragraph("Depreciation", subheading_style))
    dep = depreciation_annual_table(s)
    dep_display = dep[["Year", "Depreciation (COP)"]].copy()
    if currency == "USD":
        dep_display["Depreciation (USD)"] = dep_display.apply(
            lambda row: _to_usd_pdf(float(row["Depreciation (COP)"]), int(row["Year"])), axis=1
        )
        dep_display = dep_display.drop(columns=["Depreciation (COP)"])
        dep_col = "Depreciation (USD)"
    else:
        dep_col = "Depreciation (COP)"
    
    dep_table_data = [['Year', f'Depreciation ({currency})']]
    for _, row in dep_display.iterrows():
        year = int(row["Year"])
        dep_val = float(row[dep_col])
        dep_table_data.append([str(year), _fmt_money_pdf(dep_val)])
    
    dep_table = Table(dep_table_data, colWidths=[1.5*inch, 4.5*inch])
    _apply_table_style(dep_table)
    elements.append(dep_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Tax Benefits
    elements.append(Paragraph("Renewable Tax Benefits", subheading_style))
    incentives_enabled = bool(getattr(s.incentives, "enable_special_deduction", False))
    tax_data = [['Renewable Tax Benefits Applied:', 'Yes' if incentives_enabled else 'No']]
    if incentives_enabled:
        ded_pct = float(getattr(s.incentives, "special_deduction_pct_of_capex", 0.0))
        tax_data.append(['Special Deduction:', f"{ded_pct:.1f}% of CAPEX"])
    tax_table = Table(tax_data, colWidths=[2.5*inch, 3.5*inch])
    _apply_table_style(tax_table)
    elements.append(tax_table)
    
    # EBITDA Chart
    try:
        fig = _build_ebitda_chart_plotly(s, currency, _to_usd_pdf)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
            else:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Paragraph("<i>EBITDA chart unavailable (kaleido not installed)</i>", normal_style))
        else:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("<i>EBITDA data unavailable</i>", normal_style))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>EBITDA chart unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 8: CAPITAL STRUCTURE
    # ============================================
    elements.append(Paragraph("Capital Structure", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Sources & Uses
    elements.append(Paragraph("Sources & Uses of Capital", subheading_style))
    total_capex = _total_capex_from_lines(s)
    total_capex_display = _fmt_money_pdf(total_capex if currency == "COP" else _to_usd_pdf(total_capex, cod.year))
    
    annual_levered_for_equity = levered_cashflow_annual(s)
    total_equity_investment = 0.0
    for _, row in annual_levered_for_equity.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)
    total_equity_display = _fmt_money_pdf(total_equity_investment if currency == "COP" else _to_usd_pdf(total_equity_investment, cod.year))
    
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_debt = 0.0
    if debt_enabled:
        debt_pct_of_capex = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
        total_debt = (debt_pct_of_capex / 100.0) * total_capex
    total_debt_display = _fmt_money_pdf(total_debt if currency == "COP" else _to_usd_pdf(total_debt, cod.year)) if debt_enabled else "N/A"
    
    sources_uses_data = [
        ['Uses:', ''],
        ['Total CAPEX', total_capex_display],
        ['', ''],
        ['Sources:', ''],
        ['Debt', total_debt_display if debt_enabled else "N/A"],
        ['Equity', total_equity_display],
    ]
    sources_uses_table = Table(sources_uses_data, colWidths=[3*inch, 3*inch])
    _apply_table_style(sources_uses_table)
    elements.append(sources_uses_table)
    
    # Capital Structure Pie Chart
    try:
        fig = _build_capital_structure_chart_plotly(s)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=800)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=4.5*inch, height=4.5*inch, preserveAspectRatio=True))
            else:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Paragraph("<i>Capital structure chart unavailable (kaleido not installed)</i>", normal_style))
        else:
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("<i>No capital structure data</i>", normal_style))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>Capital structure chart unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 9: FINANCIAL RESULTS SUMMARY
    # ============================================
    elements.append(Paragraph("Financial Results Summary", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Key metrics summary
    elements.append(Paragraph("Key Financial Metrics", subheading_style))
    
    # Calculate IRRs (already calculated for executive summary, but recalculate for clarity)
    mm = cashflow_monthly_table(s).copy()
    for col in ["Unlevered CF Pre-tax (COP)", "Unlevered CF (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)
    
    monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist() if "Unlevered CF Pre-tax (COP)" in mm.columns else mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")
    
    annual_levered = levered_cashflow_annual(s)
    annual_cf_levered = [float(row.get("Levered CF (After-tax, COP)", 0.0)) for _, row in annual_levered.iterrows()]
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    monthly_levered = levered_cashflow_monthly(s)
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    metrics_data = [
        ['Metric', 'Value'],
        ['Project IRR (Unlevered, Pre-tax)', f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"],
        ['Equity IRR (Levered, After-tax)', f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—"],
        ['Payback Period (years)', f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—"]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
    _apply_table_style(metrics_table)
    elements.append(metrics_table)
    
    # Cash Flow Chart
    try:
        # Cumulative Cash Flow
        fig = _build_cashflow_chart_plotly(s, currency, _to_usd_pdf)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
        
        # DSCR Chart (if leveraged)
        fig = _build_dscr_chart_plotly(s)
        if fig:
            chart_img = _export_plotly_to_image(fig, width_px=1200)
            if chart_img:
                elements.append(Spacer(1, 0.2*inch))
                elements.append(Image(chart_img, width=6.5*inch, height=4.5*inch, preserveAspectRatio=True))
    except Exception:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<i>Cash flow charts unavailable</i>", normal_style))
    
    elements.append(PageBreak())
    
    # ============================================
    # PAGE 10+: APPENDIX - DETAILED TABLES
    # ============================================
    elements.append(Paragraph("Appendix - Detailed Tables", title_style))
    elements.append(Spacer(1, SECTION_SPACING))
    
    # Annual Operating Table
    elements.append(Paragraph("Annual Operating Results", subheading_style))
    op = operating_year_table(s)
    if len(op) > 0:
        op_display_cols = ["Year"]
        if "Price (COP/kWh)" in op.columns:
            op_display_cols.append("Price (COP/kWh)")
        if "Energy (MWh)" in op.columns:
            op_display_cols.append("Energy (MWh)")
        if "Revenue (COP)" in op.columns:
            op_display_col = "Revenue (COP)" if currency == "COP" else "Revenue (USD)"
            if currency == "USD":
                op["Revenue (USD)"] = op.apply(lambda row: _to_usd_pdf(float(row["Revenue (COP)"]), int(row["Year"])), axis=1)
            op_display_cols.append(op_display_col)
        
        op_display = op[op_display_cols].copy()
        op_table_data = [[col.replace(" (COP)", "").replace(" (USD)", "") for col in op_display_cols]]
        for _, row in op_display.iterrows():
            op_row = [str(int(row["Year"]))]
            for col in op_display_cols[1:]:
                if "Price" in col:
                    op_row.append(f"{float(row[col]):,.4f}")
                elif "Energy" in col:
                    op_row.append(f"{float(row[col]):,.0f}")
                else:
                    val = float(row[col]) if pd.notnull(row[col]) else 0.0
                    op_row.append(_fmt_money_pdf(val))
            op_table_data.append(op_row)
        
        if len(op_table_data) > 1:
            op_table = Table(op_table_data, colWidths=[0.8*inch] + [1.2*inch] * (len(op_display_cols) - 1))
            _apply_table_style(op_table)
            elements.append(op_table)
            elements.append(Spacer(1, 0.3*inch))
    
    elements.append(PageBreak())
    
    # Income Statement
    elements.append(Paragraph("Income Statement (Levered)", subheading_style))
    annual_view = annual_levered.copy()
    money_cols = [c for c in annual_view.columns if c != "Year" and "(COP)" in c]
    for col in money_cols:
        annual_view[col] = pd.to_numeric(annual_view[col], errors="coerce").fillna(0.0)
        if currency == "USD":
            usd_col = col.replace("(COP)", "(USD)")
            annual_view[usd_col] = annual_view.apply(
                lambda row: _to_usd_pdf(float(row[col]), int(row["Year"])), axis=1
            )
    
    display_cols = [
        "Year",
        "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Levered CAPEX Tax Deduction (COP)" if currency == "COP" else "Levered CAPEX Tax Deduction (USD)",
        "Levered Taxable Income (COP)" if currency == "COP" else "Levered Taxable Income (USD)",
        "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    income_df = annual_view[display_cols].copy()
    
    income_table_data = [[col.replace(" (COP)", "").replace(" (USD)", "") for col in display_cols]]
    for _, row in income_df.iterrows():
        income_row = []
        for col in display_cols:
            if col == "Year":
                income_row.append(str(int(row[col])))
            else:
                val = float(row[col]) if pd.notnull(row[col]) else 0.0
                income_row.append(_fmt_money_pdf(val))
        income_table_data.append(income_row)
    
    num_cols = len(display_cols)
    col_width = 5.5*inch / num_cols
    
    income_table = Table(income_table_data, colWidths=[col_width] * num_cols)
    _apply_table_style(income_table)
    elements.append(income_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Cash Flow Tables
    elements.append(PageBreak())
    elements.append(Paragraph("Unlevered Cash Flow", subheading_style))
    unlevered_annual = unlevered_base_cashflow_annual(s)
    if len(unlevered_annual) > 0:
        ucf_cols = ["Year"]
        if "Revenue (COP)" in unlevered_annual.columns:
            ucf_cols.append("Revenue (COP)" if currency == "COP" else "Revenue (USD)")
            if currency == "USD":
                unlevered_annual["Revenue (USD)"] = unlevered_annual.apply(lambda row: _to_usd_pdf(float(row["Revenue (COP)"]), int(row["Year"])), axis=1)
        if "Total OPEX (COP)" in unlevered_annual.columns:
            ucf_cols.append("Total OPEX (COP)" if currency == "COP" else "Total OPEX (USD)")
            if currency == "USD":
                unlevered_annual["Total OPEX (USD)"] = unlevered_annual.apply(lambda row: _to_usd_pdf(float(row["Total OPEX (COP)"]), int(row["Year"])), axis=1)
        if "Unlevered CF Pre-tax (COP)" in unlevered_annual.columns:
            ucf_cols.append("Unlevered CF Pre-tax (COP)" if currency == "COP" else "Unlevered CF Pre-tax (USD)")
            if currency == "USD":
                unlevered_annual["Unlevered CF Pre-tax (USD)"] = unlevered_annual.apply(lambda row: _to_usd_pdf(float(row["Unlevered CF Pre-tax (COP)"]), int(row["Year"])), axis=1)
        
        ucf_display = unlevered_annual[ucf_cols].copy()
        ucf_table_data = [[col.replace(" (COP)", "").replace(" (USD)", "") for col in ucf_cols]]
        for _, row in ucf_display.iterrows():
            ucf_row = [str(int(row["Year"]))]
            for col in ucf_cols[1:]:
                val = float(row[col]) if pd.notnull(row[col]) else 0.0
                ucf_row.append(_fmt_money_pdf(val))
            ucf_table_data.append(ucf_row)
        
        if len(ucf_table_data) > 1:
            ucf_table = Table(ucf_table_data, colWidths=[0.8*inch] + [1.4*inch] * (len(ucf_cols) - 1))
            _apply_table_style(ucf_table)
            elements.append(ucf_table)
            elements.append(Spacer(1, 0.3*inch))
    
    # Levered Cash Flow
    if debt_enabled:
        elements.append(PageBreak())
        elements.append(Paragraph("Levered Cash Flow", subheading_style))
        levered_cols = ["Year"]
        if "Levered CF (After-tax, COP)" in annual_levered.columns:
            levered_cols.append("Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)")
            if currency == "USD":
                annual_levered["Levered CF (After-tax, USD)"] = annual_levered.apply(lambda row: _to_usd_pdf(float(row["Levered CF (After-tax, COP)"]), int(row["Year"])), axis=1)
        
        levered_display = annual_levered[levered_cols].copy()
        levered_table_data = [[col.replace(" (COP)", "").replace(" (USD)", "") for col in levered_cols]]
        for _, row in levered_display.iterrows():
            levered_row = [str(int(row["Year"]))]
            for col in levered_cols[1:]:
                val = float(row[col]) if pd.notnull(row[col]) else 0.0
                levered_row.append(_fmt_money_pdf(val))
            levered_table_data.append(levered_row)
        
        if len(levered_table_data) > 1:
            levered_table = Table(levered_table_data, colWidths=[0.8*inch] + [2.5*inch])
            _apply_table_style(levered_table)
            elements.append(levered_table)
    
    # Sensitivity Analysis (if provided)
    if sensitivity_data and 'pivot_table' in sensitivity_data:
        elements.append(PageBreak())
        elements.append(Paragraph("Sensitivity Analysis", subheading_style))
        elements.append(Paragraph(f"Variable 1: {sensitivity_data.get('var1_name', 'N/A')}", normal_style))
        elements.append(Paragraph(f"Variable 2: {sensitivity_data.get('var2_name', 'N/A')}", normal_style))
        elements.append(Spacer(1, 0.1*inch))
        
        pivot = sensitivity_data['pivot_table']
        sens_table_data = [[''] + [f"V1={c:.1f}" for c in pivot.columns]]
        for idx, row in pivot.iterrows():
            sens_row = [f"V2={idx:.1f}"]
            for val in row:
                if np.isfinite(val):
                    sens_row.append(f"{val:.2f}%")
                else:
                    sens_row.append("—")
            sens_table_data.append(sens_row)
        
        # Limit table size for PDF (max 10x10)
        if len(sens_table_data) > 11:
            sens_table_data = sens_table_data[:11]
        if len(sens_table_data[0]) > 11:
            for i in range(len(sens_table_data)):
                sens_table_data[i] = sens_table_data[i][:11]
        
        sens_table = Table(sens_table_data, colWidths=[0.8*inch] + [0.5*inch] * (len(sens_table_data[0]) - 1))
        _apply_table_style(sens_table)
        elements.append(sens_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


# -----------------------------
# Excel Export Helper Functions
# -----------------------------
def _write_formula_cell(ws, cell_ref, formula, number_format=None, fill_color=None, font=None, border=None, alignment=None):
    """Write a formula to a cell with formatting."""
    cell = ws[cell_ref]
    cell.value = formula
    if number_format:
        cell.number_format = number_format
    if fill_color:
        cell.fill = fill_color
    if font:
        cell.font = font
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment

def _write_value_cell(ws, cell_ref, value, number_format=None, fill_color=None, font=None, border=None, alignment=None):
    """Write a value to a cell with formatting."""
    cell = ws[cell_ref]
    cell.value = value
    if number_format:
        cell.number_format = number_format
    if fill_color:
        cell.fill = fill_color
    if font:
        cell.font = font
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment

# -----------------------------
# Excel Export Helper Functions - Sheet Setup
# -----------------------------
def _setup_inputs_sheet(ws, s, INPUT_FILL, HEADER_FILL, HEADER_FONT, TITLE_FONT, SECTION_FILL, THIN_BORDER, CENTER_ALIGN, RIGHT_ALIGN):
    """Create Inputs sheet with all scenario data (values only, no formulas). Returns dict of named range references and cell positions."""
    row = 1
    named_range_refs = {}
    cell_positions = {}  # Track actual cell positions for direct references
    
    def write_section_header(sheet, r, title):
        sheet[f'A{r}'] = title
        sheet[f'A{r}'].font = TITLE_FONT
        sheet[f'A{r}'].fill = SECTION_FILL
        sheet.merge_cells(f'A{r}:D{r}')
        return r + 1
    
    def write_input_row(sheet, r, label, value, named_range_name=None):
        """Write an input row and track the actual cell reference for named ranges."""
        sheet[f'A{r}'] = label
        sheet[f'A{r}'].font = Font(bold=True)
        cell = sheet[f'B{r}']
        # Handle dates explicitly to ensure they're stored correctly
        if isinstance(value, date):
            # Write date as Excel serial number (openpyxl handles conversion)
            cell.value = value
            cell.number_format = "mm/dd/yyyy"
            # Ensure it's recognized as a date type
        else:
            cell.value = value
        cell.fill = INPUT_FILL
        if isinstance(value, (int, float)) and not isinstance(value, date):
            if abs(value) > 1000:
                cell.number_format = '#,##0'
            elif isinstance(value, float):
                cell.number_format = '0.00'
        # CRITICAL: Use the actual row number 'r' passed to this function
        # This ensures named ranges point to the correct cells
        # Verify: the cell we're writing to is B{r}, so the named range should reference B{r}
        cell_ref = f'$B${r}'  # Use absolute reference for named ranges
        if named_range_name:
            # Store the reference - format: "Inputs!$B$7"
            full_ref = f"Inputs!{cell_ref}"
            named_range_refs[named_range_name] = full_ref
            cell_positions[named_range_name] = full_ref  # Also track for direct reference
        return r + 1
    
    # A) Macroeconomic
    # Row 1: Section header
    row = write_section_header(ws, row, "A) Macroeconomic")  # row becomes 2
    # Row 2: Colombia CPI
    row = write_input_row(ws, row, "Colombia CPI (%)", s.macro.col_cpi / 100.0, "ColCPI")  # row becomes 3
    # Row 3: US CPI
    row = write_input_row(ws, row, "US CPI (%)", s.macro.us_cpi / 100.0)  # row becomes 4
    # Row 4: FX Rate
    row = write_input_row(ws, row, "FX Rate (COP/USD) - Start", s.macro.fx_cop_per_usd_start)  # row becomes 5
    row += 1  # row becomes 6
    
    # B) Timeline
    # Row 6: Section header
    row = write_section_header(ws, row, "B) Timeline")  # row becomes 7
    tl = build_timeline(s.timeline)
    # Parse start date and ensure it's a proper date object
    try:
        start_date = date.fromisoformat(s.timeline.start_date)
    except (ValueError, AttributeError):
        # Fallback: try parsing as string or use today's date
        try:
            from datetime import datetime
            start_date = datetime.strptime(s.timeline.start_date, "%Y-%m-%d").date()
        except:
            start_date = date.today()
    # Row 7: Start Date
    row = write_input_row(ws, row, "Start Date", start_date, "StartDate")  # row becomes 8
    # Row 8: Development Months
    row = write_input_row(ws, row, "Development Months", s.timeline.dev_months, "DevMonths")  # row becomes 9
    # Row 9: Construction Months
    row = write_input_row(ws, row, "Construction Months", s.timeline.capex_months, "ConMonths")  # row becomes 10
    # Row 10: Operation Years
    row = write_input_row(ws, row, "Operation Years", s.timeline.operation_years, "OpYears")  # row becomes 11
    row += 1  # row becomes 12
    
    # C) Power Generation
    # Row 12: Section header
    row = write_section_header(ws, row, "C) Power Generation")  # row becomes 13
    # Row 13: Capacity MWac
    row = write_input_row(ws, row, "Capacity (MWac)", s.generation.mwac, "MWac")  # row becomes 14
    # Row 14: Capacity MWp
    row = write_input_row(ws, row, "Capacity (MWp)", s.generation.mwp, "MWp")  # row becomes 15
    # Determine which P value to use based on production choice
    if s.generation.production_choice == "P50":
        p50_value = s.generation.p50_mwh_yr
    elif s.generation.production_choice == "P75":
        p50_value = s.generation.p75_mwh_yr
    else:
        p50_value = s.generation.p90_mwh_yr
    # Row 15: P50
    row = write_input_row(ws, row, f"{s.generation.production_choice} (MWh/year)", p50_value, "P50MWh")  # row becomes 16
    # Row 16: Degradation
    row = write_input_row(ws, row, "Degradation (%/yr)", s.generation.degradation_pct / 100.0, "DegradPct")  # row becomes 17
    row += 1  # row becomes 18
    
    # D) Power Revenues
    row = write_section_header(ws, row, "D) Power Revenues")
    # Always create PPAPrice, PPATerm, and MerchantPrice named ranges
    # Even if revenue_mode is not "Standard PPA Parameters", formulas may reference them
    if s.revenue_mode == "Standard PPA Parameters":
        row = write_input_row(ws, row, "PPA Price (COP/kWh)", s.revenue1.ppa_price_cop_per_kwh, "PPAPrice")
        row = write_input_row(ws, row, "PPA Term (years)", s.revenue1.ppa_term_years, "PPATerm")
        row = write_input_row(ws, row, "Merchant Price (COP/kWh)", s.revenue1.merchant_price_cop_per_kwh, "MerchantPrice")
    else:
        # Create with default values (0) so formulas don't error
        row = write_input_row(ws, row, "PPA Price (COP/kWh)", 0.0, "PPAPrice")
        row = write_input_row(ws, row, "PPA Term (years)", 0, "PPATerm")
        row = write_input_row(ws, row, "Merchant Price (COP/kWh)", 0.0, "MerchantPrice")
    
    # If manual revenue mode, export the price table
    price_table_start_row = None  # Track where price table starts for formulas
    if s.revenue_mode == "Manual annual series":
        row += 1  # Add a spacer
        price_table_start_row = row + 1  # Table data starts after header
        # Write table header
        ws[f'A{row}'] = "Operating Year"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = HEADER_FILL
        ws[f'B{row}'] = "Price (COP/kWh)"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = HEADER_FILL
        row += 1
        # Write price table - get max operating years from timeline
        max_op_years = int(s.timeline.operation_years)
        for op_year in range(1, max_op_years + 1):
            price = float(s.revenue2.prices_constant_cop_per_kwh.get(op_year, 0.0))
            ws[f'A{row}'] = op_year
            ws[f'A{row}'].fill = INPUT_FILL
            ws[f'B{row}'] = price
            ws[f'B{row}'].fill = INPUT_FILL
            ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        price_table_end_row = row - 1
        # Create named range for price table lookup
        named_range_refs["PriceTable"] = f"Inputs!$A${price_table_start_row}:$B${price_table_end_row}"
    else:
        # Create a default empty PriceTable range to prevent formula errors
        # Use a small dummy range that won't cause issues
        row += 1
        ws[f'A{row}'] = "Operating Year"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = HEADER_FILL
        ws[f'B{row}'] = "Price (COP/kWh)"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = HEADER_FILL
        row += 1
        # Create a dummy row with 0 values
        ws[f'A{row}'] = 1
        ws[f'A{row}'].fill = INPUT_FILL
        ws[f'B{row}'] = 0.0
        ws[f'B{row}'].fill = INPUT_FILL
        named_range_refs["PriceTable"] = f"Inputs!$A${row}:$B${row}"
    row += 1
    
    # E) CAPEX - Calculate totals by phase
    row = write_section_header(ws, row, "E) CAPEX")
    capex_df = pd.DataFrame(s.capex.lines or [])
    capex_dev = 0.0
    capex_con = 0.0
    capex_cod = 0.0
    for _, line in capex_df.iterrows():
        amt = float(line.get("Amount_COP", 0.0))
        phase = str(line.get("Phase", "Construction"))
        if "Development" in phase:
            capex_dev += amt
        elif "At COD" in phase or "COD" in phase:
            capex_cod += amt
        else:
            capex_con += amt
    row = write_input_row(ws, row, "Development CAPEX (COP)", capex_dev, "CapexDev")
    row = write_input_row(ws, row, "Construction CAPEX (COP)", capex_con, "CapexCon")
    row = write_input_row(ws, row, "At COD CAPEX (COP)", capex_cod, "CapexCOD")
    row += 1
    
    # F) OPEX
    row = write_section_header(ws, row, "F) OPEX")
    row = write_input_row(ws, row, "Fixed OM (COP/MWac/Year)", s.opex.fixed_om_cop_per_mwac_year, "FixedOM")
    row = write_input_row(ws, row, "Variable OM (COP/MWh)", s.opex.variable_om_cop_per_mwh, "VarOM")
    row = write_input_row(ws, row, "Insurance (COP/MWac/Year)", s.opex.insurance_cop_per_mwac_year, "Insurance")
    row = write_input_row(ws, row, "Grid Fees (COP/MWh)", s.opex.grid_fees_cop_per_mwh, "GridFees")
    row += 1
    
    # K) Debt
    row = write_section_header(ws, row, "K) Debt & Covenants")
    debt_enabled = "Yes" if getattr(s.debt, "enabled", False) else "No"
    row = write_input_row(ws, row, "Debt Enabled", debt_enabled, "DebtEnabled")
    # Always create all debt named ranges
    # Even if debt is disabled, formulas may reference them and should return 0
    if getattr(s.debt, "enabled", False):
        row = write_input_row(ws, row, "Debt % of CAPEX", s.debt.debt_pct_of_capex / 100.0, "DebtPct")
        row = write_input_row(ws, row, "Tenor (years)", s.debt.tenor_years, "DebtTenor")
        row = write_input_row(ws, row, "Grace Period (years)", getattr(s.debt, "grace_years", 0), "DebtGrace")
        all_in_rate = (s.debt.base_rate_pct + s.debt.margin_pct) / 100.0
        row = write_input_row(ws, row, "All-in Interest Rate (%)", all_in_rate, "DebtRate")
        row = write_input_row(ws, row, "Base Rate (%)", s.debt.base_rate_pct / 100.0, "DebtBaseRate")
        row = write_input_row(ws, row, "Margin (%)", s.debt.margin_pct / 100.0, "DebtMargin")
        row = write_input_row(ws, row, "Upfront Fee (bps)", getattr(s.debt, "upfront_fee_bps", 0.0) / 10000.0, "DebtUpfrontFee")
        row = write_input_row(ws, row, "Commitment Fee (% of margin)", getattr(s.debt, "commitment_fee_pct_of_margin", 0.0) / 100.0, "DebtCommitmentFee")
    else:
        # Create with default values (0) so formulas don't error when debt is disabled
        row = write_input_row(ws, row, "Debt % of CAPEX", 0.0, "DebtPct")
        row = write_input_row(ws, row, "Tenor (years)", 0, "DebtTenor")
        row = write_input_row(ws, row, "Grace Period (years)", 0, "DebtGrace")
        row = write_input_row(ws, row, "All-in Interest Rate (%)", 0.0, "DebtRate")
        row = write_input_row(ws, row, "Base Rate (%)", 0.0, "DebtBaseRate")
        row = write_input_row(ws, row, "Margin (%)", 0.0, "DebtMargin")
        row = write_input_row(ws, row, "Upfront Fee (bps)", 0.0, "DebtUpfrontFee")
        row = write_input_row(ws, row, "Commitment Fee (% of margin)", 0.0, "DebtCommitmentFee")
    row += 1
    
    # L) Depreciation
    row = write_section_header(ws, row, "L) Depreciation")
    row = write_input_row(ws, row, "Depreciation % of CAPEX", s.depreciation.pct_of_capex_depreciated / 100.0, "DepPct")
    row = write_input_row(ws, row, "Depreciation Years", s.depreciation.dep_years, "DepYears")
    row += 1
    
    # M) Tax
    row = write_section_header(ws, row, "M) Tax")
    row = write_input_row(ws, row, "Corporate Tax Rate (%)", s.tax.corporate_tax_rate_pct / 100.0, "TaxRate")
    row = write_input_row(ws, row, "Allow Loss Carryforward", "Yes" if s.tax.allow_loss_carryforward else "No", "AllowNOL")
    row += 1
    
    # N) Renewable Tax Benefits
    row = write_section_header(ws, row, "N) Renewable Tax Benefits")
    row = write_input_row(ws, row, "Enable Special Deduction", "Yes" if getattr(s.renewable_tax, "enable_special_deduction", True) else "No", "EnableSpecialDed")
    row = write_input_row(ws, row, "Special Deduction % of CAPEX", getattr(s.renewable_tax, "special_deduction_pct_of_capex", 50.0) / 100.0, "SpecialDedPct")
    row = write_input_row(ws, row, "Special Deduction Years", getattr(s.renewable_tax, "special_deduction_years", 15), "SpecialDedYears")
    row = write_input_row(ws, row, "Special Deduction Max % of Taxable Income", getattr(s.renewable_tax, "special_deduction_max_pct_of_taxable_income", 50.0) / 100.0, "SpecialDedMaxPct")
    row = write_input_row(ws, row, "Enable VAT Refund", "Yes" if getattr(s.renewable_tax, "enable_vat_refund", True) else "No", "EnableVATRefund")
    # Calculate total CAPEX for VAT calculation
    capex_df = pd.DataFrame(s.capex.lines or [])
    total_capex_for_vat = float(capex_df["Amount_COP"].fillna(0).sum()) if (not capex_df.empty and "Amount_COP" in capex_df.columns) else 0.0
    vat_mode = getattr(s.renewable_tax, "vat_refund_mode", "percent")
    if vat_mode == "percent":
        # Store as decimal percentage - will be multiplied by TotalCAPEX in formula
        vat_value = getattr(s.renewable_tax, "vat_pct_of_capex", 19.0) / 100.0
    else:
        # Store as fixed amount (already in COP)
        vat_value = getattr(s.renewable_tax, "vat_fixed_cop", 0.0)
    row = write_input_row(ws, row, "VAT Amount (decimal if %, COP if fixed)", vat_value, "VATAmount")
    row = write_input_row(ws, row, "VAT Mode", vat_mode, "VATMode")
    row = write_input_row(ws, row, "VAT Refund Year (after COD)", getattr(s.renewable_tax, "vat_refund_year", 1), "VATRefundYear")
    row += 1
    
    # J) Working Capital
    row = write_section_header(ws, row, "J) Working Capital")
    row = write_input_row(ws, row, "AR Days (revenue collection lag)", getattr(s.wc, "ar_days", 90), "ARDays")
    row = write_input_row(ws, row, "AP Days (expense payment lag)", getattr(s.wc, "ap_days", 60), "APDays")
    row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    return named_range_refs, cell_positions

def _create_named_ranges(wb, ws_inputs, s, named_range_refs):
    """Create named ranges from Inputs sheet for easy formula references."""
    from openpyxl.workbook.defined_name import DefinedName
    
    # Create named ranges from the tracked references
    # The references are already in format: "Inputs!$B$7"
    # We need to convert to Excel format: "'Inputs'!$B$7"
    for name, ref in named_range_refs.items():
        try:
            # Parse the reference: "Inputs!$B$7" -> sheet_name="Inputs", cell_ref="$B$7"
            if "!" in ref:
                sheet_name, cell_ref = ref.split("!", 1)
                # Excel format requires quotes around sheet name: 'Inputs'!$B$7
                excel_ref = f"'{sheet_name}'!{cell_ref}"
            else:
                # If no sheet name, assume current sheet
                excel_ref = ref
            
            # Create DefinedName with the value attribute (most reliable method)
            dn = DefinedName(name)
            dn.value = excel_ref
            wb.defined_names[name] = dn
        except Exception as e:
            # If named range creation fails, log but continue
            # Formulas can still use direct cell references as fallback
            print(f"Warning: Could not create named range '{name}' with ref '{ref}': {e}")
            pass

def _setup_timeline_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows, start_date_cell=None):
    """Create Timeline_M sheet with monthly date formulas."""
    # Headers
    headers = ["Month", "Year", "Phase", "MonthNum"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Row 2 - first month
    # Use direct cell reference if provided, otherwise use named range
    if start_date_cell:
        ws["A2"] = f"={start_date_cell}"
    else:
        ws["A2"] = "=StartDate"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["A2"].fill = FORMULA_FILL
    ws["B2"] = "=YEAR(A2)"
    ws["B2"].fill = FORMULA_FILL
    ws["C2"] = '=IF(D2<=DevMonths,"Development",IF(D2<=DevMonths+ConMonths,"Construction","Operation"))'
    ws["C2"].fill = FORMULA_FILL
    ws["D2"] = "1"
    ws["D2"].fill = FORMULA_FILL
    
    # Copy formulas down
    for row in range(3, max_rows + 1):
        ws[f"A{row}"] = f"=EDATE(A{row-1},1)"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"A{row}"].fill = FORMULA_FILL
        ws[f"B{row}"] = f"=YEAR(A{row})"
        ws[f"B{row}"].fill = FORMULA_FILL
        ws[f"C{row}"] = f'=IF(D{row}<=DevMonths,"Development",IF(D{row}<=DevMonths+ConMonths,"Construction","Operation"))'
        ws[f"C{row}"].fill = FORMULA_FILL
        ws[f"D{row}"] = f"={row-1}"
        ws[f"D{row}"].fill = FORMULA_FILL
    
    # Format columns - make Month column wider to display dates properly
    ws.column_dimensions["A"].width = 15  # Increased from 12 to display dates properly
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10

def _setup_capex_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows):
    """Create CAPEX_M sheet with CAPEX calculation formulas."""
    headers = ["Month", "Year", "Phase", "CAPEX_Dev", "CAPEX_Con", "CAPEX_COD", "CAPEX_Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Link to Timeline_M
    for row in range(2, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        
        # CAPEX formulas
        ws[f"D{row}"] = f'=IF(C{row}="Development",CapexDev/DevMonths,0)'
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"D{row}"].number_format = "#,##0"
        
        ws[f"E{row}"] = f'=IF(C{row}="Construction",CapexCon/ConMonths,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"E{row}"].number_format = "#,##0"
        
        ws[f"F{row}"] = f'=IF(AND(C{row}="Operation",Timeline_M!D{row}=DevMonths+ConMonths+1),CapexCOD,0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"F{row}"].number_format = "#,##0"
        
        ws[f"G{row}"] = f"=SUM(D{row}:F{row})"
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"G{row}"].number_format = "#,##0"
    
    # Format columns
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

def _setup_revenue_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows):
    """Create Revenue_M sheet with revenue calculation formulas."""
    headers = ["Month", "Year", "Phase", "OperatingYear", "Energy_MWh", "Price_COP_kWh", "Revenue_COP"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # COD month number
    cod_month_num = int(s.timeline.dev_months) + int(s.timeline.capex_months) + 1
    
    for row in range(2, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        
        # OperatingYear
        ws[f"D{row}"] = f'=IF(C{row}="Operation",MAX(0,YEAR(A{row})-YEAR(EDATE(StartDate,DevMonths+ConMonths))+1),0)'
        ws[f"D{row}"].fill = FORMULA_FILL
        
        # Energy
        ws[f"E{row}"] = f'=IF(C{row}="Operation",(P50MWh*(1-DegradPct)^(D{row}-1))/12,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"E{row}"].number_format = "#,##0.00"
        
        # Price
        if s.revenue_mode == "Standard PPA Parameters":
            ws[f"F{row}"] = f'=IF(C{row}="Operation",IF(D{row}<=PPATerm,PPAPrice,MerchantPrice)*(1+ColCPI)^(D{row}-1),0)'
        else:
            # For manual mode, lookup price from table using INDEX/MATCH, then apply indexation
            # INDEX/MATCH: =INDEX(PriceTable,MATCH(OperatingYear,ColumnA,0),2)
            # This looks up the price for the operating year from the price table
            # Use IFERROR to handle cases where PriceTable doesn't exist or lookup fails
            ws[f"F{row}"] = f'=IF(C{row}="Operation",IFERROR(INDEX(PriceTable,MATCH(D{row},INDEX(PriceTable,0,1),0),2),PPAPrice)*(1+ColCPI)^(D{row}-1),0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"F{row}"].number_format = "#,##0.00"
        
        # Revenue
        ws[f"G{row}"] = f"=E{row}*F{row}*1000"
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"G{row}"].number_format = "#,##0"
    
    # Format columns
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18

def _setup_opex_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows):
    """Create OPEX_M sheet with OPEX calculation formulas."""
    headers = ["Month", "Year", "Phase", "OperatingYear", "FixedOM", "VarOM", "Insurance", "GridFees", "OPEX_Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    for row in range(2, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        ws[f"D{row}"] = f"=Revenue_M!D{row}"
        
        # Fixed OM
        ws[f"E{row}"] = f'=IF(C{row}="Operation",(FixedOM*MWac*(1+ColCPI)^(D{row}-1))/12,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"E{row}"].number_format = "#,##0"
        
        # Variable OM
        ws[f"F{row}"] = f'=IF(C{row}="Operation",Revenue_M!E{row}*VarOM*(1+ColCPI)^(D{row}-1),0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"F{row}"].number_format = "#,##0"
        
        # Insurance
        ws[f"G{row}"] = f'=IF(C{row}="Operation",(Insurance*MWac*(1+ColCPI)^(D{row}-1))/12,0)'
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"G{row}"].number_format = "#,##0"
        
        # Grid Fees
        ws[f"H{row}"] = f'=IF(C{row}="Operation",Revenue_M!E{row}*GridFees,0)'
        ws[f"H{row}"].fill = FORMULA_FILL
        ws[f"H{row}"].number_format = "#,##0"
        
        # Total
        ws[f"I{row}"] = f"=SUM(E{row}:H{row})"
        ws[f"I{row}"].fill = FORMULA_FILL
        ws[f"I{row}"].number_format = "#,##0"
    
    # Format columns
    for col in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 15

def _setup_sga_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows):
    """Create SG&A_M sheet with SG&A calculation formulas."""
    headers = ["Month", "Year", "Phase", "SG&A_Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Calculate total annual SG&A by phase for simplified monthly calculation
    # In a full implementation, this would calculate each SG&A item separately with indexation
    sga_items = s.sga.items or []
    sga_by_phase = {"Development": 0.0, "Construction": 0.0, "Operation": 0.0}
    for item in sga_items:
        phase = str(item.get("Phase", "Development"))
        amt = float(item.get("Amount_COP_per_year", 0.0) or 0.0)
        if phase in sga_by_phase:
            sga_by_phase[phase] += amt
    
    for row in range(2, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        
        # SG&A Total - simplified: annual amount / 12 for the phase, with CPI indexation
        # For Development phase
        dev_sga = sga_by_phase.get("Development", 0.0) / 12.0
        # For Construction phase  
        con_sga = sga_by_phase.get("Construction", 0.0) / 12.0
        # For Operation phase - with CPI indexation
        op_sga_base = sga_by_phase.get("Operation", 0.0) / 12.0
        
        # Formula: IF phase = Development, dev_sga; IF Construction, con_sga; IF Operation, op_sga with indexation
        # Operation SG&A needs CPI indexation: base * (1+ColCPI)^(OperatingYear-1)
        ws[f"D{row}"] = f'=IF(C{row}="Development",{dev_sga},IF(C{row}="Construction",{con_sga},IF(C{row}="Operation",{op_sga_base}*(1+ColCPI)^(Revenue_M!D{row}-1),0)))'
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"D{row}"].number_format = "#,##0"
    
    # Format columns
    ws.column_dimensions['D'].width = 15

def _setup_debt_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows):
    """Create Debt_M sheet with debt schedule formulas using linear debt draw approach."""
    headers = ["Month", "Year", "Phase", "DebtDraw", "UpfrontFee", "CommitmentFee", "Interest", "Principal", "DebtService", "DebtBalance"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Helper cells for calculations
    ws["K1"] = "TotalDebt"
    # Use specific range instead of entire column
    ws["K2"] = f"=SUM(D2:D{max_rows+1})"
    ws["K2"].fill = FORMULA_FILL
    ws["L1"] = "CODMonth"
    ws["L2"] = "=DevMonths+ConMonths+1"
    ws["L2"].fill = FORMULA_FILL
    ws["M1"] = "GraceMonths"
    ws["M2"] = "=DebtGrace*12"
    ws["M2"].fill = FORMULA_FILL
    ws["N1"] = "AmortMonths"
    ws["N2"] = "=DebtTenor*12"
    ws["N2"].fill = FORMULA_FILL
    ws["O1"] = "MonthlyPrincipal"
    # Calculate monthly principal after grace period (equal principal amortization)
    ws["O2"] = '=IF(AND(DebtEnabled="Yes",N2>M2),TotalDebt/(N2-M2),0)'
    ws["O2"].fill = FORMULA_FILL
    
    # Row 2
    ws["A2"] = "=Timeline_M!A2"
    ws["A2"].number_format = "mm/dd/yyyy"
    ws["B2"] = "=Timeline_M!B2"
    ws["C2"] = "=Timeline_M!C2"
    # DebtDraw: Linear draw during construction only (CAPEX_t * DebtPct), no draws after COD
    ws["D2"] = '=IF(AND(C2<>"Operation",DebtEnabled="Yes",Timeline_M!D2<=L2),CAPEX_M!G2*DebtPct,0)'
    ws["D2"].fill = FORMULA_FILL
    ws["D2"].number_format = "#,##0"
    # Upfront Fee: Paid at COD (first operation month) as % of total debt
    ws["E2"] = '=IF(AND(C2="Operation",DebtEnabled="Yes",Timeline_M!D2=L2),K2*DebtUpfrontFee,0)'
    ws["E2"].fill = FORMULA_FILL
    ws["E2"].number_format = "#,##0"
    # Commitment Fee: On undrawn amount during construction (% of margin on undrawn)
    # Undrawn = TotalDebt - CumulativeDraws
    ws["F2"] = '=IF(AND(C2<>"Operation",DebtEnabled="Yes",Timeline_M!D2<=L2),(K2-SUM($D$2:D2))*DebtCommitmentFee*DebtMargin/12,0)'
    ws["F2"].fill = FORMULA_FILL
    ws["F2"].number_format = "#,##0"
    ws["G2"] = "0"  # Interest starts in operation
    ws["G2"].fill = FORMULA_FILL
    ws["G2"].number_format = "#,##0"
    ws["H2"] = "0"  # Principal starts after grace period
    ws["H2"].fill = FORMULA_FILL
    ws["H2"].number_format = "#,##0"
    ws["I2"] = "=E2+F2+G2+H2"  # Total debt service (fees + interest + principal)
    ws["I2"].fill = FORMULA_FILL
    ws["I2"].number_format = "#,##0"
    ws["J2"] = "=D2"  # Debt balance = cumulative draws
    ws["J2"].fill = FORMULA_FILL
    ws["J2"].number_format = "#,##0"
    
    # Row 3 onwards
    for row in range(3, max_rows + 1):
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        # DebtDraw: Linear draw during construction only, no draws after COD
        ws[f"D{row}"] = f'=IF(AND(C{row}<>"Operation",DebtEnabled="Yes",Timeline_M!D{row}<=L2),CAPEX_M!G{row}*DebtPct,0)'
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"D{row}"].number_format = "#,##0"
        # Upfront Fee: Only at COD
        ws[f"E{row}"] = f'=IF(AND(C{row}="Operation",DebtEnabled="Yes",Timeline_M!D{row}=L2),K2*DebtUpfrontFee,0)'
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"E{row}"].number_format = "#,##0"
        # Commitment Fee: On undrawn during construction
        ws[f"F{row}"] = f'=IF(AND(C{row}<>"Operation",DebtEnabled="Yes",Timeline_M!D{row}<=L2),(K2-SUM($D$2:D{row}))*DebtCommitmentFee*DebtMargin/12,0)'
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"F{row}"].number_format = "#,##0"
        # Interest: During operation, on outstanding balance
        ws[f"G{row}"] = f'=IF(AND(C{row}="Operation",DebtEnabled="Yes"),J{row-1}*DebtRate/12,0)'
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"G{row}"].number_format = "#,##0"
        # Principal: After grace period, equal monthly payments
        ws[f"H{row}"] = f'=IF(AND(C{row}="Operation",DebtEnabled="Yes",Timeline_M!D{row}>L2+M2,Timeline_M!D{row}<=L2+N2,J{row-1}>0),MIN(O2,J{row-1}),0)'
        ws[f"H{row}"].fill = FORMULA_FILL
        ws[f"H{row}"].number_format = "#,##0"
        # Debt Service: Fees + Interest + Principal
        ws[f"I{row}"] = f"=E{row}+F{row}+G{row}+H{row}"
        ws[f"I{row}"].fill = FORMULA_FILL
        ws[f"I{row}"].number_format = "#,##0"
        # Debt Balance: Cumulative draws minus principal payments
        ws[f"J{row}"] = f"=J{row-1}+D{row}-H{row}"
        ws[f"J{row}"].fill = FORMULA_FILL
        ws[f"J{row}"].number_format = "#,##0"
    
    # Format columns
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 15

def _setup_cashflow_sheet(ws, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows):
    """Create Cashflow_M sheet with complete cash flow formulas matching app calculations."""
    headers = ["Month", "Year", "Phase", "Revenue", "OPEX", "SG&A", "CAPEX", "EBITDA", "Depreciation", 
               "TaxableIncome", "SpecialDed", "TaxesPayable", "VATRefund", "DeltaNWC", "UnleveredCF", "DebtDraw", "DebtService", "Equity CF After Tax"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    
    # Helper cells for tax calculations (in row 1, columns beyond headers)
    # Calculate COD month number for depreciation and tax benefit timing
    # COD occurs at the end of construction, so COD month = DevMonths + ConMonths
    # But depreciation starts in the first month of operation, which is COD month + 1
    ws["Q1"] = "CODMonthNum"
    ws["Q2"] = "=DevMonths+ConMonths"
    ws["Q2"].fill = FORMULA_FILL
    
    # Total CAPEX for depreciation and tax benefits
    ws["R1"] = "TotalCAPEX"
    # Use a specific range instead of entire column to avoid potential issues
    ws["R2"] = f"=SUM(CAPEX_M!G2:G{max_rows+1})"
    ws["R2"].fill = FORMULA_FILL
    
    # Depreciation base and monthly amount
    ws["S1"] = "DepBase"
    ws["S2"] = "=R2*DepPct"  # Use R2 (TotalCAPEX) instead of named range
    ws["S2"].fill = FORMULA_FILL
    ws["T1"] = "MonthlyDep"
    ws["T2"] = "=IF(DepYears>0,S2/(DepYears*12),0)"  # Use S2 instead of DepBase
    ws["T2"].fill = FORMULA_FILL
    
    # Special deduction pool
    ws["U1"] = "SpecialDedPool"
    ws["U2"] = "=IF(EnableSpecialDed=\"Yes\",R2*SpecialDedPct,0)"  # Use R2 instead of TotalCAPEX
    ws["U2"].fill = FORMULA_FILL
    
    # Helper cells for VAT Refund target year calculation
    # Calculate target year once in row 2, then reference it
    ws["V1"] = "VATTargetYear"
    # Calculate target date: StartDate + (COD month + VATRefundYear - 1) months
    # Use direct cell references: StartDate in Inputs!$B$8, DevMonths in Inputs!$B$9, ConMonths in Inputs!$B$10
    ws["V2"] = "=YEAR(EDATE(Inputs!$B$8,Inputs!$B$9+Inputs!$B$10+VATRefundYear-1))"
    ws["V2"].fill = FORMULA_FILL
    
    for row in range(2, max_rows + 1):
        # Basic references
        ws[f"A{row}"] = f"=Timeline_M!A{row}"
        ws[f"A{row}"].number_format = "mm/dd/yyyy"
        ws[f"B{row}"] = f"=Timeline_M!B{row}"
        ws[f"C{row}"] = f"=Timeline_M!C{row}"
        
        # Revenue
        ws[f"D{row}"] = f"=Revenue_M!G{row}"
        ws[f"D{row}"].fill = FORMULA_FILL
        ws[f"D{row}"].number_format = "#,##0"
        
        # OPEX
        ws[f"E{row}"] = f"=OPEX_M!I{row}"
        ws[f"E{row}"].fill = FORMULA_FILL
        ws[f"E{row}"].number_format = "#,##0"
        
        # SG&A - reference SG&A monthly sheet (escape & with single quotes)
        ws[f"F{row}"] = f"='SG&A_M'!D{row}"
        ws[f"F{row}"].fill = FORMULA_FILL
        ws[f"F{row}"].number_format = "#,##0"
        
        # CAPEX
        ws[f"G{row}"] = f"=CAPEX_M!G{row}"
        ws[f"G{row}"].fill = FORMULA_FILL
        ws[f"G{row}"].number_format = "#,##0"
        
        # EBITDA = Revenue - OPEX - SG&A
        ws[f"H{row}"] = f"=D{row}-E{row}-F{row}"
        ws[f"H{row}"].fill = FORMULA_FILL
        ws[f"H{row}"].number_format = "#,##0"
        
        # Depreciation - monthly linear depreciation starting at first month of operation
        # IMPORTANT: Depreciation is a non-cash expense that reduces taxable income (and thus taxes)
        # It does NOT reduce cash flow - it's only used for tax calculations
        # Operation starts at month DevMonths+ConMonths+1 (COD month + 1)
        # Depreciation continues for DepYears*12 months from the start of operation
        # Use month number from Timeline_M column D (MonthNum)
        # Depreciation - use direct cell references to avoid named range parsing issues
        # DevMonths in Inputs!$B$9, ConMonths in Inputs!$B$10
        # First operation month is DevMonths+ConMonths+1, so use >= to include it
        # T2 contains the monthly depreciation amount (DepBase / (DepYears * 12))
        ws[f"I{row}"] = f'=IF(AND(C{row}="Operation",Timeline_M!D{row}>=Inputs!$B$9+Inputs!$B$10+1,Timeline_M!D{row}<=Inputs!$B$9+Inputs!$B$10+DepYears*12),T2,0)'
        ws[f"I{row}"].fill = FORMULA_FILL
        ws[f"I{row}"].number_format = "#,##0"
        
        # Taxable Income (before special deduction) = EBITDA - Depreciation
        # IMPORTANT: Depreciation reduces taxable income for tax purposes (reduces taxes payable)
        # Depreciation is a non-cash expense, so it does NOT reduce cash flow
        # Special deduction and NOL will be handled in subsequent columns
        ws[f"J{row}"] = f"=H{row}-I{row}"
        ws[f"J{row}"].fill = FORMULA_FILL
        ws[f"J{row}"].number_format = "#,##0"
        
        # Special Deduction - break down into helper cells to simplify formula
        # Helper cell W: Pool remaining (U2 - SUM of previous deductions)
        ws[f"W{row}"] = f"=MAX(0,U2-SUM($K$2:K{row-1}))"
        ws[f"W{row}"].fill = FORMULA_FILL
        ws[f"W{row}"].number_format = "#,##0"
        
        # Helper cell X: Annual cap (SpecialDedMaxPct * Taxable Income)
        ws[f"X{row}"] = f"=J{row}*SpecialDedMaxPct"
        ws[f"X{row}"].fill = FORMULA_FILL
        ws[f"X{row}"].number_format = "#,##0"
        
        # Special Deduction - simplified formula using helper cells
        # Pool starts in the year AFTER COD (first month of second operating year = COD month + 12)
        # Usable for SpecialDedYears (in years, convert to months)
        # Use direct cell references: DevMonths in Inputs!$B$9, ConMonths in Inputs!$B$10
        ws[f"K{row}"] = f'=IF(AND(EnableSpecialDed="Yes",C{row}="Operation",J{row}>0,Timeline_M!D{row}>=Inputs!$B$9+Inputs!$B$10+12,Timeline_M!D{row}<=Inputs!$B$9+Inputs!$B$10+12+SpecialDedYears*12),MIN(W{row},X{row}),0)'
        ws[f"K{row}"].fill = FORMULA_FILL
        ws[f"K{row}"].number_format = "#,##0"
        
        # Taxes Payable = Taxable Income After Deduction * Tax Rate
        # Tax calculation flow:
        # 1. EBITDA (H) = Revenue - OPEX - SG&A
        # 2. Depreciation (I) reduces taxable income (non-cash expense, but reduces taxes)
        # 3. Taxable Income (J) = EBITDA - Depreciation
        # 4. Special Deduction (K) further reduces taxable income
        # 5. Taxable After Deduction = MAX(0, Taxable Income - Special Deduction)
        # 6. Taxes Payable = Taxable After Deduction * Tax Rate
        # Note: Depreciation is NOT subtracted from cash flow, only from taxable income for tax purposes
        # NOL (Net Operating Loss) logic would be added here in a more complete implementation
        ws[f"L{row}"] = f"=MAX(0,J{row}-K{row})*TaxRate"
        ws[f"L{row}"].fill = FORMULA_FILL
        ws[f"L{row}"].number_format = "#,##0"
        
        # VAT Refund - simplified formula using helper cell V2 for target year
        # VATAmount: if VATMode="percent", it's a decimal (e.g., 0.19 for 19%), multiply by R2 (TotalCAPEX)
        # If VATMode="fixed", it's already the final amount in COP
        # Use helper cell $V$2 (absolute reference) which contains the target year calculation
        ws[f"M{row}"] = f'=IF(EnableVATRefund="Yes",IF(C{row}="Operation",IF(YEAR(A{row})=$V$2,IF(VATMode="percent",VATAmount*R2,VATAmount),0),0),0)'
        ws[f"M{row}"].fill = FORMULA_FILL
        ws[f"M{row}"].number_format = "#,##0"
        
        # Working Capital Calculation (ΔNWC)
        # AR/AP balances are calculated from cash timing (AR/AP lags)
        # AR Balance = Previous AR + Revenue - Cash Collected (where Cash Collected = Revenue shifted by AR lag)
        # AP Balance = Previous AP + OPEX + SG&A - Cash Paid (where Cash Paid = OPEX+SG&A shifted by AP lag)
        # Net Working Capital = AR Balance - AP Balance
        # ΔNWC = Current NWC - Previous NWC
        
        # Calculate AR lag in months (helper column Y)
        ws[f"Y{row}"] = f"=ROUND(ARDays/30,0)"
        ws[f"Y{row}"].fill = FORMULA_FILL
        
        # Calculate AP lag in months (helper column Z)
        ws[f"Z{row}"] = f"=ROUND(APDays/30,0)"
        ws[f"Z{row}"].fill = FORMULA_FILL
        
        # Calculate AR Balance (helper column AB): Previous AR + Revenue - Revenue from AR lag months ago
        if row == 2:
            # First row: AR Balance = Revenue - Revenue from AR lag months ago (or 0 if lag is 0 or row doesn't exist)
            ws[f"AB{row}"] = f"=D{row}-IF(AND(Y{row}>0,ROW()-Y{row}>=2),INDEX(D:D,ROW()-Y{row}),D{row})"
        else:
            # Subsequent rows: Previous AR + Revenue - Revenue from AR lag months ago
            ws[f"AB{row}"] = f"=AB{row-1}+D{row}-IF(AND(Y{row}>0,ROW()-Y{row}>=2),INDEX(D:D,ROW()-Y{row}),D{row})"
        ws[f"AB{row}"].fill = FORMULA_FILL
        ws[f"AB{row}"].number_format = "#,##0"
        
        # Calculate AP Balance (helper column AC): Previous AP + OPEX+SG&A - (OPEX+SG&A) from AP lag months ago
        if row == 2:
            # First row: AP Balance = OPEX+SG&A - (OPEX+SG&A) from AP lag months ago
            ws[f"AC{row}"] = f"=(E{row}+F{row})-IF(AND(Z{row}>0,ROW()-Z{row}>=2),INDEX(E:E,ROW()-Z{row})+INDEX(F:F,ROW()-Z{row}),E{row}+F{row})"
        else:
            # Subsequent rows: Previous AP + OPEX+SG&A - (OPEX+SG&A) from AP lag months ago
            ws[f"AC{row}"] = f"=AC{row-1}+(E{row}+F{row})-IF(AND(Z{row}>0,ROW()-Z{row}>=2),INDEX(E:E,ROW()-Z{row})+INDEX(F:F,ROW()-Z{row}),E{row}+F{row})"
        ws[f"AC{row}"].fill = FORMULA_FILL
        ws[f"AC{row}"].number_format = "#,##0"
        
        # Net Working Capital = AR Balance - AP Balance (helper column AD)
        ws[f"AD{row}"] = f"=AB{row}-AC{row}"
        ws[f"AD{row}"].fill = FORMULA_FILL
        ws[f"AD{row}"].number_format = "#,##0"
        
        # ΔNWC = Current NWC - Previous NWC (column N)
        if row == 2:
            ws[f"N{row}"] = f"=AD{row}"
        else:
            ws[f"N{row}"] = f"=AD{row}-AD{row-1}"
        ws[f"N{row}"].fill = FORMULA_FILL
        ws[f"N{row}"].number_format = "#,##0"
        
        # Unlevered CF After Tax = Revenue - OPEX - SG&A - CAPEX - Taxes + VAT Refund - ΔNWC
        # IMPORTANT: Depreciation is NOT subtracted from cash flow (it's a non-cash expense)
        # Depreciation only affects cash flow indirectly by reducing taxes (via reduced taxable income)
        # The tax reduction is already captured in column L (Taxes Payable)
        # ΔNWC is subtracted because an increase in working capital is a cash outflow
        ws[f"O{row}"] = f"=D{row}-E{row}-F{row}-G{row}-L{row}+M{row}-N{row}"
        ws[f"N{row}"].fill = FORMULA_FILL
        ws[f"N{row}"].number_format = "#,##0"
        
        # Debt Draw (moved to column P since Unlevered CF is now in column O)
        ws[f"P{row}"] = f"=Debt_M!D{row}"
        ws[f"P{row}"].fill = FORMULA_FILL
        ws[f"P{row}"].number_format = "#,##0"
        
        # Debt Service (moved to column Q)
        ws[f"Q{row}"] = f"=Debt_M!I{row}"
        ws[f"Q{row}"].fill = FORMULA_FILL
        ws[f"Q{row}"].number_format = "#,##0"
        
        # Equity CF = Unlevered CF After Tax + Debt Draw - Debt Service
        # Note: Unlevered CF already includes -ΔNWC, so Equity CF automatically includes it
        ws[f"R{row}"] = f"=O{row}+P{row}-Q{row}"
        ws[f"Q{row}"].fill = FORMULA_FILL
        ws[f"Q{row}"].number_format = "#,##0"
    
    # Format columns (including helper columns W, X, Y, Z, AB, AC, AD, but hide helper columns)
    for col in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "W", "X", "Y", "Z", "AB", "AC", "AD"]:
        ws.column_dimensions[col].width = 15
    # Hide helper columns (W, X, Y, Z, AB, AC, AD)
    ws.column_dimensions["W"].hidden = True
    ws.column_dimensions["X"].hidden = True
    ws.column_dimensions["Y"].hidden = True
    ws.column_dimensions["Z"].hidden = True
    ws.column_dimensions["AB"].hidden = True
    ws.column_dimensions["AC"].hidden = True
    ws.column_dimensions["AD"].hidden = True
    # Hide helper columns W and X (they're just for calculation)
    ws.column_dimensions["W"].hidden = True
    ws.column_dimensions["X"].hidden = True

def _setup_summary_sheet(ws, project_name, scenario_name, s, HEADER_FILL, HEADER_FONT, TITLE_FONT, INPUT_FILL, FORMULA_FILL, THIN_BORDER, CENTER_ALIGN, RIGHT_ALIGN, LEFT_ALIGN, max_rows):
    """Create Summary sheet with project info, color legend, and key metrics with formulas."""
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    
    row = 1
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Project Financial Summary"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    row += 2
    
    # Project info
    ws[f'A{row}'] = "Project:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = project_name
    ws[f'B{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Scenario:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'] = scenario_name
    ws[f'B{row}'].font = Font(bold=True, size=12)
    row += 1
    
    ws[f'A{row}'] = "Export Date:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    row += 2
    
    # Color Legend
    ws[f'A{row}'] = "Color Legend:"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    ws[f'A{row}'] = "Yellow fill:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "Input values (hardcoded, can be modified)"
    ws[f'B{row}'].fill = INPUT_FILL
    ws[f'B{row}'].border = THIN_BORDER
    row += 1
    
    ws[f'A{row}'] = "Grey fill:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "Calculated values (formulas, do not modify)"
    ws[f'B{row}'].fill = FORMULA_FILL
    ws[f'B{row}'].border = THIN_BORDER
    row += 2
    
    # Key metrics
    ws[f'A{row}'] = "Key Financial Metrics"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # Total CAPEX
    ws[f'A{row}'] = "Total CAPEX (COP):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=SUM(CAPEX_M!G2:G{max_rows+1})"
    ws[f'B{row}'].fill = FORMULA_FILL
    ws[f'B{row}'].number_format = "#,##0"
    row += 1
    
    # Unlevered IRR - need to set up helper ranges first
    # Calculate max_rows from timeline
    tl = build_timeline(s.timeline)
    total_months = int(s.timeline.dev_months) + int(s.timeline.capex_months) + (int(s.timeline.operation_years) * 12)
    max_rows_cf = min(500, total_months + 10)  # Match the max_rows used in cashflow sheet
    
    ws[f'D1'] = "UnleveredCF_Dates"
    ws[f'E1'] = "UnleveredCF_Values"
    # Set up helper ranges with proper date formatting and limit to actual data
    for r in range(2, max_rows_cf + 2):  # +2 to include header row
        ws[f'D{r}'] = f"=Cashflow_M!A{r}"
        ws[f'D{r}'].number_format = "mm/dd/yyyy"  # Format as date
        ws[f'D{r}'].fill = FORMULA_FILL
        ws[f'E{r}'] = f"=Cashflow_M!O{r}"  # Unlevered CF After Tax (now includes -ΔNWC)
        ws[f'E{r}'].fill = FORMULA_FILL
    # Fill remaining with empty to avoid XIRR errors
    for r in range(max_rows_cf + 2, 502):
        ws[f'D{r}'] = ""
        ws[f'E{r}'] = 0
    
    ws[f'A{row}'] = "Unlevered IRR (Pre-tax):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=XIRR(E2:E{max_rows_cf+1},D2:D{max_rows_cf+1})"
    ws[f'B{row}'].fill = FORMULA_FILL
    ws[f'B{row}'].number_format = "0.00%"
    row += 1
    
    # Equity IRR - need to set up helper ranges
    ws[f'F1'] = "EquityCF_Dates"
    ws[f'G1'] = "EquityCF_Values"
    for r in range(2, max_rows_cf + 2):
        ws[f'F{r}'] = f"=Cashflow_M!A{r}"
        ws[f'F{r}'].number_format = "mm/dd/yyyy"  # Format as date
        ws[f'F{r}'].fill = FORMULA_FILL
        ws[f'G{r}'] = f"=Cashflow_M!R{r}"  # Equity CF After Tax (now includes -ΔNWC)
        ws[f'G{r}'].fill = FORMULA_FILL
    # Fill remaining with empty
    for r in range(max_rows_cf + 2, 502):
        ws[f'F{r}'] = ""
        ws[f'G{r}'] = 0
    
    ws[f'A{row}'] = "Equity IRR (After-tax):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"=XIRR(G2:G{max_rows_cf+1},F2:F{max_rows_cf+1})"
    ws[f'B{row}'].fill = FORMULA_FILL
    ws[f'B{row}'].number_format = "0.00%"

def _setup_outputs_sheet(ws, s, HEADER_FILL, HEADER_FONT, TITLE_FONT, FORMULA_FILL, THIN_BORDER, CENTER_ALIGN, RIGHT_ALIGN, max_rows):
    """Create Outputs sheet with annual summaries and formulas."""
    tl = build_timeline(s.timeline)
    row = 1
    ws[f'A{row}'] = "Annual Outputs"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:J{row}')
    row += 2
    
    # Annual cash flow summary
    ws[f'A{row}'] = "Year"
    ws[f'B{row}'] = "Unlevered CF"
    ws[f'C{row}'] = "Equity CF"
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{row}']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
    row += 1
    
    # Annual aggregation formulas
    cod_year = tl["cod"].year
    for year_offset in range(0, min(30, int(s.timeline.operation_years) + 5)):
        year_val = cod_year + year_offset
        ws[f'A{row}'] = year_val
        ws[f'B{row}'] = f"=SUMIF(Cashflow_M!B:B,{year_val},Cashflow_M!G:G)"
        ws[f'B{row}'].fill = FORMULA_FILL
        ws[f'B{row}'].number_format = "#,##0"
        ws[f'C{row}'] = f"=SUMIF(Cashflow_M!B:B,{year_val},Cashflow_M!R:R)"
        ws[f'C{row}'].fill = FORMULA_FILL
        ws[f'C{row}'].number_format = "#,##0"
        row += 1
    
    # Format columns
    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].width = 15

# -----------------------------
# Excel Export Function
# -----------------------------
def generate_excel_report(project_name: str, scenario_name: str, s: ScenarioInputs) -> BytesIO:
    """Generate a comprehensive bankable Excel model with formulas throughout."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    from openpyxl.workbook.defined_name import DefinedName
    
    wb = Workbook()
    
    # Enable iterative calculations for circular references
    # This allows formulas with cumulative sums to work automatically
    # Users won't need to manually enable circular references in Excel
    wb.calculation.iterate = True
    wb.calculation.iterateCount = 100  # Maximum number of iterations
    wb.calculation.iterateDelta = 0.001  # Convergence tolerance
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14)
    SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    FORMULA_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    
    # Calculate total months needed
    tl = build_timeline(s.timeline)
    total_months = int(s.timeline.dev_months) + int(s.timeline.capex_months) + (int(s.timeline.operation_years) * 12)
    max_rows = min(500, total_months + 10)  # Cap at 500 rows
    
    # Create all sheets in order
    ws_summary = wb.create_sheet("Summary", 0)
    ws_inputs = wb.create_sheet("Inputs", 1)
    ws_timeline = wb.create_sheet("Timeline_M", 2)
    ws_capex = wb.create_sheet("CAPEX_M", 3)
    ws_revenue = wb.create_sheet("Revenue_M", 4)
    ws_opex = wb.create_sheet("OPEX_M", 5)
    ws_sga = wb.create_sheet("SG&A_M", 6)  # Create SG&A sheet before Debt
    ws_debt = wb.create_sheet("Debt_M", 7)
    ws_cashflow = wb.create_sheet("Cashflow_M", 8)
    ws_outputs = wb.create_sheet("Outputs", 9)
    
    # ==================== SETUP INPUTS SHEET FIRST (needed for named ranges) ====================
    named_range_refs, cell_positions = _setup_inputs_sheet(ws_inputs, s, INPUT_FILL, HEADER_FILL, HEADER_FONT, TITLE_FONT, SECTION_FILL, THIN_BORDER, CENTER_ALIGN, RIGHT_ALIGN)
    
    # Create named ranges from Inputs sheet
    _create_named_ranges(wb, ws_inputs, s, named_range_refs)
    
    # ==================== SETUP TIMELINE SHEET ====================
    # Pass StartDate cell reference for direct reference (more reliable than named range)
    # StartDate should be in Timeline section, which comes after Macro section (3 inputs + 1 header = ~row 8)
    start_date_ref = cell_positions.get("StartDate")
    if not start_date_ref:
        # Calculate expected position: Macro header(2) + 3 inputs(3,4,5) + spacer(6) + Timeline header(7) + StartDate(8) = B8
        # But this is approximate - better to track it properly
        start_date_ref = "Inputs!$B$8"  # Approximate fallback
    _setup_timeline_sheet(ws_timeline, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows, start_date_cell=start_date_ref)
    
    # ==================== SETUP CAPEX SHEET ====================
    _setup_capex_sheet(ws_capex, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows)
    
    # ==================== SETUP REVENUE SHEET ====================
    _setup_revenue_sheet(ws_revenue, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows)
    
    # ==================== SETUP OPEX SHEET ====================
    _setup_opex_sheet(ws_opex, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows)
    
    # ==================== SETUP SG&A SHEET ====================
    # Note: SG&A sheet was already created above to maintain proper order
    _setup_sga_sheet(ws_sga, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows)
    
    # ==================== SETUP DEBT SHEET ====================
    _setup_debt_sheet(ws_debt, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows)
    
    # ==================== SETUP CASHFLOW SHEET ====================
    _setup_cashflow_sheet(ws_cashflow, s, FORMULA_FILL, HEADER_FILL, HEADER_FONT, THIN_BORDER, max_rows)
    
    # ==================== SETUP SUMMARY SHEET ====================
    _setup_summary_sheet(ws_summary, project_name, scenario_name, s, HEADER_FILL, HEADER_FONT, TITLE_FONT, INPUT_FILL, FORMULA_FILL, THIN_BORDER, CENTER_ALIGN, RIGHT_ALIGN, LEFT_ALIGN, max_rows)
    
    # ==================== SETUP OUTPUTS SHEET ====================
    _setup_outputs_sheet(ws_outputs, s, HEADER_FILL, HEADER_FONT, TITLE_FONT, FORMULA_FILL, THIN_BORDER, CENTER_ALIGN, RIGHT_ALIGN, max_rows)
    
    # Validate workbook structure before saving
    # Check that all referenced sheets exist
    required_sheets = ["Inputs", "Timeline_M", "CAPEX_M", "Revenue_M", "OPEX_M", "SG&A_M", "Debt_M", "Cashflow_M", "Summary", "Outputs"]
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Required sheet '{sheet_name}' is missing from workbook")
    
    # Additional validation: Check for potential issues
    # Verify all sheet references in formulas are valid
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Check for any cells with formulas that might reference invalid sheets
        # This is a basic check - openpyxl will handle most validation
    
    # Save to buffer with proper error handling
    buffer = BytesIO()
    try:
        # Ensure workbook is in a valid state before saving
        wb.save(buffer)
        buffer.seek(0)
    except Exception as e:
        # Provide more detailed error information
        error_msg = f"Error saving Excel workbook: {e}\n"
        error_msg += f"Workbook has {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}"
        raise ValueError(error_msg)
    
    return buffer


# -----------------------------
# N) Summary
# -----------------------------
with tab_summary:
    st.subheader("Project Summary")
    
    # Export buttons at the top
    st.markdown("### Export Reports")
    export_col1, export_col2 = st.columns(2)
    
    with export_col1:
        if REPORTLAB_AVAILABLE:
            # PDF generation temporarily disabled due to plotly/kaleido issues
            if False and st.button("📄 Generate PDF Report", type="primary", width='stretch', key="pdf_export_btn"):
                try:
                    sensitivity_data = None  # Could be enhanced to capture from sensitivity tab
                    pdf_buffer = generate_summary_pdf(
                        project_name=project_name,
                        scenario_name=scenario_name,
                        s=s,
                        currency="COP",  # PDF uses base currency
                        sensitivity_data=sensitivity_data
                    )
                    st.download_button(
                        label="⬇️ Download PDF",
                        data=pdf_buffer,
                        file_name=f"{project_name}_{scenario_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        width='stretch',
                        key="pdf_download_btn"
                    )
                    st.success("PDF generated successfully!")
                except Exception as e:
                    st.error(f"Error generating PDF: {str(e)}")
        else:
            st.warning("PDF export requires reportlab. Install: `pip install reportlab`")
    
    with export_col2:
        if OPENPYXL_AVAILABLE:
            if st.button("📊 Generate Excel Report", type="primary", width='stretch', key="excel_export_btn"):
                try:
                    excel_buffer = generate_excel_report(
                        project_name=project_name,
                        scenario_name=scenario_name,
                        s=s
                    )
                    st.download_button(
                        label="⬇️ Download Excel",
                        data=excel_buffer,
                        file_name=f"{project_name}_{scenario_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width='stretch',
                        key="excel_download_btn"
                    )
                    st.success("Excel file generated successfully!")
                except Exception as e:
                    st.error(f"Error generating Excel: {str(e)}")
        else:
            st.warning("Excel export requires openpyxl. Install: `pip install openpyxl`")
    
    st.divider()
    
    currency = st.radio("Display currency", ["COP", "USD"], horizontal=True, index=0, key="currency_summary")
    
    # Get timeline and FX path (for proper USD conversion using indexation)
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    
    # Get FX series for all years (accounts for CPI indexation differences)
    # This ensures USD conversion reflects actual FX path, not just fixed rate
    a = unlevered_base_cashflow_annual(s)
    years_all = list(a["Year"].astype(int).tolist()) if len(a) > 0 else [cod.year]
    fx_series_all = fx_series(s.macro, cod.year, years_all)
    
    # Helper function to convert COP to USD using year-specific FX
    def _to_usd(value_cop: float, year: int = None) -> float:
        if currency == "COP":
            return value_cop
        if year is not None and year in fx_series_all.index:
            return value_cop / float(fx_series_all.loc[year])
        # Fallback to starting FX if year not found
        return value_cop / float(s.macro.fx_cop_per_usd_start)
    
    # For single values (like total CAPEX), use COD year FX
    fx_cod = float(fx_series_all.loc[cod.year]) if cod.year in fx_series_all.index else float(s.macro.fx_cop_per_usd_start)
    
    # 1. Power Generation & Revenue Metrics (combined in one row)
    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    base_mwh = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    degr_pct = float(gen.degradation_pct)
    
    if s.revenue_mode == "Standard PPA Parameters":
        r = s.revenue1
        ppa_price = float(r.ppa_price_cop_per_kwh)
        indexation = str(r.indexation)
        ppa_term = int(r.ppa_term_years)
        merchant_price = float(r.merchant_price_cop_per_kwh)
        
        gen_col1, gen_col2, gen_col3, gen_col4, gen_col5 = st.columns(5)
        with gen_col1:
            st.metric("Production", f"{str(gen.production_choice)} - {base_mwh:,.0f} MWh/yr")
        with gen_col2:
            st.metric("Degradation", f"{degr_pct:.2f}%")
        with gen_col3:
            st.metric("PPA Price", f"{ppa_price:,.4f} COP/kWh")
        with gen_col4:
            st.metric("PPA Term", f"{ppa_term}yr ({indexation})")
        with gen_col5:
            st.metric("Merchant Price", f"{merchant_price:,.4f} COP/kWh")
    else:
        r = s.revenue2
        indexation = str(r.indexation)
        gen_col1, gen_col2, gen_col3 = st.columns(3)
        with gen_col1:
            st.metric("Production", f"{str(gen.production_choice)} - {base_mwh:,.0f} MWh/yr")
        with gen_col2:
            st.metric("Degradation", f"{degr_pct:.2f}%")
        with gen_col3:
            st.metric("Indexation", indexation)
        st.info("Custom price schedule (see Power Revenues tab)")
    
    st.divider()
    
    # 3. CAPEX (compact)
    total_capex = _total_capex_from_lines(s)
    st.markdown("### CAPEX")
    st.metric("Total CAPEX", _fmt_cop(total_capex) if currency == "COP" else _fmt_usd(_to_usd(total_capex, cod.year)))
    
    # CAPEX Breakdown Chart
    st.markdown("#### CAPEX Breakdown by Line Item")
    capex_df = pd.DataFrame(s.capex.lines or [])
    if "Amount_COP" not in capex_df.columns:
        capex_df["Amount_COP"] = 0.0
    capex_pie = capex_df.copy()
    capex_pie["Item"] = capex_pie["Item"].fillna("").astype(str).str.strip()
    capex_pie["Amount_COP"] = pd.to_numeric(capex_pie["Amount_COP"], errors="coerce").fillna(0.0)
    capex_pie = capex_pie[capex_pie["Amount_COP"] > 0].copy()
    
    if capex_pie.empty:
        st.info("Enter CAPEX amounts to see the breakdown chart.")
    else:
        share = capex_pie["Amount_COP"] / capex_pie["Amount_COP"].sum()
        small = share < 0.03
        if small.any() and (~small).any():
            other_amt = float(capex_pie.loc[small, "Amount_COP"].sum())
            capex_pie = capex_pie.loc[~small, ["Item", "Amount_COP"]]
            capex_pie = pd.concat(
                [capex_pie, pd.DataFrame([{"Item": "Other (<3% each)", "Amount_COP": other_amt}])],
                ignore_index=True
            )
        fig_pie = px.pie(capex_pie, names="Item", values="Amount_COP", hole=0.45)
        fig_pie.update_traces(textinfo="percent+label")
        fig_pie.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
        st.plotly_chart(fig_pie, width='stretch', key="summary_capex_pie")
    
    st.divider()
    
    # 4. Operating Costs & Depreciation (side by side)
    op = operating_year_table(s)
    om = opex_monthly_schedule(s)
    annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
    annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
    
    # Merge with revenue
    if "Revenue (COP)" in op.columns and "Total OPEX (COP)" in annual_opex.columns:
        merged = op[["Year", "Revenue (COP)"]].merge(annual_opex[["Year", "Total OPEX (COP)"]], on="Year", how="inner")
        total_revenue = merged["Revenue (COP)"].sum()
        total_opex = merged["Total OPEX (COP)"].sum()
        avg_opex_pct = (total_opex / total_revenue * 100.0) if total_revenue > 0 else 0.0
    else:
        avg_opex_pct = 0.0
    
    # 5. Average SG&A over Revenue
    annual_sga = sga_annual_by_item(s)
    if "Total SG&A (COP)" in annual_sga.columns and "Revenue (COP)" in op.columns:
        merged_sga = op[["Year", "Revenue (COP)"]].merge(annual_sga[["Year", "Total SG&A (COP)"]], on="Year", how="inner")
        total_revenue_sga = merged_sga["Revenue (COP)"].sum()
        total_sga = merged_sga["Total SG&A (COP)"].sum()
        avg_sga_pct = (total_sga / total_revenue_sga * 100.0) if total_revenue_sga > 0 else 0.0
    else:
        avg_sga_pct = 0.0
    
    opdep_col1, opdep_col2 = st.columns(2)
    
    with opdep_col1:
        st.markdown("### Operating Costs")
        opex_col1, opex_col2 = st.columns(2)
        with opex_col1:
            st.metric("Average OPEX / Revenue", f"{avg_opex_pct:.2f}%")
        with opex_col2:
            st.metric("Average SG&A / Revenue", f"{avg_sga_pct:.2f}%")
    
    with opdep_col2:
        st.markdown("### Depreciation Schedule")
        dep = depreciation_annual_table(s)
        dep_display = dep[["Year", "Depreciation (COP)"]].copy()
        if currency == "USD":
            # Convert using year-specific FX rates
            dep_display["Depreciation (USD)"] = dep_display.apply(
                lambda row: _to_usd(float(row["Depreciation (COP)"]), int(row["Year"])), axis=1
            )
            dep_display = dep_display.drop(columns=["Depreciation (COP)"])
        dep_display = _df_format_money(dep_display, [c for c in dep_display.columns if c != "Year"], decimals=0)
        dep_display = _transpose_annual_table(dep_display)
        st.dataframe(dep_display, width='stretch', hide_index=True)
    
    st.divider()
    
    # 7. Renewable Tax Benefits & Debt/Equity (side by side)
    taxdebt_col1, taxdebt_col2 = st.columns(2)
    
    with taxdebt_col1:
        st.markdown("### Renewable Tax Benefits")
        incentives_enabled = bool(getattr(s.incentives, "enable_special_deduction", False))
        st.metric("Renewable Tax Benefits Applied", "Yes" if incentives_enabled else "No")
        
        if incentives_enabled:
            ded_pct = float(getattr(s.incentives, "special_deduction_pct_of_capex", 0.0))
            st.metric("Special Deduction", f"{ded_pct:.1f}%")
    
    with taxdebt_col2:
        st.markdown("### Debt & Equity")
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    
    # Calculate Total Equity Investment (After Fees) - sum of all negative levered CF
    annual_levered_for_equity = levered_cashflow_annual(s)
    total_equity_investment = 0.0
    for _, row in annual_levered_for_equity.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)  # Sum of absolute values of negative CFs
    
    if debt_enabled:
        total_capex = _total_capex_from_lines(s)
        debt_pct_of_capex = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
        debt_amount = (debt_pct_of_capex / 100.0) * total_capex
        debt_pct = (debt_amount / total_capex * 100.0) if total_capex > 0 else 0.0
        
        debt_col1, debt_col2, debt_col3 = st.columns(3)
        with debt_col1:
            st.metric("Debt", _fmt_cop(debt_amount) if currency == "COP" else _fmt_usd(_to_usd(debt_amount, cod.year)))
        with debt_col2:
            st.metric("Debt %", f"{debt_pct:.1f}%")
        with debt_col3:
            st.metric("Equity Investment (After Fees)", _fmt_cop(total_equity_investment) if currency == "COP" else _fmt_usd(_to_usd(total_equity_investment, cod.year)))
    else:
        st.metric("Equity Investment (After Fees)", _fmt_cop(total_equity_investment) if currency == "COP" else _fmt_usd(_to_usd(total_equity_investment, cod.year)))
    
    st.divider()
    
    # 9. Key Metrics & IRR Table
    # Unlevered IRR Pre-tax
    a = unlevered_base_cashflow_annual(s)
    mm = cashflow_monthly_table(s).copy()
    for col in ["CAPEX (COP)", "Unlevered CF (COP)", "Unlevered CF Pre-tax (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)
    
    if "Unlevered CF Pre-tax (COP)" in mm.columns:
        monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist()
    else:
        monthly_cf_pre_tax = mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")
    
    # Levered Equity IRR After-tax
    annual_levered = levered_cashflow_annual(s)
    annual_cf_levered = []
    for _, row in annual_levered.iterrows():
        annual_cf_levered.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
    
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    # Payback
    monthly_levered = levered_cashflow_monthly(s)
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    # 9. Key Metrics & IRR Table (side by side)
    metrics_irr_col1, metrics_irr_col2 = st.columns([1, 1.2])
    
    with metrics_irr_col1:
        st.markdown("### Key Financial Metrics")
        metrics_col1, metrics_col2 = st.columns(2)
        with metrics_col1:
            st.metric("Unlevered IRR (Pre-tax)", f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—")
        with metrics_col2:
            st.metric("Levered Equity IRR (After-tax)", f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—")
        st.metric("Payback (years, after-tax)", f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—")
    
    with metrics_irr_col2:
        with st.expander("Equity IRR Calculation (for verification)", expanded=False):
            irr_table = annual_levered[["Year", "Levered CF (After-tax, COP)"]].copy()
            if currency == "USD":
                irr_table["Levered CF (After-tax, USD)"] = irr_table.apply(
                    lambda row: _to_usd(float(row["Levered CF (After-tax, COP)"]), int(row["Year"])), axis=1
                )
                irr_table = irr_table.drop(columns=["Levered CF (After-tax, COP)"])
            
            # Format the table
            money_cols = [c for c in irr_table.columns if c != "Year"]
            irr_table = _df_format_money(irr_table, money_cols, decimals=0)
            irr_table = _transpose_annual_table(irr_table)
            st.dataframe(irr_table, width='stretch', hide_index=True)
            
            # Show summary row
            total_negative = sum([cf for cf in annual_cf_levered if cf < 0])
            total_positive = sum([cf for cf in annual_cf_levered if cf > 0])
            summary_row = pd.DataFrame([{
                "Year": "Total",
                "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)": 
                    _fmt_cop(sum(annual_cf_levered)) if currency == "COP" else _fmt_usd(_to_usd(sum(annual_cf_levered), cod.year))
            }])
            st.dataframe(summary_row, width='stretch', hide_index=True)
    
    st.divider()
    
    # 10. Charts
    st.markdown("### Levered Free Cash Flow (Equity After-Tax)")
    
    # Prepare annual view for charts
    annual_view = annual_levered.copy()
    money_cols = [c for c in annual_view.columns if c != "Year" and "(COP)" in c]
    for col in money_cols:
        annual_view[col] = pd.to_numeric(annual_view[col], errors="coerce").fillna(0.0)
        if currency == "USD":
            # Convert using year-specific FX rates (accounts for CPI indexation)
            usd_col = col.replace("(COP)", "(USD)")
            annual_view[usd_col] = annual_view.apply(
                lambda row: _to_usd(float(row[col]), int(row["Year"])), axis=1
            )
    
    y_fcf = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if y_fcf not in annual_view.columns:
        y_fcf = "Levered CF (After-tax, COP)"
    
    if PLOTLY_AVAILABLE and px is not None:
        fig_fcf = px.bar(annual_view, x="Year", y=y_fcf)
        
        # Add period indicators
        rtb_year = tl["rtb"].year
        cod_year = tl["cod"].year
        end_op_year = tl["end_op"].year
        
        fig_fcf.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
        fig_fcf.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
        fig_fcf.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
        
        fig_fcf.update_layout(
            height=320,
            margin=dict(l=10, r=10, t=40, b=10),
            title="Equity/Levered After-Tax Free Cash Flow"
        )
        st.plotly_chart(fig_fcf, width='stretch', key="summary_fcf_chart")
    else:
        st.warning("⚠️ Chart unavailable")
    
    st.markdown("### Income Statement Overview")
    
    # Income Statement Graph
    graph_cols = {
        "Revenue": "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA": "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Interest Expense": "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Taxes Payable": "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Net Income After Tax": "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    }
    
    # Create a melted dataframe for the graph
    graph_df = annual_view[["Year"]].copy()
    for label, col in graph_cols.items():
        if col in annual_view.columns:
            graph_df[label] = annual_view[col]
    
    # Melt for grouped bar chart
    graph_long = graph_df.melt(
        id_vars=["Year"],
        value_vars=[label for label, col in graph_cols.items() if col in annual_view.columns],
        var_name="Metric",
        value_name="Amount"
    )
    
    fig_income = px.bar(
        graph_long,
        x="Year",
        y="Amount",
        color="Metric",
        barmode="group",
        title="Revenue, EBITDA, Interest, Taxes, and Net Income"
    )
    
    # Add period indicators
    fig_income.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig_income.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig_income.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig_income.update_layout(height=320, margin=dict(l=10, r=10, t=40, b=10))
    st.plotly_chart(fig_income, width='stretch', key="summary_income_chart")
    
    st.divider()
    
    # 11. Comprehensive Annual Results Table
    st.markdown("### Comprehensive Annual Results")
    st.caption("Complete annual financial and operational metrics with detailed breakdown")
    
    # Gather all required data
    # Use annual_levered as base to include ALL project years (development, construction, operation)
    # Get annual_levered (may already be calculated earlier in this tab, but recalculate to be safe)
    annual_levered_full = levered_cashflow_annual(s).copy()
    
    # Base dataframe with all years from project start to end
    comprehensive = annual_levered_full[["Year"]].copy()
    
    # 1. Operating year table (PPA price, power generation, revenue) - only exists for operation years
    op_table = operating_year_table(s)
    
    # Merge operating data (only exists for operation years)
    required_op_cols = ["Year", "Price (COP/kWh)", "Energy (MWh)", "Revenue (COP)"]
    if not op_table.empty and all(col in op_table.columns for col in required_op_cols):
        comprehensive = comprehensive.merge(
            op_table[required_op_cols],
            on="Year",
            how="left"
        )
        # Fill missing values for non-operation years (development/construction)
        comprehensive["Price (COP/kWh)"] = comprehensive["Price (COP/kWh)"].fillna(0.0)
        comprehensive["Energy (MWh)"] = comprehensive["Energy (MWh)"].fillna(0.0)
        comprehensive["Revenue (COP)"] = comprehensive["Revenue (COP)"].fillna(0.0)
    else:
        # If op_table is missing columns, initialize with zeros
        comprehensive["Price (COP/kWh)"] = 0.0
        comprehensive["Energy (MWh)"] = 0.0
        comprehensive["Revenue (COP)"] = 0.0
    
    # 2. Annual OPEX (aggregate from monthly)
    om_monthly = opex_monthly_schedule(s)
    if not om_monthly.empty and "OPEX subtotal" in om_monthly.columns and "GMF" in om_monthly.columns:
        annual_opex = om_monthly.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
        annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
    else:
        # Create empty OPEX dataframe
        annual_opex = pd.DataFrame(columns=["Year", "Total OPEX (COP)"])
    
    # 3. Annual SG&A
    annual_sga = sga_annual_by_item(s)
    
    # 4. Levered cash flow annual data (already have annual_levered)
    
    # Add OPEX
    comprehensive = comprehensive.merge(
        annual_opex[["Year", "Total OPEX (COP)"]], 
        on="Year", 
        how="left"
    )
    comprehensive["Total OPEX (COP)"] = comprehensive["Total OPEX (COP)"].fillna(0.0)
    
    # Add SG&A
    if "Total SG&A (COP)" in annual_sga.columns:
        comprehensive = comprehensive.merge(
            annual_sga[["Year", "Total SG&A (COP)"]], 
            on="Year", 
            how="left"
        )
        comprehensive["Total SG&A (COP)"] = comprehensive["Total SG&A (COP)"].fillna(0.0)
    else:
        comprehensive["Total SG&A (COP)"] = 0.0
    
    # Add levered cash flow data
    # Use suffixes to avoid column name conflicts, but preserve all columns from comprehensive
    comprehensive = comprehensive.merge(
        annual_levered_full,
        on="Year",
        how="left",
        suffixes=("", "_levered")
    )
    
    # Calculate additional metrics
    # Convert Energy from MWh to kWh (check if column exists first)
    if "Energy (MWh)" in comprehensive.columns:
        comprehensive["Power Generation (kWh/year)"] = comprehensive["Energy (MWh)"].fillna(0.0) * 1000.0
    elif not comprehensive.empty:
        # If comprehensive has data but Energy column is missing, try to calculate from other sources
        # This shouldn't happen, but handle gracefully
        st.warning("Energy (MWh) column missing from comprehensive table. Power Generation will be set to 0.")
        comprehensive["Power Generation (kWh/year)"] = 0.0
    else:
        # Empty dataframe - set to 0
        comprehensive["Power Generation (kWh/year)"] = 0.0
    
    # Calculate OPEX/Revenue %
    comprehensive["OPEX/Revenue %"] = (
        (comprehensive["Total OPEX (COP)"] / comprehensive["Revenue (COP)"]) * 100.0
    ).fillna(0.0)
    comprehensive.loc[comprehensive["Revenue (COP)"] == 0, "OPEX/Revenue %"] = 0.0
    
    # Calculate SG&A/Revenue %
    comprehensive["SG&A/Revenue %"] = (
        (comprehensive["Total SG&A (COP)"] / comprehensive["Revenue (COP)"]) * 100.0
    ).fillna(0.0)
    comprehensive.loc[comprehensive["Revenue (COP)"] == 0, "SG&A/Revenue %"] = 0.0
    
    # Build the comprehensive table with ordered columns
    comp_cols = ["Year"]
    comp_labels = []
    
    # Define the column order and labels with operation symbols showing complete cash flow logic
    # This shows the ENTIRE step-by-step calculation from Revenue to Final Cash Flow
    column_mapping = [
        # Revenue Calculation:
        ("Price (COP/kWh)", "PPA Price (COP/kWh)"),  # Input (no operation symbol)
        ("Power Generation (kWh/year)", "x Power Generation (kWh/year)"),  # Multiplier
        ("Revenue (COP)", "= Revenue"),  # Result: Price × Generation
        
        # Operating Expenses:
        ("Total OPEX (COP)", "(-) OPEX"),  # Subtract operating expenses
        ("OPEX/Revenue %", "OPEX/Revenue %"),  # Percentage (informational)
        ("Total SG&A (COP)", "(-) SG&A"),  # Subtract SG&A expenses
        ("SG&A/Revenue %", "SG&A/Revenue %"),  # Percentage (informational)
        
        # EBITDA Calculation:
        ("EBITDA (COP)", "= EBITDA"),  # Result: Revenue - OPEX - SG&A
        
        # Depreciation & Interest (for tax calculation):
        ("Depreciation (COP)", "(-) Depreciation"),  # Subtract depreciation (for tax)
        ("Interest (COP)", "(-) Interest Expense"),  # Subtract interest payments (tax-deductible)
        
        # Tax Calculations:
        ("Levered CAPEX Tax Deduction (COP)", "(-) CAPEX Tax Deduction"),  # Tax benefit (reduces taxable income)
        ("Levered Loss Carryforward End (COP)", "(-) Loss Carryforward"),  # Tax adjustment (reduces taxable income)
        ("Levered Taxable Income (COP)", "= Taxable Income"),  # Result: EBITDA - Depreciation - Interest - Deductions
        ("Levered Taxes Payable (COP)", "(-) Taxes Payable"),  # Subtract taxes
        
        # Debt Service (cash flow only, not tax-deductible):
        ("Debt Service (COP)", "(-) Debt Service"),  # Subtract total debt service (interest + principal, cash flow only)
        
        # Net Income:
        ("Levered Net Income After Tax (COP)", "= Net Income After Tax"),  # Result: Taxable Income - Taxes
        
        # Cash Flow Adjustments:
        ("CAPEX (COP)", "(-) CAPEX"),  # Subtract capital expenditures
        ("Debt Draw (COP)", "(+) Debt Draw"),  # Add debt proceeds
        ("Principal (COP)", "(-) Principal Payments"),  # Subtract principal repayments
        ("Debt Fees (COP)", "(-) Debt Fees"),  # Subtract debt fees
        ("VAT Refund (COP)", "(+) VAT Refund"),  # Add VAT refund
        ("ΔNWC (COP)", "(-) Working Capital Change"),  # Subtract working capital change
        
        # Final Cash Flow:
        ("Levered CF (After-tax, COP)", "= Levered CF After-tax"),  # Final result: Net Income + Adjustments
    ]
    
    # Handle currency conversion
    currency_suffix = " (COP)" if currency == "COP" else " (USD)"
    
    for col_name, label in column_mapping:
        # Check if column exists in comprehensive dataframe
        if col_name in comprehensive.columns:
            comp_cols.append(col_name)
            comp_labels.append(label)
        # Also check for USD version if currency is USD
        elif currency == "USD":
            usd_col = col_name.replace(" (COP)", " (USD)")
            if usd_col in comprehensive.columns:
                comp_cols.append(usd_col)
                comp_labels.append(label)
    
    # Handle currency conversion BEFORE building comp_df (much simpler)
    if currency == "USD":
        # Convert money columns to USD (but keep PPA Price and Power Generation as-is)
        money_cols_to_convert = [col for col in comp_cols if col != "Year" and 
                                col not in ["Price (COP/kWh)", "Power Generation (kWh/year)", "OPEX/Revenue %", "SG&A/Revenue %"] 
                                and "(COP)" in col]
        
        for col in money_cols_to_convert:
            if col in comprehensive.columns:
                usd_col = col.replace(" (COP)", " (USD)")
                comprehensive[usd_col] = comprehensive.apply(
                    lambda row: _to_usd(float(row[col]), int(row["Year"])),
                    axis=1
                )
                # Update comp_cols to use USD column
                if col in comp_cols:
                    idx = comp_cols.index(col)
                    comp_cols[idx] = usd_col
    
    # Create the comprehensive table with updated columns
    comp_df = comprehensive[comp_cols].copy()
    
    # Convert to numeric where needed (for formatting) - BEFORE transposing
    for col in comp_df.columns:
        if col != "Year":
            comp_df[col] = pd.to_numeric(comp_df[col], errors="coerce").fillna(0.0)
    
    # Transpose the table (years as columns) - MUST happen before renaming
    # _transpose_annual_table expects "Year" column and creates "Metric" column
    comp_df = _transpose_annual_table(comp_df)
    
    # Now rename the metric labels (the "Metric" column contains the original column names)
    # We need to map the original column names to the new labels with operation symbols
    if "Metric" in comp_df.columns:
        # Create a mapping from original column names to labels
        label_map = {}
        for col_name, label in column_mapping:
            label_map[col_name] = label
            # Also handle USD versions if currency is USD
            if currency == "USD":
                usd_col = col_name.replace(" (COP)", " (USD)")
                label_map[usd_col] = label
        
        # Update the Metric column with the new labels
        comp_df["Metric"] = comp_df["Metric"].map(label_map).fillna(comp_df["Metric"])
    
    # Format the table
    # After transposition, columns are years and "Metric" column contains the row labels
    # We need to format by row based on the Metric value
    
    # Ensure all column names are strings (years might be integers) to avoid mixed type warnings
    comp_df.columns = [str(c) for c in comp_df.columns]
    
    # Convert year columns to object dtype to allow string formatting
    year_cols = [c for c in comp_df.columns if c != "Metric"]
    for col in year_cols:
        # Convert column to object dtype to allow string values
        comp_df[col] = comp_df[col].astype(object)
    
    # Format each row based on its Metric value
    for idx, row in comp_df.iterrows():
        metric_value = str(row["Metric"])
        
        # Format percentage rows
        if "%" in metric_value:
            for col in year_cols:
                if col in comp_df.columns:
                    val = row[col]
                    try:
                        comp_df.at[idx, col] = f"{float(val):.2f}%" if pd.notna(val) and val != "" and str(val) != "—" else ("—" if pd.isna(val) or val == "" else str(val))
                    except (ValueError, TypeError):
                        comp_df.at[idx, col] = "—"
        
        # Format PPA Price row (keep 4 decimals)
        elif "PPA Price" in metric_value:
            for col in year_cols:
                if col in comp_df.columns:
                    val = row[col]
                    try:
                        comp_df.at[idx, col] = f"{float(val):,.4f}" if pd.notna(val) and val != "" and str(val) != "—" else "—"
                    except (ValueError, TypeError):
                        comp_df.at[idx, col] = "—"
        
        # Format Power Generation row (no decimals, with commas)
        elif "Power Generation" in metric_value:
            for col in year_cols:
                if col in comp_df.columns:
                    val = row[col]
                    try:
                        comp_df.at[idx, col] = f"{float(val):,.0f}" if pd.notna(val) and val != "" and str(val) != "—" else "—"
                    except (ValueError, TypeError):
                        comp_df.at[idx, col] = "—"
        
        # Format all other rows as money (with commas, no decimals)
        else:
            for col in year_cols:
                if col in comp_df.columns:
                    val = row[col]
                    try:
                        if pd.notna(val) and val != "" and str(val) != "—":
                            comp_df.at[idx, col] = _fmt_num(float(val), 0)
                        else:
                            comp_df.at[idx, col] = "—"
                    except (ValueError, TypeError):
                        comp_df.at[idx, col] = "—"
    
    # Display the comprehensive table
    st.dataframe(comp_df, width='stretch', hide_index=True)


# -----------------------------
# O) Sensitivity Analysis
# -----------------------------
with tab_sensitivity:
    st.subheader("Sensitivity Analysis - Equity IRR (After-Tax)")
    
    st.info("Select two variables to analyze how Equity IRR (After-Tax) changes with different values.")
    
    # Variable selection
    col1, col2 = st.columns(2)
    
    with col1:
        var1_name = st.selectbox(
            "Variable 1 (X-axis)",
            ["CAPEX", "PPA Price", "Production (MWh/yr)", "OPEX", "Debt %"],
            index=0,
            key="sens_var1"
        )
        
        var1_base = None
        var1_min = None
        var1_max = None
        var1_steps = None
        
        if var1_name == "CAPEX":
            total_capex = _total_capex_from_lines(s)
            var1_base = total_capex
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "PPA Price":
            if s.revenue_mode == "Standard PPA Parameters":
                var1_base = float(s.revenue1.ppa_price_cop_per_kwh)
            else:
                st.warning("PPA Price sensitivity requires Standard PPA Parameters mode")
                var1_base = 0.0
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "Production (MWh/yr)":
            p_map = {"P50": s.generation.p50_mwh_yr, "P75": s.generation.p75_mwh_yr, "P90": s.generation.p90_mwh_yr}
            var1_base = float(p_map.get(s.generation.production_choice, s.generation.p50_mwh_yr))
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "OPEX":
            # Use average OPEX per year as base
            om = opex_monthly_schedule(s)
            annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
            annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
            var1_base = float(annual_opex["Total OPEX (COP)"].mean()) if len(annual_opex) > 0 else 0.0
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "Debt %":
            var1_base = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
            var1_min = st.number_input("Variable 1 Min (%)", value=0.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var1_min")
            var1_max = st.number_input("Variable 1 Max (%)", value=80.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var1_max")
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
    
    with col2:
        var2_name = st.selectbox(
            "Variable 2 (Y-axis)",
            ["CAPEX", "PPA Price", "Production (MWh/yr)", "OPEX", "Debt %"],
            index=1,
            key="sens_var2"
        )
        
        var2_base = None
        var2_min = None
        var2_max = None
        var2_steps = None
        
        if var2_name == "CAPEX":
            total_capex = _total_capex_from_lines(s)
            var2_base = total_capex
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "PPA Price":
            if s.revenue_mode == "Standard PPA Parameters":
                var2_base = float(s.revenue1.ppa_price_cop_per_kwh)
            else:
                st.warning("PPA Price sensitivity requires Standard PPA Parameters mode")
                var2_base = 0.0
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "Production (MWh/yr)":
            p_map = {"P50": s.generation.p50_mwh_yr, "P75": s.generation.p75_mwh_yr, "P90": s.generation.p90_mwh_yr}
            var2_base = float(p_map.get(s.generation.production_choice, s.generation.p50_mwh_yr))
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "OPEX":
            om = opex_monthly_schedule(s)
            annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
            annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
            var2_base = float(annual_opex["Total OPEX (COP)"].mean()) if len(annual_opex) > 0 else 0.0
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "Debt %":
            var2_base = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
            var2_min = st.number_input("Variable 2 Min (%)", value=0.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var2_min")
            var2_max = st.number_input("Variable 2 Max (%)", value=80.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var2_max")
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
    
    # Ensure variables are different
    if var1_name == var2_name:
        st.error("Please select two different variables for sensitivity analysis.")
    elif var1_base is None or var2_base is None or var1_base == 0 or var2_base == 0:
        st.warning("Cannot perform sensitivity analysis. Please check variable inputs.")
    else:
        # Generate value ranges
        if var1_name == "Debt %":
            var1_values = np.linspace(var1_min, var1_max, int(var1_steps))
        else:
            var1_values = np.linspace(var1_base * var1_min, var1_base * var1_max, int(var1_steps))
        
        if var2_name == "Debt %":
            var2_values = np.linspace(var2_min, var2_max, int(var2_steps))
        else:
            var2_values = np.linspace(var2_base * var2_min, var2_base * var2_max, int(var2_steps))
        
        # Run sensitivity analysis
        st.markdown("### Running Sensitivity Analysis...")
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        results = []
        total_runs = len(var1_values) * len(var2_values)
        run_count = 0
        
        for v1 in var1_values:
            for v2 in var2_values:
                # Create a copy of the scenario
                s_mod = _scenario_from_dict(_scenario_to_dict(s))
                
                # Modify variable 1
                if var1_name == "CAPEX":
                    # Scale all CAPEX line items proportionally
                    scale_factor = v1 / var1_base
                    for item in s_mod.capex.lines:
                        if "Amount_COP" in item:
                            item["Amount_COP"] = float(item.get("Amount_COP", 0.0)) * scale_factor
                elif var1_name == "PPA Price":
                    if s_mod.revenue_mode == "Standard PPA Parameters":
                        s_mod.revenue1.ppa_price_cop_per_kwh = v1
                elif var1_name == "Production (MWh/yr)":
                    # Scale all production values
                    scale_factor = v1 / var1_base
                    s_mod.generation.p50_mwh_yr = float(s_mod.generation.p50_mwh_yr) * scale_factor
                    s_mod.generation.p75_mwh_yr = float(s_mod.generation.p75_mwh_yr) * scale_factor
                    s_mod.generation.p90_mwh_yr = float(s_mod.generation.p90_mwh_yr) * scale_factor
                elif var1_name == "OPEX":
                    # Scale OPEX components
                    scale_factor = v1 / var1_base
                    s_mod.opex.fixed_om_cop_per_mwac_year = float(s_mod.opex.fixed_om_cop_per_mwac_year) * scale_factor
                    s_mod.opex.variable_om_cop_per_mwh = float(s_mod.opex.variable_om_cop_per_mwh) * scale_factor
                    s_mod.opex.insurance_cop_per_mwac_year = float(s_mod.opex.insurance_cop_per_mwac_year) * scale_factor
                    s_mod.opex.grid_fees_cop_per_mwh = float(s_mod.opex.grid_fees_cop_per_mwh) * scale_factor
                elif var1_name == "Debt %":
                    s_mod.debt.debt_pct_of_capex = v1
                
                # Modify variable 2
                if var2_name == "CAPEX":
                    scale_factor = v2 / var2_base
                    for item in s_mod.capex.lines:
                        if "Amount_COP" in item:
                            item["Amount_COP"] = float(item.get("Amount_COP", 0.0)) * scale_factor
                elif var2_name == "PPA Price":
                    if s_mod.revenue_mode == "Standard PPA Parameters":
                        s_mod.revenue1.ppa_price_cop_per_kwh = v2
                elif var2_name == "Production (MWh/yr)":
                    scale_factor = v2 / var2_base
                    s_mod.generation.p50_mwh_yr = float(s_mod.generation.p50_mwh_yr) * scale_factor
                    s_mod.generation.p75_mwh_yr = float(s_mod.generation.p75_mwh_yr) * scale_factor
                    s_mod.generation.p90_mwh_yr = float(s_mod.generation.p90_mwh_yr) * scale_factor
                elif var2_name == "OPEX":
                    scale_factor = v2 / var2_base
                    s_mod.opex.fixed_om_cop_per_mwac_year = float(s_mod.opex.fixed_om_cop_per_mwac_year) * scale_factor
                    s_mod.opex.variable_om_cop_per_mwh = float(s_mod.opex.variable_om_cop_per_mwh) * scale_factor
                    s_mod.opex.insurance_cop_per_mwac_year = float(s_mod.opex.insurance_cop_per_mwac_year) * scale_factor
                    s_mod.opex.grid_fees_cop_per_mwh = float(s_mod.opex.grid_fees_cop_per_mwh) * scale_factor
                elif var2_name == "Debt %":
                    s_mod.debt.debt_pct_of_capex = v2
                
                # Calculate Equity IRR for modified scenario
                try:
                    annual_levered_mod = levered_cashflow_annual(s_mod)
                    annual_cf_levered_mod = []
                    for _, row in annual_levered_mod.iterrows():
                        annual_cf_levered_mod.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
                    
                    has_pos = any(cf > 0 for cf in annual_cf_levered_mod)
                    has_neg = any(cf < 0 for cf in annual_cf_levered_mod)
                    irr = _irr_bisection(annual_cf_levered_mod) if (has_pos and has_neg) else float("nan")
                    irr_pct = irr * 100.0 if np.isfinite(irr) else float("nan")
                except Exception as e:
                    irr_pct = float("nan")
                
                results.append({
                    "Var1": v1,
                    "Var2": v2,
                    "IRR": irr_pct
                })
                
                run_count += 1
                progress_bar.progress(run_count / total_runs)
                status_text.text(f"Completed {run_count}/{total_runs} scenarios...")
        
        progress_bar.empty()
        status_text.empty()
        
        # Create sensitivity table
        st.markdown("### Sensitivity Table - Equity IRR (After-Tax) %")
        
        # Convert to pivot table format
        df_results = pd.DataFrame(results)
        pivot_table = df_results.pivot_table(
            values="IRR",
            index="Var2",
            columns="Var1",
            aggfunc="first"
        )
        
        # Format for display
        pivot_display = pivot_table.copy()
        pivot_display = pivot_display.round(2)
        
        # Format row and column labels
        if var1_name == "Debt %":
            pivot_display.columns = [f"{c:.1f}%" for c in pivot_display.columns]
        elif var1_name == "CAPEX":
            pivot_display.columns = [_fmt_cop(c) for c in pivot_display.columns]
        elif var1_name == "PPA Price":
            pivot_display.columns = [f"{c:.4f}" for c in pivot_display.columns]
        elif var1_name == "Production (MWh/yr)":
            pivot_display.columns = [f"{c:,.0f}" for c in pivot_display.columns]
        else:
            pivot_display.columns = [_fmt_cop(c) for c in pivot_display.columns]
        
        if var2_name == "Debt %":
            pivot_display.index = [f"{i:.1f}%" for i in pivot_display.index]
        elif var2_name == "CAPEX":
            pivot_display.index = [_fmt_cop(i) for i in pivot_display.index]
        elif var2_name == "PPA Price":
            pivot_display.index = [f"{i:.4f}" for i in pivot_display.index]
        elif var2_name == "Production (MWh/yr)":
            pivot_display.index = [f"{i:,.0f}" for i in pivot_display.index]
        else:
            pivot_display.index = [_fmt_cop(i) for i in pivot_display.index]
        
        # Add row and column headers
        pivot_display.index.name = var2_name
        pivot_display.columns.name = var1_name
        
        st.dataframe(pivot_display, width='stretch')
        
        # Create heatmap with text annotations
        st.markdown("### Sensitivity Heatmap")
        
        # Prepare axis labels - use percentages with actual values below
        def format_x_label(val, var_name, base_val):
            if var_name == "Debt %":
                # Debt % is already a percentage value
                return f"{val:.1f}%"
            else:
                pct = (val / base_val) * 100 if base_val > 0 else 0
                if var_name == "CAPEX":
                    # Format as percentage on top, actual value below
                    return f"{(pct):.0f}%<br>{val/1e9:.2f}B COP"
                elif var_name == "PPA Price":
                    return f"{(pct):.0f}%<br>{val:.0f} COP/kWh"
                elif var_name == "Production (MWh/yr)":
                    return f"{(pct):.0f}%<br>{val:,.0f} MWh"
                elif var_name == "OPEX":
                    return f"{(pct):.0f}%<br>{val/1e6:.1f}M COP"
                else:
                    return f"{(pct):.0f}%"
        
        def format_y_label(val, var_name, base_val):
            if var_name == "Debt %":
                # Debt % is already a percentage value
                return f"{val:.1f}%"
            else:
                pct = (val / base_val) * 100 if base_val > 0 else 0
                if var_name == "CAPEX":
                    return f"{(pct):.0f}%<br>{val/1e9:.2f}B COP"
                elif var_name == "PPA Price":
                    return f"{(pct):.0f}%<br>{val:.0f} COP/kWh"
                elif var_name == "Production (MWh/yr)":
                    return f"{(pct):.0f}%<br>{val:,.0f} MWh"
                elif var_name == "OPEX":
                    return f"{(pct):.0f}%<br>{val/1e6:.1f}M COP"
                else:
                    return f"{(pct):.0f}%"
        
        x_labels = [format_x_label(c, var1_name, var1_base) for c in pivot_table.columns]
        y_labels = [format_y_label(r, var2_name, var2_base) for r in pivot_table.index]
        
        # Create text matrix for annotations (IRR values)
        text_matrix = []
        for row in pivot_table.values:
            text_row = []
            for val in row:
                if np.isfinite(val):
                    text_row.append(f"{val:.2f}%")
                else:
                    text_row.append("—")
            text_matrix.append(text_row)
        
        # Create heatmap using graph_objects (supports text parameter)
        if PLOTLY_AVAILABLE and go is not None:
            fig = go.Figure(data=go.Heatmap(
                z=pivot_table.values,
                x=x_labels,
                y=y_labels,
                text=text_matrix,
                texttemplate="%{text}",
                textfont={"size": 18, "color": "black", "family": "Arial Black"},  # Much larger font
                colorscale="RdYlGn",
                colorbar=dict(title="Equity IRR (%)"),
                hovertemplate="%{y}<br>%{x}<br>IRR: %{z:.2f}%<extra></extra>"
            ))
            
            fig.update_layout(
                title=f"Equity IRR (After-Tax) Sensitivity: {var1_name} vs {var2_name}",
                xaxis_title=var1_name,
                yaxis_title=var2_name,
                height=500,
                margin=dict(l=10, r=10, t=50, b=10)
            )
            st.plotly_chart(fig, width='stretch', key="sensitivity_heatmap")
        else:
            st.warning("⚠️ Chart unavailable")


# Persist scenario on each run
proj["scenarios"][scenario_name] = _scenario_to_dict(s)
_save_db(db)
