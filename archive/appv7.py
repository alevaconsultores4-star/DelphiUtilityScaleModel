# Delphi Utility-Scale Financial Model (No Excel)
# Streamlit single-file app with: Projects + Scenarios, Macro, Timeline, Generation, Revenues,
# CAPEX, OPEX, SG&A, Depreciation, Debt & Covenants, Unlevered Base Cash Flow, Compare
# All inputs in COP; outputs selectable COP/USD (USD via FX path).

from __future__ import annotations

import json
import os
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from typing import Dict, List
from io import BytesIO

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# -----------------------------
# Storage
# -----------------------------
# Check for data folder first, then root directory
if os.path.exists("data/delphi_projects.json"):
    PROJECTS_FILE = "data/delphi_projects.json"
elif os.path.exists("delphi_projects.json"):
    PROJECTS_FILE = "delphi_projects.json"
else:
    PROJECTS_FILE = "delphi_projects.json"  # Default location


def _load_db() -> dict:
    if os.path.exists(PROJECTS_FILE):
        try:
            with open(PROJECTS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"projects": {}}
    return {"projects": {}}


def _save_db(db: dict) -> None:
    with open(PROJECTS_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2, ensure_ascii=False)


# -----------------------------
# Formatting helpers (thousand separators)
# -----------------------------
def _fmt_num(x: float, decimals: int = 0) -> str:
    if x is None:
        return "—"
    try:
        if isinstance(x, float) and not np.isfinite(x):
            return "—"
    except Exception:
        pass
    return f"{x:,.{decimals}f}"


def _fmt_cop(x: float) -> str:
    return f"COP {_fmt_num(float(x), 0)}"


def _fmt_usd(x: float) -> str:
    return f"USD {_fmt_num(float(x), 0)}"


def _df_format_money(df: pd.DataFrame, cols: List[str], decimals: int = 0) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda v: _fmt_num(float(v), decimals) if pd.notnull(v) else "")
    return out


def _metric_row(items):
    # Filter out empty labels to avoid warnings
    filtered_items = [(k, v) for k, v in items if k and v]
    if not filtered_items:
        return
    cols = st.columns(len(items))
    for i, (k, v) in enumerate(items):
        if k and v:
            cols[i].metric(k, v)
        # Skip empty metrics (they were just for spacing)


# -----------------------------
# Date utilities
# -----------------------------
def _today() -> date:
    return date.today()


def _parse_date_iso(s: str) -> date:
    return date.fromisoformat(s)


def _add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    day = min(
        d.day,
        [31, 29 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1],
    )
    return date(y, m, day)


def _month_starts(start: date, months: int) -> List[date]:
    return [_add_months(start, i) for i in range(months)]


def build_timeline(t: "TimelineInputs") -> dict:
    start = _parse_date_iso(t.start_date) if t.start_date else _today()
    rtb = _add_months(start, int(t.dev_months))
    cod = _add_months(rtb, int(t.capex_months))
    end_op = _add_months(cod, int(t.operation_years) * 12)
    return {"start": start, "rtb": rtb, "cod": cod, "end_op": end_op}


# -----------------------------
# Inputs (dataclasses)
# -----------------------------
INDEX_CHOICES = ["Colombia CPI", "Colombia PPI", "US CPI", "Custom"]


@dataclass
class MacroInputs:
    col_cpi: float = 6.0
    col_ppi: float = 5.0
    us_cpi: float = 3.0
    custom_index_rate: float = 5.0

    fx_cop_per_usd_start: float = 4000.0
    fx_method: str = "Inflation differential (PPP approx.)"  # or "Flat"
    fx_flat: float = 4000.0


@dataclass
class TimelineInputs:
    start_date: str = str(_today())
    dev_months: int = 18
    capex_months: int = 12
    operation_years: int = 25


@dataclass
class GenerationInputs:
    mwac: float = 100.0
    mwp: float = 130.0
    p50_mwh_yr: float = 220000.0
    p75_mwh_yr: float = 210000.0
    p90_mwh_yr: float = 200000.0
    production_choice: str = "P50"
    degradation_pct: float = 0.5  # %/yr


@dataclass
class RevenueOption1PPA:
    ppa_price_cop_per_kwh: float = 320.0
    ppa_term_years: int = 12
    merchant_price_cop_per_kwh: float = 250.0
    indexation: str = "Colombia CPI"


def _default_manual_prices() -> Dict[int, float]:
    return {i: 300.0 for i in range(1, 26)}


@dataclass
class RevenueOption2Manual:
    prices_constant_cop_per_kwh: Dict[int, float] = field(default_factory=_default_manual_prices)
    indexation: str = "Colombia CPI"


CAPEX_PHASES = ["Development", "Construction", "At COD"]
CAPEX_DISTS = ["Straight-line (monthly)", "Front-loaded", "Back-loaded"]


def _default_capex_lines() -> List[Dict[str, object]]:
    return [
        {"Item": "EPC (modules + BOS + installation)", "Amount_COP": 0.0, "Phase": "Construction"},
        {"Item": "Interconnection / Substation / Line", "Amount_COP": 0.0, "Phase": "Construction"},
        {"Item": "Development costs", "Amount_COP": 0.0, "Phase": "Development"},
    ]


@dataclass
class CapexInputs:
    lines: List[Dict[str, object]] = field(default_factory=_default_capex_lines)
    distribution: str = "Straight-line (monthly)"


PHASES = ["Development", "Construction", "Operation"]


def _default_opex_other_items() -> List[Dict[str, object]]:
    return [
        {"Item": "Other OPEX item", "Amount_COP_per_year": 0.0, "Phase": "Operation", "Indexation": "Colombia CPI"},
    ]


@dataclass
class OpexInputs:
    fixed_om_cop_per_mwac_year: float = 0.0
    fixed_om_indexation: str = "Colombia CPI"

    variable_om_cop_per_mwh: float = 0.0
    variable_om_indexation: str = "Colombia CPI"

    insurance_cop_per_mwac_year: float = 0.0
    insurance_indexation: str = "Colombia CPI"

    grid_fees_cop_per_mwh: float = 0.0

    land_hectares: float = 0.0
    land_price_dev_cop_per_ha_year: float = 0.0
    land_price_con_cop_per_ha_year: float = 0.0
    land_price_op_cop_per_ha_year: float = 0.0
    land_indexation: str = "Colombia CPI"

    ica_pct_of_revenue: float = 0.0
    gmf_pct_of_outflows: float = 0.4  # default 0.4%

    other_items: List[Dict[str, object]] = field(default_factory=_default_opex_other_items)


def _default_sga_items() -> List[Dict[str, object]]:
    return [
        {"Item": "Project management / Owner team", "Amount_COP_per_year": 0.0, "Phase": "Development", "Indexation": "Colombia CPI"},
        {"Item": "Permitting / legal / compliance", "Amount_COP_per_year": 0.0, "Phase": "Development", "Indexation": "Colombia CPI"},
        {"Item": "Corporate overhead allocation", "Amount_COP_per_year": 0.0, "Phase": "Operation", "Indexation": "Colombia CPI"},
    ]


@dataclass
class SgaInputs:
    items: List[Dict[str, object]] = field(default_factory=_default_sga_items)


@dataclass
class DepreciationInputs:
    pct_of_capex_depreciated: float = 100.0
    dep_years: int = 20  # 3–25


@dataclass
class TaxInputs:
    corporate_tax_rate_pct: float = 35.0
    allow_loss_carryforward: bool = True

@dataclass
class RenewableIncentivesInputs:
    # Special deduction (Ley 1715 / 2099 style): up to 50% of eligible investment
    # usable over up to 15 years, but each year capped at 50% of taxable income.
    enable_special_deduction: bool = True
    special_deduction_pct_of_capex: float = 50.0          # 0–50 (%)
    special_deduction_years: int = 15                     # typically 15
    special_deduction_max_pct_of_taxable_income: float = 50.0  # cap per year

    # VAT: model as either excluded (default) or refunded as a one-time cash inflow
    vat_mode: str = "Excluded"                            # "Excluded" or "Refund"
    vat_pct_of_capex: float = 0.0                         # if you prefer % input
    vat_fixed_cop: float = 0.0                            # or fixed COP input
    vat_refund_year_index: int = 1                        # refund in op Year 1 by default

@dataclass
class WorkingCapitalInputs:
    ar_days: int = 90   # revenue collection lag
    ap_days: int = 60   # expense payment lag
    apply_ap_to_opex: bool = True
    apply_ap_to_sga: bool = True

@dataclass
class DebtInputs:
    enabled: bool = False
    debt_pct_of_capex: float = 70.0  # %
    tenor_years: int = 7            # allow 5–10
    grace_years: int = 0            # 0–(tenor-1)

    # Pricing (Natural COP)
    base_rate_pct: float = 9.0      # e.g., IBR
    margin_pct: float = 7.0         # spread
    # Fees
    upfront_fee_bps: float = 175.0  # bps of total debt, paid at COD/first draw
    commitment_fee_pct_of_margin: float = 30.0  # % of margin, on undrawn during construction

    # Covenants (for indicators)
    target_dscr: float = 1.20
    min_dscr_covenant: float = 1.20
    lockup_dscr: float = 1.15

    amortization_type: str = "Sculpted to DSCR"  # "Sculpted to DSCR", "Equal principal", or "Typical amortization"
    balloon_pct: float = 0.0   # NEW: % of original debt left as balloon (0 = fully amortizing)

@dataclass
class RenewableTaxInputs:
    enable_special_deduction: bool = True
    special_deduction_pct_of_capex: float = 50.0   # % of CAPEX
    special_deduction_years: int = 15               # max carryforward

    enable_vat_refund: bool = True
    vat_refund_mode: str = "percent"                # "percent" or "fixed"
    vat_pct_of_capex: float = 19.0                  # if percent
    vat_fixed_cop: float = 0.0                      # if fixed
    vat_refund_year: int = 1                         # years after COD


@dataclass
class ScenarioInputs:
    name: str = "Base"
    macro: MacroInputs = field(default_factory=MacroInputs)
    timeline: TimelineInputs = field(default_factory=TimelineInputs)
    generation: GenerationInputs = field(default_factory=GenerationInputs)

    revenue_mode: str = "Standard PPA Parameters"
    revenue1: RevenueOption1PPA = field(default_factory=RevenueOption1PPA)
    revenue2: RevenueOption2Manual = field(default_factory=RevenueOption2Manual)

    capex: CapexInputs = field(default_factory=CapexInputs)
    opex: OpexInputs = field(default_factory=OpexInputs)
    sga: SgaInputs = field(default_factory=SgaInputs)

    depreciation: DepreciationInputs = field(default_factory=DepreciationInputs)
    tax: TaxInputs = field(default_factory=TaxInputs)
    wc: WorkingCapitalInputs = field(default_factory=WorkingCapitalInputs)

    debt: DebtInputs = field(default_factory=DebtInputs)
    renewable_tax: RenewableTaxInputs = field(default_factory=RenewableTaxInputs)
    incentives: RenewableIncentivesInputs = field(default_factory=RenewableIncentivesInputs)



def _scenario_to_dict(s: ScenarioInputs) -> dict:
    return asdict(s)


def _scenario_from_dict(d: dict) -> ScenarioInputs:
    macro = MacroInputs(**d.get("macro", {}))
    timeline = TimelineInputs(**d.get("timeline", {}))
    generation = GenerationInputs(**d.get("generation", {}))

    revenue_mode = d.get("revenue_mode", "Standard PPA Parameters")
    revenue1 = RevenueOption1PPA(**d.get("revenue1", {}))
    # Handle revenue2 with proper integer key conversion for prices_constant_cop_per_kwh
    revenue2_dict = d.get("revenue2", {})
    if "prices_constant_cop_per_kwh" in revenue2_dict:
        # Convert string keys back to integers (JSON serializes dict keys as strings)
        prices_dict = revenue2_dict["prices_constant_cop_per_kwh"]
        if isinstance(prices_dict, dict) and len(prices_dict) > 0:
            # Check if keys are strings and convert to int
            first_key = next(iter(prices_dict.keys()))
            if isinstance(first_key, str):
                revenue2_dict["prices_constant_cop_per_kwh"] = {int(k): float(v) for k, v in prices_dict.items()}
    revenue2 = RevenueOption2Manual(**revenue2_dict)

    capex = CapexInputs(**d.get("capex", {})) if "capex" in d else CapexInputs()
    opex = OpexInputs(**d.get("opex", {})) if "opex" in d else OpexInputs()
    sga = SgaInputs(**d.get("sga", {})) if "sga" in d else SgaInputs()

    depreciation = DepreciationInputs(**d.get("depreciation", {})) if "depreciation" in d else DepreciationInputs()
    tax = TaxInputs(**d.get("tax", {})) if "tax" in d else TaxInputs()
    wc = WorkingCapitalInputs(**d.get("wc", {})) if "wc" in d else WorkingCapitalInputs()
    debt = DebtInputs(**d.get("debt", {})) if "debt" in d else DebtInputs()
    incentives = RenewableIncentivesInputs(**d.get("incentives", {}))

    return ScenarioInputs(
        name=d.get("name", "Base"),
        macro=macro,
        timeline=timeline,
        generation=generation,
        revenue_mode=revenue_mode,
        revenue1=revenue1,
        revenue2=revenue2,
        capex=capex,
        opex=opex,
        sga=sga,
        depreciation=depreciation,
        tax=tax,
        wc=wc,
        debt=debt,
        incentives=incentives,
    )


# -----------------------------
# Macro series
# -----------------------------
def _idx_key(choice: str) -> str:
    if choice == "Colombia CPI":
        return "col_cpi"
    if choice == "Colombia PPI":
        return "col_ppi"
    if choice == "US CPI":
        return "us_cpi"
    return "custom_index_rate"


def annual_index_series(macro: MacroInputs, base_year: int, years: List[int], index_key: str) -> pd.Series:
    r = float(getattr(macro, index_key)) / 100.0
    out = {}
    for y in years:
        n = y - base_year
        out[y] = (1.0 + r) ** n
    return pd.Series(out)


def fx_series(macro: MacroInputs, base_year: int, years: List[int]) -> pd.Series:
    fx0 = float(macro.fx_cop_per_usd_start)
    if macro.fx_method == "Flat":
        return pd.Series({y: float(macro.fx_flat or fx0) for y in years})
    col = float(macro.col_cpi) / 100.0
    us = float(macro.us_cpi) / 100.0
    drift = (1.0 + col) / (1.0 + us) - 1.0
    return pd.Series({y: fx0 * ((1.0 + drift) ** (y - base_year)) for y in years})


def index_factor_for_year(macro: MacroInputs, year: int, base_year: int, index_choice: str) -> float:
    years = list(range(min(base_year, year), max(base_year, year) + 1))
    s = annual_index_series(macro, base_year, years, _idx_key(index_choice))
    return float(s.loc[year])


# -----------------------------
# Generation & Revenues (annual)
# -----------------------------
def operating_year_table(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    end_op = tl["end_op"]
    op_years = int(s.timeline.operation_years)

    # Get all calendar years from COD to end of operation
    years = list(range(cod.year, end_op.year + 1))
    base_year = cod.year

    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    base_mwh = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    degr = float(gen.degradation_pct) / 100.0
    
    # Calculate prorated generation for each calendar year
    mwh_list = []
    price_base_list = []
    
    for year in years:
        # Determine start and end dates for this calendar year
        if year == cod.year:
            # First year: from COD date to end of year
            year_start = cod
            year_end = date(year, 12, 31)
        elif year == end_op.year:
            # Last year: from start of year to end_op date
            year_start = date(year, 1, 1)
            year_end = end_op
        else:
            # Full year
            year_start = date(year, 1, 1)
            year_end = date(year, 12, 31)
        
        # Calculate months of operation in this calendar year
        months_in_year = (year_end.year - year_start.year) * 12 + (year_end.month - year_start.month) + 1
        fraction_of_year = months_in_year / 12.0
        
        # Calculate which operating year the middle of this calendar year falls into
        # Use the midpoint of the calendar year to determine operating year
        if year == cod.year:
            mid_month = (cod.month + 12) / 2.0
        elif year == end_op.year:
            mid_month = (1 + end_op.month) / 2.0
        else:
            mid_month = 6.5  # Middle of year
        
        # Calculate months from COD to midpoint of this calendar year
        months_from_cod = (year - cod.year) * 12 + (mid_month - cod.month)
        operating_year_num = int(months_from_cod // 12)  # 0-indexed operating year
        
        # Calculate annual generation for this operating year (with degradation)
        annual_mwh = base_mwh * ((1.0 - degr) ** operating_year_num)
        
        # Prorate by fraction of year
        prorated_mwh = annual_mwh * fraction_of_year
        
        # Calculate price based on operating year
        if s.revenue_mode == "Standard PPA Parameters":
            r = s.revenue1
            term = int(r.ppa_term_years)
            p0 = float(r.ppa_price_cop_per_kwh)
            pm = float(r.merchant_price_cop_per_kwh)
            price = p0 if (operating_year_num < term) else pm
        else:
            r = s.revenue2
            # Look up price by operating year (1-indexed)
            # operating_year_num is 0-indexed (0 = first operating year, 1 = second, etc.)
            # Manual prices are stored with 1-indexed keys (1 = first operating year, 2 = second, etc.)
            op_year_key = operating_year_num + 1
            price = float(r.prices_constant_cop_per_kwh.get(op_year_key, 0.0))
            
            # Debug: if price is 0, check if the key exists in the dictionary
            # This helps identify if it's a lookup issue or missing data
            if price == 0.0 and op_year_key not in r.prices_constant_cop_per_kwh:
                # Price not found for this operating year - will show warning in UI
                pass
        
        mwh_list.append(prorated_mwh)
        price_base_list.append(price)
    
    # Apply indexation
    index_choice = s.revenue1.indexation if s.revenue_mode == "Standard PPA Parameters" else s.revenue2.indexation
    idx = annual_index_series(s.macro, base_year, years, _idx_key(index_choice))
    price_indexed = [price_base_list[i] * float(idx.loc[years[i]]) for i in range(len(years))]

    df = pd.DataFrame({"Year": years, "Energy (MWh)": mwh_list})
    df["Price (COP/kWh)"] = price_indexed
    df["Revenue (COP)"] = df["Energy (MWh)"] * 1000.0 * df["Price (COP/kWh)"]
    return df


# -----------------------------
# CAPEX
# -----------------------------
def _weights(n: int, mode: str) -> np.ndarray:
    if n <= 0:
        return np.array([])
    if mode == "Straight-line (monthly)":
        w = np.ones(n)
    elif mode == "Front-loaded":
        w = np.linspace(n, 1, n)
    else:
        w = np.linspace(1, n, n)
    return w / w.sum()


def capex_monthly_schedule(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    start = tl["start"]
    rtb = tl["rtb"]
    cod = tl["cod"]

    dev_n = int(s.timeline.dev_months)
    con_n = int(s.timeline.capex_months)

    dev_months = _month_starts(start, dev_n)
    con_months = _month_starts(rtb, con_n)

    lines = s.capex.lines or []
    dev_total = sum(float(x.get("Amount_COP", 0.0) or 0.0) for x in lines if x.get("Phase") == "Development")
    con_total = sum(float(x.get("Amount_COP", 0.0) or 0.0) for x in lines if x.get("Phase") == "Construction")
    cod_total = sum(float(x.get("Amount_COP", 0.0) or 0.0) for x in lines if x.get("Phase") == "At COD")

    w_dev = _weights(dev_n, s.capex.distribution)
    w_con = _weights(con_n, s.capex.distribution)

    rows = []
    for i, m in enumerate(dev_months):
        # Normalize to 1st of month to match opex_monthly_schedule format
        month_normalized = date(m.year, m.month, 1)
        weight = w_dev[i] if (len(w_dev) > i) else (1.0 / dev_n if dev_n > 0 else 0.0)
        capex_value = dev_total * weight
        rows.append({"Month": month_normalized, "Phase": "Development", "CAPEX (COP)": capex_value})
    
    for i, m in enumerate(con_months):
        # Normalize to 1st of month to match opex_monthly_schedule format
        month_normalized = date(m.year, m.month, 1)
        weight = w_con[i] if (len(w_con) > i) else (1.0 / con_n if con_n > 0 else 0.0)
        capex_value = con_total * weight
        rows.append({"Month": month_normalized, "Phase": "Construction", "CAPEX (COP)": capex_value})
    
    rows.append({"Month": date(cod.year, cod.month, 1), "Phase": "At COD", "CAPEX (COP)": cod_total})

    df = pd.DataFrame(rows).sort_values("Month").reset_index(drop=True)
    df["Year"] = df["Month"].apply(lambda d: d.year)
    return df


# -----------------------------
# OPEX monthly schedule (includes Revenue + Energy + CAPEX merge)
# -----------------------------
def _phase_for_month(tl: dict, m: date) -> str:
    if m < date(tl["rtb"].year, tl["rtb"].month, 1):
        return "Development"
    if m < date(tl["cod"].year, tl["cod"].month, 1):
        return "Construction"
    return "Operation"


def opex_monthly_schedule(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    start = tl["start"]
    end_op = tl["end_op"]

    months = []
    cur = date(start.year, start.month, 1)
    endm = date(end_op.year, end_op.month, 1)
    while cur < endm:
        months.append(cur)
        cur = _add_months(cur, 1)

    df = pd.DataFrame({"Month": months})
    df["Year"] = df["Month"].apply(lambda d: d.year)
    df["Phase"] = df["Month"].apply(lambda m: _phase_for_month(tl, m))

    op = operating_year_table(s)
    op_map_mwh = {int(r["Year"]): float(r["Energy (MWh)"]) for _, r in op.iterrows()}
    op_map_rev = {int(r["Year"]): float(r["Revenue (COP)"]) for _, r in op.iterrows()}
    
    # Calculate operating months per calendar year for proper proration
    cod = tl["cod"]
    operating_months_per_year = {}
    for year in op["Year"].unique():
        year_int = int(year)
        if year_int == cod.year:
            # First year: from COD month to end of year
            operating_months_per_year[year_int] = 13 - cod.month
        elif year_int == end_op.year:
            # Last year: from start of year to end_op month
            operating_months_per_year[year_int] = end_op.month
        else:
            # Full year
            operating_months_per_year[year_int] = 12

    def get_monthly_value(row, annual_map, months_map):
        if row["Phase"] != "Operation":
            return 0.0
        year = int(row["Year"])
        annual_val = annual_map.get(year, 0.0)
        months = months_map.get(year, 12)
        return annual_val / months if months > 0 else 0.0

    df["Energy (MWh)"] = df.apply(lambda r: get_monthly_value(r, op_map_mwh, operating_months_per_year), axis=1)
    df["Revenue (COP)"] = df.apply(lambda r: get_monthly_value(r, op_map_rev, operating_months_per_year), axis=1)

    cap = capex_monthly_schedule(s)[["Month", "CAPEX (COP)"]].copy()
    # Ensure Month column is normalized to 1st of month for both dataframes
    df["Month"] = df["Month"].apply(lambda d: date(d.year, d.month, 1) if isinstance(d, date) else d)
    cap["Month"] = cap["Month"].apply(lambda d: date(d.year, d.month, 1) if isinstance(d, date) else d)
    df = df.merge(cap, on="Month", how="left")
    df["CAPEX (COP)"] = df["CAPEX (COP)"].fillna(0.0)

    mwac = float(s.generation.mwac or 0.0)

    base_year = tl["cod"].year
    idx_fixed = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, s.opex.fixed_om_indexation))
    idx_ins = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, s.opex.insurance_indexation))
    idx_land = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, s.opex.land_indexation))

    df["Fixed O&M"] = 0.0
    df.loc[df["Phase"] == "Operation", "Fixed O&M"] = (float(s.opex.fixed_om_cop_per_mwac_year) * mwac / 12.0) * idx_fixed[df["Phase"] == "Operation"].values

    df["Insurance"] = 0.0
    df.loc[df["Phase"] == "Operation", "Insurance"] = (float(s.opex.insurance_cop_per_mwac_year) * mwac / 12.0) * idx_ins[df["Phase"] == "Operation"].values

    df["Variable O&M"] = float(s.opex.variable_om_cop_per_mwh) * df["Energy (MWh)"]
    df["Grid fees"] = float(s.opex.grid_fees_cop_per_mwh) * df["Energy (MWh)"]

    ha = float(s.opex.land_hectares or 0.0)
    df["Land lease"] = 0.0
    df.loc[df["Phase"] == "Development", "Land lease"] = (ha * float(s.opex.land_price_dev_cop_per_ha_year) / 12.0) * idx_land[df["Phase"] == "Development"].values
    df.loc[df["Phase"] == "Construction", "Land lease"] = (ha * float(s.opex.land_price_con_cop_per_ha_year) / 12.0) * idx_land[df["Phase"] == "Construction"].values
    df.loc[df["Phase"] == "Operation", "Land lease"] = (ha * float(s.opex.land_price_op_cop_per_ha_year) / 12.0) * idx_land[df["Phase"] == "Operation"].values

    other = pd.DataFrame(s.opex.other_items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in other.columns:
            other[col] = "" if col != "Amount_COP_per_year" else 0.0
    other["Phase"] = other["Phase"].where(other["Phase"].isin(PHASES), "Operation")
    other["Indexation"] = other["Indexation"].where(other["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    for _, r in other.iterrows():
        name = str(r.get("Item", "")).strip() or "Other"
        amt = float(r.get("Amount_COP_per_year", 0.0) or 0.0)
        ph = str(r.get("Phase", "Operation"))
        idx_choice = str(r.get("Indexation", "Colombia CPI"))
        if name not in df.columns:
            df[name] = 0.0
        idx_series = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, idx_choice))
        df.loc[df["Phase"] == ph, name] = (amt / 12.0) * idx_series[df["Phase"] == ph].values

    df["ICA"] = 0.0
    df.loc[df["Phase"] == "Operation", "ICA"] = (float(s.opex.ica_pct_of_revenue) / 100.0) * df.loc[df["Phase"] == "Operation", "Revenue (COP)"]

    meta = {"Month", "Year", "Phase", "Energy (MWh)", "Revenue (COP)", "CAPEX (COP)"}
    fixed_cols = ["Fixed O&M", "Insurance", "Variable O&M", "Grid fees", "Land lease", "ICA"]
    dyn_cols = [c for c in df.columns if c not in meta and c not in fixed_cols and c not in {"OPEX subtotal", "GMF"}]
    opex_cols = fixed_cols + dyn_cols
    df["OPEX subtotal"] = df[opex_cols].sum(axis=1) if opex_cols else 0.0

    # GMF on outgoing cash: CAPEX + OPEX subtotal
    gmf = float(s.opex.gmf_pct_of_outflows) / 100.0
    df["GMF"] = gmf * (df["CAPEX (COP)"].fillna(0.0) + df["OPEX subtotal"].fillna(0.0))

    return df


# -----------------------------
# SG&A schedules
# -----------------------------
def sga_monthly_schedule(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    start = tl["start"]
    end_op = tl["end_op"]

    months = []
    cur = date(start.year, start.month, 1)
    endm = date(end_op.year, end_op.month, 1)
    while cur < endm:
        months.append(cur)
        cur = _add_months(cur, 1)

    df = pd.DataFrame({"Month": months})
    df["Year"] = df["Month"].apply(lambda d: d.year)
    df["Phase"] = df["Month"].apply(lambda m: _phase_for_month(tl, m))

    items = pd.DataFrame(s.sga.items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in items.columns:
            items[col] = "" if col != "Amount_COP_per_year" else 0.0
    items["Phase"] = items["Phase"].where(items["Phase"].isin(PHASES), "Development")
    items["Indexation"] = items["Indexation"].where(items["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    base_year = tl["cod"].year
    for _, r in items.iterrows():
        name = str(r.get("Item", "")).strip() or "SGA"
        amt = float(r.get("Amount_COP_per_year", 0.0) or 0.0)
        ph = str(r.get("Phase", "Development"))
        idx_choice = str(r.get("Indexation", "Colombia CPI"))
        if name not in df.columns:
            df[name] = 0.0
        idx_series = df["Year"].apply(lambda y: index_factor_for_year(s.macro, int(y), base_year, idx_choice))
        df.loc[df["Phase"] == ph, name] = (amt / 12.0) * idx_series[df["Phase"] == ph].values

    return df.fillna(0.0)


def sga_annual_by_item(s: ScenarioInputs) -> pd.DataFrame:
    dfm = sga_monthly_schedule(s).copy()
    meta = {"Month", "Year", "Phase"}
    item_cols = [c for c in dfm.columns if c not in meta]
    annual = dfm.groupby("Year", as_index=False)[item_cols].sum() if item_cols else dfm.groupby("Year", as_index=False).size()[["Year"]]
    annual["Total SG&A (COP)"] = annual[item_cols].sum(axis=1) if item_cols else 0.0
    return annual


# -----------------------------
# Depreciation (annual)
# -----------------------------
def depreciation_annual_table(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    cod_year = cod.year

    cap_df = pd.DataFrame(s.capex.lines or [])
    total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0

    pct = float(s.depreciation.pct_of_capex_depreciated or 0.0) / 100.0
    dep_years = int(s.depreciation.dep_years)
    dep_months = dep_years * 12  # Total months of depreciation

    dep_base = total_capex * pct
    monthly_dep = dep_base / dep_months if dep_months > 0 else 0.0
    annual_dep = monthly_dep * 12.0  # Full year equivalent

    # Calculate end date of depreciation period
    dep_end = _add_months(cod, dep_months)
    
    # Get all calendar years from COD to end of depreciation
    years = list(range(cod_year, dep_end.year + 1))
    
    # Calculate prorated depreciation for each calendar year
    dep_list = []
    cumulative_dep = 0.0
    
    for year in years:
        # Determine start and end dates for this calendar year
        if year == cod_year:
            # First year: from COD date to end of year
            year_start = cod
            year_end = date(year, 12, 31)
        elif year == dep_end.year:
            # Last year: from start of year to dep_end date
            year_start = date(year, 1, 1)
            year_end = dep_end
        else:
            # Full year
            year_start = date(year, 1, 1)
            year_end = date(year, 12, 31)
        
        # Calculate months of depreciation in this calendar year
        months_in_year = (year_end.year - year_start.year) * 12 + (year_end.month - year_start.month) + 1
        
        # Calculate depreciation for this year
        year_dep = monthly_dep * months_in_year
        dep_list.append(year_dep)
        cumulative_dep += year_dep
    
    df = pd.DataFrame({
        "Year": years,
        "Depreciable CAPEX (COP)": [dep_base] * len(years),
        "Depreciation (COP)": dep_list
    })
    df["Cumulative Depreciation (COP)"] = df["Depreciation (COP)"].cumsum()
    df["Remaining Book Value (COP)"] = np.maximum(dep_base - df["Cumulative Depreciation (COP)"], 0.0)
    return df

# -----------------------------
# CAPEX deduction for renewable incentives
# -----------------------------

def renewable_tax_annual_table(s: ScenarioInputs) -> pd.DataFrame:
    tl = build_timeline(s.timeline)
    cod_year = tl["cod"].year

    cf = cashflow_annual_table(s).copy()

    # Base taxable income BEFORE incentives
    cf["Taxable Income (Pre-Incentives)"] = (
        cf["Operating CF (COP)"]
        - cf.get("Depreciation (COP)", 0.0)
    )

    # -------------------------
    # Special CAPEX deduction
    # -------------------------
    eligible_capex = _eligible_capex_for_tax(s)
    deduction_pool = eligible_capex
    max_years = int(s.renewable_tax.special_deduction_years)

    deductions = []
    remaining = deduction_pool

    for _, r in cf.iterrows():
        year = int(r["Year"])
        ti = max(0.0, float(r["Taxable Income (Pre-Incentives)"]))

        if (
            not s.renewable_tax.enable_special_deduction
            or remaining <= 0
            or year < cod_year
            or year >= cod_year + max_years
        ):
            used = 0.0
        else:
            used = min(remaining, 0.5 * ti)

        remaining -= used
        deductions.append(used)

    cf["Special Deduction Used (COP)"] = deductions
    cf["Remaining Deduction Balance (COP)"] = (
        deduction_pool - pd.Series(deductions).cumsum()
    ).clip(lower=0.0)

    # -------------------------
    # Final taxable income
    # -------------------------
    cf["Taxable Income (After Incentives)"] = (
        cf["Taxable Income (Pre-Incentives)"]
        - cf["Special Deduction Used (COP)"]
    ).clip(lower=0.0)

    tax_rate = float(s.tax.corporate_tax_rate_pct) / 100.0
    cf["Income Tax (COP)"] = cf["Taxable Income (After Incentives)"] * tax_rate

    # -------------------------
    # VAT refund (cash inflow)
    # -------------------------
    cf["VAT Refund (COP)"] = 0.0

    if s.renewable_tax.enable_vat_refund:
        if s.renewable_tax.vat_refund_mode == "percent":
            vat_amt = (
                _total_capex_from_lines(s)
                * float(s.renewable_tax.vat_pct_of_capex)
                / 100.0
            )
        else:
            vat_amt = float(s.renewable_tax.vat_fixed_cop)

        refund_year = cod_year + int(s.renewable_tax.vat_refund_year)
        cf.loc[cf["Year"] == refund_year, "VAT Refund (COP)"] = vat_amt

    return cf


# -----------------------------
# CASH FLOW (monthly + annual)
# -----------------------------
def _sga_monthly_total(s: ScenarioInputs) -> pd.DataFrame:
    df = sga_monthly_schedule(s).copy()
    meta = {"Month", "Year", "Phase"}
    item_cols = [c for c in df.columns if c not in meta]
    df["SG&A (COP)"] = df[item_cols].sum(axis=1) if item_cols else 0.0
    return df[["Month", "Year", "Phase", "SG&A (COP)"]]

def _shift_by_months(values: pd.Series, months: int) -> pd.Series:
    """Shift a monthly series forward by N months (cash received/paid later).
    Missing months are filled with 0.0."""
    months = int(max(0, months))
    if months == 0:
        return values.astype(float)
    return values.shift(months, fill_value=0.0).astype(float)

def cashflow_monthly_table(s: ScenarioInputs) -> pd.DataFrame:
    base = opex_monthly_schedule(s).copy()

    sga_m = _sga_monthly_total(s).copy()
    base = base.merge(sga_m[["Month", "SG&A (COP)"]], on="Month", how="left")
    base["SG&A (COP)"] = base["SG&A (COP)"].fillna(0.0)

    base["Total OPEX (COP)"] = base["OPEX subtotal"].fillna(0.0) + base["GMF"].fillna(0.0)

    # --- Accrual (non-cash timing) operating CF + unlevered CF (this is what debt sizing will use initially) ---
    base["Operating CF (COP)"] = (
        base["Revenue (COP)"].fillna(0.0)
        - base["Total OPEX (COP)"].fillna(0.0)
        - base["SG&A (COP)"].fillna(0.0)
    )

    base["Unlevered CF (COP)"] = base["Operating CF (COP)"] - base["CAPEX (COP)"].fillna(0.0)
    base["Cumulative Unlevered CF (COP)"] = base["Unlevered CF (COP)"].cumsum()


    # -----------------------------
    # Working Capital (simple AR/AP lags) + Cash CF
    # -----------------------------
    # You said you use s.wc (not s.working_capital)
    ar_days = float(getattr(s.wc, "ar_days", 90.0))   # revenue collected in ~90 days
    ap_days = float(getattr(s.wc, "ap_days", 60.0))   # expenses paid in ~60 days

    # Convert days to whole months (simple approximation)
    ar_lag_m = int(round(ar_days / 30.0))
    ap_lag_m = int(round(ap_days / 30.0))

    # What "cash paid expenses" means here: OPEX + SG&A (exclude CAPEX)
    base["Total OpEx+SGA (COP)"] = base["Total OPEX (COP)"].fillna(0.0) + base["SG&A (COP)"].fillna(0.0)

    # Cash timing via shifts (simple, fast, stable)
    base["Cash Collected (COP)"]   = base["Revenue (COP)"].fillna(0.0).shift(ar_lag_m, fill_value=0.0)
    base["Cash Paid OPEX (COP)"]   = base["Total OPEX (COP)"].fillna(0.0).shift(ap_lag_m, fill_value=0.0)
    base["Cash Paid SG&A (COP)"]   = base["SG&A (COP)"].fillna(0.0).shift(ap_lag_m, fill_value=0.0)

    # AR/AP balances implied by the lag model
    # AR(t) = AR(t-1) + Revenue(t) - CashCollected(t)
    # AP(t) = AP(t-1) + Expenses(t) - CashPaid(t)
    ar = []
    ap = []
    ar_prev = 0.0
    ap_prev = 0.0

    rev = base["Revenue (COP)"].fillna(0.0).to_numpy()
    cash_in = base["Cash Collected (COP)"].fillna(0.0).to_numpy()

    opex = base["Total OPEX (COP)"].fillna(0.0).to_numpy()
    sga  = base["SG&A (COP)"].fillna(0.0).to_numpy()
    cash_opex = base["Cash Paid OPEX (COP)"].fillna(0.0).to_numpy()
    cash_sga  = base["Cash Paid SG&A (COP)"].fillna(0.0).to_numpy()

    for i in range(len(base)):
        ar_now = ar_prev + rev[i] - cash_in[i]
        ap_now = ap_prev + (opex[i] + sga[i]) - (cash_opex[i] + cash_sga[i])
        ar.append(ar_now)
        ap.append(ap_now)
        ar_prev, ap_prev = ar_now, ap_now

    base["AR Balance (COP)"] = ar
    base["AP Balance (COP)"] = ap

    base["Net Working Capital (COP)"] = base["AR Balance (COP)"] - base["AP Balance (COP)"]
    base["ΔNWC (COP)"] = base["Net Working Capital (COP)"].diff().fillna(base["Net Working Capital (COP)"])

    # Cash operating CF and cash unlevered CF
    base["Operating CF (Cash, COP)"] = (
        base["Cash Collected (COP)"].fillna(0.0)
        - base["Cash Paid OPEX (COP)"].fillna(0.0)
        - base["Cash Paid SG&A (COP)"].fillna(0.0)
    )

    # CAPEX assumed paid when incurred (no lag)
    base["Unlevered CF (Cash, COP)"] = base["Operating CF (Cash, COP)"] - base["CAPEX (COP)"].fillna(0.0)
    base["Cumulative Unlevered CF (Cash, COP)"] = base["Unlevered CF (Cash, COP)"].cumsum()

    # -----------------------------
    # Debt fees (cash) — only when there is debt
    # -----------------------------
    base["Debt Fees (COP)"] = 0.0

    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * _total_capex_from_lines(s)
    debt_amt = max(debt_amt, 0.0)

    if debt_amt > 0:
        # 1) Commitment fee (annual -> spread monthly)
        cf_ann = debt_commitment_fee_annual(s).copy()  # expects columns: Year, Commitment Fee (COP)
        if (not cf_ann.empty) and ("Year" in cf_ann.columns) and ("Commitment Fee (COP)" in cf_ann.columns):
            fee_map = {int(r["Year"]): float(r["Commitment Fee (COP)"]) for _, r in cf_ann.iterrows()}
            base["Debt Fees (COP)"] += base["Year"].map(lambda y: fee_map.get(int(y), 0.0) / 12.0)

        # 2) Upfront fee at COD month (if you have upfront_fee_pct in s.debt)
        tl = build_timeline(s.timeline)
        cod_m = date(tl["cod"].year, tl["cod"].month, 1)
        upfront_pct = float(getattr(s.debt, "upfront_fee_pct", 0.0)) / 100.0
        if upfront_pct > 0:
            base.loc[base["Month"] == cod_m, "Debt Fees (COP)"] += debt_amt * upfront_pct

    # Apply fees to CASH unlevered CF (equity cash)
    base["Unlevered CF (Cash, COP)"] = base["Unlevered CF (Cash, COP)"].fillna(0.0) - base["Debt Fees (COP)"].fillna(0.0)
    base["Cumulative Unlevered CF (Cash, COP)"] = base["Unlevered CF (Cash, COP)"].cumsum()


    cols = [
        "Month", "Year", "Phase",
        "Energy (MWh)", "Revenue (COP)", "Cash Collected (COP)",
        "CAPEX (COP)", "Debt Fees (COP)",
        "Total OPEX (COP)", "Cash Paid OPEX (COP)",
        "SG&A (COP)", "Cash Paid SG&A (COP)",
        "Operating CF (COP)", "Operating CF (Cash, COP)",
        "AR Balance (COP)", "AP Balance (COP)", "Net Working Capital (COP)", "ΔNWC (COP)",
        "Unlevered CF (COP)", "Unlevered CF (Cash, COP)",
        "Cumulative Unlevered CF (COP)", "Cumulative Unlevered CF (Cash, COP)",
    ]

    # IMPORTANT: prevents KeyError if some columns are not created yet
    cols = [c for c in cols if c in base.columns]

    return base.loc[:, cols].copy()


def cashflow_annual_table(s: ScenarioInputs) -> pd.DataFrame:
    m = cashflow_monthly_table(s).copy()
    annual = (
        m.groupby("Year", as_index=False)[
            ["Energy (MWh)", "Revenue (COP)", "CAPEX (COP)", "Total OPEX (COP)", "SG&A (COP)", "Operating CF (COP)", "Unlevered CF (COP)"]
        ]
        .sum()
    )
    annual["Cumulative Unlevered CF (COP)"] = annual["Unlevered CF (COP)"].cumsum()
    return annual


def unlevered_base_cashflow_annual(s: ScenarioInputs) -> pd.DataFrame:
    a = cashflow_annual_table(s).copy()
    dep = depreciation_annual_table(s)[["Year", "Depreciation (COP)"]].copy()
    out = a.merge(dep, on="Year", how="left").fillna({"Depreciation (COP)": 0.0})
    
    # Ensure CAPEX is preserved in the output (it should already be in 'a' from cashflow_annual_table)
    # The "Unlevered CF (COP)" column already includes CAPEX: Operating CF - CAPEX
    # Verify CAPEX column exists, if not add it from the annual table
    if "CAPEX (COP)" not in out.columns and "CAPEX (COP)" in a.columns:
        out["CAPEX (COP)"] = a["CAPEX (COP)"]

    out["EBITDA (COP)"] = out["Operating CF (COP)"]
    out["Taxable Income (COP)"] = out["EBITDA (COP)"] - out["Depreciation (COP)"]

    rate = float(s.tax.corporate_tax_rate_pct) / 100.0
    allow_nol = bool(s.tax.allow_loss_carryforward)

    # ---- Renewable incentive (Law 1715 / 2099): special deduction ----
    # Pool = 50% of total investment (CAPEX). Usable up to 15 years starting year after COD.
    # Annual cap: deduction <= 50% of renta líquida (before this deduction).

    tl = build_timeline(s.timeline)
    cod_year = int(tl["cod"].year)

    total_capex = float(_total_capex_from_lines(s))  # you already have this helper for debt
    special_pool_total = 0.50 * total_capex          # 50% of total investment
    special_pool_remaining = special_pool_total

    special_years = 15
    special_start_year = cod_year + 1
    special_end_year = special_start_year + special_years - 1

    # ---- Incentives + tax calc ----
    inc = getattr(s, "incentives", RenewableIncentivesInputs())

    # Total eligible CAPEX (same base you already use for depreciation)
    cap_df = pd.DataFrame(s.capex.lines or [])
    total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0

    # Special deduction pool
    pool_total = (float(inc.special_deduction_pct_of_capex) / 100.0) * total_capex if bool(inc.enable_special_deduction) else 0.0
    pool_remaining = pool_total

    # VAT refund (cash only)
    vat_amount = (float(inc.vat_pct_of_capex) / 100.0) * total_capex + float(inc.vat_fixed_cop)
    vat_amount = max(0.0, vat_amount) if inc.vat_mode == "Refund" else 0.0

    nol = 0.0
    taxes = []
    nol_end = []

    ded_used_list = []
    ded_rem_list = []
    taxable_after_all_list = []
    vat_refund_list = []

    # Operating-year index so "refund in op Year #1" maps correctly
    # Operating year 1 = COD calendar year, operating year 2 = COD+1, etc.
    # Years before COD are not operating years (op_year = 0)
    # This calculation works regardless of when COD occurs (early/late in year, any calendar year)
    max_years = int(max(1, inc.special_deduction_years))

    for _, r in out.iterrows():
        year_i = int(r["Year"])
        # Calculate operating year: 1 for COD year, 2 for COD+1, etc. 0 for pre-COD years
        # This works for any COD date (e.g., COD in Jan 2029 = op_year 1 for 2029, op_year 2 for 2030)
        op_year = (year_i - cod_year + 1) if year_i >= cod_year else 0

        # Base taxable income = EBITDA - Depreciation
        ti_base = float(r["Taxable Income (COP)"])

        # Apply CAPEX deduction FIRST (to minimize taxes)
        # The deduction applies in the FIRST year with positive taxable income (regardless of timeline)
        ded_used = 0.0
        if (
            bool(inc.enable_special_deduction) 
            and (1 <= op_year <= max_years)  # Only during operating years 1 through max_years
            and ti_base > 0  # Apply deduction if there's any taxable income (before NOL)
            and pool_remaining > 0  # Only if pool hasn't been exhausted
        ):
            # Annual cap is based on taxable income BEFORE NOL (as per tax law)
            # The cap is a percentage of taxable income, ensuring we don't exceed the annual limit
            annual_cap = (float(inc.special_deduction_max_pct_of_taxable_income) / 100.0) * ti_base
            ded_used = min(pool_remaining, annual_cap)  # Use the smaller of: remaining pool or annual cap
            pool_remaining -= ded_used

        # Taxable income after CAPEX deduction: EBITDA - Depreciation - CAPEX Tax Deduction
        ti_after_ded = ti_base - ded_used

        # Loss carryforward: accumulate losses (negative taxable income after deduction)
        # If negative, add to NOL. If positive, use NOL to offset it.
        if allow_nol:
            if ti_after_ded < 0:
                # Loss: add to NOL carryforward
                nol = nol + (-ti_after_ded)
                taxable_after_all = 0.0  # No taxable income when there's a loss
            else:
                # Profit: use NOL to offset taxable income
                used = min(nol, ti_after_ded)
                nol -= used
                taxable_after_all = max(ti_after_ded - used, 0.0)
        else:
            # No NOL: taxable income is what's left after deduction (can be negative)
            taxable_after_all = max(ti_after_ded, 0.0)
            if ti_after_ded < 0:
                nol = nol + (-ti_after_ded)  # Still track losses even if NOL not allowed (for display)

        taxes_payable = taxable_after_all * rate
        taxes.append(taxes_payable)
        nol_end.append(nol)

        ded_used_list.append(ded_used)
        ded_rem_list.append(pool_remaining)
        taxable_after_all_list.append(taxable_after_all)

        # VAT refund: one-time cash inflow in operating year N (default year 1)
        vat_refund = vat_amount if (vat_amount > 0 and op_year == int(inc.vat_refund_year_index)) else 0.0
        vat_refund_list.append(vat_refund)

    out["CAPEX Tax Deduction (COP)"] = ded_used_list
    out["Loss Carryforward End (COP)"] = nol_end  # Moved before Taxable Income
    out["Special Deduction Remaining (COP)"] = ded_rem_list
    # Taxable Income is what remains after: EBITDA - Depreciation - CAPEX Deduction - NOL
    out["Taxable Income (COP)"] = taxable_after_all_list
    out["Taxable Income After Incentives (COP)"] = taxable_after_all_list

    out["Taxes Payable (COP)"] = taxes

    # After-tax CF: taxes reduce cash; VAT refund adds cash (does not affect taxes)
    out["VAT Refund (COP)"] = vat_refund_list
    # Ensure "Unlevered CF (COP)" includes CAPEX (it should from cashflow_annual_table)
    # If for some reason CAPEX is missing, recalculate: Operating CF - CAPEX
    if "CAPEX (COP)" in out.columns:
        # Verify Unlevered CF includes CAPEX: it should be Operating CF - CAPEX
        out["Unlevered CF (COP)"] = out["Operating CF (COP)"] - out["CAPEX (COP)"].fillna(0.0)
    # Rename pre-tax column for clarity (this already includes CAPEX from the calculation above)
    out["Unlevered CF Pre-tax (COP)"] = out["Unlevered CF (COP)"]
    out["Unlevered CF After Tax (COP)"] = out["Unlevered CF Pre-tax (COP)"] - out["Taxes Payable (COP)"] + out["VAT Refund (COP)"]

    return out


# -----------------------------
# IRR / Payback helpers
# -----------------------------
def _irr_bisection(cashflows: List[float], low=-0.95, high=5.0, tol=1e-7, max_iter=200) -> float:
    def npv(rate: float) -> float:
        denom_base = 1.0 + rate
        if denom_base <= 1e-12:
            return float("inf")
        total = 0.0
        for i, cf in enumerate(cashflows):
            if i == 0:
                total += cf
                continue
            den = denom_base ** i
            if den == 0.0:
                return float("inf")
            total += cf / den
        return total

    if len(cashflows) < 2:
        return float("nan")

    f_low = npv(low)
    f_high = npv(high)

    tries = 0
    while f_low * f_high > 0 and tries < 20:
        high *= 2
        f_high = npv(high)
        tries += 1

    if f_low * f_high > 0:
        return float("nan")

    mid = float("nan")
    for _ in range(max_iter):
        mid = (low + high) / 2
        f_mid = npv(mid)
        if abs(f_mid) < tol:
            return mid
        if f_low * f_mid <= 0:
            high = mid
            f_high = f_mid
        else:
            low = mid
            f_low = f_mid
    return mid


def _payback_months(months: List[date], unlevered_cf: List[float]) -> float:
    """
    Calculate payback period in months using cumulative cash flows.
    Payback is the time (in months) from the start until cumulative CF becomes positive.
    Uses linear interpolation in the period where cumulative crosses zero.
    
    Args:
        months: List of dates (not used in calculation, but kept for consistency)
        unlevered_cf: List of monthly cash flows (should be after-tax for after-tax payback)
    
    Returns:
        Payback period in months (as float, can be fractional), or NaN if never pays back
    """
    cum = 0.0
    for i, cf in enumerate(unlevered_cf):
        prev = cum  # Cumulative at end of previous period
        cum += cf   # Cumulative at end of current period
        if cum >= 0 and i > 0:  # Cumulative has crossed zero
            if cf == 0:
                # Edge case: zero cash flow, payback is at end of this period
                return float(i)
            # Linear interpolation: what fraction of current period's CF is needed
            # to go from prev (negative) to 0?
            # frac = amount needed / cash flow in period = (0 - prev) / cf
            frac = (0 - prev) / cf
            # Payback = (i-1) full periods + fraction of current period
            return (i - 1) + frac
    return float("nan")  # Never pays back


# -----------------------------
# Debt engine (annual sculpting + fees)
# -----------------------------
def _total_capex_from_lines(s: ScenarioInputs) -> float:
    cap_df = pd.DataFrame(s.capex.lines or [])
    if cap_df.empty or "Amount_COP" not in cap_df.columns:
        return 0.0
    return float(pd.to_numeric(cap_df["Amount_COP"], errors="coerce").fillna(0.0).sum())

def _eligible_capex_for_tax(s: ScenarioInputs) -> float:
    total_capex = _total_capex_from_lines(s)
    pct = max(0.0, min(100.0, s.renewable_tax.special_deduction_pct_of_capex))
    return total_capex * pct / 100.0

def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)

def debt_commitment_fee_annual(s: ScenarioInputs) -> pd.DataFrame:
    """
    Commitment fee on undrawn debt during construction:
    Commitment rate = margin * commitment_fee_pct_of_margin
    Applied monthly on undrawn = TotalDebt - CumulativeDrawn
    """
    tl = build_timeline(s.timeline)
    cod_m = date(tl["cod"].year, tl["cod"].month, 1)

    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex
    debt_amt = max(debt_amt, 0.0)

    # Monthly CAPEX schedule up to COD month (exclude COD month)
    capm = capex_monthly_schedule(s).copy()
    capm["Month"] = pd.to_datetime(capm["Month"])
    capm = capm[capm["Month"] < pd.to_datetime(cod_m)].copy()
    if capm.empty or debt_amt <= 0:
        return pd.DataFrame({"Year": [], "Commitment Fee (COP)": []})

    # Assume debt draws pro-rata with CAPEX spend during dev+construction (pre-COD)
    draw_m = capm.copy()
    draw_m["Draw (COP)"] = (float(s.debt.debt_pct_of_capex) / 100.0) * pd.to_numeric(draw_m["CAPEX (COP)"], errors="coerce").fillna(0.0)

    draw_m["Cum Draw (COP)"] = draw_m["Draw (COP)"].cumsum()
    draw_m["Undrawn (COP)"] = np.maximum(debt_amt - draw_m["Cum Draw (COP)"], 0.0)

    margin = float(s.debt.margin_pct) / 100.0
    commit_mult = float(s.debt.commitment_fee_pct_of_margin) / 100.0
    commit_rate_annual = margin * commit_mult
    commit_rate_monthly = commit_rate_annual / 12.0

    draw_m["Commitment Fee (COP)"] = draw_m["Undrawn (COP)"] * commit_rate_monthly
    draw_m["Year"] = draw_m["Month"].dt.year

    ann = draw_m.groupby("Year", as_index=False)[["Commitment Fee (COP)"]].sum()
    return ann


def debt_schedule_annual(s: ScenarioInputs) -> pd.DataFrame:
    """
    Annual debt schedule starting at COD year for tenor years.
    Amortization:
      - Sculpted to Target DSCR,
      - Equal principal, or
      - Typical amortization (fixed payment/annuity)
    DSCR calculated as: Operating CF / Debt Service
    """
    tl = build_timeline(s.timeline)
    cod_year = tl["cod"].year

    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex
    debt_amt = max(debt_amt, 0.0)

    tenor = int(s.debt.tenor_years)
    tenor = max(5, min(10, tenor))
    grace = int(s.debt.grace_years)
    grace = max(0, min(tenor - 1, grace))

    years = list(range(cod_year, cod_year + tenor))

    # Operating CF annual from model (EBITDA proxy)
    a = cashflow_annual_table(s).copy()
    opcf_map = {int(r["Year"]): float(r["Operating CF (COP)"]) for _, r in a.iterrows()}
    cf = [float(opcf_map.get(y, 0.0)) for y in years]

    all_in = (float(s.debt.base_rate_pct) + float(s.debt.margin_pct)) / 100.0
    target = float(s.debt.target_dscr) if float(s.debt.target_dscr) > 0 else 1.20

    # NEW: balloon target (COP). % of original debt you want left outstanding at maturity.
    balloon_pct = float(getattr(s.debt, "balloon_pct", 0.0)) / 100.0
    balloon_target = debt_amt * balloon_pct
    balloon_target = max(0.0, min(balloon_target, debt_amt))

    outstanding = debt_amt
    rows = []
    
    # Pre-calculate fixed payment for typical amortization (if applicable)
    fixed_payment_typical = 0.0
    if s.debt.amortization_type == "Typical amortization":
        # Calculate fixed payment to amortize (debt_amt - balloon_target) over (tenor - grace) years
        principal_to_amortize = debt_amt - balloon_target
        amortization_years = tenor - grace
        if amortization_years > 0 and all_in > 0 and principal_to_amortize > 1e-9:
            r = all_in
            n = amortization_years
            if (1.0 + r) ** n > 1.0:
                fixed_payment_typical = principal_to_amortize * (r * (1.0 + r) ** n) / ((1.0 + r) ** n - 1.0)
            else:
                fixed_payment_typical = principal_to_amortize / n

    for i, y in enumerate(years, start=1):
        operating_cf = float(cf[i - 1])
        interest_due = outstanding * all_in  # Interest due on current outstanding

        if outstanding <= 1e-9:
            principal = 0.0
            debt_service = 0.0
            dscr = float("nan")
            rows.append({
                "Year": y,
                "Operating CF (COP)": operating_cf,
                "Interest (COP)": 0.0,
                "Principal (COP)": principal,
                "Debt Service (COP)": debt_service,
                "DSCR": dscr,
                "Outstanding End (COP)": 0.0,
            })
            outstanding = 0.0
            continue

        if i <= grace:
            principal = 0.0
            # During grace period:
            # - No principal payments
            # - Interest is capitalized (added to principal) if cash flow is insufficient
            # - Interest is only paid in cash if there's sufficient operating cash flow
            if operating_cf >= interest_due:
                # Sufficient cash flow: pay all interest in cash
                interest_paid = interest_due
                interest_capitalized = 0.0
            else:
                # Insufficient cash flow: pay what we can, capitalize the rest
                interest_paid = max(0.0, operating_cf)  # Pay what we can from operating CF (can't be negative)
                interest_capitalized = interest_due - interest_paid  # Capitalize the rest
                outstanding = outstanding + interest_capitalized  # Add capitalized interest to principal
            
            # Interest paid in cash (for levered CF calculation)
            interest = interest_paid
        else:
            # After grace period: pay full interest in cash
            interest = interest_due
            
            # Calculate remaining amortization years (tenor minus grace, minus years already amortized)
            # Years already amortized = (i - 1 - grace) since grace years don't count
            years_amortized = max(0, i - 1 - grace)
            remaining_years = max(tenor - grace - years_amortized, 1)

            if s.debt.amortization_type == "Equal principal":
                # Amortize evenly down to balloon_target over remaining years
                principal = max(0.0, (outstanding - balloon_target) / remaining_years)

                # Final year: force ending outstanding = balloon_target
                if i == tenor:
                    principal = max(0.0, outstanding - balloon_target)

            elif s.debt.amortization_type == "Typical amortization":
                # Fixed payment (annuity) amortization
                # Use pre-calculated fixed payment, but only during amortization period (after grace)
                if fixed_payment_typical > 0:
                    # Payment = Interest + Principal, so Principal = Payment - Interest
                    principal = max(0.0, fixed_payment_typical - interest)
                    
                    # Ensure we don't pay more than outstanding (minus balloon)
                    principal = min(principal, max(0.0, outstanding - balloon_target))
                else:
                    principal = 0.0
                
                # Final year: force ending outstanding = balloon_target
                if i == tenor:
                    principal = max(0.0, outstanding - balloon_target)

            else:
                # Sculpt to target DSCR (CFADS proxy = Operating CF)
                max_ds = operating_cf / target if target > 0 else 0.0
                principal = max(0.0, max_ds - interest)

                # Constrain: don't pay more than what would amortize evenly over remaining years
                # This ensures we respect the tenor (don't pay off too early)
                max_principal_by_tenor = max(0.0, (outstanding - balloon_target) / remaining_years) if remaining_years > 0 else outstanding - balloon_target
                principal = min(principal, max(0.0, max_principal_by_tenor))

                # Do not amortize below balloon target (unless final year)
                if i < tenor:
                    principal = min(principal, max(0.0, outstanding - balloon_target))
                else:
                    principal = max(0.0, outstanding - balloon_target)

        principal = min(principal, outstanding)
        debt_service = interest + principal
        dscr = (operating_cf / debt_service) if debt_service > 0 else float("nan")

        outstanding = outstanding - principal

        rows.append({
            "Year": y,
            "Operating CF (COP)": operating_cf,
            "Interest (COP)": interest,
            "Principal (COP)": principal,
            "Debt Service (COP)": debt_service,
            "DSCR": dscr,
            "Outstanding End (COP)": outstanding,
        })

    df = pd.DataFrame(rows)

    balloon = float(df["Outstanding End (COP)"].iloc[-1]) if len(df) else 0.0
    df["Balloon at Maturity (COP)"] = 0.0
    if len(df):
        df.loc[df.index[-1], "Balloon at Maturity (COP)"] = balloon

    # Upfront fee (paid at COD, show in first year)
    upfront_fee = debt_amt * (float(s.debt.upfront_fee_bps) / 10_000.0)
    df["Upfront Fee (COP)"] = 0.0
    if len(df):
        df.loc[df.index[0], "Upfront Fee (COP)"] = upfront_fee

    # Covenant / lock-up flags
    min_covenant = float(s.debt.min_dscr_covenant)
    lockup = float(s.debt.lockup_dscr)
    df["Lock-up? (DSCR < lock-up)"] = df["DSCR"].apply(lambda x: bool(np.isfinite(x) and x < lockup))
    df["Breach? (DSCR < min)"] = df["DSCR"].apply(lambda x: bool(np.isfinite(x) and x < min_covenant))

    return df


# -----------------------------
# Levered Cash Flow (Equity Cash Flow)
# -----------------------------
def _recalculate_taxes_with_interest(s: ScenarioInputs, unlevered_df: pd.DataFrame, interest_by_year: dict) -> pd.DataFrame:
    """
    Recalculate tax benefits and taxes for levered cash flow.
    Taxable income = EBITDA - Depreciation - Interest (interest is tax-deductible).
    Then recalculate CAPEX deduction, loss carryforward, and taxes payable.
    """
    out = unlevered_df.copy()
    
    rate = float(s.tax.corporate_tax_rate_pct) / 100.0
    allow_nol = bool(s.tax.allow_loss_carryforward)
    
    # Get incentives
    inc = getattr(s, "incentives", RenewableIncentivesInputs())
    tl = build_timeline(s.timeline)
    cod_year = int(tl["cod"].year)
    
    # Total eligible CAPEX for deduction pool
    cap_df = pd.DataFrame(s.capex.lines or [])
    total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0
    
    # Special deduction pool
    pool_total = (float(inc.special_deduction_pct_of_capex) / 100.0) * total_capex if bool(inc.enable_special_deduction) else 0.0
    pool_remaining = pool_total
    
    # VAT refund (same as unlevered)
    vat_amount = (float(inc.vat_pct_of_capex) / 100.0) * total_capex + float(inc.vat_fixed_cop)
    vat_amount = max(0.0, vat_amount) if inc.vat_mode == "Refund" else 0.0
    
    nol = 0.0
    taxes = []
    nol_end = []
    ded_used_list = []
    taxable_after_all_list = []
    
    max_years = int(max(1, inc.special_deduction_years))
    
    for _, r in out.iterrows():
        year_i = int(r["Year"])
        op_year = (year_i - cod_year + 1) if year_i >= cod_year else 0
        
        # Levered taxable income = EBITDA - Depreciation - Interest
        ebitda = float(r.get("EBITDA (COP)", 0.0))
        depreciation = float(r.get("Depreciation (COP)", 0.0))
        interest = float(interest_by_year.get(year_i, 0.0))
        
        ti_base = ebitda - depreciation - interest  # Interest reduces taxable income
        
        # Apply CAPEX deduction FIRST (to minimize taxes)
        ded_used = 0.0
        if (
            bool(inc.enable_special_deduction) 
            and (1 <= op_year <= max_years) 
            and ti_base > 0  # Apply deduction if there's any taxable income (before NOL)
            and pool_remaining > 0
        ):
            annual_cap = (float(inc.special_deduction_max_pct_of_taxable_income) / 100.0) * ti_base
            ded_used = min(pool_remaining, annual_cap)
            pool_remaining -= ded_used
        
        # Taxable income after CAPEX deduction
        ti_after_ded = ti_base - ded_used
        
        # Apply NOL on remaining taxable income (after deduction)
        if allow_nol:
            if ti_after_ded < 0:
                nol = nol + (-ti_after_ded)
                taxable_after_all = 0.0
            else:
                used = min(nol, ti_after_ded)
                nol -= used
                taxable_after_all = max(ti_after_ded - used, 0.0)
        else:
            taxable_after_all = max(ti_after_ded, 0.0)
            if ti_after_ded < 0:
                nol = nol + (-ti_after_ded)  # Still track losses even if NOL not allowed
        
        taxes_payable = taxable_after_all * rate
        taxes.append(taxes_payable)
        nol_end.append(nol)
        ded_used_list.append(ded_used)
        taxable_after_all_list.append(taxable_after_all)
    
    # Update columns with recalculated values
    out["Levered Taxable Income (COP)"] = taxable_after_all_list
    out["Levered CAPEX Tax Deduction (COP)"] = ded_used_list
    out["Levered Loss Carryforward End (COP)"] = nol_end
    out["Levered Taxes Payable (COP)"] = taxes
    
    # Calculate Net Income After Tax = Taxable Income - Taxes Payable
    # (Taxable Income is already after all deductions and NOL)
    out["Levered Net Income After Tax (COP)"] = out["Levered Taxable Income (COP)"] - out["Levered Taxes Payable (COP)"]
    
    return out


def levered_cashflow_annual(s: ScenarioInputs) -> pd.DataFrame:
    """
    Calculate levered (equity) cash flow after-tax.
    Recalculates taxes with interest expense included (interest is tax-deductible).
    If debt is disabled or not enabled, assumes no debt (returns unlevered CF).
    """
    # Start with unlevered base cash flow (pre-tax)
    unlevered = unlevered_base_cashflow_annual(s).copy()
    
    # Check if debt is enabled
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex if debt_enabled else 0.0
    debt_amt = max(debt_amt, 0.0)
    
    # Initialize debt-related columns
    unlevered["Debt Draw (COP)"] = 0.0
    unlevered["Interest (COP)"] = 0.0
    unlevered["Principal (COP)"] = 0.0
    unlevered["Debt Service (COP)"] = 0.0
    unlevered["Debt Fees (COP)"] = 0.0
    
    if debt_enabled and debt_amt > 0:
        # Get debt schedule (annual interest, principal, upfront fee)
        ds = debt_schedule_annual(s).copy()
        
        if not ds.empty and "Year" in ds.columns:
            # Merge debt schedule - drop initialized columns first to avoid suffix conflicts
            merge_cols = ["Year"]
            for col in ["Interest (COP)", "Principal (COP)", "Upfront Fee (COP)"]:
                if col in ds.columns:
                    merge_cols.append(col)
                    # Drop initialized column if it exists to avoid merge conflicts
                    if col in unlevered.columns:
                        unlevered = unlevered.drop(columns=[col])
            
            if len(merge_cols) > 1:  # More than just "Year"
                unlevered = unlevered.merge(ds[merge_cols], on="Year", how="left")
                
                # Ensure columns exist and fill NaN values (after merge, columns should exist)
                for col in ["Interest (COP)", "Principal (COP)"]:
                    if col in unlevered.columns:
                        unlevered[col] = unlevered[col].fillna(0.0)
                    else:
                        unlevered[col] = 0.0
                
                unlevered["Debt Service (COP)"] = unlevered["Interest (COP)"] + unlevered["Principal (COP)"]
                
                # Upfront fee
                if "Upfront Fee (COP)" in unlevered.columns:
                    unlevered["Debt Fees (COP)"] = unlevered["Debt Fees (COP)"] + unlevered["Upfront Fee (COP)"].fillna(0.0)
            
            # Recalculate taxes with interest expense included
            # Create interest by year map
            interest_by_year = {}
            for _, row in unlevered.iterrows():
                year = int(row["Year"])
                interest_by_year[year] = float(row.get("Interest (COP)", 0.0))
            
            # Recalculate tax benefits and taxes with interest expense
            unlevered = _recalculate_taxes_with_interest(s, unlevered, interest_by_year)
            
            # Debt draws: pro-rata with CAPEX during construction (up to and including COD)
            tl = build_timeline(s.timeline)
            cod_year = tl["cod"].year
            cod_m = date(tl["cod"].year, tl["cod"].month, 1)
            
            # Get monthly CAPEX schedule
            capm = capex_monthly_schedule(s).copy()
            capm["Year"] = capm["Month"].apply(lambda d: d.year)
            
            # Calculate debt draws (pro-rata with CAPEX, up to and including COD month)
            if not capm.empty:
                # Include CAPEX up to and including COD month
                up_to_cod = capm[capm["Month"] <= cod_m].copy()
                if not up_to_cod.empty:
                    up_to_cod["Debt Draw (COP)"] = (float(s.debt.debt_pct_of_capex) / 100.0) * pd.to_numeric(up_to_cod["CAPEX (COP)"], errors="coerce").fillna(0.0)
                    
                    # Ensure total debt draws equal debt amount (adjust last draw if needed)
                    total_draws = up_to_cod["Debt Draw (COP)"].sum()
                    if total_draws > 0 and abs(total_draws - debt_amt) > 1.0:  # Allow small rounding differences
                        # Adjust proportionally or add remainder to COD month
                        adjustment = debt_amt - total_draws
                        if cod_m in up_to_cod["Month"].values:
                            # Add adjustment to COD month
                            cod_idx = up_to_cod[up_to_cod["Month"] == cod_m].index[0]
                            up_to_cod.loc[cod_idx, "Debt Draw (COP)"] += adjustment
                        else:
                            # Scale proportionally
                            if total_draws > 0:
                                scale_factor = debt_amt / total_draws
                                up_to_cod["Debt Draw (COP)"] = up_to_cod["Debt Draw (COP)"] * scale_factor
                    
                    # Aggregate to annual
                    annual_draws = up_to_cod.groupby("Year", as_index=False)["Debt Draw (COP)"].sum()
                    # Merge into unlevered (drop the initialized column first to avoid suffix conflicts)
                    if "Debt Draw (COP)" in unlevered.columns:
                        unlevered = unlevered.drop(columns=["Debt Draw (COP)"])
                    unlevered = unlevered.merge(annual_draws, on="Year", how="left")
                    # Ensure column exists and fill NaN
                    if "Debt Draw (COP)" in unlevered.columns:
                        unlevered["Debt Draw (COP)"] = unlevered["Debt Draw (COP)"].fillna(0.0)
                    else:
                        unlevered["Debt Draw (COP)"] = 0.0
                # If no CAPEX up to COD, Debt Draw stays at initialized 0.0
            
            # Commitment fees (during construction)
            cf_ann = debt_commitment_fee_annual(s).copy()
            if not cf_ann.empty and "Year" in cf_ann.columns and "Commitment Fee (COP)" in cf_ann.columns:
                unlevered = unlevered.merge(cf_ann[["Year", "Commitment Fee (COP)"]], on="Year", how="left")
                unlevered["Debt Fees (COP)"] = unlevered["Debt Fees (COP)"] + unlevered["Commitment Fee (COP)"].fillna(0.0)
    
    # Calculate levered (equity) cash flow
    # If debt is enabled, use recalculated taxes (with interest expense)
    # If no debt, use unlevered after-tax CF
    if debt_enabled and debt_amt > 0:
        # Levered CF = Unlevered Pre-tax CF + Debt Draws - Interest - Principal - Debt Fees - Levered Taxes + VAT Refund
        unlevered["Levered CF (After-tax, COP)"] = (
            unlevered["Unlevered CF Pre-tax (COP)"].fillna(0.0)
            + unlevered["Debt Draw (COP)"].fillna(0.0)
            - unlevered["Interest (COP)"].fillna(0.0)
            - unlevered["Principal (COP)"].fillna(0.0)
            - unlevered["Debt Fees (COP)"].fillna(0.0)
            - unlevered["Levered Taxes Payable (COP)"].fillna(0.0)
            + unlevered["VAT Refund (COP)"].fillna(0.0)
        )
    else:
        # No debt: levered CF = unlevered after-tax CF
        # Also set levered tax columns to match unlevered (no interest, so taxes are same)
        unlevered["Levered CF (After-tax, COP)"] = unlevered["Unlevered CF After Tax (COP)"].fillna(0.0)
        unlevered["Levered Taxable Income (COP)"] = unlevered["Taxable Income (COP)"].fillna(0.0)
        unlevered["Levered CAPEX Tax Deduction (COP)"] = unlevered["CAPEX Tax Deduction (COP)"].fillna(0.0)
        unlevered["Levered Loss Carryforward End (COP)"] = unlevered["Loss Carryforward End (COP)"].fillna(0.0)
        unlevered["Levered Taxes Payable (COP)"] = unlevered["Taxes Payable (COP)"].fillna(0.0)
        unlevered["Levered Net Income After Tax (COP)"] = unlevered["Levered Taxable Income (COP)"] - unlevered["Levered Taxes Payable (COP)"]
    
    unlevered["Cumulative Levered CF (COP)"] = unlevered["Levered CF (After-tax, COP)"].cumsum()
    
    return unlevered


def levered_cashflow_monthly(s: ScenarioInputs) -> pd.DataFrame:
    """
    Calculate monthly levered (equity) cash flow after-tax.
    Starts with monthly unlevered CF, adds debt draws, subtracts debt service and fees.
    """
    # Start with monthly cash flow table
    monthly = cashflow_monthly_table(s).copy()
    
    # Get annual levered CF to extract debt components
    annual_levered = levered_cashflow_annual(s).copy()
    
    # Initialize debt columns
    monthly["Debt Draw (COP)"] = 0.0
    monthly["Interest (COP)"] = 0.0
    monthly["Principal (COP)"] = 0.0
    monthly["Debt Service (COP)"] = 0.0
    monthly["Debt Fees (COP)"] = 0.0
    
    # Check if debt is enabled
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex if debt_enabled else 0.0
    debt_amt = max(debt_amt, 0.0)
    
    if debt_enabled and debt_amt > 0:
        tl = build_timeline(s.timeline)
        cod_m = date(tl["cod"].year, tl["cod"].month, 1)
        
        # Debt draws: pro-rata with CAPEX during construction (pre-COD)
        capm = capex_monthly_schedule(s).copy()
        if not capm.empty:
            capm["Debt Draw (COP)"] = (float(s.debt.debt_pct_of_capex) / 100.0) * pd.to_numeric(capm["CAPEX (COP)"], errors="coerce").fillna(0.0)
            # Only draws before COD
            capm.loc[capm["Month"] >= cod_m, "Debt Draw (COP)"] = 0.0
            # Merge into monthly (drop initialized column first to avoid suffix conflicts)
            if "Debt Draw (COP)" in monthly.columns:
                monthly = monthly.drop(columns=["Debt Draw (COP)"])
            monthly = monthly.merge(capm[["Month", "Debt Draw (COP)"]], on="Month", how="left")
            # Ensure column exists and fill NaN
            if "Debt Draw (COP)" in monthly.columns:
                monthly["Debt Draw (COP)"] = monthly["Debt Draw (COP)"].fillna(0.0)
            else:
                monthly["Debt Draw (COP)"] = 0.0
        
        # Debt service: spread annual interest and principal evenly across months
        if not annual_levered.empty:
            # Create a map of annual debt service components
            interest_map = {int(r["Year"]): float(r["Interest (COP)"]) for _, r in annual_levered.iterrows()}
            principal_map = {int(r["Year"]): float(r["Principal (COP)"]) for _, r in annual_levered.iterrows()}
            
            monthly["Interest (COP)"] = monthly["Year"].map(lambda y: interest_map.get(int(y), 0.0) / 12.0)
            monthly["Principal (COP)"] = monthly["Year"].map(lambda y: principal_map.get(int(y), 0.0) / 12.0)
            monthly["Debt Service (COP)"] = monthly["Interest (COP)"] + monthly["Principal (COP)"]
            
            # Only apply debt service during operation (after COD)
            monthly.loc[monthly["Month"] < cod_m, "Interest (COP)"] = 0.0
            monthly.loc[monthly["Month"] < cod_m, "Principal (COP)"] = 0.0
            monthly.loc[monthly["Month"] < cod_m, "Debt Service (COP)"] = 0.0
        
        # Debt fees: upfront fee at COD, commitment fees during construction
        # Initialize Debt Fees column if not exists
        if "Debt Fees (COP)" not in monthly.columns:
            monthly["Debt Fees (COP)"] = 0.0
        
        # Upfront fee
        upfront_fee = debt_amt * (float(s.debt.upfront_fee_bps) / 10_000.0)
        monthly.loc[monthly["Month"] == cod_m, "Debt Fees (COP)"] = upfront_fee
        
        # Commitment fees: get from annual and spread
        cf_ann = debt_commitment_fee_annual(s).copy()
        if not cf_ann.empty and "Year" in cf_ann.columns:
            cf_map = {int(r["Year"]): float(r["Commitment Fee (COP)"]) for _, r in cf_ann.iterrows()}
            monthly["Commitment Fee (COP)"] = monthly["Year"].map(lambda y: cf_map.get(int(y), 0.0) / 12.0)
            monthly["Debt Fees (COP)"] = monthly["Debt Fees (COP)"].fillna(0.0) + monthly["Commitment Fee (COP)"].fillna(0.0)
        
        monthly["Debt Fees (COP)"] = monthly["Debt Fees (COP)"].fillna(0.0)
    
    # Calculate monthly levered CF from annual levered CF
    # Use the annual levered CF (which has taxes recalculated with interest) and spread proportionally
    if not annual_levered.empty:
        # Get annual levered CF (this already includes all debt components and recalculated taxes)
        levered_cf_annual_map = {int(r["Year"]): float(r["Levered CF (After-tax, COP)"]) for _, r in annual_levered.iterrows()}
        
        # Calculate monthly levered CF by prorating annual levered CF
        # For operation months: prorate based on operating months
        # For construction months: use monthly unlevered pre-tax CF + debt draws - debt service - fees
        tl = build_timeline(s.timeline)
        cod = tl["cod"]
        end_op = tl["end_op"]
        
        monthly["Levered CF (After-tax, COP)"] = 0.0
        
        for _, row in monthly.iterrows():
            month_date = row["Month"]
            year = month_date.year
            phase = _phase_for_month(tl, month_date)
            
            if phase == "Operation" and year in levered_cf_annual_map:
                # Operation months: prorate annual levered CF
                if year == cod.year:
                    operating_months = 13 - cod.month
                elif year == end_op.year:
                    operating_months = end_op.month
                else:
                    operating_months = 12
                
                if operating_months > 0:
                    annual_levered_cf = levered_cf_annual_map[year]
                    monthly.loc[monthly.index == row.name, "Levered CF (After-tax, COP)"] = annual_levered_cf / operating_months
            else:
                # Construction/non-operation months: calculate from components
                # Use unlevered pre-tax CF + debt draws - debt service - fees
                # (No taxes during construction, so this is simpler)
                monthly.loc[monthly.index == row.name, "Levered CF (After-tax, COP)"] = (
                    row.get("Unlevered CF (COP)", 0.0)
                    + row.get("Debt Draw (COP)", 0.0)
                    - row.get("Debt Service (COP)", 0.0)
                    - row.get("Debt Fees (COP)", 0.0)
                )
    else:
        # Fallback: use unlevered CF + debt components
        monthly["Levered CF (After-tax, COP)"] = (
            monthly["Unlevered CF (COP)"].fillna(0.0)
            + monthly["Debt Draw (COP)"].fillna(0.0)
            - monthly["Debt Service (COP)"].fillna(0.0)
            - monthly["Debt Fees (COP)"].fillna(0.0)
        )
    
    monthly["Cumulative Levered CF (COP)"] = monthly["Levered CF (After-tax, COP)"].cumsum()
    
    return monthly


# -----------------------------
# APP UI
# -----------------------------
st.set_page_config(page_title="Delphi Utility-Scale Model", page_icon="⚡", layout="wide")
st.title("Delphi Utility-Scale Project Model (COP inputs, COP/USD outputs)")
st.caption(f"Projects + scenarios stored at: {PROJECTS_FILE}")

db = _load_db()

# Sidebar: project & scenario management
with st.sidebar:
    st.header("Project & Scenario")

    projects = sorted(list(db.get("projects", {}).keys()))
    project_name = st.selectbox("Project", ["(New project)"] + projects, index=0)

    if project_name == "(New project)":
        new_project = st.text_input("New project name", value="")
        if st.button("Create project", type="primary", use_container_width=True):
            if new_project.strip():
                db.setdefault("projects", {}).setdefault(new_project.strip(), {"scenarios": {}})
                _save_db(db)
                st.rerun()
            else:
                st.warning("Enter a project name.")
        st.stop()

    proj = db["projects"].setdefault(project_name, {"scenarios": {}})

    scen_names = sorted(list(proj.get("scenarios", {}).keys()))
    scenario_name = st.selectbox("Scenario", ["(New scenario)"] + scen_names, index=0)

    if scenario_name == "(New scenario)":
        new_s = st.text_input("New scenario name", value="Base")
        if st.button("Create scenario", type="primary", use_container_width=True):
            nm = new_s.strip() or "Base"
            proj["scenarios"][nm] = _scenario_to_dict(ScenarioInputs(name=nm))
            _save_db(db)
            st.rerun()
        st.stop()

    # Load scenario
    s = _scenario_from_dict(proj["scenarios"][scenario_name])
    s.name = scenario_name

    st.divider()
    cdel1, cdel2 = st.columns(2)
    with cdel1:
        if st.button("Save scenario", use_container_width=True):
            proj["scenarios"][scenario_name] = _scenario_to_dict(s)
            _save_db(db)
            st.success("Saved.")
    with cdel2:
        if st.button("Delete scenario", use_container_width=True):
            try:
                del proj["scenarios"][scenario_name]
                _save_db(db)
                st.rerun()
            except Exception:
                st.error("Could not delete.")

    st.divider()
    st.subheader("Compare")
    compare_scenarios = st.multiselect("Select scenarios", scen_names, default=[])


# Tabs (top-down)
tab_macro, tab_timeline, tab_gen, tab_rev, tab_capex, tab_opex, tab_sga, tab_dep, tab_incent, tab_ucf, tab_debt, tab_levered, tab_compare, tab_sensitivity, tab_summary = st.tabs(
    [
        "A) Macroeconomic",
        "B) Timeline",
        "C) Power Generation",
        "D) Power Revenues",
        "E) CAPEX",
        "F) OPEX",
        "G) SG&A",
        "H) Depreciation",
        "I) Renewable tax benefits",
        "J) Unlevered Base Cash Flow",
        "K) Debt & Covenants",
        "L) Levered Cash Flow",
        "M) Compare",
        "N) Sensitivity",
        "O) Summary",
    ]
)

# -----------------------------
# A) Macro
# -----------------------------
with tab_macro:
    st.subheader("Macroeconomic inputs (annual rates, %)")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        s.macro.col_cpi = st.number_input("Colombian CPI (%)", value=float(s.macro.col_cpi), step=0.1, format="%.2f")
    with c2:
        s.macro.col_ppi = st.number_input("Colombian PPI (%)", value=float(s.macro.col_ppi), step=0.1, format="%.2f")
    with c3:
        s.macro.us_cpi = st.number_input("US CPI (%)", value=float(s.macro.us_cpi), step=0.1, format="%.2f")
    with c4:
        s.macro.custom_index_rate = st.number_input("Custom index (%)", value=float(s.macro.custom_index_rate), step=0.1, format="%.2f")

    st.markdown("#### FX (COP per USD)")
    fx1, fx2, fx3 = st.columns([1.2, 1, 1])
    with fx1:
        s.macro.fx_cop_per_usd_start = st.number_input("Starting FX (COP/USD)", value=float(s.macro.fx_cop_per_usd_start), step=10.0, format="%.0f")
        st.caption(f"Formatted: {_fmt_cop(s.macro.fx_cop_per_usd_start)} / USD")
    with fx2:
        s.macro.fx_method = st.selectbox("FX method", ["Inflation differential (PPP approx.)", "Flat"], index=0 if s.macro.fx_method != "Flat" else 1)
    with fx3:
        s.macro.fx_flat = st.number_input("Flat FX (if selected)", value=float(s.macro.fx_flat or s.macro.fx_cop_per_usd_start), step=10.0, format="%.0f")
        st.caption(f"Formatted: {_fmt_cop(s.macro.fx_flat)} / USD")

    st.info("FX path default uses a simple PPP approximation: FX grows with (Col CPI / US CPI). You can switch to Flat FX.")


# -----------------------------
# B) Timeline
# -----------------------------
with tab_timeline:
    st.subheader("Project timeline (Development → CAPEX → Operation)")

    c1, c2, c3 = st.columns(3)
    with c1:
        s.timeline.start_date = st.date_input("Project start date", value=_parse_date_iso(s.timeline.start_date)).isoformat()
    with c2:
        s.timeline.dev_months = int(st.number_input("Development duration (months)", value=int(s.timeline.dev_months), min_value=0, step=1, format="%d"))
    with c3:
        s.timeline.capex_months = int(st.number_input("Construction/CAPEX duration (months)", value=int(s.timeline.capex_months), min_value=0, step=1, format="%d"))

    s.timeline.operation_years = int(st.number_input("Operation duration (years)", value=int(s.timeline.operation_years), min_value=1, step=1, format="%d"))

    tl = build_timeline(s.timeline)
    _metric_row(
        [
            ("Start", str(tl["start"])),
            ("RTB (end of dev)", str(tl["rtb"])),
            ("COD (start ops)", str(tl["cod"])),
            ("End of operation", str(tl["end_op"])),
        ]
    )

    st.markdown("#### Timeline visual (calendar, years)")

    gantt = pd.DataFrame(
        [
            {"Stage": "Development", "Start": date(tl["start"].year, tl["start"].month, 1), "Finish": date(tl["rtb"].year, tl["rtb"].month, 1)},
            {"Stage": "Construction", "Start": date(tl["rtb"].year, tl["rtb"].month, 1), "Finish": date(tl["cod"].year, tl["cod"].month, 1)},
            {"Stage": "Operation", "Start": date(tl["cod"].year, tl["cod"].month, 1), "Finish": date(tl["end_op"].year, tl["end_op"].month, 1)},
        ]
    )

    fig = px.timeline(
        gantt,
        x_start="Start",
        x_end="Finish",
        y="Stage",
        color="Stage",
        category_orders={"Stage": ["Development", "Construction", "Operation"]},
    )
    fig.update_yaxes(autorange="reversed")
    fig.update_xaxes(dtick="M12", tickformat="%Y")
    fig.update_layout(height=280, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
    st.plotly_chart(fig, use_container_width=True)


# -----------------------------
# C) Generation
# -----------------------------
with tab_gen:
    st.subheader("Power generation inputs")
    
    # Show current revenue mode and note that it affects prices/revenues here
    revenue_mode_display = s.revenue_mode if hasattr(s, 'revenue_mode') else "Standard PPA Parameters"
    st.info(f"ℹ️ **Revenue mode:** {revenue_mode_display}. Prices and revenues shown below are calculated based on the revenue mode selected in the 'Power Revenues' tab. Change the revenue mode there to update prices and revenues here.")

    c1, c2, c3 = st.columns(3)
    with c1:
        s.generation.mwac = st.number_input("Capacity (MWac)", value=float(s.generation.mwac), step=1.0, format="%.2f")
    with c2:
        s.generation.mwp = st.number_input("Capacity (MWp)", value=float(s.generation.mwp), step=1.0, format="%.2f")
    with c3:
        s.generation.degradation_pct = st.number_input("Annual degradation (%/yr)", value=float(s.generation.degradation_pct), step=0.05, format="%.2f")

    c4, c5, c6, c7 = st.columns(4)
    with c4:
        s.generation.p50_mwh_yr = st.number_input("P50 (MWh/year)", value=float(s.generation.p50_mwh_yr), step=1000.0, format="%.0f")
    with c5:
        s.generation.p75_mwh_yr = st.number_input("P75 (MWh/year)", value=float(s.generation.p75_mwh_yr), step=1000.0, format="%.0f")
    with c6:
        s.generation.p90_mwh_yr = st.number_input("P90 (MWh/year)", value=float(s.generation.p90_mwh_yr), step=1000.0, format="%.0f")
    with c7:
        s.generation.production_choice = st.selectbox("Production choice", ["P50", "P75", "P90"], index=["P50", "P75", "P90"].index(s.generation.production_choice))

    # Force recalculation by calling operating_year_table with current scenario state
    op = operating_year_table(s)
    
    # Check if prices are zero when in manual mode (might indicate prices not set)
    if revenue_mode_display == "Manual annual series":
        # Check if prices dictionary has any non-zero values
        r = s.revenue2
        has_prices = any(v > 0 for v in r.prices_constant_cop_per_kwh.values()) if r.prices_constant_cop_per_kwh else False
        zero_prices = op["Price (COP/kWh)"].fillna(0.0) == 0.0
        if zero_prices.any():
            if not has_prices:
                st.warning(f"⚠️ **Warning:** Prices are showing as 0.0 because no prices have been entered yet. Please go to the 'Power Revenues' tab and enter prices for all operating years in the Manual annual series table. The table will update automatically after you save the prices.")
            else:
                # Show which operating years have prices vs which are missing
                missing_years = []
                for idx, row in op.iterrows():
                    if row["Price (COP/kWh)"] == 0.0:
                        # Try to determine which operating year this corresponds to
                        year = int(row["Year"])
                        # This is approximate - we can't easily reverse-engineer the operating year from calendar year
                        missing_years.append(year)
                if missing_years:
                    st.warning(f"⚠️ **Warning:** Some prices are showing as 0.0 for years {missing_years[:5]}{'...' if len(missing_years) > 5 else ''}. Please check that prices are entered for all operating years in the 'Power Revenues' tab. After entering prices, navigate back to this tab to see the updated values.")
    
    fig = px.line(op, x="Year", y="Energy (MWh)")
    fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)


# -----------------------------
# D) Revenues
# -----------------------------
with tab_rev:
    st.subheader("Power revenues (indexed, annual)")
    
    st.warning("⚠️ **Important:** The revenue mode selected here affects prices and revenues throughout the entire model, including the Power Generation tab, Unlevered Cash Flow, and all downstream calculations.")

    s.revenue_mode = st.radio("Revenue mode", ["Standard PPA Parameters", "Manual annual series"], horizontal=True, index=0 if s.revenue_mode == "Standard PPA Parameters" else 1)

    if s.revenue_mode == "Standard PPA Parameters":
        r = s.revenue1
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            r.ppa_price_cop_per_kwh = st.number_input("PPA price at COD (COP/kWh, current COP)", value=float(r.ppa_price_cop_per_kwh), step=1.0, format="%.2f")
        with c2:
            r.ppa_term_years = int(st.number_input("PPA term (years)", value=int(r.ppa_term_years), min_value=1, step=1, format="%d"))
        with c3:
            r.merchant_price_cop_per_kwh = st.number_input("Post-term merchant price (COP/kWh)", value=float(r.merchant_price_cop_per_kwh), step=1.0, format="%.2f")
        with c4:
            r.indexation = st.selectbox("Indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(r.indexation) if r.indexation in INDEX_CHOICES else 0)
    else:
        r = s.revenue2
        c1, c2 = st.columns([1, 2])
        with c1:
            r.indexation = st.selectbox("Indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(r.indexation) if r.indexation in INDEX_CHOICES else 0)
        with c2:
            st.info("Enter constant COP/kWh by operating year (Year 1 = first year at COD).")

        prices_df = pd.DataFrame({"OpYear": list(range(1, int(s.timeline.operation_years) + 1))})
        prices_df["COP_per_kWh_constant"] = prices_df["OpYear"].apply(lambda i: float(r.prices_constant_cop_per_kwh.get(int(i), 0.0)))
        edited = st.data_editor(
            prices_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key="manual_prices_editor",  # Add key to ensure proper state management
            column_config={
                "OpYear": st.column_config.NumberColumn("Operating year", format="%d", disabled=True),
                "COP_per_kWh_constant": st.column_config.NumberColumn("Price (COP/kWh, constant)", step=1.0, format="%.2f"),
            },
        )
        # Save prices immediately after editing
        r.prices_constant_cop_per_kwh = {int(row.OpYear): float(row.COP_per_kWh_constant) for row in edited.itertuples(index=False)}
        
        # Show confirmation that prices are saved
        if len(r.prices_constant_cop_per_kwh) > 0:
            non_zero_prices = sum(1 for v in r.prices_constant_cop_per_kwh.values() if v > 0)
            if non_zero_prices > 0:
                st.success(f"✓ Prices saved for {non_zero_prices} operating year(s). Scenario will be auto-saved and prices will persist when you reload.")

    op = operating_year_table(s)

    c1, c2 = st.columns(2)
    with c1:
        fig1 = px.bar(op, x="Year", y="Energy (MWh)")
        fig1.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig1, use_container_width=True)
    with c2:
        fig2 = px.line(op, x="Year", y="Revenue (COP)")
        fig2.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig2, use_container_width=True)

    disp = op.copy()
    disp = _df_format_money(disp, ["Energy (MWh)", "Price (COP/kWh)", "Revenue (COP)"], decimals=0)
    st.dataframe(disp, use_container_width=True, hide_index=True)


# -----------------------------
# E) CAPEX
# -----------------------------
with tab_capex:
    st.subheader("CAPEX (COP)")

    s.capex.distribution = st.selectbox("Construction spend distribution", CAPEX_DISTS, index=CAPEX_DISTS.index(s.capex.distribution) if s.capex.distribution in CAPEX_DISTS else 0)

    capex_df = pd.DataFrame(s.capex.lines or [])
    for col in ["Item", "Amount_COP", "Phase"]:
        if col not in capex_df.columns:
            capex_df[col] = "" if col != "Amount_COP" else 0.0
    capex_df = capex_df[["Item", "Amount_COP", "Phase"]].copy()
    capex_df["Phase"] = capex_df["Phase"].where(capex_df["Phase"].isin(CAPEX_PHASES), "Construction")

    edited = st.data_editor(
        capex_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Item": st.column_config.TextColumn("CAPEX line item"),
            "Amount_COP": st.column_config.NumberColumn("Amount (COP)", min_value=0.0, step=1_000_000.0, format="%.0f"),
            "Phase": st.column_config.SelectboxColumn("Phase", options=CAPEX_PHASES),
        },
    )
    s.capex.lines = edited.to_dict(orient="records")

    total_capex = float(pd.to_numeric(edited["Amount_COP"], errors="coerce").fillna(0.0).sum())
    mwac = float(s.generation.mwac or 0.0)
    mwp = float(s.generation.mwp or 0.0)
    capex_per_mwac = (total_capex / mwac) if mwac > 0 else np.nan
    capex_per_mwp = (total_capex / mwp) if mwp > 0 else np.nan

    c1, c2, c3 = st.columns(3)
    c1.metric("Total CAPEX (COP)", _fmt_cop(total_capex))
    c2.metric("CAPEX / MWac (COP)", _fmt_cop(capex_per_mwac) if np.isfinite(capex_per_mwac) else "—")
    c3.metric("CAPEX / MWp (COP)", _fmt_cop(capex_per_mwp) if np.isfinite(capex_per_mwp) else "—")

    st.markdown("#### CAPEX breakdown (by line item)")
    capex_pie = edited.copy()
    capex_pie["Item"] = capex_pie["Item"].fillna("").astype(str).str.strip()
    capex_pie["Amount_COP"] = pd.to_numeric(capex_pie["Amount_COP"], errors="coerce").fillna(0.0)
    capex_pie = capex_pie[capex_pie["Amount_COP"] > 0].copy()

    if capex_pie.empty:
        st.info("Enter CAPEX amounts to see the breakdown chart.")
    else:
        share = capex_pie["Amount_COP"] / capex_pie["Amount_COP"].sum()
        small = share < 0.03
        if small.any() and (~small).any():
            other_amt = float(capex_pie.loc[small, "Amount_COP"].sum())
            capex_pie = capex_pie.loc[~small, ["Item", "Amount_COP"]]
            capex_pie = pd.concat(
                [capex_pie, pd.DataFrame([{"Item": "Other (<3% each)", "Amount_COP": other_amt}])],
                ignore_index=True
            )
        fig_pie = px.pie(capex_pie, names="Item", values="Amount_COP", hole=0.45)
        fig_pie.update_traces(textinfo="percent+label")
        fig_pie.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
        st.plotly_chart(fig_pie, use_container_width=True, key="capex_breakdown_pie")
    
    st.markdown("#### CAPEX schedule (monthly, aligned to timeline)")
    sched = capex_monthly_schedule(s)
    # Ensure CAPEX (COP) column is numeric and fill any NaN with 0
    sched["CAPEX (COP)"] = pd.to_numeric(sched["CAPEX (COP)"], errors="coerce").fillna(0.0)
    sched_disp = _df_format_money(sched.copy(), ["CAPEX (COP)"], decimals=0)
    st.dataframe(sched_disp[["Month", "Phase", "CAPEX (COP)"]], use_container_width=True, hide_index=True)

    fig = px.bar(sched, x="Month", y="CAPEX (COP)", color="Phase")
    fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("#### Annual CAPEX (calendar years)")
    ann = sched.groupby("Year", as_index=False)["CAPEX (COP)"].sum()
    ann_disp = _df_format_money(ann.copy(), ["CAPEX (COP)"], decimals=0)
    st.dataframe(ann_disp, use_container_width=True, hide_index=True)


# -----------------------------
# F) OPEX
# -----------------------------
with tab_opex:
    st.subheader("OPEX (COP) — operating costs, land lease, taxes & levies")

    c1, c2, c3 = st.columns(3)
    with c1:
        s.opex.fixed_om_cop_per_mwac_year = st.number_input("Fixed O&M (COP/MWac-year)", value=float(s.opex.fixed_om_cop_per_mwac_year), step=1_000_000.0, format="%.0f")
        s.opex.fixed_om_indexation = st.selectbox("Fixed O&M indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(s.opex.fixed_om_indexation) if s.opex.fixed_om_indexation in INDEX_CHOICES else 0)
    with c2:
        s.opex.variable_om_cop_per_mwh = st.number_input("Variable O&M (COP/MWh)", value=float(s.opex.variable_om_cop_per_mwh), step=1_000.0, format="%.0f")
        s.opex.grid_fees_cop_per_mwh = st.number_input("Grid fees (COP/MWh)", value=float(s.opex.grid_fees_cop_per_mwh), step=1_000.0, format="%.0f")
    with c3:
        s.opex.insurance_cop_per_mwac_year = st.number_input("Insurance (COP/MWac-year)", value=float(s.opex.insurance_cop_per_mwac_year), step=1_000_000.0, format="%.0f")
        s.opex.insurance_indexation = st.selectbox("Insurance indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(s.opex.insurance_indexation) if s.opex.insurance_indexation in INDEX_CHOICES else 0)

    st.markdown("#### Land lease")
    lc1, lc2, lc3, lc4, lc5 = st.columns([1, 1, 1, 1, 1.2])
    with lc1:
        s.opex.land_hectares = st.number_input("Hectares leased (Ha)", value=float(s.opex.land_hectares), step=1.0, format="%.2f")
    with lc2:
        s.opex.land_price_dev_cop_per_ha_year = st.number_input("Dev lease (COP/Ha-year)", value=float(s.opex.land_price_dev_cop_per_ha_year), step=1_000_000.0, format="%.0f")
    with lc3:
        s.opex.land_price_con_cop_per_ha_year = st.number_input("Const lease (COP/Ha-year)", value=float(s.opex.land_price_con_cop_per_ha_year), step=1_000_000.0, format="%.0f")
    with lc4:
        s.opex.land_price_op_cop_per_ha_year = st.number_input("Op lease (COP/Ha-year)", value=float(s.opex.land_price_op_cop_per_ha_year), step=1_000_000.0, format="%.0f")
    with lc5:
        s.opex.land_indexation = st.selectbox("Land lease indexation", INDEX_CHOICES, index=INDEX_CHOICES.index(s.opex.land_indexation) if s.opex.land_indexation in INDEX_CHOICES else 0)

    st.markdown("#### Taxes & levies")
    tc1, tc2 = st.columns(2)
    with tc1:
        s.opex.ica_pct_of_revenue = st.number_input("ICA (% of revenue)", value=float(s.opex.ica_pct_of_revenue), step=0.01, format="%.3f")
    with tc2:
        s.opex.gmf_pct_of_outflows = st.number_input("GMF (% of outgoing cash)", value=float(s.opex.gmf_pct_of_outflows), step=0.01, format="%.3f")

    st.markdown("#### Other OPEX items (dynamic)")
    o_df = pd.DataFrame(s.opex.other_items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in o_df.columns:
            o_df[col] = "" if col != "Amount_COP_per_year" else 0.0
    o_df = o_df[["Item", "Amount_COP_per_year", "Phase", "Indexation"]].copy()
    o_df["Phase"] = o_df["Phase"].where(o_df["Phase"].isin(PHASES), "Operation")
    o_df["Indexation"] = o_df["Indexation"].where(o_df["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    o_edited = st.data_editor(
        o_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Item": st.column_config.TextColumn("OPEX item"),
            "Amount_COP_per_year": st.column_config.NumberColumn("Amount (COP/year)", min_value=0.0, step=1_000_000.0, format="%.0f"),
            "Phase": st.column_config.SelectboxColumn("Phase", options=PHASES),
            "Indexation": st.column_config.SelectboxColumn("Indexation", options=INDEX_CHOICES),
        },
    )
    s.opex.other_items = o_edited.to_dict(orient="records")

    st.divider()
    st.markdown("## Outputs")

    om_full = opex_monthly_schedule(s).copy()

    meta_cols = {"Month", "Year", "Phase", "Energy (MWh)", "Revenue (COP)", "CAPEX (COP)", "OPEX subtotal", "GMF"}
    fixed_cols = {"Fixed O&M", "Insurance", "Variable O&M", "Grid fees", "Land lease", "ICA"}
    dyn_cols = [c for c in om_full.columns if c not in meta_cols and c not in fixed_cols]
    opex_item_cols = list(fixed_cols) + dyn_cols

    annual_items = om_full.groupby("Year", as_index=False)[opex_item_cols + ["GMF"]].sum()
    annual_items["Total OPEX (COP)"] = annual_items[opex_item_cols].sum(axis=1) + annual_items["GMF"]

    long = annual_items.melt(id_vars=["Year"], value_vars=opex_item_cols + ["GMF"], var_name="Item", value_name="OPEX (COP)")
    fig = px.bar(long, x="Year", y="OPEX (COP)", color="Item", barmode="stack")
    fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
    st.plotly_chart(fig, use_container_width=True)

    annual = om_full.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
    annual["Total OPEX (COP)"] = annual["OPEX subtotal"] + annual["GMF"]

    op = operating_year_table(s)[["Year", "Energy (MWh)"]].copy()
    annual = annual.merge(op, on="Year", how="left").fillna({"Energy (MWh)": 0.0})
    annual["OPEX per MWh (COP/MWh)"] = np.where(annual["Energy (MWh)"] > 0, annual["Total OPEX (COP)"] / annual["Energy (MWh)"], 0.0)

    disp = _df_format_money(annual.copy(), ["OPEX subtotal", "GMF", "Total OPEX (COP)", "OPEX per MWh (COP/MWh)", "Energy (MWh)"], decimals=0)
    st.dataframe(disp, use_container_width=True, hide_index=True)


# -----------------------------
# G) SG&A
# -----------------------------
with tab_sga:
    st.subheader("SG&A (COP) — Development, Construction, and Operation")

    sga_df = pd.DataFrame(s.sga.items or [])
    for col in ["Item", "Amount_COP_per_year", "Phase", "Indexation"]:
        if col not in sga_df.columns:
            sga_df[col] = "" if col != "Amount_COP_per_year" else 0.0
    sga_df = sga_df[["Item", "Amount_COP_per_year", "Phase", "Indexation"]].copy()
    sga_df["Phase"] = sga_df["Phase"].where(sga_df["Phase"].isin(PHASES), "Development")
    sga_df["Indexation"] = sga_df["Indexation"].where(sga_df["Indexation"].isin(INDEX_CHOICES), "Colombia CPI")

    sga_edited = st.data_editor(
        sga_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Item": st.column_config.TextColumn("SG&A item"),
            "Amount_COP_per_year": st.column_config.NumberColumn("Amount (COP/year)", min_value=0.0, step=1_000_000.0, format="%.0f"),
            "Phase": st.column_config.SelectboxColumn("Phase", options=PHASES),
            "Indexation": st.column_config.SelectboxColumn("Indexation", options=INDEX_CHOICES),
        },
    )
    s.sga.items = sga_edited.to_dict(orient="records")

    st.divider()
    st.markdown("## Outputs")

    annual_sga = sga_annual_by_item(s)
    item_cols = [c for c in annual_sga.columns if c not in ["Year", "Total SG&A (COP)"]]
    if item_cols:
        annual_long = annual_sga.melt(id_vars=["Year"], value_vars=item_cols, var_name="Item", value_name="SG&A (COP)")
        fig = px.bar(annual_long, x="Year", y="SG&A (COP)", color="Item", barmode="stack")
        fig.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Add SG&A line items to see the chart.")

    annual_disp = _df_format_money(annual_sga.copy(), [c for c in annual_sga.columns if c != "Year"], decimals=0)
    st.dataframe(annual_disp, use_container_width=True, hide_index=True)


# -----------------------------
# H) Depreciation
# -----------------------------
with tab_dep:
    st.subheader("Depreciation (linear, starting at COD)")

    c1, c2 = st.columns(2)
    with c1:
        s.depreciation.pct_of_capex_depreciated = st.number_input(
            "% of CAPEX depreciated",
            value=float(s.depreciation.pct_of_capex_depreciated),
            min_value=0.0,
            max_value=100.0,
            step=1.0,
            format="%.1f",
        )
    with c2:
        s.depreciation.dep_years = int(
            st.number_input(
                "Depreciation period (years after COD)",
                value=int(s.depreciation.dep_years),
                min_value=3,
                max_value=25,
                step=1,
                format="%d",
            )
        )

    dep = depreciation_annual_table(s)

    fig = px.bar(dep, x="Year", y="Depreciation (COP)")
    fig.update_layout(height=340, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)

    dep_disp = _df_format_money(dep.copy(), [c for c in dep.columns if c != "Year"], decimals=0)
    st.dataframe(dep_disp, use_container_width=True, hide_index=True)


# -----------------------------
# K) Debt & Covenants
# -----------------------------
with tab_debt:
    st.subheader("Debt & Covenants (tenor 5–10 years, sculpted amortization)")

    s.debt.enabled = st.checkbox("Enable debt", value=bool(s.debt.enabled))
    s.debt.balloon_pct = st.slider("Balloon (% of original debt at maturity)", 0.0, 50.0, float(getattr(s.debt, "balloon_pct", 0.0)), 1.0)

    if not s.debt.enabled:
        st.info("Debt is disabled. Enable it to see debt sizing, amortization, DSCR, and covenant indicators.")
    else:
        total_capex = _total_capex_from_lines(s)
        s.debt.debt_pct_of_capex = st.slider("Debt as % of CAPEX", min_value=0.0, max_value=90.0, value=float(s.debt.debt_pct_of_capex), step=1.0)
        debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex

        c1, c2, c3 = st.columns(3)
        with c1:
            s.debt.tenor_years = int(st.number_input("Tenor (years)", value=int(s.debt.tenor_years), min_value=5, max_value=10, step=1, format="%d"))
        with c2:
            s.debt.grace_years = int(st.number_input("Grace period (years)", value=int(s.debt.grace_years), min_value=0, max_value=max(0, int(s.debt.tenor_years) - 1), step=1, format="%d"))
        with c3:
            amort_options = ["Sculpted to DSCR", "Equal principal", "Typical amortization"]
            current_index = 0
            if s.debt.amortization_type == "Equal principal":
                current_index = 1
            elif s.debt.amortization_type == "Typical amortization":
                current_index = 2
            s.debt.amortization_type = st.selectbox("Amortization", amort_options, index=current_index)

        st.markdown("#### Pricing (Natural COP)")
        p1, p2, p3 = st.columns(3)
        with p1:
            s.debt.base_rate_pct = st.number_input("Base rate (e.g., IBR) %", value=float(s.debt.base_rate_pct), step=0.25, format="%.2f")
        with p2:
            s.debt.margin_pct = st.number_input("Margin (spread) %", value=float(s.debt.margin_pct), step=0.25, format="%.2f")
        with p3:
            all_in = float(s.debt.base_rate_pct) + float(s.debt.margin_pct)
            st.metric("All-in interest rate", f"{all_in:,.2f}%")

        st.markdown("#### Fees")
        f1, f2 = st.columns(2)
        with f1:
            s.debt.upfront_fee_bps = st.number_input("Upfront / structuring fee (bps of debt)", value=float(s.debt.upfront_fee_bps), step=5.0, format="%.0f")
        with f2:
            s.debt.commitment_fee_pct_of_margin = st.number_input("Commitment fee (% of margin) on undrawn", value=float(s.debt.commitment_fee_pct_of_margin), step=1.0, format="%.0f")

        st.markdown("#### Covenants")
        k1, k2, k3 = st.columns(3)
        with k1:
            s.debt.target_dscr = st.number_input("Target DSCR (for sculpting)", value=float(s.debt.target_dscr), step=0.01, format="%.2f")
        with k2:
            s.debt.min_dscr_covenant = st.number_input("Minimum DSCR covenant", value=float(s.debt.min_dscr_covenant), step=0.01, format="%.2f")
        with k3:
            s.debt.lockup_dscr = st.number_input("Lock-up DSCR threshold", value=float(s.debt.lockup_dscr), step=0.01, format="%.2f")

        _metric_row([
            ("Total CAPEX", _fmt_cop(total_capex)),
            ("Debt amount", _fmt_cop(debt_amt)),
            ("Equity (implied)", _fmt_cop(max(total_capex - debt_amt, 0.0))),
        ])

        st.divider()

        ds = debt_schedule_annual(s)
        com = debt_commitment_fee_annual(s)

        # KPIs over debt life where debt service > 0
        ds_valid = ds[pd.to_numeric(ds["Debt Service (COP)"], errors="coerce").fillna(0.0) > 0].copy()
        min_dscr = float(ds_valid["DSCR"].min()) if not ds_valid.empty else float("nan")
        lockup_years = int(ds["Lock-up? (DSCR < lock-up)"].sum()) if "Lock-up? (DSCR < lock-up)" in ds.columns else 0
        breach_years = int(ds["Breach? (DSCR < min)"].sum()) if "Breach? (DSCR < min)" in ds.columns else 0
        balloon = float(ds["Balloon at Maturity (COP)"].sum()) if "Balloon at Maturity (COP)" in ds.columns else 0.0
        upfront_fee = float(ds["Upfront Fee (COP)"].sum()) if "Upfront Fee (COP)" in ds.columns else 0.0
        commit_total = float(com["Commitment Fee (COP)"].sum()) if (not com.empty and "Commitment Fee (COP)" in com.columns) else 0.0

        # Calculate breach years with warning
        breach_warning = breach_years > 1
        
        # Display metrics - highlight breach years if > 1
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Min DSCR (debt life)", f"{min_dscr:,.2f}x" if np.isfinite(min_dscr) else "—")
        with col2:
            st.metric("Lock-up years", str(lockup_years))
        with col3:
            if breach_warning:
                # Display in red box for visibility
                st.markdown(
                    f'<div style="background-color: #ffebee; padding: 8px; border-radius: 5px; border-left: 4px solid #f44336; margin-top: 8px;">'
                    f'<div style="font-size: 0.8rem; color: #666;">Breach years</div>'
                    f'<div style="font-size: 1.5rem; font-weight: bold; color: #c62828;">{breach_years}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            else:
                st.metric("Breach years", str(breach_years))
        with col4:
            st.metric("Balloon", _fmt_cop(balloon))
        
        # Warning for DSCR breaches exceeding 1 year
        if breach_warning:
            st.error(f"⚠️ **WARNING: DSCR covenant breaches exceed 1 year ({breach_years} years).** The project may be overleveraged and unable to service debt from operating cash flow.")
        
        _metric_row([
            ("Upfront fee", _fmt_cop(upfront_fee)),
            ("Commitment fees (total)", _fmt_cop(commit_total)),
            ("Tenor", f"{int(s.debt.tenor_years)} years"),
            ("Grace", f"{int(s.debt.grace_years)} years"),
        ])
        
        # Calculate actual equity investment from levered cash flow and compare to implied equity
        try:
            annual_levered = levered_cashflow_annual(s)
            actual_equity_investment = 0.0
            for _, row in annual_levered.iterrows():
                levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
                if levered_cf < 0:
                    actual_equity_investment += abs(levered_cf)
            
            implied_equity = max(total_capex - debt_amt, 0.0)
            equity_difference = actual_equity_investment - implied_equity
            equity_difference_pct = (equity_difference / implied_equity * 100.0) if implied_equity > 0 else 0.0
            
            # Warning if actual equity significantly exceeds implied equity (more than 10% difference)
            if equity_difference > 0 and equity_difference_pct > 10.0:
                st.warning(
                    f"⚠️ **EQUITY SHORTFALL WARNING:**\n\n"
                    f"- **Implied Equity** (from debt tab): {_fmt_cop(implied_equity)}\n"
                    f"- **Actual Equity Investment** (from levered cash flow): {_fmt_cop(actual_equity_investment)}\n"
                    f"- **Additional Equity Required**: {_fmt_cop(equity_difference)} ({equity_difference_pct:.1f}% more)\n\n"
                    f"This indicates that debt service and fees exceed operating cash flow, requiring the investor to contribute more equity than initially expected. "
                    f"The project may be overleveraged."
                )
        except Exception as e:
            # If levered cash flow calculation fails, skip the warning
            pass

        st.markdown("### Debt schedule (annual)")
        disp = ds.copy()
        money_cols = ["Operating CF (COP)", "Interest (COP)", "Principal (COP)", "Debt Service (COP)", "Outstanding End (COP)", "Balloon at Maturity (COP)", "Upfront Fee (COP)"]
        disp = _df_format_money(disp, money_cols, decimals=0)
        st.dataframe(disp, use_container_width=True, hide_index=True)

        st.markdown("### DSCR vs covenant thresholds")
        ds_plot = ds.copy()
        ds_plot["Min covenant"] = float(s.debt.min_dscr_covenant)
        ds_plot["Lock-up"] = float(s.debt.lockup_dscr)
        ds_long = ds_plot.melt(id_vars=["Year"], value_vars=["DSCR", "Min covenant", "Lock-up"], var_name="Line", value_name="Value")
        fig = px.line(ds_long, x="Year", y="Value", color="Line")
        fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("### Commitment fees (during construction)")
        if com.empty:
            st.info("No commitment fees computed (likely no construction months or debt amount is 0).")
        else:
            com_disp = _df_format_money(com.copy(), ["Commitment Fee (COP)"], decimals=0)
            st.dataframe(com_disp, use_container_width=True, hide_index=True)
            figc = px.bar(com, x="Year", y="Commitment Fee (COP)")
            figc.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(figc, use_container_width=True)


# -----------------------------
# I) Renewable tax benefits (moved before Unlevered CF)
# -----------------------------
with tab_incent:
    st.subheader("Renewable tax benefits (Colombia)")

    inc = s.incentives

    st.markdown("### Special deduction (up to 50% of eligible CAPEX over up to 15 years)")
    c1, c2, c3 = st.columns(3)
    with c1:
        inc.enable_special_deduction = st.checkbox("Enable special deduction", value=bool(inc.enable_special_deduction))
    with c2:
        inc.special_deduction_pct_of_capex = st.number_input(
            "Deduction pool (% of CAPEX, max 50%)",
            value=float(inc.special_deduction_pct_of_capex),
            min_value=0.0, max_value=50.0, step=1.0, format="%.1f"
        )
    with c3:
        inc.special_deduction_years = int(st.number_input(
            "Carryforward window (years)",
            value=int(inc.special_deduction_years),
            min_value=1, max_value=30, step=1, format="%d"
        ))

    inc.special_deduction_max_pct_of_taxable_income = st.number_input(
        "Annual cap (% of taxable income, typically 50%)",
        value=float(inc.special_deduction_max_pct_of_taxable_income),
        min_value=0.0, max_value=100.0, step=1.0, format="%.1f"
    )

    # Calculate and display total CAPEX deduction pool
    if bool(inc.enable_special_deduction):
        cap_df = pd.DataFrame(s.capex.lines or [])
        total_capex = float(cap_df["Amount_COP"].fillna(0).sum()) if (not cap_df.empty and "Amount_COP" in cap_df.columns) else 0.0
        pool_total = (float(inc.special_deduction_pct_of_capex) / 100.0) * total_capex
        st.info(f"**Total CAPEX Deduction Pool:** {_fmt_cop(pool_total)} ({inc.special_deduction_pct_of_capex:.1f}% of {_fmt_cop(total_capex)} total CAPEX)")
    else:
        st.info("Special deduction is disabled. Enable it above to see the deduction pool.")

    st.divider()
    st.markdown("### VAT treatment")
    inc.vat_mode = st.selectbox("VAT mode", ["Excluded", "Refund"], index=0 if inc.vat_mode != "Refund" else 1)

    c4, c5, c6 = st.columns(3)
    with c4:
        inc.vat_pct_of_capex = st.number_input(
            "VAT as % of CAPEX (optional)",
            value=float(inc.vat_pct_of_capex),
            min_value=0.0, max_value=100.0, step=0.5, format="%.2f"
        )
    with c5:
        inc.vat_fixed_cop = st.number_input(
            "VAT fixed COP (optional)",
            value=float(inc.vat_fixed_cop),
            min_value=0.0, step=1_000_000.0, format="%.0f"
        )
    with c6:
        inc.vat_refund_year_index = int(st.number_input(
            "Refund in operating year #",
            value=int(inc.vat_refund_year_index),
            min_value=1, max_value=10, step=1, format="%d"
        ))

    st.caption("If VAT is Excluded, leave VAT inputs at 0. If Refund, you can use % or fixed; both will be added (so use one).")
    st.info("ℹ️ These tax benefits are automatically applied in the Unlevered Base Cash Flow calculation below.")


# -----------------------------
# J) Unlevered Base Cash Flow
# -----------------------------
with tab_ucf:
    st.subheader("Unlevered Base Cash Flow (includes tax benefits, pre-debt)")

    st.info("This tab calculates unlevered cash flow before and after tax. Tax benefits from the Renewable tax benefits tab are automatically included. This is the base for calculating debt capacity and equity returns.")

    tx1, tx2 = st.columns([1, 2])
    with tx1:
        s.tax.corporate_tax_rate_pct = st.number_input("Corporate income tax rate (%)", value=float(s.tax.corporate_tax_rate_pct), step=0.5, format="%.2f")
    with tx2:
        s.tax.allow_loss_carryforward = st.checkbox("Apply loss carryforward (NOL) so taxes are zero until losses are used", value=bool(s.tax.allow_loss_carryforward))

    st.markdown("### Working Capital (timing)")
    wc1, wc2, wc3, wc4 = st.columns([1, 1, 1, 1])
    with wc1:
        s.wc.ar_days = int(st.number_input("AR days (revenue collection)", value=int(s.wc.ar_days), min_value=0, step=15, format="%d"))
    with wc2:
        s.wc.ap_days = int(st.number_input("AP days (expense payment)", value=int(s.wc.ap_days), min_value=0, step=15, format="%d"))
    with wc3:
        s.wc.apply_ap_to_opex = st.checkbox("Apply AP lag to OPEX", value=bool(s.wc.apply_ap_to_opex))
    with wc4:
        s.wc.apply_ap_to_sga = st.checkbox("Apply AP lag to SG&A", value=bool(s.wc.apply_ap_to_sga))

    currency = st.radio("Display currency", ["COP", "USD"], horizontal=True, index=0, key="currency_unlevered")

    m = cashflow_monthly_table(s)
    a = unlevered_base_cashflow_annual(s)

    tl = build_timeline(s.timeline)
    years = list(a["Year"].astype(int).tolist())
    fx = fx_series(s.macro, tl["cod"].year, years)

    def _conv(series: pd.Series, year_col: pd.Series) -> pd.Series:
        if currency == "COP":
            return series
        return series / year_col.map(lambda y: float(fx.loc[int(y)]) if int(y) in fx.index else float(s.macro.fx_cop_per_usd_start))

    st.divider()

    # --- KPIs ---
    mm = cashflow_monthly_table(s).copy()
    for col in ["CAPEX (COP)", "Unlevered CF (COP)", "Unlevered CF Pre-tax (COP)", "Cumulative Unlevered CF (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)

    # Calculate total CAPEX from line items (more reliable than summing monthly table)
    cap_total = _total_capex_from_lines(s)
    
    # Pre-tax cash flows - use Pre-tax column if available, otherwise fall back to original
    if "Unlevered CF Pre-tax (COP)" in mm.columns:
        monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist()
    else:
        monthly_cf_pre_tax = mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    monthly_dates = mm["Month"].tolist() if "Month" in mm.columns else []

    # After-tax cash flows: build from annual "Unlevered CF After Tax" values
    # Create a map of annual after-tax CF by year
    annual_after_tax_map = {}
    annual_pre_tax_map = {}
    for _, row in a.iterrows():
        year = int(row["Year"])
        annual_after_tax_map[year] = float(row.get("Unlevered CF After Tax (COP)", 0.0))
        annual_pre_tax_map[year] = float(row.get("Unlevered CF Pre-tax (COP)", row.get("Unlevered CF (COP)", 0.0)))
    
    # Calculate operating months per year for proration
    cod = tl["cod"]
    end_op = tl["end_op"]
    operating_months_map = {}
    for year in annual_after_tax_map.keys():
        if year == cod.year:
            operating_months_map[year] = 13 - cod.month
        elif year == end_op.year:
            operating_months_map[year] = end_op.month
        else:
            operating_months_map[year] = 12
    
    # Build monthly after-tax CF series from annual "Unlevered CF After Tax" values
    # This matches Excel: use annual after-tax CF values, prorated to monthly
    # For operation months: prorate annual after-tax CF evenly across operating months
    # For construction months: use pre-tax CF (no taxes during construction)
    monthly_cf_after_tax = []
    for i, month_date in enumerate(monthly_dates):
        year = month_date.year
        phase = _phase_for_month(tl, month_date)
        
        if phase == "Operation" and year in annual_after_tax_map:
            # Operation months: use prorated annual after-tax CF from "Unlevered CF After Tax" column
            operating_months = operating_months_map.get(year, 12)
            if operating_months > 0:
                # Distribute annual after-tax CF evenly across operating months in the calendar year
                # This ensures the sum of monthly values equals the annual "Unlevered CF After Tax"
                monthly_after_tax = annual_after_tax_map[year] / operating_months
            else:
                monthly_after_tax = 0.0
            monthly_cf_after_tax.append(monthly_after_tax)
        else:
            # Construction/non-operation months: use pre-tax CF (no tax impact during construction)
            # This ensures all negative cashflows (CAPEX) are included in the IRR calculation
            monthly_cf_after_tax.append(monthly_cf_pre_tax[i] if i < len(monthly_cf_pre_tax) else 0.0)

    # Pre-tax IRR
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")

    # After-tax IRR - calculate from monthly cash flows
    has_pos_after = any(cf > 0 for cf in monthly_cf_after_tax)
    has_neg_after = any(cf < 0 for cf in monthly_cf_after_tax)
    irr_m_after = _irr_bisection(monthly_cf_after_tax) if (has_pos_after and has_neg_after) else float("nan")
    irr_annual_after = (1.0 + irr_m_after) ** 12 - 1.0 if np.isfinite(irr_m_after) else float("nan")
    
    # Also calculate annual IRR directly from annual "Unlevered CF After Tax" values for verification
    # This should match Excel when using annual cash flows
    annual_cf_after_tax = []
    for _, row in a.iterrows():
        # Get the annual "Unlevered CF After Tax (COP)" value
        annual_cf_after_tax.append(float(row.get("Unlevered CF After Tax (COP)", 0.0)))
    
    # Calculate annual IRR directly (this matches Excel annual calculation)
    has_pos_annual = any(cf > 0 for cf in annual_cf_after_tax)
    has_neg_annual = any(cf < 0 for cf in annual_cf_after_tax)
    irr_annual_direct = _irr_bisection(annual_cf_after_tax) if (has_pos_annual and has_neg_annual) else float("nan")
    
    # Use the annual direct calculation (matches Excel)
    if np.isfinite(irr_annual_direct):
        irr_annual_after = irr_annual_direct

    # Payback (after-tax)
    payback_m_after = _payback_months(monthly_dates, monthly_cf_after_tax) if monthly_cf_after_tax else float("nan")
    payback_years_after = payback_m_after / 12.0 if np.isfinite(payback_m_after) else float("nan")

    # Peak funding (from pre-tax cumulative CF)
    cum = mm["Cumulative Unlevered CF (COP)"].astype(float) if "Cumulative Unlevered CF (COP)" in mm.columns else pd.Series([0.0])
    min_cum = float(cum.min()) if len(cum) else 0.0
    peak_funding = float(max(0.0, -min_cum)) if np.isfinite(min_cum) else 0.0

    if currency == "COP":
        _metric_row([
            ("Total Investment (CAPEX)", _fmt_cop(cap_total)),
            ("Unlevered IRR (annualized, pre-tax)", f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"),
            ("Unlevered IRR (annualized, after-tax)", f"{irr_annual_after*100:,.2f}%" if np.isfinite(irr_annual_after) else "—"),
            ("Payback (years, after-tax)", f"{payback_years_after:,.2f}" if np.isfinite(payback_years_after) else "—"),
        ])
        _metric_row([
            ("Peak Funding Need", _fmt_cop(peak_funding)),
            ("", ""),
            ("", ""),
            ("", ""),
        ])
    else:
        fx0 = float(s.macro.fx_cop_per_usd_start)
        _metric_row([
            ("Total Investment (CAPEX)", _fmt_usd(cap_total / fx0)),
            ("Unlevered IRR (annualized, pre-tax)", f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"),
            ("Unlevered IRR (annualized, after-tax)", f"{irr_annual_after*100:,.2f}%" if np.isfinite(irr_annual_after) else "—"),
            ("Payback (years, after-tax)", f"{payback_years_after:,.2f}" if np.isfinite(payback_years_after) else "—"),
        ])
        _metric_row([
            ("Peak Funding Need", _fmt_usd(peak_funding / fx0)),
            ("", ""),
            ("", ""),
            ("", ""),
        ])

    # Annual table (with currency switch)
    annual_view = a.copy()
    money_cols = [
        "Revenue (COP)", "Total OPEX (COP)", "SG&A (COP)", "EBITDA (COP)", "Depreciation (COP)",
        "CAPEX Tax Deduction (COP)", "Loss Carryforward End (COP)", "Taxable Income (COP)", "Taxes Payable (COP)", "CAPEX (COP)", "Unlevered CF Pre-tax (COP)", "Unlevered CF After Tax (COP)",
        "Cumulative Unlevered CF (COP)",
    ]
    for c in money_cols:
        if c in annual_view.columns:
            annual_view[c] = _conv(annual_view[c], annual_view["Year"])

    if currency == "USD":
        ren = {c: c.replace("(COP)", "(USD)") for c in annual_view.columns if "(COP)" in c}
        annual_view = annual_view.rename(columns=ren)

    st.markdown("### Annual summary (calendar years)")
    display_cols = [
        "Year",
        "CAPEX (COP)" if currency == "COP" else "CAPEX (USD)",
        "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "Total OPEX (COP)" if currency == "COP" else "Total OPEX (USD)",
        "SG&A (COP)" if currency == "COP" else "SG&A (USD)",
        "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "CAPEX Tax Deduction (COP)" if currency == "COP" else "CAPEX Tax Deduction (USD)",
        "Loss Carryforward End (COP)" if currency == "COP" else "Loss Carryforward End (USD)",
        "Taxable Income (COP)" if currency == "COP" else "Taxable Income (USD)",
        "Taxes Payable (COP)" if currency == "COP" else "Taxes Payable (USD)",
        "Unlevered CF Pre-tax (COP)" if currency == "COP" else "Unlevered CF Pre-tax (USD)",
        "Unlevered CF After Tax (COP)" if currency == "COP" else "Unlevered CF After Tax (USD)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    disp = annual_view[display_cols].copy()
    disp = _df_format_money(disp, [c for c in disp.columns if c != "Year"], decimals=0)
    st.dataframe(disp, use_container_width=True, hide_index=True)

    y_after = "Unlevered CF After Tax (COP)" if currency == "COP" else "Unlevered CF After Tax (USD)"
    fig = px.bar(annual_view, x="Year", y=y_after)
    fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("### Monthly cash flow (detailed, pre-tax)")
    m_disp = mm.copy()
    m_money = [c for c in m_disp.columns if c not in ["Month", "Year", "Phase"]]
    m_disp = _df_format_money(m_disp, m_money, decimals=0)
    st.dataframe(m_disp, use_container_width=True, hide_index=True)


# -----------------------------
# K) Debt & Covenants
# -----------------------------
with tab_levered:
    st.subheader("Levered Cash Flow (Equity Cash Flow After-Tax)")
    
    st.info("This tab calculates the levered after-tax cash flow available to equity investors. "
            "If debt is disabled or not enabled, the model assumes no debt and returns unlevered cash flow. "
            "This cash flow will be used to calculate Investor Equity IRR in the summary tab.")
    
    currency = st.radio("Display currency", ["COP", "USD"], horizontal=True, index=0, key="currency_levered")
    
    # Calculate levered cash flows
    annual_levered = levered_cashflow_annual(s)
    monthly_levered = levered_cashflow_monthly(s)
    
    # Check debt status
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    total_capex = _total_capex_from_lines(s)
    debt_amt = (float(s.debt.debt_pct_of_capex) / 100.0) * total_capex if debt_enabled else 0.0
    
    if not debt_enabled or debt_amt <= 0:
        st.warning("⚠️ Debt is not enabled or debt amount is zero. Showing unlevered cash flow (no debt impact).")
    
    # FX conversion helper
    tl = build_timeline(s.timeline)
    years = list(annual_levered["Year"].astype(int).tolist())
    fx = fx_series(s.macro, tl["cod"].year, years)
    
    def _conv(series: pd.Series, year_col: pd.Series) -> pd.Series:
        if currency == "COP":
            return series
        return series / year_col.map(lambda y: float(fx.loc[int(y)]) if int(y) in fx.index else float(s.macro.fx_cop_per_usd_start))
    
    # Key metrics
    total_debt_draws = float(annual_levered["Debt Draw (COP)"].sum())
    total_debt_service = float(annual_levered["Debt Service (COP)"].sum())
    total_debt_fees = float(annual_levered["Debt Fees (COP)"].sum())
    
    # Calculate all-in interest rate (same as debt tab)
    all_in_rate = float(s.debt.base_rate_pct) + float(s.debt.margin_pct) if debt_enabled else 0.0
    
    if currency == "COP":
        _metric_row([
            ("Total Debt Draws", _fmt_cop(total_debt_draws)),
            ("Total Debt Service", _fmt_cop(total_debt_service)),
            ("Total Debt Fees", _fmt_cop(total_debt_fees)),
            ("All-in interest rate", f"{all_in_rate:,.2f}%" if debt_enabled else "—"),
        ])
        _metric_row([
            ("Debt Status", "Enabled" if debt_enabled and debt_amt > 0 else "No Debt"),
            ("Debt Amount", _fmt_cop(debt_amt) if debt_enabled else "—"),
            ("", ""),
            ("", ""),
        ])
    else:
        fx0 = float(s.macro.fx_cop_per_usd_start)
        _metric_row([
            ("Total Debt Draws", _fmt_usd(total_debt_draws / fx0)),
            ("Total Debt Service", _fmt_usd(total_debt_service / fx0)),
            ("Total Debt Fees", _fmt_usd(total_debt_fees / fx0)),
            ("All-in interest rate", f"{all_in_rate:,.2f}%" if debt_enabled else "—"),
        ])
        _metric_row([
            ("Debt Status", "Enabled" if debt_enabled and debt_amt > 0 else "No Debt"),
            ("Debt Amount", _fmt_usd(debt_amt / fx0) if debt_enabled else "—"),
            ("", ""),
            ("", ""),
        ])
    
    st.divider()
    
    # Calculate Equity IRR from annual levered CF after-tax (matches table values)
    annual_cf_levered = []
    for _, row in annual_levered.iterrows():
        annual_cf_levered.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
    
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    # Payback - still use monthly for precision
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    # Total equity investment (sum of all negative cash flows from annual table, after fees)
    # This represents total equity contributions over the project life
    # This should match the sum of all negative values in the "Levered CF After-tax" column
    total_equity_investment = 0.0
    for _, row in annual_levered.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)  # Sum of absolute values of negative CFs
    
    _metric_row([
        ("Equity IRR (annualized, after-tax)", f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—"),
        ("Payback (years, after-tax)", f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—"),
        ("Total Equity Investment (After Fees)", _fmt_cop(total_equity_investment) if currency == "COP" else _fmt_usd(total_equity_investment / fx0)),
    ])
    
    st.divider()
    
    # Annual table - Income Statement format
    annual_view = annual_levered.copy()
    money_cols = [
        "Revenue (COP)", "EBITDA (COP)", "Depreciation (COP)", "Interest (COP)",
        "Levered CAPEX Tax Deduction (COP)", "Levered Loss Carryforward End (COP)",
        "Levered Taxable Income (COP)", "Levered Taxes Payable (COP)", "Levered Net Income After Tax (COP)",
        "CAPEX (COP)", "Debt Draw (COP)", "Principal (COP)", "Debt Fees (COP)", "VAT Refund (COP)",
        "Levered CF (After-tax, COP)", "Cumulative Levered CF (COP)",
    ]
    for c in money_cols:
        if c in annual_view.columns:
            annual_view[c] = _conv(annual_view[c], annual_view["Year"])
    
    if currency == "USD":
        ren = {c: c.replace("(COP)", "(USD)") for c in annual_view.columns if "(COP)" in c}
        annual_view = annual_view.rename(columns=ren)
    
    st.markdown("### Income Statement (levered, with debt)")
    display_cols = [
        "Year",
        "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Levered CAPEX Tax Deduction (COP)" if currency == "COP" else "Levered CAPEX Tax Deduction (USD)",
        "Levered Loss Carryforward End (COP)" if currency == "COP" else "Levered Loss Carryforward End (USD)",
        "Levered Taxable Income (COP)" if currency == "COP" else "Levered Taxable Income (USD)",
        "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    disp = annual_view[display_cols].copy()
    disp = _df_format_money(disp, [c for c in disp.columns if c != "Year"], decimals=0)
    st.dataframe(disp, use_container_width=True, hide_index=True)
    
    # Detailed Levered Free Cash Flow Calculation Table
    st.markdown("### Levered Free Cash Flow Calculation (Step-by-Step)")
    st.caption("This table shows the detailed calculation of levered after-tax free cash flow from Revenue to final cash flow.")
    
    fcf_cols = ["Year"]
    fcf_labels = []
    
    # Build the calculation table step by step
    step_cols = {
        "Revenue": "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA": "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation": "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "Interest Expense": "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Debt Service": "Debt Service (COP)" if currency == "COP" else "Debt Service (USD)",
        "CAPEX Tax Deduction": "Levered CAPEX Tax Deduction (COP)" if currency == "COP" else "Levered CAPEX Tax Deduction (USD)",
        "Loss Carryforward": "Levered Loss Carryforward End (COP)" if currency == "COP" else "Levered Loss Carryforward End (USD)",
        "Taxable Income": "Levered Taxable Income (COP)" if currency == "COP" else "Levered Taxable Income (USD)",
        "Taxes Payable": "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Levered CF After-tax": "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)",
    }
    
    for label, col in step_cols.items():
        if col in annual_view.columns:
            fcf_cols.append(col)
            fcf_labels.append(label)
    
    # Create the detailed calculation table
    fcf_df = annual_view[fcf_cols].copy()
    
    # Rename columns for clarity
    rename_dict = {"Year": "Year"}
    for i, col in enumerate(fcf_cols[1:], 1):  # Skip "Year"
        if i <= len(fcf_labels):
            rename_dict[col] = fcf_labels[i-1]
    fcf_df = fcf_df.rename(columns=rename_dict)
    
    # Format the table
    fcf_df = _df_format_money(fcf_df, [c for c in fcf_df.columns if c != "Year"], decimals=0)
    st.dataframe(fcf_df, use_container_width=True, hide_index=True)
    
    # Income Statement Graph
    st.markdown("### Income Statement Overview")
    graph_cols = {
        "Revenue": "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA": "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Interest Expense": "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Taxes Payable": "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Net Income After Tax": "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    }
    
    # Create a melted dataframe for the graph
    graph_df = annual_view[["Year"]].copy()
    for label, col in graph_cols.items():
        if col in annual_view.columns:
            graph_df[label] = annual_view[col]
    
    # Melt for grouped bar chart
    graph_long = graph_df.melt(
        id_vars=["Year"],
        value_vars=[label for label, col in graph_cols.items() if col in annual_view.columns],
        var_name="Metric",
        value_name="Amount"
    )
    
    fig_income = px.bar(
        graph_long,
        x="Year",
        y="Amount",
        color="Metric",
        barmode="group",
        title="Revenue, EBITDA, Interest, Taxes, and Net Income"
    )
    
    # Add period indicators
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    # Add vertical lines for period boundaries
    fig_income.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig_income.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig_income.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig_income.update_layout(height=400, margin=dict(l=10, r=10, t=40, b=10))
    st.plotly_chart(fig_income, use_container_width=True)
    
    # Levered Free Cash Flow (Equity Cash Flow)
    st.markdown("### Levered Free Cash Flow (Equity/After-Tax)")
    
    # Graph of levered free cash flow
    y_fcf = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if y_fcf not in annual_view.columns:
        y_fcf = "Levered CF (After-tax, COP)"
    
    fig_fcf = px.bar(annual_view, x="Year", y=y_fcf)
    
    # Add period indicators
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    # Add vertical lines for period boundaries
    fig_fcf.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig_fcf.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig_fcf.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig_fcf.update_layout(
        height=400,
        margin=dict(l=10, r=10, t=40, b=10),
        title="Equity/Levered After-Tax Free Cash Flow"
    )
    st.plotly_chart(fig_fcf, use_container_width=True)
    
    # Calculation table showing how levered FCF is derived
    st.markdown("#### Levered Free Cash Flow Calculation")
    
    # Build calculation table
    calc_cols = ["Year"]
    calc_labels = []
    
    # Net Income After Tax (starting point)
    net_income_col = "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)"
    if net_income_col in annual_view.columns:
        calc_cols.append(net_income_col)
        calc_labels.append("Net Income After Tax")
    
    # Add back Depreciation (non-cash)
    dep_col = "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)"
    if dep_col in annual_view.columns:
        calc_cols.append(dep_col)
        calc_labels.append("+ Depreciation (add back)")
    
    # Subtract CAPEX
    capex_col = "CAPEX (COP)" if currency == "COP" else "CAPEX (USD)"
    if capex_col in annual_view.columns:
        calc_cols.append(capex_col)
        calc_labels.append("- CAPEX")
    
    # Add Debt Draws
    debt_draw_col = "Debt Draw (COP)" if currency == "COP" else "Debt Draw (USD)"
    if debt_draw_col in annual_view.columns:
        calc_cols.append(debt_draw_col)
        calc_labels.append("+ Debt Draws")
    
    # Subtract Principal payments
    principal_col = "Principal (COP)" if currency == "COP" else "Principal (USD)"
    if principal_col in annual_view.columns:
        calc_cols.append(principal_col)
        calc_labels.append("- Principal Payments")
    
    # Subtract Debt Fees
    debt_fees_col = "Debt Fees (COP)" if currency == "COP" else "Debt Fees (USD)"
    if debt_fees_col in annual_view.columns:
        calc_cols.append(debt_fees_col)
        calc_labels.append("- Debt Fees")
    
    # Add VAT Refund
    vat_col = "VAT Refund (COP)" if currency == "COP" else "VAT Refund (USD)"
    if vat_col in annual_view.columns:
        calc_cols.append(vat_col)
        calc_labels.append("+ VAT Refund")
    
    # Final: Levered Free Cash Flow
    calc_cols.append(y_fcf)
    calc_labels.append("= Levered Free Cash Flow (After-Tax)")
    
    # Create calculation dataframe
    calc_df = annual_view[calc_cols].copy()
    
    # Rename columns for display
    rename_dict = {}
    for i, col in enumerate(calc_cols):
        if col != "Year":
            rename_dict[col] = calc_labels[i - 1] if i > 0 else calc_labels[i]
    calc_df = calc_df.rename(columns=rename_dict)
    
    # Format the table
    calc_df = _df_format_money(calc_df, [c for c in calc_df.columns if c != "Year"], decimals=0)
    st.dataframe(calc_df, use_container_width=True, hide_index=True)
    
    # Cumulative Levered CF Chart (larger)
    y_cum = "Cumulative Levered CF (COP)" if currency == "COP" else "Cumulative Levered CF (USD)"
    # Verify column exists, fallback to COP version if USD version doesn't exist
    if y_cum not in annual_view.columns:
        y_cum = "Cumulative Levered CF (COP)"
    fig2 = px.line(annual_view, x="Year", y=y_cum)
    
    # Add period indicators (vertical lines for key dates)
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    # Add vertical lines for period boundaries
    fig2.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB (Construction Start)", annotation_position="top")
    fig2.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD (Operation Start)", annotation_position="top")
    fig2.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End of Operation", annotation_position="top")
    
    fig2.update_layout(height=500, margin=dict(l=10, r=10, t=40, b=10), title="Cumulative Levered Cash Flow (After-Tax)")
    st.plotly_chart(fig2, use_container_width=True)
    
    # Comparison chart: Unlevered After-Tax vs Levered After-Tax
    st.markdown("### Unlevered After-Tax vs Levered After-Tax Cash Flow Comparison")
    compare_df = annual_view[["Year"]].copy()
    unlevered_col = "Unlevered CF After Tax (COP)" if currency == "COP" else "Unlevered CF After Tax (USD)"
    if unlevered_col not in annual_view.columns:
        unlevered_col = "Unlevered CF After Tax (COP)"
    levered_col = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if levered_col not in annual_view.columns:
        levered_col = "Levered CF (After-tax, COP)"
    compare_df["Unlevered After-Tax"] = annual_view[unlevered_col]
    compare_df["Levered After-Tax"] = annual_view[levered_col]
    compare_long = compare_df.melt(id_vars=["Year"], value_vars=["Unlevered After-Tax", "Levered After-Tax"], var_name="Type", value_name="Cash Flow")
    fig3 = px.bar(compare_long, x="Year", y="Cash Flow", color="Type", barmode="group")
    
    # Add period indicators
    tl = build_timeline(s.timeline)
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    # Add vertical lines for period boundaries
    fig3.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig3.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig3.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig3.update_layout(height=360, margin=dict(l=10, r=10, t=40, b=10))
    st.plotly_chart(fig3, use_container_width=True)
    
    # Monthly table (optional, collapsed)
    with st.expander("Monthly levered cash flow (detailed)"):
        m_disp = monthly_levered.copy()
        m_money = [c for c in m_disp.columns if c not in ["Month", "Year", "Phase"] and "(COP)" in c]
        for c in m_money:
            if c in m_disp.columns:
                m_disp[c] = _conv(m_disp[c], m_disp["Year"])
                if currency == "USD":
                    new_name = c.replace("(COP)", "(USD)")
                    m_disp = m_disp.rename(columns={c: new_name})
        
        m_disp = _df_format_money(m_disp, [c for c in m_disp.columns if c not in ["Month", "Year", "Phase"]], decimals=0)
        st.dataframe(m_disp, use_container_width=True, hide_index=True)


# -----------------------------
# M) Compare
# -----------------------------
with tab_compare:
    st.subheader("Scenario comparison")

    if not compare_scenarios:
        st.warning("Select scenarios to compare in the sidebar.")
    else:
        rows = []
        for nm in compare_scenarios:
            sd = _scenario_from_dict(proj["scenarios"][nm])

            rev = operating_year_table(sd)
            total_cap = _total_capex_from_lines(sd)

            om = opex_monthly_schedule(sd)
            annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
            annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
            total_opex_op = float(annual_opex["Total OPEX (COP)"].sum())

            rows.append(
                {
                    "Scenario": nm,
                    "Total CAPEX (COP)": total_cap,
                    "CAPEX/MWac (COP)": total_cap / float(sd.generation.mwac) if float(sd.generation.mwac) > 0 else np.nan,
                    "Total OPEX (COP)": total_opex_op,
                    "Total Revenue (COP)": float(rev["Revenue (COP)"].sum()),
                }
            )

        summary = pd.DataFrame(rows).sort_values("Scenario")
        disp = summary.copy()
        for col in ["Total CAPEX (COP)", "CAPEX/MWac (COP)", "Total OPEX (COP)", "Total Revenue (COP)"]:
            if col in disp.columns:
                disp[col] = disp[col].apply(lambda v: _fmt_num(float(v), 0) if pd.notnull(v) else "")
        st.dataframe(disp, use_container_width=True, hide_index=True)


# -----------------------------
# PDF Export Function
# -----------------------------
def generate_summary_pdf(project_name: str, scenario_name: str, s: ScenarioInputs, 
                         currency: str, sensitivity_data: dict = None) -> BytesIO:
    """Generate a professional PDF report of the project summary."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                           rightMargin=0.75*inch, leftMargin=0.75*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f4e79'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1f4e79'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title and metadata
    elements.append(Paragraph("Project Financial Summary", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Project info table
    info_data = [
        ['Project:', project_name],
        ['Scenario:', scenario_name],
        ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Currency:', currency]
    ]
    info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Get timeline and FX
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    a = unlevered_base_cashflow_annual(s)
    years_all = list(a["Year"].astype(int).tolist()) if len(a) > 0 else [cod.year]
    fx_series_all = fx_series(s.macro, cod.year, years_all)
    
    def _to_usd_pdf(value_cop: float, year: int = None) -> float:
        if currency == "COP":
            return value_cop
        if year is not None and year in fx_series_all.index:
            return value_cop / float(fx_series_all.loc[year])
        return value_cop / float(s.macro.fx_cop_per_usd_start)
    
    # Helper to format money
    def _fmt_money_pdf(val: float) -> str:
        if currency == "COP":
            return f"COP {val:,.0f}"
        else:
            return f"USD {val:,.0f}"
    
    # 1. Power Generation
    elements.append(Paragraph("Power Generation", heading_style))
    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    base_mwh = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    degr_pct = float(gen.degradation_pct)
    
    gen_data = [
        ['Production:', f"{gen.production_choice} - {base_mwh:,.0f} MWh/yr"],
        ['Degradation:', f"{degr_pct:.2f}%/yr"]
    ]
    
    if s.revenue_mode == "Standard PPA Parameters":
        r = s.revenue1
        ppa_price = float(r.ppa_price_cop_per_kwh)
        indexation = str(r.indexation)
        ppa_term = int(r.ppa_term_years)
        merchant_price = float(r.merchant_price_cop_per_kwh)
        gen_data.append(['PPA Price:', f"{ppa_price:,.4f} COP/kWh ({ppa_term}yr, {indexation})"])
        gen_data.append(['Merchant Price:', f"{merchant_price:,.4f} COP/kWh"])
    else:
        r = s.revenue2
        indexation = str(r.indexation)
        gen_data.append(['Indexation:', indexation])
        gen_data.append(['Price Schedule:', 'Custom (see Power Revenues tab)'])
    
    gen_table = Table(gen_data, colWidths=[2*inch, 4*inch])
    gen_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(gen_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 2. CAPEX
    elements.append(Paragraph("CAPEX", heading_style))
    total_capex = _total_capex_from_lines(s)
    capex_display = _fmt_money_pdf(total_capex if currency == "COP" else _to_usd_pdf(total_capex, cod.year))
    elements.append(Paragraph(f"Total CAPEX: {capex_display}", styles['Normal']))
    elements.append(Spacer(1, 0.1*inch))
    
    # CAPEX breakdown
    capex_df = pd.DataFrame(s.capex.lines or [])
    if "Amount_COP" in capex_df.columns and len(capex_df) > 0:
        capex_df["Amount_COP"] = pd.to_numeric(capex_df["Amount_COP"], errors="coerce").fillna(0.0)
        capex_df = capex_df[capex_df["Amount_COP"] > 0].copy()
        if len(capex_df) > 0:
            capex_table_data = [['Item', 'Amount']]
            for _, row in capex_df.iterrows():
                item = str(row.get("Item", ""))
                amt = float(row.get("Amount_COP", 0.0))
                amt_display = _fmt_money_pdf(amt if currency == "COP" else _to_usd_pdf(amt, cod.year))
                capex_table_data.append([item, amt_display])
            
            capex_table = Table(capex_table_data, colWidths=[4*inch, 2*inch])
            capex_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(capex_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 3. Operating Costs
    elements.append(Paragraph("Operating Costs", heading_style))
    op = operating_year_table(s)
    om = opex_monthly_schedule(s)
    annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
    annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
    
    if "Revenue (COP)" in op.columns and "Total OPEX (COP)" in annual_opex.columns:
        merged = op[["Year", "Revenue (COP)"]].merge(annual_opex[["Year", "Total OPEX (COP)"]], on="Year", how="inner")
        total_revenue = merged["Revenue (COP)"].sum()
        total_opex = merged["Total OPEX (COP)"].sum()
        avg_opex_pct = (total_opex / total_revenue * 100.0) if total_revenue > 0 else 0.0
    else:
        avg_opex_pct = 0.0
    
    annual_sga = sga_annual_by_item(s)
    if "Total SG&A (COP)" in annual_sga.columns and "Revenue (COP)" in op.columns:
        merged_sga = op[["Year", "Revenue (COP)"]].merge(annual_sga[["Year", "Total SG&A (COP)"]], on="Year", how="inner")
        total_revenue_sga = merged_sga["Revenue (COP)"].sum()
        total_sga = merged_sga["Total SG&A (COP)"].sum()
        avg_sga_pct = (total_sga / total_revenue_sga * 100.0) if total_revenue_sga > 0 else 0.0
    else:
        avg_sga_pct = 0.0
    
    opex_data = [
        ['Average OPEX / Revenue:', f"{avg_opex_pct:.2f}%"],
        ['Average SG&A / Revenue:', f"{avg_sga_pct:.2f}%"]
    ]
    opex_table = Table(opex_data, colWidths=[2.5*inch, 3.5*inch])
    opex_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(opex_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 4. Depreciation
    elements.append(Paragraph("Depreciation Schedule", heading_style))
    dep = depreciation_annual_table(s)
    dep_display = dep[["Year", "Depreciation (COP)"]].copy()
    if currency == "USD":
        dep_display["Depreciation (USD)"] = dep_display.apply(
            lambda row: _to_usd_pdf(float(row["Depreciation (COP)"]), int(row["Year"])), axis=1
        )
        dep_display = dep_display.drop(columns=["Depreciation (COP)"])
        dep_col = "Depreciation (USD)"
    else:
        dep_col = "Depreciation (COP)"
    
    dep_table_data = [['Year', 'Depreciation']]
    for _, row in dep_display.iterrows():
        year = int(row["Year"])
        dep_val = float(row[dep_col])
        dep_table_data.append([str(year), _fmt_money_pdf(dep_val)])
    
    dep_table = Table(dep_table_data, colWidths=[1*inch, 2*inch])
    dep_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(dep_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 5. Tax Benefits
    elements.append(Paragraph("Renewable Tax Benefits", heading_style))
    incentives_enabled = bool(getattr(s.incentives, "enable_special_deduction", False))
    tax_data = [['Renewable Tax Benefits Applied:', 'Yes' if incentives_enabled else 'No']]
    if incentives_enabled:
        ded_pct = float(getattr(s.incentives, "special_deduction_pct_of_capex", 0.0))
        tax_data.append(['Special Deduction:', f"{ded_pct:.1f}%"])
    tax_table = Table(tax_data, colWidths=[2.5*inch, 3.5*inch])
    tax_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(tax_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 6. Debt & Equity
    elements.append(Paragraph("Debt & Equity", heading_style))
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    annual_levered_for_equity = levered_cashflow_annual(s)
    total_equity_investment = 0.0
    for _, row in annual_levered_for_equity.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)
    
    debt_data = []
    if debt_enabled:
        total_capex = _total_capex_from_lines(s)
        debt_pct_of_capex = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
        debt_amount = (debt_pct_of_capex / 100.0) * total_capex
        debt_pct = (debt_amount / total_capex * 100.0) if total_capex > 0 else 0.0
        debt_data.append(['Debt:', _fmt_money_pdf(debt_amount if currency == "COP" else _to_usd_pdf(debt_amount, cod.year))])
        debt_data.append(['Debt %:', f"{debt_pct:.1f}%"])
    debt_data.append(['Equity Investment (After Fees):', 
                     _fmt_money_pdf(total_equity_investment if currency == "COP" else _to_usd_pdf(total_equity_investment, cod.year))])
    
    debt_table = Table(debt_data, colWidths=[2.5*inch, 3.5*inch])
    debt_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(debt_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 7. Key Financial Metrics
    elements.append(Paragraph("Key Financial Metrics", heading_style))
    
    # Calculate IRRs
    mm = cashflow_monthly_table(s).copy()
    for col in ["CAPEX (COP)", "Unlevered CF (COP)", "Unlevered CF Pre-tax (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)
    
    if "Unlevered CF Pre-tax (COP)" in mm.columns:
        monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist()
    else:
        monthly_cf_pre_tax = mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")
    
    annual_levered = levered_cashflow_annual(s)
    annual_cf_levered = []
    for _, row in annual_levered.iterrows():
        annual_cf_levered.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
    
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    monthly_levered = levered_cashflow_monthly(s)
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    metrics_data = [
        ['Metric', 'Value'],
        ['Unlevered IRR (Annualized, Pre-tax)', f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"],
        ['Levered Equity IRR (Annualized, After-tax)', f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—"],
        ['Payback (years, after-tax)', f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—"]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(metrics_table)
    elements.append(PageBreak())
    
    # 8. Income Statement
    elements.append(Paragraph("Income Statement (Levered, with Debt)", heading_style))
    annual_view = annual_levered.copy()
    money_cols = [c for c in annual_view.columns if c != "Year" and "(COP)" in c]
    for col in money_cols:
        annual_view[col] = pd.to_numeric(annual_view[col], errors="coerce").fillna(0.0)
        if currency == "USD":
            usd_col = col.replace("(COP)", "(USD)")
            annual_view[usd_col] = annual_view.apply(
                lambda row: _to_usd_pdf(float(row[col]), int(row["Year"])), axis=1
            )
    
    display_cols = [
        "Year",
        "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Depreciation (COP)" if currency == "COP" else "Depreciation (USD)",
        "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Levered CAPEX Tax Deduction (COP)" if currency == "COP" else "Levered CAPEX Tax Deduction (USD)",
        "Levered Taxable Income (COP)" if currency == "COP" else "Levered Taxable Income (USD)",
        "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    income_df = annual_view[display_cols].copy()
    
    # Create income statement table
    income_table_data = [[col.replace(" (COP)", "").replace(" (USD)", "") for col in display_cols]]
    for _, row in income_df.iterrows():
        income_row = []
        for col in display_cols:
            if col == "Year":
                income_row.append(str(int(row[col])))
            else:
                val = float(row[col]) if pd.notnull(row[col]) else 0.0
                income_row.append(_fmt_money_pdf(val))
        income_table_data.append(income_row)
    
    # Adjust column widths based on number of columns
    num_cols = len(display_cols)
    col_width = 5.5*inch / num_cols
    
    income_table = Table(income_table_data, colWidths=[col_width] * num_cols)
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(income_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # 9. Sensitivity Analysis (if available)
    if sensitivity_data:
        elements.append(PageBreak())
        elements.append(Paragraph("Sensitivity Analysis", heading_style))
        elements.append(Paragraph(f"Variable 1: {sensitivity_data.get('var1_name', 'N/A')}", styles['Normal']))
        elements.append(Paragraph(f"Variable 2: {sensitivity_data.get('var2_name', 'N/A')}", styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Add sensitivity table if available
        if 'pivot_table' in sensitivity_data:
            pivot = sensitivity_data['pivot_table']
            sens_table_data = [[''] + [f"V1={c:.1f}" for c in pivot.columns]]
            for idx, row in pivot.iterrows():
                sens_row = [f"V2={idx:.1f}"]
                for val in row:
                    if np.isfinite(val):
                        sens_row.append(f"{val:.2f}%")
                    else:
                        sens_row.append("—")
                sens_table_data.append(sens_row)
            
            # Limit table size for PDF (max 10x10)
            if len(sens_table_data) > 11:
                sens_table_data = sens_table_data[:11]
            if len(sens_table_data[0]) > 11:
                for i in range(len(sens_table_data)):
                    sens_table_data[i] = sens_table_data[i][:11]
            
            sens_table = Table(sens_table_data, colWidths=[0.8*inch] + [0.5*inch] * (len(sens_table_data[0]) - 1))
            sens_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(sens_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


# -----------------------------
# Excel Export Function
# -----------------------------
def generate_excel_report(project_name: str, scenario_name: str, s: ScenarioInputs) -> BytesIO:
    """Generate a comprehensive Excel report with Summary, Inputs, and Outputs sheets."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ==================== SUMMARY SHEET ====================
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    row = 1
    ws_summary['A1'] = "Project Financial Summary"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:B1')
    row += 2
    
    # Project info
    info_data = [
        ['Project:', project_name],
        ['Scenario:', scenario_name],
        ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for label, value in info_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # Key metrics
    ws_summary[f'A{row}'] = "Key Financial Metrics"
    ws_summary[f'A{row}'].font = title_font
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Calculate metrics
    mm = cashflow_monthly_table(s).copy()
    for col in ["CAPEX (COP)", "Unlevered CF (COP)", "Unlevered CF Pre-tax (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)
    
    if "Unlevered CF Pre-tax (COP)" in mm.columns:
        monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist()
    else:
        monthly_cf_pre_tax = mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")
    
    annual_levered = levered_cashflow_annual(s)
    annual_cf_levered = []
    for _, row_data in annual_levered.iterrows():
        annual_cf_levered.append(float(row_data.get("Levered CF (After-tax, COP)", 0.0)))
    
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    monthly_levered = levered_cashflow_monthly(s)
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    total_capex = _total_capex_from_lines(s)
    
    metrics_data = [
        ['Total CAPEX (COP):', f"{total_capex:,.0f}"],
        ['Unlevered IRR (Annualized, Pre-tax):', f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—"],
        ['Levered Equity IRR (Annualized, After-tax):', f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—"],
        ['Payback (years, after-tax):', f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—"]
    ]
    
    for label, value in metrics_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1
    
    # ==================== INPUTS SHEET ====================
    ws_inputs = wb.create_sheet("Inputs", 1)
    current_row = 1
    
    def write_section_header(sheet, row, title):
        sheet[f'A{row}'] = title
        sheet[f'A{row}'].font = title_font
        sheet[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sheet.merge_cells(f'A{row}:D{row}')
        return row + 1
    
    def write_input_row(sheet, row, label, value):
        sheet[f'A{row}'] = label
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = value
        return row + 1
    
    # A) Macroeconomic
    current_row = write_section_header(ws_inputs, current_row, "A) Macroeconomic")
    current_row = write_input_row(ws_inputs, current_row, "FX Rate (COP/USD) - Start", f"{s.macro.fx_cop_per_usd_start:,.2f}")
    current_row = write_input_row(ws_inputs, current_row, "Colombia CPI (%)", f"{s.macro.col_cpi:,.2f}")
    current_row = write_input_row(ws_inputs, current_row, "US CPI (%)", f"{s.macro.us_cpi:,.2f}")
    current_row += 1
    
    # B) Timeline
    current_row = write_section_header(ws_inputs, current_row, "B) Timeline")
    # Calculate RTB and COD dates from timeline
    tl = build_timeline(s.timeline)
    current_row = write_input_row(ws_inputs, current_row, "Start Date", s.timeline.start_date)
    current_row = write_input_row(ws_inputs, current_row, "RTB Date", tl["rtb"].isoformat())
    current_row = write_input_row(ws_inputs, current_row, "COD Date", tl["cod"].isoformat())
    current_row = write_input_row(ws_inputs, current_row, "Development Months", s.timeline.dev_months)
    current_row = write_input_row(ws_inputs, current_row, "CAPEX Months", s.timeline.capex_months)
    current_row = write_input_row(ws_inputs, current_row, "Operation Years", s.timeline.operation_years)
    current_row += 1
    
    # C) Power Generation
    current_row = write_section_header(ws_inputs, current_row, "C) Power Generation")
    current_row = write_input_row(ws_inputs, current_row, "Capacity (MWac)", f"{s.generation.mwac:,.2f}")
    current_row = write_input_row(ws_inputs, current_row, "Capacity (MWp)", f"{s.generation.mwp:,.2f}")
    current_row = write_input_row(ws_inputs, current_row, "P50 (MWh/year)", f"{s.generation.p50_mwh_yr:,.0f}")
    current_row = write_input_row(ws_inputs, current_row, "P75 (MWh/year)", f"{s.generation.p75_mwh_yr:,.0f}")
    current_row = write_input_row(ws_inputs, current_row, "P90 (MWh/year)", f"{s.generation.p90_mwh_yr:,.0f}")
    current_row = write_input_row(ws_inputs, current_row, "Production Choice", s.generation.production_choice)
    current_row = write_input_row(ws_inputs, current_row, "Degradation (%/yr)", f"{s.generation.degradation_pct:.2f}")
    current_row += 1
    
    # D) Power Revenues
    current_row = write_section_header(ws_inputs, current_row, "D) Power Revenues")
    current_row = write_input_row(ws_inputs, current_row, "Revenue Mode", s.revenue_mode)
    if s.revenue_mode == "Standard PPA Parameters":
        current_row = write_input_row(ws_inputs, current_row, "PPA Price (COP/kWh)", f"{s.revenue1.ppa_price_cop_per_kwh:,.2f}")
        current_row = write_input_row(ws_inputs, current_row, "PPA Term (years)", s.revenue1.ppa_term_years)
        current_row = write_input_row(ws_inputs, current_row, "Merchant Price (COP/kWh)", f"{s.revenue1.merchant_price_cop_per_kwh:,.2f}")
        current_row = write_input_row(ws_inputs, current_row, "Indexation", s.revenue1.indexation)
    else:
        current_row = write_input_row(ws_inputs, current_row, "Indexation", s.revenue2.indexation)
        current_row = write_input_row(ws_inputs, current_row, "Custom Price Schedule", "See Power Revenues tab")
    current_row += 1
    
    # E) CAPEX
    current_row = write_section_header(ws_inputs, current_row, "E) CAPEX")
    current_row = write_input_row(ws_inputs, current_row, "Distribution", s.capex.distribution)
    capex_df = pd.DataFrame(s.capex.lines or [])
    if len(capex_df) > 0:
        current_row += 1
        ws_inputs[f'A{current_row}'] = "Item"
        ws_inputs[f'B{current_row}'] = "Amount (COP)"
        ws_inputs[f'C{current_row}'] = "Phase"
        for col in ['A', 'B', 'C']:
            ws_inputs[f'{col}{current_row}'].font = header_font
            ws_inputs[f'{col}{current_row}'].fill = header_fill
            ws_inputs[f'{col}{current_row}'].alignment = center_align
        current_row += 1
        for _, row_data in capex_df.iterrows():
            ws_inputs[f'A{current_row}'] = str(row_data.get("Item", ""))
            ws_inputs[f'B{current_row}'] = float(row_data.get("Amount_COP", 0.0))
            ws_inputs[f'B{current_row}'].number_format = '#,##0'
            ws_inputs[f'C{current_row}'] = str(row_data.get("Phase", ""))
            current_row += 1
    current_row += 1
    
    # F) OPEX
    current_row = write_section_header(ws_inputs, current_row, "F) OPEX")
    current_row = write_input_row(ws_inputs, current_row, "Land Hectares", f"{s.opex.land_hectares:,.2f}")
    current_row += 1
    
    # G) SG&A
    current_row = write_section_header(ws_inputs, current_row, "G) SG&A")
    current_row += 1
    
    # H) Depreciation
    current_row = write_section_header(ws_inputs, current_row, "H) Depreciation")
    current_row = write_input_row(ws_inputs, current_row, "Depreciation Period (years)", s.depreciation.dep_years)
    current_row += 1
    
    # I) Renewable Tax Benefits
    current_row = write_section_header(ws_inputs, current_row, "I) Renewable Tax Benefits")
    current_row = write_input_row(ws_inputs, current_row, "Special Deduction Enabled", "Yes" if getattr(s.incentives, "enable_special_deduction", False) else "No")
    if getattr(s.incentives, "enable_special_deduction", False):
        current_row = write_input_row(ws_inputs, current_row, "Special Deduction (% of CAPEX)", f"{getattr(s.incentives, 'special_deduction_pct_of_capex', 0.0):.1f}")
    current_row += 1
    
    # K) Debt & Covenants
    current_row = write_section_header(ws_inputs, current_row, "K) Debt & Covenants")
    current_row = write_input_row(ws_inputs, current_row, "Debt Enabled", "Yes" if getattr(s.debt, "enabled", False) else "No")
    if getattr(s.debt, "enabled", False):
        current_row = write_input_row(ws_inputs, current_row, "Debt % of CAPEX", f"{getattr(s.debt, 'debt_pct_of_capex', 0.0):.1f}%")
        current_row = write_input_row(ws_inputs, current_row, "Tenor (years)", s.debt.tenor_years)
        current_row = write_input_row(ws_inputs, current_row, "Grace Period (years)", s.debt.grace_years)
        current_row = write_input_row(ws_inputs, current_row, "Amortization Type", s.debt.amortization_type)
        current_row = write_input_row(ws_inputs, current_row, "Base Rate (%)", f"{s.debt.base_rate_pct:.2f}")
        current_row = write_input_row(ws_inputs, current_row, "Margin (%)", f"{s.debt.margin_pct:.2f}")
        current_row = write_input_row(ws_inputs, current_row, "Target DSCR", f"{s.debt.target_dscr:.2f}")
        current_row = write_input_row(ws_inputs, current_row, "Min DSCR Covenant", f"{s.debt.min_dscr_covenant:.2f}")
    
    # Adjust column widths
    ws_inputs.column_dimensions['A'].width = 30
    ws_inputs.column_dimensions['B'].width = 25
    ws_inputs.column_dimensions['C'].width = 20
    ws_inputs.column_dimensions['D'].width = 20
    
    # ==================== OUTPUTS SHEET ====================
    ws_outputs = wb.create_sheet("Outputs", 2)
    
    # Income Statement (P&L)
    row = 1
    ws_outputs['A1'] = "Income Statement (Levered, with Debt)"
    ws_outputs['A1'].font = title_font
    ws_outputs.merge_cells('A1:J1')
    row += 1
    
    annual_view = levered_cashflow_annual(s)
    display_cols = [
        "Year",
        "Revenue (COP)",
        "EBITDA (COP)",
        "Depreciation (COP)",
        "Interest (COP)",
        "Levered CAPEX Tax Deduction (COP)",
        "Levered Taxable Income (COP)",
        "Levered Taxes Payable (COP)",
        "Levered Net Income After Tax (COP)",
    ]
    display_cols = [c for c in display_cols if c in annual_view.columns]
    income_df = annual_view[display_cols].copy()
    
    # Write headers
    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws_outputs.cell(row=row, column=col_idx)
        cell.value = col_name.replace(" (COP)", "")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1
    
    # Write data
    for _, row_data in income_df.iterrows():
        for col_idx, col_name in enumerate(display_cols, start=1):
            cell = ws_outputs.cell(row=row, column=col_idx)
            if col_name == "Year":
                cell.value = int(row_data[col_name])
            else:
                cell.value = float(row_data[col_name]) if pd.notnull(row_data[col_name]) else 0.0
                cell.number_format = '#,##0'
            cell.border = border
            if col_idx > 1:
                cell.alignment = right_align
        row += 1
    
    row += 2
    
    # Levered Cash Flow
    ws_outputs[f'A{row}'] = "Levered Cash Flow (After-Tax)"
    ws_outputs[f'A{row}'].font = title_font
    ws_outputs.merge_cells(f'A{row}:J{row}')
    row += 1
    
    levered_cols = ["Year", "Levered CF (After-tax, COP)"]
    if "Levered CF (After-tax, COP)" in annual_view.columns:
        levered_df = annual_view[levered_cols].copy()
        
        # Headers
        for col_idx, col_name in enumerate(levered_cols, start=1):
            cell = ws_outputs.cell(row=row, column=col_idx)
            cell.value = col_name.replace(" (COP)", "")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        # Data
        for _, row_data in levered_df.iterrows():
            for col_idx, col_name in enumerate(levered_cols, start=1):
                cell = ws_outputs.cell(row=row, column=col_idx)
                if col_name == "Year":
                    cell.value = int(row_data[col_name])
                else:
                    cell.value = float(row_data[col_name]) if pd.notnull(row_data[col_name]) else 0.0
                    cell.number_format = '#,##0'
                cell.border = border
                if col_idx > 1:
                    cell.alignment = right_align
            row += 1
    
    row += 2
    
    # Unlevered Cash Flow
    ws_outputs[f'A{row}'] = "Unlevered Cash Flow"
    ws_outputs[f'A{row}'].font = title_font
    ws_outputs.merge_cells(f'A{row}:J{row}')
    row += 1
    
    unlevered_df = unlevered_base_cashflow_annual(s)
    unlevered_cols = ["Year", "Unlevered CF Pre-tax (COP)", "Unlevered CF (COP)"]
    unlevered_cols = [c for c in unlevered_cols if c in unlevered_df.columns]
    
    if unlevered_cols:
        # Headers
        for col_idx, col_name in enumerate(unlevered_cols, start=1):
            cell = ws_outputs.cell(row=row, column=col_idx)
            cell.value = col_name.replace(" (COP)", "")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        # Data
        for _, row_data in unlevered_df.iterrows():
            for col_idx, col_name in enumerate(unlevered_cols, start=1):
                cell = ws_outputs.cell(row=row, column=col_idx)
                if col_name == "Year":
                    cell.value = int(row_data[col_name])
                else:
                    cell.value = float(row_data[col_name]) if pd.notnull(row_data[col_name]) else 0.0
                    cell.number_format = '#,##0'
                cell.border = border
                if col_idx > 1:
                    cell.alignment = right_align
            row += 1
    
    row += 2
    
    # IRR Calculation Tables
    ws_outputs[f'A{row}'] = "IRR Calculation - Levered Equity (After-Tax)"
    ws_outputs[f'A{row}'].font = title_font
    ws_outputs.merge_cells(f'A{row}:C{row}')
    row += 1
    
    # Headers
    ws_outputs[f'A{row}'] = "Year"
    ws_outputs[f'B{row}'] = "Levered CF (After-tax, COP)"
    ws_outputs[f'C{row}'] = "Cumulative"
    for col in ['A', 'B', 'C']:
        cell = ws_outputs[f'{col}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    row += 1
    
    # Data with cumulative
    cumulative = 0.0
    for _, row_data in annual_levered.iterrows():
        year = int(row_data["Year"])
        cf = float(row_data.get("Levered CF (After-tax, COP)", 0.0))
        cumulative += cf
        
        ws_outputs[f'A{row}'] = year
        ws_outputs[f'B{row}'] = cf
        ws_outputs[f'B{row}'].number_format = '#,##0'
        ws_outputs[f'C{row}'] = cumulative
        ws_outputs[f'C{row}'].number_format = '#,##0'
        
        for col in ['A', 'B', 'C']:
            ws_outputs[f'{col}{row}'].border = border
            if col != 'A':
                ws_outputs[f'{col}{row}'].alignment = right_align
        row += 1
    
    # Add IRR result
    row += 1
    ws_outputs[f'A{row}'] = "Equity IRR (Annualized, After-tax):"
    ws_outputs[f'A{row}'].font = Font(bold=True)
    ws_outputs[f'B{row}'] = f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—"
    ws_outputs[f'B{row}'].font = Font(bold=True)
    
    # Adjust column widths
    for col in range(1, 11):
        ws_outputs.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# -----------------------------
# N) Summary
# -----------------------------
with tab_summary:
    st.subheader("Project Summary")
    
    # Export buttons at the top
    st.markdown("### Export Reports")
    export_col1, export_col2 = st.columns(2)
    
    with export_col1:
        if REPORTLAB_AVAILABLE:
            if st.button("📄 Generate PDF Report", type="primary", use_container_width=True, key="pdf_export_btn"):
                try:
                    sensitivity_data = None  # Could be enhanced to capture from sensitivity tab
                    pdf_buffer = generate_summary_pdf(
                        project_name=project_name,
                        scenario_name=scenario_name,
                        s=s,
                        currency="COP",  # PDF uses base currency
                        sensitivity_data=sensitivity_data
                    )
                    st.download_button(
                        label="⬇️ Download PDF",
                        data=pdf_buffer,
                        file_name=f"{project_name}_{scenario_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key="pdf_download_btn"
                    )
                    st.success("PDF generated successfully!")
                except Exception as e:
                    st.error(f"Error generating PDF: {str(e)}")
        else:
            st.warning("PDF export requires reportlab. Install: `pip install reportlab`")
    
    with export_col2:
        if OPENPYXL_AVAILABLE:
            if st.button("📊 Generate Excel Report", type="primary", use_container_width=True, key="excel_export_btn"):
                try:
                    excel_buffer = generate_excel_report(
                        project_name=project_name,
                        scenario_name=scenario_name,
                        s=s
                    )
                    st.download_button(
                        label="⬇️ Download Excel",
                        data=excel_buffer,
                        file_name=f"{project_name}_{scenario_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="excel_download_btn"
                    )
                    st.success("Excel file generated successfully!")
                except Exception as e:
                    st.error(f"Error generating Excel: {str(e)}")
        else:
            st.warning("Excel export requires openpyxl. Install: `pip install openpyxl`")
    
    st.divider()
    
    currency = st.radio("Display currency", ["COP", "USD"], horizontal=True, index=0, key="currency_summary")
    
    # Get timeline and FX path (for proper USD conversion using indexation)
    tl = build_timeline(s.timeline)
    cod = tl["cod"]
    
    # Get FX series for all years (accounts for CPI indexation differences)
    # This ensures USD conversion reflects actual FX path, not just fixed rate
    a = unlevered_base_cashflow_annual(s)
    years_all = list(a["Year"].astype(int).tolist()) if len(a) > 0 else [cod.year]
    fx_series_all = fx_series(s.macro, cod.year, years_all)
    
    # Helper function to convert COP to USD using year-specific FX
    def _to_usd(value_cop: float, year: int = None) -> float:
        if currency == "COP":
            return value_cop
        if year is not None and year in fx_series_all.index:
            return value_cop / float(fx_series_all.loc[year])
        # Fallback to starting FX if year not found
        return value_cop / float(s.macro.fx_cop_per_usd_start)
    
    # For single values (like total CAPEX), use COD year FX
    fx_cod = float(fx_series_all.loc[cod.year]) if cod.year in fx_series_all.index else float(s.macro.fx_cop_per_usd_start)
    
    # 1. Power Generation (single line)
    gen = s.generation
    p_map = {"P50": gen.p50_mwh_yr, "P75": gen.p75_mwh_yr, "P90": gen.p90_mwh_yr}
    base_mwh = float(p_map.get(gen.production_choice, gen.p50_mwh_yr))
    degr_pct = float(gen.degradation_pct)
    
    gen_col1, gen_col2, gen_col3 = st.columns(3)
    with gen_col1:
        st.metric("Production", f"{str(gen.production_choice)} - {base_mwh:,.0f} MWh/yr")
    with gen_col2:
        st.metric("Degradation", f"{degr_pct:.2f}%")
    
    # 2. PPA Price and Indexation (single line)
    if s.revenue_mode == "Standard PPA Parameters":
        r = s.revenue1
        ppa_price = float(r.ppa_price_cop_per_kwh)
        indexation = str(r.indexation)
        ppa_term = int(r.ppa_term_years)
        merchant_price = float(r.merchant_price_cop_per_kwh)
        
        with gen_col3:
            st.metric("PPA Price", f"{ppa_price:,.4f} COP/kWh ({ppa_term}yr, {indexation})")
        st.metric("Merchant Price", f"{merchant_price:,.4f} COP/kWh")
    else:
        r = s.revenue2
        indexation = str(r.indexation)
        with gen_col3:
            st.metric("Indexation", indexation)
        st.info("Custom price schedule (see Power Revenues tab)")
    
    st.divider()
    
    # 3. CAPEX (compact)
    total_capex = _total_capex_from_lines(s)
    st.markdown("### CAPEX")
    st.metric("Total CAPEX", _fmt_cop(total_capex) if currency == "COP" else _fmt_usd(_to_usd(total_capex, cod.year)))
    
    # CAPEX Breakdown Chart
    st.markdown("#### CAPEX Breakdown by Line Item")
    capex_df = pd.DataFrame(s.capex.lines or [])
    if "Amount_COP" not in capex_df.columns:
        capex_df["Amount_COP"] = 0.0
    capex_pie = capex_df.copy()
    capex_pie["Item"] = capex_pie["Item"].fillna("").astype(str).str.strip()
    capex_pie["Amount_COP"] = pd.to_numeric(capex_pie["Amount_COP"], errors="coerce").fillna(0.0)
    capex_pie = capex_pie[capex_pie["Amount_COP"] > 0].copy()
    
    if capex_pie.empty:
        st.info("Enter CAPEX amounts to see the breakdown chart.")
    else:
        share = capex_pie["Amount_COP"] / capex_pie["Amount_COP"].sum()
        small = share < 0.03
        if small.any() and (~small).any():
            other_amt = float(capex_pie.loc[small, "Amount_COP"].sum())
            capex_pie = capex_pie.loc[~small, ["Item", "Amount_COP"]]
            capex_pie = pd.concat(
                [capex_pie, pd.DataFrame([{"Item": "Other (<3% each)", "Amount_COP": other_amt}])],
                ignore_index=True
            )
        fig_pie = px.pie(capex_pie, names="Item", values="Amount_COP", hole=0.45)
        fig_pie.update_traces(textinfo="percent+label")
        fig_pie.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10), legend_title_text="")
        st.plotly_chart(fig_pie, use_container_width=True, key="summary_capex_pie")
    
    st.divider()
    
    # 4. Average OPEX over Revenue
    st.markdown("### Operating Costs")
    op = operating_year_table(s)
    om = opex_monthly_schedule(s)
    annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
    annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
    
    # Merge with revenue
    if "Revenue (COP)" in op.columns and "Total OPEX (COP)" in annual_opex.columns:
        merged = op[["Year", "Revenue (COP)"]].merge(annual_opex[["Year", "Total OPEX (COP)"]], on="Year", how="inner")
        total_revenue = merged["Revenue (COP)"].sum()
        total_opex = merged["Total OPEX (COP)"].sum()
        avg_opex_pct = (total_opex / total_revenue * 100.0) if total_revenue > 0 else 0.0
    else:
        avg_opex_pct = 0.0
    
    # 5. Average SG&A over Revenue
    annual_sga = sga_annual_by_item(s)
    if "Total SG&A (COP)" in annual_sga.columns and "Revenue (COP)" in op.columns:
        merged_sga = op[["Year", "Revenue (COP)"]].merge(annual_sga[["Year", "Total SG&A (COP)"]], on="Year", how="inner")
        total_revenue_sga = merged_sga["Revenue (COP)"].sum()
        total_sga = merged_sga["Total SG&A (COP)"].sum()
        avg_sga_pct = (total_sga / total_revenue_sga * 100.0) if total_revenue_sga > 0 else 0.0
    else:
        avg_sga_pct = 0.0
    
    opex_col1, opex_col2 = st.columns(2)
    with opex_col1:
        st.metric("Average OPEX / Revenue", f"{avg_opex_pct:.2f}%")
    with opex_col2:
        st.metric("Average SG&A / Revenue", f"{avg_sga_pct:.2f}%")
    
    st.divider()
    
    # 6. Depreciation Schedule
    st.markdown("### Depreciation Schedule")
    dep = depreciation_annual_table(s)
    dep_display = dep[["Year", "Depreciation (COP)"]].copy()
    if currency == "USD":
        # Convert using year-specific FX rates
        dep_display["Depreciation (USD)"] = dep_display.apply(
            lambda row: _to_usd(float(row["Depreciation (COP)"]), int(row["Year"])), axis=1
        )
        dep_display = dep_display.drop(columns=["Depreciation (COP)"])
    dep_display = _df_format_money(dep_display, [c for c in dep_display.columns if c != "Year"], decimals=0)
    st.dataframe(dep_display, use_container_width=True, hide_index=True)
    
    st.divider()
    
    # 7. Renewable Tax Benefits
    st.markdown("### Renewable Tax Benefits")
    incentives_enabled = bool(getattr(s.incentives, "enable_special_deduction", False))
    st.metric("Renewable Tax Benefits Applied", "Yes" if incentives_enabled else "No")
    
    if incentives_enabled:
        ded_pct = float(getattr(s.incentives, "special_deduction_pct_of_capex", 0.0))
        st.metric("Special Deduction", f"{ded_pct:.1f}%")
    
    st.divider()
    
    # 8. Debt Information
    # 8. Debt & Equity (compact)
    st.markdown("### Debt & Equity")
    debt_enabled = bool(getattr(s.debt, "enabled", False))
    
    # Calculate Total Equity Investment (After Fees) - sum of all negative levered CF
    annual_levered_for_equity = levered_cashflow_annual(s)
    total_equity_investment = 0.0
    for _, row in annual_levered_for_equity.iterrows():
        levered_cf = float(row.get("Levered CF (After-tax, COP)", 0.0))
        if levered_cf < 0:
            total_equity_investment += abs(levered_cf)  # Sum of absolute values of negative CFs
    
    if debt_enabled:
        total_capex = _total_capex_from_lines(s)
        debt_pct_of_capex = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
        debt_amount = (debt_pct_of_capex / 100.0) * total_capex
        debt_pct = (debt_amount / total_capex * 100.0) if total_capex > 0 else 0.0
        
        debt_col1, debt_col2, debt_col3 = st.columns(3)
        with debt_col1:
            st.metric("Debt", _fmt_cop(debt_amount) if currency == "COP" else _fmt_usd(_to_usd(debt_amount, cod.year)))
        with debt_col2:
            st.metric("Debt %", f"{debt_pct:.1f}%")
        with debt_col3:
            st.metric("Equity Investment (After Fees)", _fmt_cop(total_equity_investment) if currency == "COP" else _fmt_usd(_to_usd(total_equity_investment, cod.year)))
    else:
        st.metric("Equity Investment (After Fees)", _fmt_cop(total_equity_investment) if currency == "COP" else _fmt_usd(_to_usd(total_equity_investment, cod.year)))
    
    st.divider()
    
    # 9. Key Metrics
    st.markdown("### Key Financial Metrics")
    
    # Unlevered IRR Pre-tax
    a = unlevered_base_cashflow_annual(s)
    mm = cashflow_monthly_table(s).copy()
    for col in ["CAPEX (COP)", "Unlevered CF (COP)", "Unlevered CF Pre-tax (COP)"]:
        if col in mm.columns:
            mm[col] = pd.to_numeric(mm[col], errors="coerce").fillna(0.0)
    
    if "Unlevered CF Pre-tax (COP)" in mm.columns:
        monthly_cf_pre_tax = mm["Unlevered CF Pre-tax (COP)"].astype(float).tolist()
    else:
        monthly_cf_pre_tax = mm["Unlevered CF (COP)"].astype(float).tolist() if "Unlevered CF (COP)" in mm.columns else []
    
    has_pos_pre = any(cf > 0 for cf in monthly_cf_pre_tax)
    has_neg_pre = any(cf < 0 for cf in monthly_cf_pre_tax)
    irr_m_pre = _irr_bisection(monthly_cf_pre_tax) if (has_pos_pre and has_neg_pre) else float("nan")
    irr_annual_pre = (1.0 + irr_m_pre) ** 12 - 1.0 if np.isfinite(irr_m_pre) else float("nan")
    
    # Levered Equity IRR After-tax
    annual_levered = levered_cashflow_annual(s)
    annual_cf_levered = []
    for _, row in annual_levered.iterrows():
        annual_cf_levered.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
    
    has_pos_annual = any(cf > 0 for cf in annual_cf_levered)
    has_neg_annual = any(cf < 0 for cf in annual_cf_levered)
    irr_annual_equiv = _irr_bisection(annual_cf_levered) if (has_pos_annual and has_neg_annual) else float("nan")
    
    # Payback
    monthly_levered = levered_cashflow_monthly(s)
    monthly_cf = monthly_levered["Levered CF (After-tax, COP)"].astype(float).tolist() if "Levered CF (After-tax, COP)" in monthly_levered.columns else []
    monthly_dates = monthly_levered["Month"].tolist() if "Month" in monthly_levered.columns else []
    payback_m = _payback_months(monthly_dates, monthly_cf) if monthly_cf else float("nan")
    payback_years = payback_m / 12.0 if np.isfinite(payback_m) else float("nan")
    
    metrics_col1, metrics_col2, metrics_col3 = st.columns(3)
    with metrics_col1:
        st.metric("Unlevered IRR (Annualized, Pre-tax)", f"{irr_annual_pre*100:,.2f}%" if np.isfinite(irr_annual_pre) else "—")
    with metrics_col2:
        st.metric("Levered Equity IRR (Annualized, After-tax)", f"{irr_annual_equiv*100:,.2f}%" if np.isfinite(irr_annual_equiv) else "—")
    with metrics_col3:
        st.metric("Payback (years, after-tax)", f"{payback_years:,.2f}" if np.isfinite(payback_years) else "—")
    
    # Table showing annual levered CF for IRR verification
    st.markdown("#### Equity IRR Calculation (for verification)")
    irr_table = annual_levered[["Year", "Levered CF (After-tax, COP)"]].copy()
    if currency == "USD":
        irr_table["Levered CF (After-tax, USD)"] = irr_table.apply(
            lambda row: _to_usd(float(row["Levered CF (After-tax, COP)"]), int(row["Year"])), axis=1
        )
        irr_table = irr_table.drop(columns=["Levered CF (After-tax, COP)"])
    
    # Format the table
    money_cols = [c for c in irr_table.columns if c != "Year"]
    irr_table = _df_format_money(irr_table, money_cols, decimals=0)
    st.dataframe(irr_table, use_container_width=True, hide_index=True)
    
    # Show summary row
    total_negative = sum([cf for cf in annual_cf_levered if cf < 0])
    total_positive = sum([cf for cf in annual_cf_levered if cf > 0])
    summary_row = pd.DataFrame([{
        "Year": "Total",
        "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)": 
            _fmt_cop(sum(annual_cf_levered)) if currency == "COP" else _fmt_usd(_to_usd(sum(annual_cf_levered), cod.year))
    }])
    st.dataframe(summary_row, use_container_width=True, hide_index=True)
    
    st.divider()
    
    # 10. Charts
    st.markdown("### Levered Free Cash Flow (Equity After-Tax)")
    
    # Prepare annual view for charts
    annual_view = annual_levered.copy()
    money_cols = [c for c in annual_view.columns if c != "Year" and "(COP)" in c]
    for col in money_cols:
        annual_view[col] = pd.to_numeric(annual_view[col], errors="coerce").fillna(0.0)
        if currency == "USD":
            # Convert using year-specific FX rates (accounts for CPI indexation)
            usd_col = col.replace("(COP)", "(USD)")
            annual_view[usd_col] = annual_view.apply(
                lambda row: _to_usd(float(row[col]), int(row["Year"])), axis=1
            )
    
    y_fcf = "Levered CF (After-tax, COP)" if currency == "COP" else "Levered CF (After-tax, USD)"
    if y_fcf not in annual_view.columns:
        y_fcf = "Levered CF (After-tax, COP)"
    
    fig_fcf = px.bar(annual_view, x="Year", y=y_fcf)
    
    # Add period indicators
    rtb_year = tl["rtb"].year
    cod_year = tl["cod"].year
    end_op_year = tl["end_op"].year
    
    fig_fcf.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig_fcf.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig_fcf.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig_fcf.update_layout(
        height=400,
        margin=dict(l=10, r=10, t=40, b=10),
        title="Equity/Levered After-Tax Free Cash Flow"
    )
    st.plotly_chart(fig_fcf, use_container_width=True, key="summary_fcf_chart")
    
    st.markdown("### Income Statement Overview")
    
    # Income Statement Graph
    graph_cols = {
        "Revenue": "Revenue (COP)" if currency == "COP" else "Revenue (USD)",
        "EBITDA": "EBITDA (COP)" if currency == "COP" else "EBITDA (USD)",
        "Interest Expense": "Interest (COP)" if currency == "COP" else "Interest (USD)",
        "Taxes Payable": "Levered Taxes Payable (COP)" if currency == "COP" else "Levered Taxes Payable (USD)",
        "Net Income After Tax": "Levered Net Income After Tax (COP)" if currency == "COP" else "Levered Net Income After Tax (USD)",
    }
    
    # Create a melted dataframe for the graph
    graph_df = annual_view[["Year"]].copy()
    for label, col in graph_cols.items():
        if col in annual_view.columns:
            graph_df[label] = annual_view[col]
    
    # Melt for grouped bar chart
    graph_long = graph_df.melt(
        id_vars=["Year"],
        value_vars=[label for label, col in graph_cols.items() if col in annual_view.columns],
        var_name="Metric",
        value_name="Amount"
    )
    
    fig_income = px.bar(
        graph_long,
        x="Year",
        y="Amount",
        color="Metric",
        barmode="group",
        title="Revenue, EBITDA, Interest, Taxes, and Net Income"
    )
    
    # Add period indicators
    fig_income.add_vline(x=rtb_year, line_dash="dash", line_color="orange", annotation_text="RTB", annotation_position="top")
    fig_income.add_vline(x=cod_year, line_dash="dash", line_color="green", annotation_text="COD", annotation_position="top")
    fig_income.add_vline(x=end_op_year, line_dash="dash", line_color="red", annotation_text="End Op", annotation_position="top")
    
    fig_income.update_layout(height=400, margin=dict(l=10, r=10, t=40, b=10))
    st.plotly_chart(fig_income, use_container_width=True, key="summary_income_chart")


# -----------------------------
# O) Sensitivity Analysis
# -----------------------------
with tab_sensitivity:
    st.subheader("Sensitivity Analysis - Equity IRR (After-Tax)")
    
    st.info("Select two variables to analyze how Equity IRR (After-Tax) changes with different values.")
    
    # Variable selection
    col1, col2 = st.columns(2)
    
    with col1:
        var1_name = st.selectbox(
            "Variable 1 (X-axis)",
            ["CAPEX", "PPA Price", "Production (MWh/yr)", "OPEX", "Debt %"],
            index=0,
            key="sens_var1"
        )
        
        var1_base = None
        var1_min = None
        var1_max = None
        var1_steps = None
        
        if var1_name == "CAPEX":
            total_capex = _total_capex_from_lines(s)
            var1_base = total_capex
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "PPA Price":
            if s.revenue_mode == "Standard PPA Parameters":
                var1_base = float(s.revenue1.ppa_price_cop_per_kwh)
            else:
                st.warning("PPA Price sensitivity requires Standard PPA Parameters mode")
                var1_base = 0.0
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "Production (MWh/yr)":
            p_map = {"P50": s.generation.p50_mwh_yr, "P75": s.generation.p75_mwh_yr, "P90": s.generation.p90_mwh_yr}
            var1_base = float(p_map.get(s.generation.production_choice, s.generation.p50_mwh_yr))
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "OPEX":
            # Use average OPEX per year as base
            om = opex_monthly_schedule(s)
            annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
            annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
            var1_base = float(annual_opex["Total OPEX (COP)"].mean()) if len(annual_opex) > 0 else 0.0
            var1_min = st.number_input("Variable 1 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var1_min") / 100.0
            var1_max = st.number_input("Variable 1 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var1_max") / 100.0
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
        elif var1_name == "Debt %":
            var1_base = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
            var1_min = st.number_input("Variable 1 Min (%)", value=0.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var1_min")
            var1_max = st.number_input("Variable 1 Max (%)", value=80.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var1_max")
            var1_steps = st.number_input("Variable 1 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var1_steps")
    
    with col2:
        var2_name = st.selectbox(
            "Variable 2 (Y-axis)",
            ["CAPEX", "PPA Price", "Production (MWh/yr)", "OPEX", "Debt %"],
            index=1,
            key="sens_var2"
        )
        
        var2_base = None
        var2_min = None
        var2_max = None
        var2_steps = None
        
        if var2_name == "CAPEX":
            total_capex = _total_capex_from_lines(s)
            var2_base = total_capex
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "PPA Price":
            if s.revenue_mode == "Standard PPA Parameters":
                var2_base = float(s.revenue1.ppa_price_cop_per_kwh)
            else:
                st.warning("PPA Price sensitivity requires Standard PPA Parameters mode")
                var2_base = 0.0
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "Production (MWh/yr)":
            p_map = {"P50": s.generation.p50_mwh_yr, "P75": s.generation.p75_mwh_yr, "P90": s.generation.p90_mwh_yr}
            var2_base = float(p_map.get(s.generation.production_choice, s.generation.p50_mwh_yr))
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "OPEX":
            om = opex_monthly_schedule(s)
            annual_opex = om.groupby("Year", as_index=False)[["OPEX subtotal", "GMF"]].sum()
            annual_opex["Total OPEX (COP)"] = annual_opex["OPEX subtotal"] + annual_opex["GMF"]
            var2_base = float(annual_opex["Total OPEX (COP)"].mean()) if len(annual_opex) > 0 else 0.0
            var2_min = st.number_input("Variable 2 Min (% of base)", value=80.0, min_value=50.0, max_value=100.0, step=5.0, key="sens_var2_min") / 100.0
            var2_max = st.number_input("Variable 2 Max (% of base)", value=120.0, min_value=100.0, max_value=150.0, step=5.0, key="sens_var2_max") / 100.0
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
        elif var2_name == "Debt %":
            var2_base = float(getattr(s.debt, "debt_pct_of_capex", 0.0))
            var2_min = st.number_input("Variable 2 Min (%)", value=0.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var2_min")
            var2_max = st.number_input("Variable 2 Max (%)", value=80.0, min_value=0.0, max_value=90.0, step=10.0, key="sens_var2_max")
            var2_steps = st.number_input("Variable 2 Steps", value=5, min_value=3, max_value=10, step=1, key="sens_var2_steps")
    
    # Ensure variables are different
    if var1_name == var2_name:
        st.error("Please select two different variables for sensitivity analysis.")
    elif var1_base is None or var2_base is None or var1_base == 0 or var2_base == 0:
        st.warning("Cannot perform sensitivity analysis. Please check variable inputs.")
    else:
        # Generate value ranges
        if var1_name == "Debt %":
            var1_values = np.linspace(var1_min, var1_max, int(var1_steps))
        else:
            var1_values = np.linspace(var1_base * var1_min, var1_base * var1_max, int(var1_steps))
        
        if var2_name == "Debt %":
            var2_values = np.linspace(var2_min, var2_max, int(var2_steps))
        else:
            var2_values = np.linspace(var2_base * var2_min, var2_base * var2_max, int(var2_steps))
        
        # Run sensitivity analysis
        st.markdown("### Running Sensitivity Analysis...")
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        results = []
        total_runs = len(var1_values) * len(var2_values)
        run_count = 0
        
        for v1 in var1_values:
            for v2 in var2_values:
                # Create a copy of the scenario
                s_mod = _scenario_from_dict(_scenario_to_dict(s))
                
                # Modify variable 1
                if var1_name == "CAPEX":
                    # Scale all CAPEX line items proportionally
                    scale_factor = v1 / var1_base
                    for item in s_mod.capex.lines:
                        if "Amount_COP" in item:
                            item["Amount_COP"] = float(item.get("Amount_COP", 0.0)) * scale_factor
                elif var1_name == "PPA Price":
                    if s_mod.revenue_mode == "Standard PPA Parameters":
                        s_mod.revenue1.ppa_price_cop_per_kwh = v1
                elif var1_name == "Production (MWh/yr)":
                    # Scale all production values
                    scale_factor = v1 / var1_base
                    s_mod.generation.p50_mwh_yr = float(s_mod.generation.p50_mwh_yr) * scale_factor
                    s_mod.generation.p75_mwh_yr = float(s_mod.generation.p75_mwh_yr) * scale_factor
                    s_mod.generation.p90_mwh_yr = float(s_mod.generation.p90_mwh_yr) * scale_factor
                elif var1_name == "OPEX":
                    # Scale OPEX components
                    scale_factor = v1 / var1_base
                    s_mod.opex.fixed_om_cop_per_mwac_year = float(s_mod.opex.fixed_om_cop_per_mwac_year) * scale_factor
                    s_mod.opex.variable_om_cop_per_mwh = float(s_mod.opex.variable_om_cop_per_mwh) * scale_factor
                    s_mod.opex.insurance_cop_per_mwac_year = float(s_mod.opex.insurance_cop_per_mwac_year) * scale_factor
                    s_mod.opex.grid_fees_cop_per_mwh = float(s_mod.opex.grid_fees_cop_per_mwh) * scale_factor
                elif var1_name == "Debt %":
                    s_mod.debt.debt_pct_of_capex = v1
                
                # Modify variable 2
                if var2_name == "CAPEX":
                    scale_factor = v2 / var2_base
                    for item in s_mod.capex.lines:
                        if "Amount_COP" in item:
                            item["Amount_COP"] = float(item.get("Amount_COP", 0.0)) * scale_factor
                elif var2_name == "PPA Price":
                    if s_mod.revenue_mode == "Standard PPA Parameters":
                        s_mod.revenue1.ppa_price_cop_per_kwh = v2
                elif var2_name == "Production (MWh/yr)":
                    scale_factor = v2 / var2_base
                    s_mod.generation.p50_mwh_yr = float(s_mod.generation.p50_mwh_yr) * scale_factor
                    s_mod.generation.p75_mwh_yr = float(s_mod.generation.p75_mwh_yr) * scale_factor
                    s_mod.generation.p90_mwh_yr = float(s_mod.generation.p90_mwh_yr) * scale_factor
                elif var2_name == "OPEX":
                    scale_factor = v2 / var2_base
                    s_mod.opex.fixed_om_cop_per_mwac_year = float(s_mod.opex.fixed_om_cop_per_mwac_year) * scale_factor
                    s_mod.opex.variable_om_cop_per_mwh = float(s_mod.opex.variable_om_cop_per_mwh) * scale_factor
                    s_mod.opex.insurance_cop_per_mwac_year = float(s_mod.opex.insurance_cop_per_mwac_year) * scale_factor
                    s_mod.opex.grid_fees_cop_per_mwh = float(s_mod.opex.grid_fees_cop_per_mwh) * scale_factor
                elif var2_name == "Debt %":
                    s_mod.debt.debt_pct_of_capex = v2
                
                # Calculate Equity IRR for modified scenario
                try:
                    annual_levered_mod = levered_cashflow_annual(s_mod)
                    annual_cf_levered_mod = []
                    for _, row in annual_levered_mod.iterrows():
                        annual_cf_levered_mod.append(float(row.get("Levered CF (After-tax, COP)", 0.0)))
                    
                    has_pos = any(cf > 0 for cf in annual_cf_levered_mod)
                    has_neg = any(cf < 0 for cf in annual_cf_levered_mod)
                    irr = _irr_bisection(annual_cf_levered_mod) if (has_pos and has_neg) else float("nan")
                    irr_pct = irr * 100.0 if np.isfinite(irr) else float("nan")
                except Exception as e:
                    irr_pct = float("nan")
                
                results.append({
                    "Var1": v1,
                    "Var2": v2,
                    "IRR": irr_pct
                })
                
                run_count += 1
                progress_bar.progress(run_count / total_runs)
                status_text.text(f"Completed {run_count}/{total_runs} scenarios...")
        
        progress_bar.empty()
        status_text.empty()
        
        # Create sensitivity table
        st.markdown("### Sensitivity Table - Equity IRR (After-Tax) %")
        
        # Convert to pivot table format
        df_results = pd.DataFrame(results)
        pivot_table = df_results.pivot_table(
            values="IRR",
            index="Var2",
            columns="Var1",
            aggfunc="first"
        )
        
        # Format for display
        pivot_display = pivot_table.copy()
        pivot_display = pivot_display.round(2)
        
        # Format row and column labels
        if var1_name == "Debt %":
            pivot_display.columns = [f"{c:.1f}%" for c in pivot_display.columns]
        elif var1_name == "CAPEX":
            pivot_display.columns = [_fmt_cop(c) for c in pivot_display.columns]
        elif var1_name == "PPA Price":
            pivot_display.columns = [f"{c:.4f}" for c in pivot_display.columns]
        elif var1_name == "Production (MWh/yr)":
            pivot_display.columns = [f"{c:,.0f}" for c in pivot_display.columns]
        else:
            pivot_display.columns = [_fmt_cop(c) for c in pivot_display.columns]
        
        if var2_name == "Debt %":
            pivot_display.index = [f"{i:.1f}%" for i in pivot_display.index]
        elif var2_name == "CAPEX":
            pivot_display.index = [_fmt_cop(i) for i in pivot_display.index]
        elif var2_name == "PPA Price":
            pivot_display.index = [f"{i:.4f}" for i in pivot_display.index]
        elif var2_name == "Production (MWh/yr)":
            pivot_display.index = [f"{i:,.0f}" for i in pivot_display.index]
        else:
            pivot_display.index = [_fmt_cop(i) for i in pivot_display.index]
        
        # Add row and column headers
        pivot_display.index.name = var2_name
        pivot_display.columns.name = var1_name
        
        st.dataframe(pivot_display, use_container_width=True)
        
        # Create heatmap with text annotations
        st.markdown("### Sensitivity Heatmap")
        
        # Prepare axis labels - use percentages with actual values below
        def format_x_label(val, var_name, base_val):
            if var_name == "Debt %":
                # Debt % is already a percentage value
                return f"{val:.1f}%"
            else:
                pct = (val / base_val) * 100 if base_val > 0 else 0
                if var_name == "CAPEX":
                    # Format as percentage on top, actual value below
                    return f"{(pct):.0f}%<br>{val/1e9:.2f}B COP"
                elif var_name == "PPA Price":
                    return f"{(pct):.0f}%<br>{val:.0f} COP/kWh"
                elif var_name == "Production (MWh/yr)":
                    return f"{(pct):.0f}%<br>{val:,.0f} MWh"
                elif var_name == "OPEX":
                    return f"{(pct):.0f}%<br>{val/1e6:.1f}M COP"
                else:
                    return f"{(pct):.0f}%"
        
        def format_y_label(val, var_name, base_val):
            if var_name == "Debt %":
                # Debt % is already a percentage value
                return f"{val:.1f}%"
            else:
                pct = (val / base_val) * 100 if base_val > 0 else 0
                if var_name == "CAPEX":
                    return f"{(pct):.0f}%<br>{val/1e9:.2f}B COP"
                elif var_name == "PPA Price":
                    return f"{(pct):.0f}%<br>{val:.0f} COP/kWh"
                elif var_name == "Production (MWh/yr)":
                    return f"{(pct):.0f}%<br>{val:,.0f} MWh"
                elif var_name == "OPEX":
                    return f"{(pct):.0f}%<br>{val/1e6:.1f}M COP"
                else:
                    return f"{(pct):.0f}%"
        
        x_labels = [format_x_label(c, var1_name, var1_base) for c in pivot_table.columns]
        y_labels = [format_y_label(r, var2_name, var2_base) for r in pivot_table.index]
        
        # Create text matrix for annotations (IRR values)
        text_matrix = []
        for row in pivot_table.values:
            text_row = []
            for val in row:
                if np.isfinite(val):
                    text_row.append(f"{val:.2f}%")
                else:
                    text_row.append("—")
            text_matrix.append(text_row)
        
        # Create heatmap using graph_objects (supports text parameter)
        fig = go.Figure(data=go.Heatmap(
            z=pivot_table.values,
            x=x_labels,
            y=y_labels,
            text=text_matrix,
            texttemplate="%{text}",
            textfont={"size": 18, "color": "black", "family": "Arial Black"},  # Much larger font
            colorscale="RdYlGn",
            colorbar=dict(title="Equity IRR (%)"),
            hovertemplate="%{y}<br>%{x}<br>IRR: %{z:.2f}%<extra></extra>"
        ))
        
        fig.update_layout(
            title=f"Equity IRR (After-Tax) Sensitivity: {var1_name} vs {var2_name}",
            xaxis_title=var1_name,
            yaxis_title=var2_name,
            height=500,
            margin=dict(l=10, r=10, t=50, b=10)
        )
        st.plotly_chart(fig, use_container_width=True, key="sensitivity_heatmap")


# Persist scenario on each run
proj["scenarios"][scenario_name] = _scenario_to_dict(s)
_save_db(db)
